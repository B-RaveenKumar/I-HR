"""
Hierarchical Timetable API Routes
Provides REST endpoints for hierarchical timetable management with conflict detection
"""

from flask import Blueprint, request, jsonify, session, make_response
from functools import wraps
from hierarchical_timetable import HierarchicalTimetableManager
from database import get_db
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

# Create Blueprint
hierarchical_bp = Blueprint('hierarchical_timetable', __name__, url_prefix='/api/hierarchical-timetable')

# ==================== AUTHENTICATION DECORATORS ====================

def check_admin_auth(f):
    """Ensure only school admins can access route"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session['user_type'] != 'admin':
            return jsonify({'success': False, 'error': 'Unauthorized access. Admin required.'}), 401
        return f(*args, **kwargs)
    return decorated_function

def check_staff_auth(f):
    """Ensure only staff can access route"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session['user_type'] != 'staff':
            return jsonify({'success': False, 'error': 'Unauthorized access. Staff login required.'}), 401
        return f(*args, **kwargs)
    return decorated_function

def check_auth_either(f):
    """Ensure user is logged in (admin or staff)"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Unauthorized access. Login required.'}), 401
        return f(*args, **kwargs)
    return decorated_function


def _ensure_section_mentor_table(db):
    db.execute('''
        CREATE TABLE IF NOT EXISTS timetable_section_mentors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            school_id INTEGER NOT NULL,
            section_id INTEGER NOT NULL,
            staff_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(school_id, section_id)
        )
    ''')

# ==================== ORGANIZATION CONFIGURATION ====================

@hierarchical_bp.route('/organization/set-type', methods=['POST'])
@check_admin_auth
def set_organization_type():
    """Set organization type (School or College)"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        org_type = data.get('organization_type', 'school')
        
        result = HierarchicalTimetableManager.set_organization_type(school_id, org_type)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error setting organization type: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/organization/config', methods=['GET'])
@check_auth_either
def get_organization_config():
    """Get organization configuration"""
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.get_organization_config(school_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching organization config: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== ACADEMIC LEVELS ====================

@hierarchical_bp.route('/levels', methods=['GET'])
@check_auth_either
def get_academic_levels():
    """Get all academic levels for school"""
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.get_academic_levels(school_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching academic levels: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/levels/update/<int:level_id>', methods=['POST'])
@check_admin_auth
def update_academic_level(level_id):
    """Update name of an academic level"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        level_name = data.get('level_name')
        description = data.get('description')
        
        if not level_name:
            return jsonify({'success': False, 'error': 'Level name is required'}), 400
            
        result = HierarchicalTimetableManager.update_academic_level(school_id, level_id, level_name, description)
        return jsonify(result), 200 if result['success'] else 400
        
    except Exception as e:
        logger.error(f"Error updating academic level: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/levels/create', methods=['POST'])
@check_admin_auth
def create_academic_level():
    """Create a new academic level"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}

        level_name = data.get('level_name')
        level_number = data.get('level_number')
        description = data.get('description')

        if not level_name:
            return jsonify({'success': False, 'error': 'Level name is required'}), 400

        result = HierarchicalTimetableManager.create_academic_level(
            school_id,
            level_name,
            level_number,
            description
        )
        return jsonify(result), 200 if result['success'] else 400

    except Exception as e:
        logger.error(f"Error creating academic level: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/levels/<int:level_id>', methods=['DELETE'])
@check_admin_auth
def delete_academic_level(level_id):
    """Delete an academic level"""
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.delete_academic_level(school_id, level_id)
        return jsonify(result), 200 if result['success'] else 400

    except Exception as e:
        logger.error(f"Error deleting academic level: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== SECTIONS MANAGEMENT ====================

@hierarchical_bp.route('/sections/create', methods=['POST'])
@check_admin_auth
def create_section():
    """Create a new section for an academic level"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        
        level_id = data.get('level_id')
        section_name = data.get('section_name')
        capacity = data.get('capacity', 60)
        
        if not level_id or not section_name:
            return jsonify({'success': False, 'error': 'Missing required fields: level_id, section_name'}), 400
        
        result = HierarchicalTimetableManager.create_section(school_id, level_id, section_name, capacity)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error creating section: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/sections/level/<int:level_id>', methods=['GET'])
@check_auth_either
def get_sections_for_level(level_id):
    """Get all sections for a specific academic level"""
    try:
        result = HierarchicalTimetableManager.get_sections_for_level(level_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching sections: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/sections/all', methods=['GET'])
@check_auth_either
def get_all_sections():
    """Get all sections for school"""
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.get_all_sections(school_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching all sections: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/sections/<int:section_id>', methods=['DELETE'])
@check_admin_auth
def delete_section(section_id):
    """Delete a section"""
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.delete_section(school_id, section_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error deleting section: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/sections/<int:section_id>', methods=['PUT'])
@check_admin_auth
def update_section(section_id):
    """Update a section"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        
        section_name = data.get('section_name')
        capacity = data.get('capacity')
        
        result = HierarchicalTimetableManager.update_section(
            school_id, section_id, section_name, capacity
        )
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error updating section: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== AVAILABILITY & CONFLICT CHECK ====================

@hierarchical_bp.route('/check-staff-availability', methods=['POST'])
@check_admin_auth
def check_staff_availability():
    """Check if staff is available for assignment (CONFLICT DETECTION)"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        
        staff_id = data.get('staff_id')
        day_of_week = data.get('day_of_week')
        period_number = data.get('period_number')
        section_id = data.get('section_id')
        
        if None in [staff_id, day_of_week, period_number]:
            return jsonify({
                'success': False,
                'error': 'Missing required fields: staff_id, day_of_week, period_number'
            }), 400
        
        result = HierarchicalTimetableManager.check_staff_availability(
            school_id, staff_id, day_of_week, period_number, section_id
        )
        
        return jsonify(result), 200
    
    except Exception as e:
        logger.error(f"Error checking staff availability: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/check-section-availability', methods=['POST'])
@check_admin_auth
def check_section_availability():
    """Check if section is available for assignment"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        
        section_id = data.get('section_id')
        day_of_week = data.get('day_of_week')
        period_number = data.get('period_number')
        
        if None in [section_id, day_of_week, period_number]:
            return jsonify({
                'success': False,
                'error': 'Missing required fields: section_id, day_of_week, period_number'
            }), 400
        
        result = HierarchicalTimetableManager.check_section_availability(
            school_id, section_id, day_of_week, period_number
        )
        
        return jsonify(result), 200
    
    except Exception as e:
        logger.error(f"Error checking section availability: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== ASSIGNMENT OPERATIONS ====================

@hierarchical_bp.route('/assign-staff', methods=['POST'])
@check_admin_auth
def assign_staff_to_period():
    """
    Assign staff to teach a section during a specific period
    CORE ASSIGNMENT ENDPOINT WITH VALIDATION
    """
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        
        staff_id = data.get('staff_id')
        section_id = data.get('section_id')
        level_id = data.get('level_id')
        day_of_week = data.get('day_of_week')
        period_number = data.get('period_number')
        subject_name = data.get('subject_name')
        room_number = data.get('room_number')
        
        if None in [staff_id, section_id, level_id, day_of_week, period_number]:
            missing = []
            if staff_id is None: missing.append('staff_id')
            if section_id is None: missing.append('section_id')
            if level_id is None: missing.append('level_id')
            if day_of_week is None: missing.append('day_of_week')
            if period_number is None: missing.append('period_number')
            logger.error(f"Assign Staff Missing Fields: {missing}. Data: {data}")
            return jsonify({
                'success': False,
                'error': f'Missing required fields: {", ".join(missing)}'
            }), 400
        
        result = HierarchicalTimetableManager.assign_staff_to_period(
            school_id, staff_id, section_id, level_id, day_of_week, 
            period_number, subject_name, room_number
        )
        
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error assigning staff: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/assignment/<int:assignment_id>', methods=['DELETE'])
@check_admin_auth
def delete_assignment(assignment_id):
    """Delete a timetable assignment"""
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.delete_assignment(school_id, assignment_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error deleting assignment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/assignment/<int:assignment_id>', methods=['PUT', 'PATCH'])
@check_admin_auth
def update_assignment(assignment_id):
    """Update a timetable assignment"""
    try:
        school_id = session.get('school_id')
        data = request.json or {}
        
        subject_name = data.get('subject_name')
        room_number = data.get('room_number')
        
        result = HierarchicalTimetableManager.update_assignment(
            school_id, assignment_id, subject_name, room_number
        )
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error updating assignment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/assignments/availability', methods=['GET'])
@check_auth_either
def get_all_staff_availability():
    """Get summarized availability (free/busy) for all staff"""
    try:
        school_id = session.get('school_id')
        day_of_week = request.args.get('day_of_week')
        if day_of_week is not None and day_of_week != '':
            day_of_week = int(day_of_week)
        else:
            day_of_week = None
            
        result = HierarchicalTimetableManager.get_all_staff_availability(school_id, day_of_week)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching availability summary: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/assignments/all', methods=['GET'])
@check_auth_either
def get_all_assignments():
    """Get all timetable assignments for school"""
    try:
        school_id = session.get('school_id')
        day_of_week = request.args.get('day_of_week')
        if day_of_week is not None:
            day_of_week = int(day_of_week)
            
        result = HierarchicalTimetableManager.get_all_assignments(school_id, day_of_week)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching all assignments: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@hierarchical_bp.route('/section-mentor/list', methods=['GET'])
@check_auth_either
def list_section_mentors():
    """List mentor assignments for class sections."""
    try:
        school_id = session.get('school_id')
        db = get_db()
        _ensure_section_mentor_table(db)
        cursor = db.cursor()

        cursor.execute('''
            SELECT sm.id, sm.section_id, sm.staff_id, sm.created_at, sm.updated_at,
                   s.section_name, l.level_name, l.level_number,
                   st.full_name, st.staff_id AS login_staff_id, st.department
            FROM timetable_section_mentors sm
            JOIN timetable_sections s ON sm.section_id = s.id
            JOIN timetable_academic_levels l ON s.level_id = l.id
            JOIN staff st ON sm.staff_id = st.id
            WHERE sm.school_id = ?
            ORDER BY l.level_number, s.section_name
        ''', (school_id,))

        mentors = []
        for row in cursor.fetchall():
            mentors.append({
                'id': row[0],
                'section_id': row[1],
                'staff_id': row[2],
                'created_at': row[3],
                'updated_at': row[4],
                'section_name': row[5],
                'level_name': row[6],
                'level_number': row[7],
                'full_name': row[8],
                'login_staff_id': row[9],
                'department': row[10]
            })

        return jsonify({'success': True, 'data': mentors})
    except Exception as e:
        logger.error(f"Error listing section mentors: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@hierarchical_bp.route('/section-mentor/save', methods=['POST'])
@check_admin_auth
def save_section_mentor():
    """Assign a mentor to a section."""
    try:
        school_id = session.get('school_id')
        data = request.get_json() or {}
        section_id = data.get('section_id')
        staff_id = data.get('staff_id')

        if not section_id or not staff_id:
            return jsonify({'success': False, 'error': 'Section and staff are required'}), 400

        db = get_db()
        _ensure_section_mentor_table(db)
        cursor = db.cursor()

        cursor.execute('''
            SELECT s.id, s.section_name, l.level_name
            FROM timetable_sections s
            JOIN timetable_academic_levels l ON s.level_id = l.id
            WHERE s.id = ? AND s.school_id = ? AND s.is_active = 1
        ''', (section_id, school_id))
        section_row = cursor.fetchone()
        if not section_row:
            return jsonify({'success': False, 'error': 'Section not found'}), 404

        cursor.execute('''
            SELECT id, staff_id, full_name
            FROM staff
            WHERE id = ? AND school_id = ?
        ''', (staff_id, school_id))
        staff_row = cursor.fetchone()
        if not staff_row:
            return jsonify({'success': False, 'error': 'Staff not found'}), 404

        cursor.execute('''
            SELECT id
            FROM timetable_section_mentors
            WHERE school_id = ? AND section_id = ?
        ''', (school_id, section_id))
        existing_row = cursor.fetchone()

        if existing_row:
            cursor.execute('''
                UPDATE timetable_section_mentors
                SET staff_id = ?, updated_at = CURRENT_TIMESTAMP
                WHERE school_id = ? AND section_id = ?
            ''', (staff_id, school_id, section_id))
        else:
            cursor.execute('''
                INSERT INTO timetable_section_mentors (school_id, section_id, staff_id)
                VALUES (?, ?, ?)
            ''', (school_id, section_id, staff_id))
        db.commit()

        return jsonify({
            'success': True,
            'message': f'{staff_row[2]} assigned as mentor for {section_row[2]} - {section_row[1]}'
        })
    except Exception as e:
        logger.error(f"Error saving section mentor: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@hierarchical_bp.route('/section-mentor/remove/<int:mentor_id>', methods=['POST', 'DELETE'])
@check_admin_auth
def remove_section_mentor(mentor_id):
    """Remove mentor assignment from a section."""
    try:
        school_id = session.get('school_id')
        db = get_db()
        _ensure_section_mentor_table(db)
        cursor = db.cursor()

        cursor.execute('DELETE FROM timetable_section_mentors WHERE id = ? AND school_id = ?', (mentor_id, school_id))
        db.commit()

        return jsonify({'success': True, 'message': 'Mentor assignment removed successfully'})
    except Exception as e:
        logger.error(f"Error removing section mentor: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@hierarchical_bp.route('/staff-period/template', methods=['GET'])
@check_admin_auth
def download_staff_period_template():
    """Download sample sheet for Staff Period Allocation Manager bulk upload."""
    try:
        import pandas as pd

        school_id = session.get('school_id')
        db = get_db()
        cursor = db.cursor()

        cursor.execute('''
            SELECT id, staff_id, full_name, department, destination, gender, phone, email
            FROM staff
            WHERE school_id = ?
            ORDER BY full_name
        ''', (school_id,))
        staff_rows = cursor.fetchall()

        cursor.execute('''
            SELECT l.id AS level_id, l.level_name, l.level_number,
                   s.id AS section_id, s.section_name
            FROM timetable_academic_levels l
            LEFT JOIN timetable_sections s
                ON s.level_id = l.id AND s.school_id = l.school_id AND s.is_active = 1
            WHERE l.school_id = ? AND l.is_active = 1
            ORDER BY l.level_number, s.section_name
        ''', (school_id,))
        level_section_rows = cursor.fetchall()

        cursor.execute('''
            SELECT l.level_name, s.section_name, p.day_of_week, p.period_number,
                   p.period_name, p.start_time, p.end_time
            FROM timetable_periods p
            JOIN timetable_academic_levels l ON p.level_id = l.id
            JOIN timetable_sections s ON p.section_id = s.id
            WHERE p.school_id = ?
            ORDER BY l.level_number, s.section_name, p.day_of_week, p.period_number
        ''', (school_id,))
        period_rows = cursor.fetchall()

        sample_staff = staff_rows[0] if staff_rows else None
        sample_level_section = next((r for r in level_section_rows if r['section_id']), None)
        sample_period = period_rows[0] if period_rows else None

        sample_template = {
            'login_staff_id': [sample_staff['staff_id'] if sample_staff else 'LOGIN001'],
            'staff_name': [sample_staff['full_name'] if sample_staff else 'Sample Staff'],
            'grade_name': [sample_level_section['level_name'] if sample_level_section else 'Class 1'],
            'section_name': [sample_level_section['section_name'] if sample_level_section else 'A'],
            'day_of_week': [sample_period['day_of_week'] if sample_period else 1],
            'day_name': ['Monday'],
            'period_number': [sample_period['period_number'] if sample_period else 1],
            'period_name': [sample_period['period_name'] if sample_period else 'Period 1'],
            'subject_name': ['Mathematics']
        }

        instructions = {
            'Field': [
                'login_staff_id', 'grade_name', 'section_name',
                'day_of_week/day_name', 'period_number', 'subject_name'
            ],
            'Required': ['Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'No'],
            'Format / Notes': [
                'Use login ID from staff table (staff.staff_id).',
                'Exact grade/class name as configured.',
                'Exact section name under the selected grade.',
                'Use day number 0-6 OR day name Sunday-Saturday.',
                'Must exist in timetable periods for same grade/section/day.',
                'Optional subject for this allocation.'
            ]
        }

        staff_lookup = {
            'login_staff_id': [r['staff_id'] for r in staff_rows],
            'staff_name': [r['full_name'] for r in staff_rows],
            'department': [r['department'] for r in staff_rows],
            'position': [r['destination'] for r in staff_rows],
            'gender': [r['gender'] for r in staff_rows],
            'phone': [r['phone'] for r in staff_rows],
            'email': [r['email'] for r in staff_rows]
        }

        grade_section_lookup = {
            'level_id': [r['level_id'] for r in level_section_rows],
            'grade_name': [r['level_name'] for r in level_section_rows],
            'level_number': [r['level_number'] for r in level_section_rows],
            'section_id': [r['section_id'] for r in level_section_rows],
            'section_name': [r['section_name'] for r in level_section_rows]
        }

        period_reference = {
            'grade_name': [r['level_name'] for r in period_rows],
            'section_name': [r['section_name'] for r in period_rows],
            'day_of_week': [r['day_of_week'] for r in period_rows],
            'period_number': [r['period_number'] for r in period_rows],
            'period_name': [r['period_name'] for r in period_rows],
            'start_time': [r['start_time'] for r in period_rows],
            'end_time': [r['end_time'] for r in period_rows]
        }

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame(sample_template).to_excel(writer, sheet_name='Assignment Template', index=False)
            pd.DataFrame(instructions).to_excel(writer, sheet_name='Instructions', index=False)
            pd.DataFrame(staff_lookup).to_excel(writer, sheet_name='Staff Lookup', index=False)
            pd.DataFrame(grade_section_lookup).to_excel(writer, sheet_name='Grade Section Lookup', index=False)
            pd.DataFrame(period_reference).to_excel(writer, sheet_name='Period Reference', index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        output.seek(0)
        workbook = load_workbook(output)
        header_font = Font(bold=True)
        wrap_center = Alignment(wrap_text=True, horizontal='center', vertical='center')

        for sheet in workbook.worksheets:
            sheet.freeze_panes = 'A2'
            if sheet.title == 'Assignment Template':
                sheet.freeze_panes = 'A2'
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = wrap_center
            for column_index, column_cells in enumerate(sheet.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(column_index)
                for cell in column_cells:
                    value = '' if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                sheet.column_dimensions[column_letter].width = min(max_length + 4, 24)

        output = BytesIO()
        workbook.save(output)

        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=staff_period_allocation_template.xlsx'
        response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
    except Exception as e:
        logger.error(f"Error generating staff period template: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@hierarchical_bp.route('/staff-period/bulk-upload', methods=['POST'])
@check_admin_auth
def bulk_upload_staff_period_allocations():
    """Bulk upload staff period allocations from Excel/CSV for hierarchical timetable."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if not file or not file.filename:
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        filename = file.filename.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
            return jsonify({'success': False, 'error': 'Unsupported file format. Use .xlsx, .xls, or .csv'}), 400

        import pandas as pd

        if filename.endswith('.csv'):
            df = pd.read_csv(file.stream)
        else:
            df = pd.read_excel(file)

        if df.empty:
            return jsonify({'success': False, 'error': 'Uploaded file is empty'}), 400

        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'error': 'School ID required'}), 400

        df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

        if 'login_staff_id' not in df.columns and 'staff_id' in df.columns:
            df['login_staff_id'] = df['staff_id']

        required = ['login_staff_id', 'grade_name', 'section_name', 'period_number']
        missing = [c for c in required if c not in df.columns]
        if missing:
            return jsonify({'success': False, 'error': f"Missing required columns: {', '.join(missing)}"}), 400

        db = get_db()
        cursor = db.cursor()

        day_map = {
            'sunday': 0, 'sun': 0,
            'monday': 1, 'mon': 1,
            'tuesday': 2, 'tue': 2, 'tues': 2,
            'wednesday': 3, 'wed': 3,
            'thursday': 4, 'thu': 4, 'thurs': 4,
            'friday': 5, 'fri': 5,
            'saturday': 6, 'sat': 6
        }

        def clean(value):
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return ''
            text = str(value).strip()
            return '' if text.lower() in {'nan', 'none', 'null'} else text

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for row_no, row in enumerate(df.to_dict(orient='records'), start=2):
            try:
                login_staff_id = clean(row.get('login_staff_id'))
                grade_name = clean(row.get('grade_name'))
                section_name = clean(row.get('section_name'))
                period_text = clean(row.get('period_number'))
                subject_name = clean(row.get('subject_name'))
                room_number = clean(row.get('room_number'))

                day_text = clean(row.get('day_of_week')) or clean(row.get('day_name'))
                if day_text.isdigit():
                    day_of_week = int(day_text)
                else:
                    day_of_week = day_map.get(day_text.lower(), None)

                if day_of_week is None or day_of_week < 0 or day_of_week > 6:
                    raise ValueError('Invalid day value. Use 0-6 or valid day name')

                if not period_text:
                    raise ValueError('period_number is required')
                period_number = int(float(period_text))

                cursor.execute('SELECT id FROM staff WHERE school_id = ? AND staff_id = ?', (school_id, login_staff_id))
                staff_row = cursor.fetchone()
                if not staff_row:
                    raise ValueError(f'Staff login ID not found: {login_staff_id}')
                staff_id = staff_row['id']

                cursor.execute('''
                    SELECT id FROM timetable_academic_levels
                    WHERE school_id = ? AND LOWER(level_name) = LOWER(?) AND is_active = 1
                ''', (school_id, grade_name))
                level_row = cursor.fetchone()
                if not level_row:
                    raise ValueError(f'Grade not found: {grade_name}')
                level_id = level_row['id']

                cursor.execute('''
                    SELECT id FROM timetable_sections
                    WHERE school_id = ? AND level_id = ? AND LOWER(section_name) = LOWER(?) AND is_active = 1
                ''', (school_id, level_id, section_name))
                section_row = cursor.fetchone()
                if not section_row:
                    raise ValueError(f'Section not found: {section_name} under {grade_name}')
                section_id = section_row['id']

                cursor.execute('''
                    SELECT id FROM timetable_periods
                    WHERE school_id = ? AND level_id = ? AND section_id = ?
                      AND day_of_week = ? AND period_number = ?
                ''', (school_id, level_id, section_id, day_of_week, period_number))
                period_row = cursor.fetchone()
                if not period_row:
                    raise ValueError('Period not found for selected grade/section/day/period_number')

                cursor.execute('''
                    SELECT id FROM timetable_hierarchical_assignments
                    WHERE school_id = ? AND staff_id = ? AND section_id = ? AND level_id = ?
                      AND day_of_week = ? AND period_number = ?
                ''', (school_id, staff_id, section_id, level_id, day_of_week, period_number))
                existing = cursor.fetchone()

                if existing:
                    cursor.execute('''
                        UPDATE timetable_hierarchical_assignments
                        SET subject_name = ?, room_number = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                    ''', (subject_name, room_number, existing['id']))
                    updated_count += 1
                else:
                    result = HierarchicalTimetableManager.assign_staff_to_period(
                        school_id, staff_id, section_id, level_id, day_of_week,
                        period_number, subject_name, room_number
                    )
                    if result.get('success'):
                        created_count += 1
                    else:
                        skipped_count += 1
                        errors.append(f"Row {row_no}: {result.get('error', 'Assignment failed')}")

            except Exception as row_err:
                skipped_count += 1
                errors.append(f'Row {row_no}: {row_err}')

        if created_count > 0 or updated_count > 0:
            db.commit()

        return jsonify({
            'success': True,
            'message': 'Staff period bulk upload completed',
            'total_rows': len(df.index),
            'created_count': created_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'errors': errors
        })

    except Exception as e:
        logger.error(f"Error in staff period bulk upload: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== SCHEDULE VIEWS ====================

@hierarchical_bp.route('/staff-schedule/<int:staff_id>', methods=['GET'])
@check_auth_either
def get_staff_schedule(staff_id):
    """
    Get complete schedule for a staff member
    STAFF VIEW ENDPOINT
    """
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.get_staff_schedule(school_id, staff_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching staff schedule: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@hierarchical_bp.route('/section-schedule/<int:section_id>', methods=['GET'])
@check_auth_either
def get_section_schedule(section_id):
    """
    Get complete schedule for a section/class
    CLASS VIEW ENDPOINT
    """
    try:
        school_id = session.get('school_id')
        result = HierarchicalTimetableManager.get_section_schedule(school_id, section_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching section schedule: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== COLOR-CODED GRID ====================

@hierarchical_bp.route('/grid/<item_type>/<int:item_id>', methods=['GET'])
@check_auth_either
def get_color_coded_grid(item_type, item_id):
    """
    Get color-coded grid for visualization
    
    item_type: 'staff' or 'section'
    item_id: Staff ID or Section ID
    """
    try:
        school_id = session.get('school_id')
        
        if item_type not in ['staff', 'section']:
            return jsonify({'success': False, 'error': 'Invalid item_type. Must be "staff" or "section"'}), 400
        
        result = HierarchicalTimetableManager.get_color_coded_grid(school_id, item_type, item_id)
        return jsonify(result), 200 if result['success'] else 400
    
    except Exception as e:
        logger.error(f"Error fetching color-coded grid: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== INITIALIZATION ====================

def register_hierarchical_timetable_routes(app):
    """Register the hierarchical timetable blueprint with the app"""
    app.register_blueprint(hierarchical_bp)
    logger.info("Hierarchical Timetable API routes registered successfully")
