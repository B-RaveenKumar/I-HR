from flask import Flask, render_template, request, redirect, url_for, session, jsonify, g, make_response, flash, send_from_directory
import sqlite3
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import datetime
import re
import calendar
import time
import logging
from database import get_db, init_db

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
from zk_biometric import sync_attendance_from_device, ZKBiometricDevice, verify_staff_biometric, process_device_attendance_automatically
from shift_management import ShiftManager
from excel_reports import ExcelReportGenerator
from staff_management_enhanced import StaffManager
from attendance_advanced import AdvancedAttendanceManager
from reporting_dashboard import ReportingDashboard
from data_visualization import DataVisualization
from notification_system import NotificationManager
from backup_manager import BackupManager
from salary_calculator import SalaryCalculator
from pf_calculator import calculate_pf_components

from apscheduler.schedulers.background import BackgroundScheduler
import atexit
import os
import json
import subprocess
import calendar
import base64
import hashlib
import hmac
import secrets
import urllib.parse
import urllib.request

# Import cloud modules (with fallback for backward compatibility)
try:
    from cloud_api import cloud_api
    from cloud_connector import start_cloud_connector, stop_cloud_connector, get_cloud_connector
    from cloud_config import get_cloud_config, get_device_config
    CLOUD_ENABLED = True
except ImportError:
    print("Cloud modules not available. Running in Ethernet-only mode.")
    CLOUD_ENABLED = False

# Create Flask app instance
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())

# Initialize CSRF protection
csrf = CSRFProtect(app)

# Initialize database with the app
init_db(app)

# Initialize APScheduler for automatic sync
scheduler = BackgroundScheduler()
scheduler.start()

# Shut down the scheduler when exiting the app
atexit.register(lambda: scheduler.shutdown())

# Auto-sync configuration (interval in minutes) - Load from database
from database import (
    DATABASE_URL,
    get_system_setting,
    set_system_setting,
    get_school_attendance_mode,
    set_school_attendance_mode,
    get_attendance_config,
    set_attendance_config,
)


def _load_auto_sync_interval(default_minutes=15):
    """Load auto-sync interval from system settings with safe fallback."""
    try:
        with app.app_context():
            raw_value = get_system_setting('auto_sync_interval_minutes', default_minutes)
        interval = int(raw_value)
        return interval if interval > 0 else int(default_minutes)
    except Exception as e:
        logger.warning(f"Failed to load auto-sync interval from settings. Using default {default_minutes}: {e}")
        return int(default_minutes)


AUTO_SYNC_INTERVAL_MINUTES = _load_auto_sync_interval(15)

# Helper function to get module settings for a school
def get_module_enabled(school_id):
    """
    Get module enabled settings for a school.
    Returns dict with module_name: bool values.
    """
    db = get_db()
    school_settings = db.execute('SELECT * FROM schools WHERE id = ?', (school_id,)).fetchone()
    
    # Helper function to safely get column value (works with both SQLite Row and dict)
    def get_setting_value(row, column_name, default=1):
        try:
            return row[column_name] if row else default
        except (KeyError, IndexError):
            return default
    
    return {
        'staff_management': bool(get_setting_value(school_settings, 'staff_management_enabled')),
        'shift_management': bool(get_setting_value(school_settings, 'shift_management_enabled')),
        'salary_management': bool(get_setting_value(school_settings, 'salary_management_enabled')),
        'payroll_processing_review': bool(get_setting_value(school_settings, 'salary_management_enabled')),
        'timetable_management': bool(get_setting_value(school_settings, 'timetable_management_enabled')),
        'reports': bool(get_setting_value(school_settings, 'reports_enabled')),
        'biometric_devices': bool(get_setting_value(school_settings, 'biometric_devices_enabled')),
        'department_shift_assignments': bool(get_setting_value(school_settings, 'department_shift_assignments_enabled')),
        'holiday_management': bool(get_setting_value(school_settings, 'holiday_management_enabled')),
        'quota_management': bool(get_setting_value(school_settings, 'quota_management_enabled')),
        'sub_admin_management': bool(get_setting_value(school_settings, 'sub_admin_management_enabled')),
        'student_management': bool(get_setting_value(school_settings, 'student_management_enabled')),
    }

# Add custom Jinja2 filters
@app.template_filter('dateformat')
def dateformat_filter(date, format='%Y-%m-%d'):
    """Format a date using strftime"""
    if date is None:
        return ""
    if isinstance(date, str):
        try:
            # Try to parse string date
            date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            try:
                # Try alternative format
                date = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').date()
            except ValueError:
                return date  # Return as-is if can't parse
    return date.strftime(format)

@app.template_filter('timeformat')
def timeformat_filter(time, format='%I:%M %p'):
    """Format a time using strftime in 12-hour format"""
    if time is None:
        return "--:--"

    # Already a datetime: format directly.
    if isinstance(time, datetime.datetime):
        return time.strftime(format)

    # Pure time object: combine with today's date for formatting.
    if isinstance(time, datetime.time):
        return datetime.datetime.combine(datetime.date.today(), time).strftime(format)

    if isinstance(time, str):
        try:
            # Try to parse string datetime with full format
            time_obj = datetime.datetime.strptime(time, '%Y-%m-%d %H:%M:%S')
            return time_obj.strftime(format)
        except ValueError:
            try:
                # Try to parse string time only
                time_obj = datetime.datetime.strptime(time, '%H:%M:%S').time()
                return datetime.datetime.combine(datetime.date.today(), time_obj).strftime(format)
            except ValueError:
                try:
                    # Try alternative format
                    time_obj = datetime.datetime.strptime(time, '%H:%M').time()
                    return datetime.datetime.combine(datetime.date.today(), time_obj).strftime(format)
                except ValueError:
                    return time  # Return as-is if can't parse

    # Fallback: return as string for unknown input types.
    return str(time)

@app.template_filter('datetimeformat')
def datetimeformat_filter(datetime_obj, format='%Y-%m-%d %I:%M %p'):
    """Format a datetime using strftime in 12-hour format"""
    if datetime_obj is None:
        return ""
    if isinstance(datetime_obj, str):
        try:
            # Try to parse string datetime
            datetime_obj = datetime.datetime.strptime(datetime_obj, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                # Try alternative format
                datetime_obj = datetime.datetime.strptime(datetime_obj, '%Y-%m-%d')
            except ValueError:
                return datetime_obj  # Return as-is if can't parse
    return datetime_obj.strftime(format)

# Context processor for dynamic institution branding
@app.context_processor
def inject_institution_branding():
    """Make institution branding data available to all templates"""
    institution = {
        'name': None,
        'logo_path': None,
        'branding_enabled': True
    }
    
    # If user is logged in and we have school_id in session
    if 'school_id' in session:
        try:
            db = get_db()
            school = db.execute(
                'SELECT name, logo_path, branding_enabled FROM schools WHERE id = ?',
                (session['school_id'],)
            ).fetchone()
            
            if school:
                institution['name'] = school['name']
                institution['logo_path'] = school['logo_path'] if school['logo_path'] else None
                institution['branding_enabled'] = bool(school['branding_enabled'])
        except Exception as e:
            logger.error(f"Error loading institution branding: {e}")
    
    return {'institution': institution}

@app.template_filter('capitalize_first')
def capitalize_first_filter(text):
    """Capitalize first letter of text"""
    if not text:
        return ""
    return text[0].upper() + text[1:] if len(text) > 1 else text.upper()

from functools import wraps

@app.before_request
def refresh_sub_admin_permissions():
    """
    Check and update sub-admin permissions on each request.
    This ensures that if an admin revokes access, it takes effect immediately
    without requiring the staff member to logout/login.
    """
    # Only check for logged-in staff members with potential sub-admin status
    if 'user_id' in session and session.get('user_type') == 'staff':
        staff_id = session.get('staff_id')
        school_id = session.get('school_id')
        
        if staff_id and school_id:
            db = get_db()
            # Check current permissions in database
            permissions = db.execute('''
                SELECT module_name, can_view, can_edit, can_delete 
                FROM sub_admin_permissions 
                WHERE staff_id = ? AND school_id = ?
            ''', (staff_id, school_id)).fetchall()
            
            # Update session based on current database state
            if len(permissions) > 0:
                # User has sub-admin permissions
                session['is_sub_admin'] = True
                session['permissions'] = {
                    p['module_name']: {
                        'view': bool(p['can_view']), 
                        'edit': bool(p['can_edit']), 
                        'delete': bool(p['can_delete'])
                    } for p in permissions
                }
                session.modified = True
            else:
                # No permissions found - remove sub-admin status
                session['is_sub_admin'] = False
                session['permissions'] = {}
                session.modified = True

def requires_permission(module_name, action='view'):
    """
    Decorator to protect routes based on sub-admin permissions.
    Super admin bypasses this check. Sub-Admins (staff with delegated access) are checked.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('index'))
                
            # Company admin can view most things or doesn't use this route normally
            if session.get('user_type') == 'company_admin':
                return f(*args, **kwargs)
                
            # Standard admin bypass (not a sub-admin)
            if session.get('user_type') == 'admin' and not session.get('is_sub_admin'):
                return f(*args, **kwargs)
                
            # Sub-Admin check — works for both user_type='admin' and user_type='staff' with is_sub_admin=True
            if session.get('is_sub_admin'):
                perms = session.get('permissions', {})
                # If the module exists and the action is permitted
                if perms.get(module_name, {}).get(action):
                    return f(*args, **kwargs)
                flash(f"Access Denied: You do not have '{action}' permission for {module_name.replace('_', ' ')}.", "danger")
                # Redirect back to staff dashboard since sub-admins live there
                return redirect(url_for('staff_dashboard'))
                
            # Standard staff member – no access to admin routes
            if session.get('user_type') == 'staff':
                return redirect(url_for('staff_dashboard'))
                
            return redirect(url_for('index'))
        return decorated_function
    return decorator

@app.template_filter('simple_date')
def simple_date_filter(date_str):
    """Convert YYYY-MM-DD to readable format"""
    if not date_str:
        return ''
    try:
        if isinstance(date_str, str):
            dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
            return dt.strftime('%B %d, %Y')
        return date_str.strftime('%B %d, %Y')
    except:
        return date_str

# Register cloud API blueprint if available
if CLOUD_ENABLED:
    app.register_blueprint(cloud_api)
    print("Cloud API endpoints registered")

# Register timetable management API blueprint
try:
    from timetable_api_routes import timetable_api
    app.register_blueprint(timetable_api)
    print("Timetable Management API endpoints registered")
except ImportError:
    print("Timetable API routes module not found - timetable features may be unavailable")

# Register hierarchical timetable API blueprint
try:
    from hierarchical_timetable_routes import register_hierarchical_timetable_routes
    register_hierarchical_timetable_routes(app)
    print("Hierarchical Timetable API endpoints registered")
except ImportError:
    print("Hierarchical Timetable API routes module not found - hierarchical timetable features may be unavailable")

# Ensure on_duty_permissions table exists
def ensure_on_duty_permission_table():
    db = get_db()
    table_exists = db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='on_duty_permissions'").fetchone()
    if not table_exists:
        db.execute('''
            CREATE TABLE IF NOT EXISTS on_duty_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER,
                school_id INTEGER,
                permission_type TEXT,
                start_datetime TEXT,
                end_datetime TEXT,
                reason TEXT,
                status TEXT DEFAULT 'pending',
                applied_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        db.commit()


def ensure_system_settings_capacity():
    """Ensure system_settings columns can hold JSON payloads (MySQL legacy fix)."""
    try:
        if not str(DATABASE_URL).startswith('mysql'):
            return

        db = get_db()
        # Older deployments use varchar(255), which truncates JSON settings.
        db.execute('ALTER TABLE system_settings MODIFY setting_value TEXT NULL')
        db.execute('ALTER TABLE system_settings MODIFY description TEXT NULL')
        db.commit()
    except Exception as exc:
        # Non-fatal: app can run, but large settings may still truncate.
        print(f'Startup system_settings capacity check warning: {exc}')

with app.app_context():
    ensure_on_duty_permission_table()
    ensure_system_settings_capacity()
    # Repair/seed shift definitions as school-scoped records.
    try:
        from database import migrate_shift_definitions
        migrate_shift_definitions()
    except Exception as _shift_mig_err:
        print(f'Startup shift definitions migration warning: {_shift_mig_err}')
    # Expand department_shift_mappings CHECK constraint + seed overtime shift
    try:
        from database import migrate_department_shift_constraint
        migrate_department_shift_constraint()
    except Exception as _mig_err:
        print(f'Startup migration warning: {_mig_err}')
    # Seed shift history table for existing staff
    try:
        from database import migrate_shift_history
        migrate_shift_history()
    except Exception as _mig_err2:
        print(f'Startup shift history migration warning: {_mig_err2}')


########################################
# HELPER: Get Primary Device for Institution
########################################

def get_institution_device():
    """
    Get the primary Direct_LAN biometric device for the logged-in institution.
    Returns tuple: (device_ip, port) or (None, None) if no device configured.
    
    This helper ensures backward compatibility while supporting the new multi-device system.
    """
    if 'school_id' not in session:
        return None, None
    
    try:
        from database import get_primary_device_for_institution
        device = get_primary_device_for_institution(session['school_id'])
        
        if device and device.get('connection_type') == 'Direct_LAN':
            return device.get('ip_address'), device.get('port', 4370)
        
        return None, None
    except Exception as e:
        logger.error(f"Error getting institution device: {e}")
        return None, None


def record_shift_history(db, staff_id, school_id, new_shift_type, effective_from):
    """
    Record a shift change in staff_shift_history.
    Closes the previous open record by setting effective_to = effective_from - 1 day,
    then inserts a new open-ended record.
    This ensures the calendar can resolve the correct shift for any historical date.
    """
    import datetime as _dt
    try:
        effective_date = (
            _dt.datetime.strptime(effective_from, '%Y-%m-%d').date()
            if isinstance(effective_from, str)
            else effective_from
        )
        day_before = (effective_date - _dt.timedelta(days=1)).strftime('%Y-%m-%d')

        # Close any currently open history record for this staff member
        db.execute('''
            UPDATE staff_shift_history
            SET effective_to = ?
            WHERE staff_id = ? AND school_id = ? AND effective_to IS NULL
        ''', (day_before, staff_id, school_id))

        # Insert the new history record (open-ended)
        db.execute('''
            INSERT INTO staff_shift_history
                (staff_id, school_id, shift_type, effective_from, effective_to)
            VALUES (?, ?, ?, ?, NULL)
        ''', (staff_id, school_id, new_shift_type, effective_from))
    except Exception as _e:
        print(f'Warning: record_shift_history failed for staff_id={staff_id}: {_e}')


# Route for admin to process OD permission (ensure only one definition)
@app.route('/process_on_duty_permission', methods=['POST'])
def process_on_duty_permission():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    permission_id = request.form.get('permission_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    db = get_db()
    permission = db.execute('SELECT * FROM on_duty_permissions WHERE id = ?', (permission_id,)).fetchone()
    if not permission:
        return jsonify({'success': False, 'error': 'Permission not found'})

    status = 'approved' if decision == 'approve' else 'rejected'
    db.execute('''
        UPDATE on_duty_permissions
        SET status = ?, processed_by = ?, processed_at = ?
        WHERE id = ?
    ''', (status, admin_id, processed_at, permission_id))

    if status == 'approved':
        staff_id = permission['staff_id']
        school_id = permission['school_id']
        # Accept both '%Y-%m-%d %H:%M' and '%Y-%m-%d %H:%M:%S' formats
        try:
            start_dt = datetime.datetime.strptime(permission['start_datetime'], '%Y-%m-%d %H:%M')
        except ValueError:
            start_dt = datetime.datetime.strptime(permission['start_datetime'], '%Y-%m-%d %H:%M:%S')
        try:
            end_dt = datetime.datetime.strptime(permission['end_datetime'], '%Y-%m-%d %H:%M')
        except ValueError:
            end_dt = datetime.datetime.strptime(permission['end_datetime'], '%Y-%m-%d %H:%M:%S')
        current_dt = start_dt
        while current_dt.date() <= end_dt.date():
            date_str = current_dt.strftime('%Y-%m-%d')
            existing = db.execute('SELECT * FROM attendance WHERE staff_id = ? AND date = ?', (staff_id, date_str)).fetchone()
            # Mark as present and OD
            if existing:
                db.execute('UPDATE attendance SET status = ? WHERE staff_id = ? AND date = ?', ('OD', staff_id, date_str))
            else:
                db.execute('INSERT INTO attendance (staff_id, school_id, date, status) VALUES (?, ?, ?, ?)', (staff_id, school_id, date_str, 'OD'))
            current_dt += datetime.timedelta(days=1)
    db.commit()
    return jsonify({'success': True})



# File upload configuration
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.teardown_appcontext
def close_db(error):
    """Close database connection at the end of request"""
    _ = error  # Suppress unused parameter warning
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

# Favicon route
@app.route('/favicon.ico')
def favicon():
    """Serve favicon from static/images directory"""
    return send_from_directory(
        os.path.join(app.root_path, 'static', 'images'),
        'favicon.ico',
        mimetype='image/vnd.microsoft.icon'
    )

# Manifest route for PWA
@app.route('/manifest.json')
def manifest():
    """Serve PWA manifest with proper MIME type"""
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        'manifest.json',
        mimetype='application/manifest+json'
    )

# Service Worker route
@app.route('/sw.js')
def service_worker():
    """Serve service worker from static directory"""
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        'sw.js',
        mimetype='application/javascript'
    )

# iClock Protocol Endpoint (Traditional ZKTeco protocol)
@csrf.exempt
@app.route('/iclock/getrequest.aspx', methods=['GET', 'POST'])
@app.route('/iclock/cdata.aspx', methods=['GET', 'POST'])
def iclock_cdata():
    """
    UNIVERSAL ADMS RECEIVER - Protocol Agnostic Endpoint
    Supports all ZK device types: Fingerprint, Face, Palm, Card
    Auto-detects format: Text/Tab-separated, JSON, XML
    """
    try:
        from universal_adms_parser import UniversalADMSParser
        
        # Get device parameters from query string
        sn = request.args.get('SN', '')
        options = request.args.get('options', '')
        table = request.args.get('table', '')  # Some devices send table=ATTLOG
        
        # Extract additional device info from query params (for handshake)
        device_model = request.args.get('model', request.args.get('~DeviceName', ''))
        firmware_ver = request.args.get('FWVersion', request.args.get('~ZKFPVersion', ''))
        platform = request.args.get('~Platform', request.args.get('platform', ''))
        
        logger.info(f"Universal ADMS request from SN: {sn}, method: {request.method}, "
                   f"model: {device_model}, fw: {firmware_ver}, options: {options}")
        
        db = get_db()
        
        # ===== STEP 1: HANDSHAKE & IDENTIFICATION (GET Request) =====
        if request.method == 'GET':
            # Device is requesting commands or performing handshake
            
            # If device provides serial number, capture/update device metadata
            if sn:
                device = db.execute(
                    'SELECT * FROM biometric_devices WHERE serial_number = ? ORDER BY is_active DESC, id DESC LIMIT 1',
                    (sn,)
                ).fetchone()
                
                if device:
                    # Update device metadata from handshake
                    device_columns = {
                        col['name']
                        for col in db.execute("PRAGMA table_info(biometric_devices)").fetchall()
                    }
                    update_fields = []
                    update_values = []
                    
                    if device_model and 'device_model' in device_columns:
                        update_fields.append('device_model = ?')
                        update_values.append(device_model)
                    
                    if firmware_ver and 'firmware_ver' in device_columns:
                        update_fields.append('firmware_ver = ?')
                        update_values.append(firmware_ver)
                    
                    if platform and 'platform' in device_columns:
                        update_fields.append('platform = ?')
                        update_values.append(platform)
                    
                    if 'last_handshake' in device_columns:
                        update_fields.append('last_handshake = ?')
                        update_values.append(datetime.datetime.now())
                    
                    # Store raw options for debugging
                    if options and 'raw_options_data' in device_columns:
                        update_fields.append('raw_options_data = ?')
                        update_values.append(options)
                    
                    if update_fields:
                        update_values.append(device['id'])
                        db.execute(
                            f"UPDATE biometric_devices SET {', '.join(update_fields)} WHERE id = ?",
                            tuple(update_values)
                        )
                        db.commit()
                        logger.info(f"✓ Updated device metadata for {sn}: {device_model}")
                else:
                    # Unknown device - log for manual registration
                    client_ip = request.remote_addr
                    
                    # Check if this unknown device has been seen before
                    unknown_device = db.execute(
                        'SELECT * FROM unknown_device_log WHERE serial_number = ?',
                        (sn,)
                    ).fetchone()
                    
                    if unknown_device:
                        # Update last seen and increment count
                        db.execute('''
                            UPDATE unknown_device_log 
                            SET last_seen = ?, attempt_count = attempt_count + 1,
                                device_model = COALESCE(?, device_model),
                                firmware_ver = COALESCE(?, firmware_ver),
                                platform = COALESCE(?, platform)
                            WHERE serial_number = ?
                            ''', (datetime.datetime.now(), device_model or None, firmware_ver or None, 
                             platform or None, sn))
                    else:
                        # New unknown device - create log entry
                        db.execute('''
                            INSERT INTO unknown_device_log 
                            (serial_number, ip_address, device_model, firmware_ver, platform, 
                             request_type, raw_payload)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (sn, client_ip, device_model or 'Unknown', firmware_ver or 'Unknown',
                             platform or 'Unknown', 'GET/handshake', options or ''))
                    
                    db.commit()
                    logger.warning(f"⚠ Unknown device {sn} ({device_model}) attempted handshake from {client_ip}")
            
            # Return OK to acknowledge device
            response = make_response("OK")
            response.headers['Content-Type'] = 'text/plain'
            return response
        
        # ===== STEP 2: DATA PUSH (POST Request) - UNIVERSAL PARSER =====
        if request.method == 'POST':
            # Get raw POST data and content type
            raw_data = request.get_data(as_text=True)
            content_type = request.content_type
            
            logger.info(f"Universal ADMS POST from {sn}: Content-Type={content_type}, "
                       f"Size={len(raw_data)} bytes")
            logger.debug(f"Raw data sample (first 300 chars): {raw_data[:300]}")
            
            if not sn:
                logger.warning("iClock request missing SN parameter")
                # Log to protocol detection for debugging (if table exists)
                try:
                    db.execute('''
                        INSERT INTO protocol_detection_log 
                        (serial_number, request_method, request_path, content_type, raw_body, 
                         detected_format, parsed_successfully, error_message)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', ('UNKNOWN', 'POST', request.path, content_type, raw_data[:1000],
                         'N/A', False, 'Missing SN parameter'))
                    db.commit()
                except Exception as log_err:
                    if 'protocol_detection_log' in str(log_err) and 'doesn\'t exist' in str(log_err):
                        logger.warning("protocol_detection_log table missing; skipping protocol debug log")
                    else:
                        raise
                return make_response("ERROR: SN required", 400)
            
            # Check if device is registered
            device = db.execute(
                'SELECT * FROM biometric_devices WHERE serial_number = ? ORDER BY is_active DESC, id DESC LIMIT 1',
                (sn,)
            ).fetchone()
            device_columns = {
                col['name']
                for col in db.execute("PRAGMA table_info(biometric_devices)").fetchall()
            }
            
            if not device:
                # Log unknown device attempt
                client_ip = request.remote_addr
                
                unknown_device = db.execute(
                    'SELECT * FROM unknown_device_log WHERE serial_number = ?',
                    (sn,)
                ).fetchone()
                
                if unknown_device:
                    db.execute('''
                        UPDATE unknown_device_log 
                        SET last_seen = ?, attempt_count = attempt_count + 1,
                            raw_payload = ?
                        WHERE serial_number = ?
                    ''', (datetime.datetime.now(), raw_data[:5000], sn))
                else:
                    db.execute('''
                        INSERT INTO unknown_device_log 
                        (serial_number, ip_address, request_type, raw_payload)
                        VALUES (?, ?, ?, ?)
                    ''', (sn, client_ip, 'POST/attendance', raw_data[:5000]))
                
                db.commit()
                logger.warning(f"⚠ Device {sn} not registered - attendance data logged for review")
                
                # Still return OK to avoid device errors
                return make_response("OK")

            # Device exists but is disabled; acknowledge and skip attendance processing.
            is_device_active = str(device.get('is_active', 1)).lower() in ('1', 'true', 'yes')
            if not is_device_active:
                logger.warning(f"⚠ Device {sn} is mapped but inactive (ID: {device.get('id')}); skipping attendance push")
                return make_response("OK")
            
            # ===== STEP 3: UNIVERSAL PARSING - THE SMART LISTENER =====
            parser = UniversalADMSParser()
            parse_result = parser.parse(raw_data, content_type)
            
            detected_format = parse_result.get('format', 'unknown')
            
            # Update device's protocol_type if detected
            if detected_format and detected_format not in ['empty', 'unknown', 'error']:
                format_map = {'text': 'Text', 'json': 'JSON', 'xml': 'XML'}
                protocol_type = format_map.get(detected_format, 'Auto')

                if 'protocol_type' in device_columns:
                    db.execute(
                        'UPDATE biometric_devices SET protocol_type = ? WHERE id = ?',
                        (protocol_type, device['id'])
                    )
            
            # Log protocol detection for debugging (if table exists)
            try:
                db.execute('''
                    INSERT INTO protocol_detection_log 
                    (device_id, serial_number, request_method, request_path, content_type, 
                     raw_body, detected_format, parsed_successfully, error_message, raw_headers)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (device['id'], sn, 'POST', request.path, content_type, 
                     raw_data[:1000], detected_format, parse_result['success'],
                     parse_result.get('error', ''), str(dict(request.headers))[:500]))
                db.commit()
            except Exception as log_err:
                if 'protocol_detection_log' in str(log_err) and 'doesn\'t exist' in str(log_err):
                    logger.warning("protocol_detection_log table missing; skipping protocol debug log")
                else:
                    raise
            
            if not parse_result['success']:
                logger.error(f"❌ Parse failed for device {sn}: {parse_result.get('error')}")
                logger.error(f"Raw data sample: {raw_data[:500]}")
                
                # Still return OK to prevent device from retrying infinitely
                return make_response("OK")
            
            # ===== STEP 4: PROCESS NORMALIZED ATTENDANCE RECORDS =====
            records = parse_result.get('records', [])
            processed_count = 0
            rejected_count = 0
            staff_columns = {
                col['name']
                for col in db.execute("PRAGMA table_info(staff)").fetchall()
            }
            has_staff_biometric_id = 'biometric_id' in staff_columns
            
            logger.info(f"✓ Parsed {len(records)} record(s) from {detected_format.upper()} format")
            
            for record in records:
                try:
                    user_id = record['user_id']
                    timestamp = record['timestamp']
                    verification_type = record['verification_type']
                    biometric_method = record['biometric_method']
                    
                    # Find staff by staff_id; include biometric_id when the column exists.
                    if has_staff_biometric_id:
                        staff = db.execute(
                            '''SELECT s.*, sch.name as school_name 
                               FROM staff s 
                               JOIN schools sch ON s.school_id = sch.id 
                               WHERE (s.staff_id = ? OR s.biometric_id = ?) 
                               AND s.school_id = ?''',
                            (user_id, user_id, device['school_id'])
                        ).fetchone()
                    else:
                        staff = db.execute(
                            '''SELECT s.*, sch.name as school_name 
                               FROM staff s 
                               JOIN schools sch ON s.school_id = sch.id 
                               WHERE s.staff_id = ? AND s.school_id = ?''',
                            (user_id, device['school_id'])
                        ).fetchone()
                    
                    if staff:
                        # Use UnifiedAttendanceProcessor for consistency
                        from zk_biometric import UnifiedAttendanceProcessor
                        
                        processor = UnifiedAttendanceProcessor()
                        punch_result = processor.process_attendance_punch(
                            device_id=device['id'],
                            user_id=user_id,
                            timestamp=timestamp,
                            punch_code=record['punch_code'],
                            verification_method=biometric_method
                        )
                        
                        if punch_result['success']:
                            processed_count += 1
                            logger.info(f"✓ Processed: {staff['full_name']} - {verification_type} "
                                      f"at {record['timestamp_str']} via {biometric_method}")
                        else:
                            rejected_count += 1
                            logger.warning(f"⚠ Rejected: {punch_result.get('message')}")
                    else:
                        rejected_count += 1
                        logger.warning(f"⚠ Staff with ID {user_id} not found for device {sn}")
                
                except Exception as e:
                    rejected_count += 1
                    logger.error(f"❌ Error processing record: {e}, record: {record}")
                    continue
            
            logger.info(f"✓ Universal ADMS processed {processed_count}/{len(records)} records "
                       f"from device {sn} (format: {detected_format.upper()})")
            
            # Update device sync status
            sync_update_fields = []
            sync_update_values = []

            if 'last_sync' in device_columns:
                sync_update_fields.append('last_sync = ?')
                sync_update_values.append(datetime.datetime.now())
            if 'sync_status' in device_columns:
                sync_update_fields.append('sync_status = ?')
                sync_update_values.append('success')

            if sync_update_fields:
                sync_update_values.append(device['id'])
                db.execute(
                    f"UPDATE biometric_devices SET {', '.join(sync_update_fields)} WHERE id = ?",
                    tuple(sync_update_values)
                )
            db.commit()
            
            # Return OK to acknowledge receipt
            response = make_response("OK")
            response.headers['Content-Type'] = 'text/plain'
            return response
    
    except Exception as e:
        logger.error(f"❌ Error in Universal ADMS handler: {e}")
        import traceback
        logger.error(traceback.format_exc())
        
        # Try to log the error
        try:
            db = get_db()
            db.execute('''
                INSERT INTO protocol_detection_log 
                (serial_number, request_method, raw_body, detected_format, 
                 parsed_successfully, error_message)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (request.args.get('SN', 'UNKNOWN'), request.method, 
                 str(request.get_data(as_text=True))[:1000], 'error', False, str(e)[:500]))
            db.commit()
        except:
            pass
        
        return make_response("ERROR", 500)

# In app.py, update the index route
@app.route('/')
def index():
    db = get_db()

    # First check if the column exists
    columns = db.execute("PRAGMA table_info(schools)").fetchall()
    has_is_hidden = any(col['name'] == 'is_hidden' for col in columns)
    has_logo_url = any(col['name'] == 'logo_url' for col in columns)
    has_student_mgmt = any(col['name'] == 'student_management_enabled' for col in columns)

    # Build the SELECT clause dynamically
    select_fields = ['id', 'name']
    if has_logo_url:
        select_fields.append('logo_url')
    if has_student_mgmt:
        select_fields.append('student_management_enabled')
    
    select_clause = ', '.join(select_fields)

    if has_is_hidden:
        schools = db.execute(f'SELECT {select_clause} FROM schools WHERE is_hidden = 0 OR is_hidden IS NULL ORDER BY name').fetchall()
    else:
        schools = db.execute(f'SELECT {select_clause} FROM schools ORDER BY name').fetchall()

    return render_template('index.html', schools=schools)

# Routes
@app.route('/company_login', methods=['GET', 'POST'])
def handle_company_login():
    if request.method == 'GET':
        return render_template('company_login.html')

    # Handle POST request
    username = request.form.get('username')
    password = request.form.get('password')

    db = get_db()
    company_admin = db.execute('''
        SELECT * FROM company_admins WHERE username = ?
    ''', (username,)).fetchone()

    if not company_admin:
        return jsonify({'error': 'Company admin not found'}), 401

    if not check_password_hash(company_admin['password'], password):
        return jsonify({'error': 'Invalid password'}), 401

    session['user_id'] = company_admin['id']
    session['user_type'] = 'company_admin'
    session['full_name'] = company_admin['full_name']
    return jsonify({'redirect': url_for('company_dashboard')})

@app.route('/login', methods=['POST'])
@csrf.exempt  # Exempt from CSRF for easier login handling
def handle_school_login():
    school_id = request.form.get('school_id')
    username = request.form.get('username')
    password = request.form.get('password')

    print(f"Login attempt - School ID: {school_id}, Username: {username}")  # Debug log

    if not school_id:
        return jsonify({'error': 'Please select a school'}), 400

    db = get_db()

    # Check school admin
    admin = db.execute('''
        SELECT * FROM admins
        WHERE school_id = ? AND username = ?
    ''', (school_id, username)).fetchone()

    if admin and check_password_hash(admin['password'], password):
        print("Admin login successful")  # Debug log
        session['user_id'] = admin['id']
        session['school_id'] = admin['school_id']
        session['user_type'] = 'admin'
        session['full_name'] = admin['full_name']
        
        # Load institution branding data
        school = db.execute(
            'SELECT name, logo_path, branding_enabled FROM schools WHERE id = ?',
            (admin['school_id'],)
        ).fetchone()
        if school:
            session['institution_name'] = school['name']
            session['institution_logo'] = school['logo_path']
            session['branding_enabled'] = school['branding_enabled']
        
        return jsonify({'redirect': url_for('admin_dashboard')})

    # Check staff - using username as staff_id
    staff = db.execute('''
        SELECT * FROM staff
        WHERE school_id = ? AND staff_id = ?
    ''', (school_id, username)).fetchone()

    if staff:
        # Support both 'password_hash' and 'password' column names
        password_hash = staff['password_hash'] if 'password_hash' in staff.keys() and staff['password_hash'] is not None else (staff['password'] if 'password' in staff.keys() else '')
        if password_hash is None:
            password_hash = ''
        print(f"Staff found: {staff['full_name']}, Has password hash: {bool(password_hash)}")  # Debug log

        # Check if password hash exists and verify password
        if password_hash and check_password_hash(password_hash, password):
            print("Staff login successful")  # Debug log
            
            # Check for Sub-Admin permissions — cast school_id to int to match DB type
            try:
                school_id_int = int(school_id)
            except (ValueError, TypeError):
                school_id_int = school_id
            permissions = db.execute('''
                SELECT module_name, can_view, can_edit, can_delete 
                FROM sub_admin_permissions 
                WHERE staff_id = ? AND school_id = ?
            ''', (username, school_id_int)).fetchall()
            
            print(f"Sub-admin permissions found: {len(permissions)} rows for staff_id={username}, school_id={school_id_int}")  # Debug
            
            is_sub_admin = len(permissions) > 0
            
            session['user_id'] = staff['id']
            session['school_id'] = staff['school_id']
            session['staff_id'] = username  # Store staff_id for all staff (needed for permission checks)
            # Always keep user_type as 'staff' - Sub-Admins are still staff members
            session['user_type'] = 'staff'
            if is_sub_admin:
                session['is_sub_admin'] = True
                session['permissions'] = {
                    p['module_name']: {
                        'view': bool(p['can_view']), 
                        'edit': bool(p['can_edit']), 
                        'delete': bool(p['can_delete'])
                    } for p in permissions
                }
                print(f"Staff logged in as Sub-Admin with permissions: {session['permissions']}")
            else:
                session['is_sub_admin'] = False
                session['permissions'] = {}
                
            session['full_name'] = staff['full_name']
            
            # Load institution branding data
            school = db.execute(
                'SELECT name, logo_path, branding_enabled FROM schools WHERE id = ?',
                (staff['school_id'],)
            ).fetchone()
            if school:
                session['institution_name'] = school['name']
                session['institution_logo'] = school['logo_path']
                session['branding_enabled'] = school['branding_enabled']
            
            # Both staff and sub-admins go to staff_dashboard
            return jsonify({'redirect': url_for('staff_dashboard')})
        elif not password_hash:
            print("Staff has no password set")  # Debug log
            return jsonify({'error': 'Password not set for this staff member. Please contact admin.'}), 401
        else:
            print("Password verification failed")  # Debug log

    # Check student - using username as student_id
    student = db.execute('''
        SELECT * FROM students
        WHERE school_id = ? AND student_id = ?
    ''', (school_id, username)).fetchone()

    if student:
        password_hash = student['password'] if 'password' in student.keys() and student['password'] is not None else ''
        print(f"Student found: {student['full_name']}, Has password hash: {bool(password_hash)}")  # Debug log

        if password_hash and check_password_hash(password_hash, password):
            print("Student login successful")  # Debug log
            session['student_id'] = student['id']
            session['school_id'] = student['school_id']
            session['user_type'] = 'student'
            session['student_username'] = username
            session['full_name'] = student['full_name']
            
            # Load institution branding data
            school = db.execute(
                'SELECT name, logo_path, branding_enabled FROM schools WHERE id = ?',
                (student['school_id'],)
            ).fetchone()
            if school:
                session['institution_name'] = school['name']
                session['institution_logo'] = school['logo_path']
                session['branding_enabled'] = school['branding_enabled']
            
            return jsonify({'redirect': url_for('student_dashboard')})
        elif not password_hash:
            print("Student has no password set")  # Debug log
            return jsonify({'error': 'Password not set for this student. Please contact admin.'}), 401

    print("Login failed - invalid credentials")  # Debug log
    return jsonify({'error': 'Invalid credentials'}), 401

# Add these new routes to app.py

@app.route('/company/school_details/<int:school_id>')
def school_details(school_id):
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return redirect(url_for('index'))

    db = get_db()

    # Get school info
    school = db.execute('SELECT * FROM schools WHERE id = ?', (school_id,)).fetchone()
    if not school:
        return redirect(url_for('company_dashboard'))

    # Get admins
    admins = db.execute('''
        SELECT id, username, full_name, email
        FROM admins
        WHERE school_id = ?
    ''', (school_id,)).fetchall()

    # Get staff
    staff = db.execute('''
        SELECT id, staff_id, full_name, department, position, email, phone
        FROM staff
        WHERE school_id = ?
        ORDER BY CAST(staff_id AS INTEGER) ASC
    ''', (school_id,)).fetchall()

    # Get attendance summary
    today = datetime.date.today()
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave,
            SUM(CASE WHEN a.status = 'on_duty' THEN 1 ELSE 0 END) as on_duty
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, school_id)).fetchone()

    # Get pending leaves
    pending_leaves = db.execute('''
        SELECT l.id, s.full_name, l.leave_type, l.start_date, l.end_date, l.reason
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ? AND l.status = 'pending'
        ORDER BY l.applied_at
    ''', (school_id,)).fetchall()

    # Create module_settings dict from school columns
    # Helper function to safely get column value
    def get_column_value(row, column_name, default=1):
        try:
            return row[column_name] if row else default
        except (KeyError, IndexError):
            return default
    
    module_settings = {
        'staff_management': get_column_value(school, 'staff_management_enabled'),
        'shift_management': get_column_value(school, 'shift_management_enabled'),
        'salary_management': get_column_value(school, 'salary_management_enabled'),
        'payroll_processing_review': get_column_value(school, 'salary_management_enabled'),
        'timetable_management': get_column_value(school, 'timetable_management_enabled'),
        'reports': get_column_value(school, 'reports_enabled'),
        'biometric_devices': get_column_value(school, 'biometric_devices_enabled'),
        'department_shift_assignments': get_column_value(school, 'department_shift_assignments_enabled'),
        'holiday_management': get_column_value(school, 'holiday_management_enabled'),
        'quota_management': get_column_value(school, 'quota_management_enabled'),
        'sub_admin_management': get_column_value(school, 'sub_admin_management_enabled'),
        'student_management': get_column_value(school, 'student_management_enabled'),
    }

    attendance_mode = get_school_attendance_mode(school_id)
    attendance_qr_config = get_attendance_config(
        f'attendance_qr_config_school_{school_id}',
        {
            'rotation_seconds': 30,
            'screen_name': 'Main Attendance Display',
            'auto_start': True
        }
    )
    attendance_id_scan_config = get_attendance_config(
        f'attendance_idscan_config_school_{school_id}',
        {
            'match_priority': 'staff_id_then_card_uid',
            'allow_continuous_scan': True,
            'kiosk_label': 'ID Scan Kiosk'
        }
    )
    attendance_otp_config = get_attendance_config(
        f'attendance_otp_config_school_{school_id}',
        {
            'ttl_seconds': 180,
            'max_attempts': 3,
            'channels': ['email', 'sms']
        }
    )
    otp_provider_config = get_attendance_config(
        'company_otp_provider_config',
        {
            'smtp_enabled': False,
            'smtp_username': '',
            'smtp_password': '',
            'smtp_use_tls': True,
            'smtp_from_email': '',
            'twilio_enabled': False,
            'twilio_sid': '',
            'twilio_from_number': ''
        }
    )

    return render_template('school_details.html',
                         school=school,
                         admins=admins,
                         staff=staff,
                         attendance_summary=attendance_summary,
                         pending_leaves=pending_leaves,
                         module_settings=module_settings,
                         attendance_mode=attendance_mode,
                         attendance_qr_config=attendance_qr_config,
                         attendance_id_scan_config=attendance_id_scan_config,
                         attendance_otp_config=attendance_otp_config,
                         otp_provider_config=otp_provider_config,
                         today=today)


@csrf.exempt
@app.route('/api/attendance/settings/<int:school_id>', methods=['GET', 'POST'])
def attendance_settings_api(school_id):
    if 'user_id' not in session or session.get('user_type') != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    db = get_db()
    school = db.execute('SELECT id FROM schools WHERE id = ?', (school_id,)).fetchone()
    if not school:
        return jsonify({'success': False, 'error': 'School not found'}), 404

    allowed_modes = {'biometric', 'dynamic_qr', 'id_scan', 'otp'}
    default_payload = {
        'mode': get_school_attendance_mode(school_id),
        'qr_config': get_attendance_config(
            f'attendance_qr_config_school_{school_id}',
            {'rotation_seconds': 30, 'screen_name': 'Main Attendance Display', 'auto_start': True}
        ),
        'id_scan_config': get_attendance_config(
            f'attendance_idscan_config_school_{school_id}',
            {'match_priority': 'staff_id_then_card_uid', 'allow_continuous_scan': True, 'kiosk_label': 'ID Scan Kiosk'}
        ),
        'otp_config': get_attendance_config(
            f'attendance_otp_config_school_{school_id}',
            {'ttl_seconds': 180, 'max_attempts': 3, 'channels': ['email', 'sms']}
        )
    }

    if request.method == 'GET':
        return jsonify({'success': True, 'settings': default_payload})

    data = request.get_json(silent=True) or {}
    mode = str(data.get('mode', default_payload['mode'])).strip().lower()
    if mode not in allowed_modes:
        return jsonify({'success': False, 'error': 'Invalid attendance mode'}), 400

    qr_config = data.get('qr_config') if isinstance(data.get('qr_config'), dict) else default_payload['qr_config']
    id_scan_config = data.get('id_scan_config') if isinstance(data.get('id_scan_config'), dict) else default_payload['id_scan_config']
    otp_config = data.get('otp_config') if isinstance(data.get('otp_config'), dict) else default_payload['otp_config']

    # Guardrails for secure defaults.
    qr_rotation = int(qr_config.get('rotation_seconds', 30) or 30)
    qr_config['rotation_seconds'] = max(30, min(120, qr_rotation))

    otp_ttl = int(otp_config.get('ttl_seconds', 180) or 180)
    otp_attempts = int(otp_config.get('max_attempts', 3) or 3)
    otp_config['ttl_seconds'] = max(60, min(600, otp_ttl))
    otp_config['max_attempts'] = max(1, min(10, otp_attempts))

    if mode == 'otp':
        provider = data.get('otp_provider_config') if isinstance(data.get('otp_provider_config'), dict) else None
        if provider is None:
            provider = get_attendance_config('company_otp_provider_config', {})

        has_email = _as_bool(provider.get('smtp_enabled'))
        has_sms = _as_bool(provider.get('twilio_enabled'))
        if not (has_email or has_sms):
            return jsonify({
                'success': False,
                'error': 'OTP mode cannot be enabled until at least one provider (SMTP or Twilio) is configured.'
            }), 400

        if has_email and not str(provider.get('smtp_from_email', '')).strip():
            return jsonify({'success': False, 'error': 'SMTP from email is required when SMTP is enabled.'}), 400
        if has_email and not str(provider.get('smtp_host', '')).strip():
            return jsonify({'success': False, 'error': 'SMTP host is required when SMTP is enabled.'}), 400
        if has_email and not int(provider.get('smtp_port', 587) or 587):
            return jsonify({'success': False, 'error': 'SMTP port is required when SMTP is enabled.'}), 400
        if has_email and not str(provider.get('smtp_password', '')).strip():
            return jsonify({'success': False, 'error': 'SMTP app password is required when SMTP is enabled.'}), 400
        if has_sms and (
            not str(provider.get('twilio_sid', '')).strip() or
            not str(provider.get('twilio_auth_token', '')).strip() or
            not str(provider.get('twilio_from_number', '')).strip()
        ):
            return jsonify({'success': False, 'error': 'Twilio SID, Auth Token, and From Number are required when Twilio is enabled.'}), 400

        # If provider config is submitted together with attendance settings,
        # persist it atomically in the same save flow.
        if isinstance(data.get('otp_provider_config'), dict):
            provider_save_ok = set_attendance_config(
                'company_otp_provider_config',
                {
                    'smtp_enabled': _as_bool(provider.get('smtp_enabled')),
                    'smtp_username': str(provider.get('smtp_username', '')).strip(),
                    'smtp_password': str(provider.get('smtp_password', '')).strip(),
                    'smtp_use_tls': _as_bool(provider.get('smtp_use_tls', True)),
                    'smtp_from_email': str(provider.get('smtp_from_email', '')).strip(),
                    'smtp_host': str(provider.get('smtp_host', '')).strip(),
                    'smtp_port': int(provider.get('smtp_port', 587) or 587),
                    'twilio_enabled': _as_bool(provider.get('twilio_enabled')),
                    'twilio_sid': str(provider.get('twilio_sid', '')).strip(),
                    'twilio_auth_token': str(provider.get('twilio_auth_token', '')).strip(),
                    'twilio_from_number': str(provider.get('twilio_from_number', '')).strip()
                },
                'Company-level OTP provider settings for attendance verification'
            )
            if not provider_save_ok:
                return jsonify({'success': False, 'error': 'Unable to persist OTP provider settings. Please try again.'}), 500

    set_school_attendance_mode(school_id, mode)
    set_attendance_config(
        f'attendance_qr_config_school_{school_id}',
        qr_config,
        'Per-school dynamic QR attendance config'
    )
    set_attendance_config(
        f'attendance_idscan_config_school_{school_id}',
        id_scan_config,
        'Per-school ID scan attendance config'
    )
    set_attendance_config(
        f'attendance_otp_config_school_{school_id}',
        otp_config,
        'Per-school OTP attendance config'
    )

    return jsonify({
        'success': True,
        'message': 'Attendance settings updated successfully',
        'settings': {
            'mode': mode,
            'qr_config': qr_config,
            'id_scan_config': id_scan_config,
            'otp_config': otp_config
        }
    })


@csrf.exempt
@app.route('/api/company/otp-provider-settings', methods=['GET', 'POST'])
def company_otp_provider_settings_api():
    if 'user_id' not in session or session.get('user_type') != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    defaults = {
        'smtp_enabled': False,
        'smtp_username': '',
        'smtp_password': '',
        'smtp_use_tls': True,
        'smtp_from_email': '',
        'smtp_host': '',
        'smtp_port': 587,
        'twilio_enabled': False,
        'twilio_sid': '',
        'twilio_auth_token': '',
        'twilio_from_number': ''
    }

    if request.method == 'GET':
        config = get_attendance_config('company_otp_provider_config', defaults)
        masked = dict(defaults)
        masked.update(config)

        sid = str(masked.get('twilio_sid') or '')
        token = str(masked.get('twilio_auth_token') or '')
        smtp_password = str(masked.get('smtp_password') or '')
        masked['twilio_sid'] = (sid[:4] + '...' + sid[-4:]) if len(sid) > 8 else sid
        masked['twilio_auth_token'] = '********' if token else ''
        masked['smtp_password_set'] = bool(smtp_password)
        masked['smtp_password'] = ''
        return jsonify({'success': True, 'settings': masked})

    data = request.get_json(silent=True) or {}
    existing_config = get_attendance_config('company_otp_provider_config', defaults)
    config = dict(defaults)
    config.update({
        'smtp_enabled': _as_bool(data.get('smtp_enabled', False)),
        'smtp_username': str(data.get('smtp_username', '')).strip(),
        'smtp_password': str(data.get('smtp_password', '')).strip(),
        'smtp_use_tls': _as_bool(data.get('smtp_use_tls', True)),
        'smtp_from_email': str(data.get('smtp_from_email', '')).strip(),
        'smtp_host': str(data.get('smtp_host', '')).strip(),
        'smtp_port': int(data.get('smtp_port', 587) or 587),
        'twilio_enabled': _as_bool(data.get('twilio_enabled', False)),
        'twilio_sid': str(data.get('twilio_sid', '')).strip(),
        'twilio_auth_token': str(data.get('twilio_auth_token', '')).strip(),
        'twilio_from_number': str(data.get('twilio_from_number', '')).strip()
    })

    # Preserve existing secrets unless explicitly replaced.
    if config['smtp_enabled'] and not config['smtp_password']:
        config['smtp_password'] = str(existing_config.get('smtp_password', '')).strip()
    if config['twilio_enabled'] and not config['twilio_auth_token']:
        config['twilio_auth_token'] = str(existing_config.get('twilio_auth_token', '')).strip()

    if config['smtp_enabled'] and not config['smtp_from_email']:
        return jsonify({'success': False, 'error': 'SMTP from email is required when SMTP is enabled.'}), 400
    if config['smtp_enabled'] and not config['smtp_host']:
        return jsonify({'success': False, 'error': 'SMTP host is required when SMTP is enabled.'}), 400
    if config['smtp_enabled'] and not config['smtp_password']:
        return jsonify({'success': False, 'error': 'SMTP app password is required when SMTP is enabled.'}), 400
    if config['twilio_enabled'] and (not config['twilio_sid'] or not config['twilio_auth_token'] or not config['twilio_from_number']):
        return jsonify({'success': False, 'error': 'Twilio SID, Auth Token, and From Number are required when Twilio is enabled.'}), 400

    save_ok = set_attendance_config(
        'company_otp_provider_config',
        config,
        'Company-level OTP provider settings for attendance verification'
    )
    if not save_ok:
        return jsonify({'success': False, 'error': 'Unable to persist OTP provider settings. Please try again.'}), 500

    return jsonify({'success': True, 'message': 'OTP provider settings saved successfully'})


def _ensure_attendance_selection_runtime_tables():
    """Ensure runtime tables for QR/OTP challenges exist."""
    db = get_db()
    if str(DATABASE_URL).startswith('mysql'):
        db.execute('''
            CREATE TABLE IF NOT EXISTS attendance_qr_tokens (
                id INT PRIMARY KEY AUTO_INCREMENT,
                school_id INT NOT NULL,
                token_jti VARCHAR(128) NOT NULL,
                token_hash VARCHAR(128) NOT NULL,
                issued_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL,
                used_by_staff_id INT,
                used_at TIMESTAMP NULL,
                status VARCHAR(32) DEFAULT 'active'
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS attendance_otp_challenges (
                id INT PRIMARY KEY AUTO_INCREMENT,
                school_id INT NOT NULL,
                staff_id INT NOT NULL,
                channel VARCHAR(24) NOT NULL,
                otp_hash VARCHAR(128) NOT NULL,
                expires_at TIMESTAMP NOT NULL,
                attempts INT DEFAULT 0,
                max_attempts INT DEFAULT 3,
                verified_at TIMESTAMP NULL,
                status VARCHAR(32) DEFAULT 'pending',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS attendance_office_otp_codes (
                id INT PRIMARY KEY AUTO_INCREMENT,
                school_id INT NOT NULL,
                otp_hash VARCHAR(128) NOT NULL,
                issued_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL,
                status VARCHAR(32) DEFAULT 'active'
            )
        ''')
    else:
        db.execute('''
            CREATE TABLE IF NOT EXISTS attendance_qr_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                token_jti TEXT NOT NULL,
                token_hash TEXT NOT NULL,
                issued_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL,
                used_by_staff_id INTEGER,
                used_at TIMESTAMP,
                status TEXT DEFAULT 'active'
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS attendance_otp_challenges (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                staff_id INTEGER NOT NULL,
                channel TEXT NOT NULL,
                otp_hash TEXT NOT NULL,
                expires_at TIMESTAMP NOT NULL,
                attempts INTEGER DEFAULT 0,
                max_attempts INTEGER DEFAULT 3,
                verified_at TIMESTAMP,
                status TEXT DEFAULT 'pending',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        db.execute('''
            CREATE TABLE IF NOT EXISTS attendance_office_otp_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                otp_hash TEXT NOT NULL,
                issued_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL,
                status TEXT DEFAULT 'active'
            )
        ''')
    db.commit()


def _hash_secure_value(raw_value):
    return hashlib.sha256(str(raw_value).encode('utf-8')).hexdigest()


def _as_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {'1', 'true', 'yes', 'on', 'enabled'}
    return bool(value)


def _smtp_provider_ready(provider_config):
    smtp_enabled = _as_bool(provider_config.get('smtp_enabled'))
    smtp_host = str(provider_config.get('smtp_host', '')).strip()
    smtp_from = str(provider_config.get('smtp_from_email', '')).strip()
    smtp_password = str(provider_config.get('smtp_password', '')).strip()
    smtp_port = int(provider_config.get('smtp_port', 587) or 587)

    complete = bool(smtp_host and smtp_from and smtp_password and smtp_port)
    return smtp_enabled or complete


def _twilio_provider_ready(provider_config):
    twilio_enabled = _as_bool(provider_config.get('twilio_enabled'))
    sid = str(provider_config.get('twilio_sid', '')).strip()
    token = str(provider_config.get('twilio_auth_token', '')).strip()
    from_number = str(provider_config.get('twilio_from_number', '')).strip()

    complete = bool(sid and token and from_number)
    return twilio_enabled or complete


def _send_smtp_email(provider_config, to_email, subject, body_text):
    """Send email via configured SMTP provider settings."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    smtp_host = str(provider_config.get('smtp_host', '')).strip()
    smtp_port = int(provider_config.get('smtp_port', 587) or 587)
    smtp_from = str(provider_config.get('smtp_from_email', '')).strip()
    smtp_username = str(provider_config.get('smtp_username', '')).strip() or smtp_from
    smtp_password = str(provider_config.get('smtp_password', '')).strip()
    smtp_use_tls = _as_bool(provider_config.get('smtp_use_tls', True))
    smtp_use_ssl = _as_bool(provider_config.get('smtp_use_ssl', False))

    # Auto-select implicit SSL when using SMTPS port unless explicitly overridden.
    if smtp_port == 465 and 'smtp_use_ssl' not in provider_config:
        smtp_use_ssl = True
        smtp_use_tls = False

    if not smtp_host or not smtp_port or not smtp_from or not smtp_password:
        return {'success': False, 'error': 'SMTP provider settings are incomplete'}

    msg = MIMEMultipart('alternative')
    msg['From'] = smtp_from
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body_text, 'plain'))

    try:
        if smtp_use_ssl:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=20)
            server.ehlo()
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=20)
            server.ehlo()
            if smtp_use_tls:
                server.starttls()
                server.ehlo()
        server.login(smtp_username, smtp_password)
        server.sendmail(smtp_from, [to_email], msg.as_string())
        server.quit()
        return {'success': True}
    except Exception as exc:
        return {'success': False, 'error': f'SMTP OTP delivery failed: {exc}'}


def _send_twilio_sms(provider_config, to_number, body_text):
    """Send SMS via Twilio REST API without requiring external SDK."""
    sid = str(provider_config.get('twilio_sid', '')).strip()
    token = str(provider_config.get('twilio_auth_token', '')).strip()
    from_number = str(provider_config.get('twilio_from_number', '')).strip()

    if not sid or not token or not from_number:
        return {'success': False, 'error': 'Twilio credentials are incomplete'}

    endpoint = f'https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json'
    form_data = urllib.parse.urlencode({
        'From': from_number,
        'To': to_number,
        'Body': body_text
    }).encode('utf-8')

    request_obj = urllib.request.Request(endpoint, data=form_data, method='POST')
    basic_token = base64.b64encode(f'{sid}:{token}'.encode('utf-8')).decode('utf-8')
    request_obj.add_header('Authorization', f'Basic {basic_token}')
    request_obj.add_header('Content-Type', 'application/x-www-form-urlencoded')

    try:
        with urllib.request.urlopen(request_obj, timeout=15) as response:
            status_code = getattr(response, 'status', 200)
            payload = response.read().decode('utf-8')
            if status_code >= 300:
                return {'success': False, 'error': f'Twilio rejected request ({status_code})', 'details': payload}
            return {'success': True, 'details': payload}
    except Exception as exc:
        return {'success': False, 'error': f'Twilio SMS delivery failed: {exc}'}


def _build_signed_qr_token(school_id, ttl_seconds):
    now = datetime.datetime.utcnow()
    expires_at = now + datetime.timedelta(seconds=int(ttl_seconds))
    payload = {
        'school_id': int(school_id),
        'mode': 'dynamic_qr',
        'jti': secrets.token_urlsafe(12),
        'iat': int(now.timestamp()),
        'exp': int(expires_at.timestamp())
    }
    payload_json = json.dumps(payload, separators=(',', ':'), sort_keys=True)
    payload_b64 = base64.urlsafe_b64encode(payload_json.encode('utf-8')).decode('utf-8')
    signature = hmac.new(
        app.secret_key.encode('utf-8'),
        payload_json.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    token = f'{payload_b64}.{signature}'
    return token, payload, expires_at


def _decode_signed_qr_token(token):
    try:
        payload_b64, signature = token.split('.', 1)
        payload_json = base64.urlsafe_b64decode(payload_b64.encode('utf-8')).decode('utf-8')
        expected = hmac.new(
            app.secret_key.encode('utf-8'),
            payload_json.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None, 'Invalid token signature'
        payload = json.loads(payload_json)
        exp = int(payload.get('exp', 0) or 0)
        if exp <= int(datetime.datetime.utcnow().timestamp()):
            return None, 'Token expired'
        return payload, None
    except Exception:
        return None, 'Malformed token'


def _build_qr_svg(raw_text):
    """Render QR as SVG bytes using server-side library fallback."""
    import qrcode
    import qrcode.image.svg

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(raw_text)
    qr.make(fit=True)
    img = qr.make_image(image_factory=qrcode.image.svg.SvgImage)
    return img.to_string()


def _issue_office_otp_code(school_id, ttl_seconds=30):
    """Issue a short-lived 4-digit office OTP shown on admin panel."""
    db = get_db()
    ttl = max(30, min(120, int(ttl_seconds or 30)))
    otp_code = f'{secrets.randbelow(10000):04d}'
    now = datetime.datetime.now()
    expires_at = now + datetime.timedelta(seconds=ttl)

    db.execute(
        'UPDATE attendance_office_otp_codes SET status = ? WHERE school_id = ? AND status = ?',
        ('expired', school_id, 'active')
    )
    db.execute(
        '''
        INSERT INTO attendance_office_otp_codes
        (school_id, otp_hash, issued_at, expires_at, status)
        VALUES (?, ?, ?, ?, 'active')
        ''',
        (school_id, _hash_secure_value(otp_code), now, expires_at)
    )
    db.commit()

    return {
        'otp_code': otp_code,
        'issued_at': now,
        'expires_at': expires_at,
        'ttl_seconds': ttl
    }


def _is_valid_office_otp_code(school_id, office_otp):
    """Validate entered office OTP against latest unexpired school code."""
    db = get_db()
    row = db.execute(
        '''
        SELECT id, otp_hash, expires_at
        FROM attendance_office_otp_codes
        WHERE school_id = ? AND status = 'active'
        ORDER BY id DESC
        LIMIT 1
        ''',
        (school_id,)
    ).fetchone()

    if not row:
        return False

    expires_at = row['expires_at']
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.datetime.fromisoformat(expires_at.replace('Z', ''))
        except ValueError:
            try:
                expires_at = datetime.datetime.strptime(expires_at, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                expires_at = datetime.datetime.now() - datetime.timedelta(seconds=1)

    if expires_at and expires_at < datetime.datetime.now():
        db.execute('UPDATE attendance_office_otp_codes SET status = ? WHERE id = ?', ('expired', row['id']))
        db.commit()
        return False

    return hmac.compare_digest(_hash_secure_value(office_otp), row['otp_hash'])


def _record_mode_attendance(staff_db_id, school_id, mode_name, event_time=None):
    db = get_db()
    timestamp = event_time or datetime.datetime.now()
    event_date = timestamp.date()
    current_time = timestamp.strftime('%H:%M:%S')

    attendance_columns = {
        col['name']
        for col in db.execute('PRAGMA table_info(attendance)').fetchall()
    }

    def _has_column(column_name):
        return column_name in attendance_columns

    def _parse_hhmmss(value):
        if not value:
            return None
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                return datetime.datetime.strptime(str(value), fmt).time()
            except ValueError:
                continue
        return None

    staff_row = db.execute('SELECT shift_type FROM staff WHERE id = ?', (staff_db_id,)).fetchone()
    staff_shift_type = staff_row['shift_type'] if staff_row and staff_row['shift_type'] else 'general'
    shift_manager = ShiftManager()

    existing = db.execute(
        '''
        SELECT id, time_in, time_out, status
        FROM attendance
        WHERE staff_id = ? AND school_id = ? AND date = ?
        ORDER BY
            CASE
                WHEN COALESCE(time_in, '') <> '' AND COALESCE(time_out, '') = '' THEN 0
                WHEN COALESCE(time_in, '') = '' THEN 1
                ELSE 2
            END,
            COALESCE(time_in, '') DESC
        LIMIT 1
        ''',
        (staff_db_id, school_id, event_date)
    ).fetchone()

    def _update_attendance_row(field_values, is_checkout=False):
        target_id = existing['id'] if existing and existing['id'] else None

        set_parts = []
        params = []
        for key, value in field_values.items():
            if value is None:
                continue
            if key in {'time_in', 'time_out', 'status', 'notes'} or _has_column(key):
                set_parts.append(f'{key} = ?')
                params.append(value)

        if not set_parts:
            return

        if target_id:
            db.execute(f"UPDATE attendance SET {', '.join(set_parts)} WHERE id = ?", tuple(params + [target_id]))
            return

        if is_checkout:
            where_clause = '''
                WHERE staff_id = ? AND school_id = ? AND date = ?
                  AND COALESCE(time_in, '') <> ''
                  AND (time_out IS NULL OR time_out = '')
            '''
        else:
            where_clause = '''
                WHERE staff_id = ? AND school_id = ? AND date = ?
                  AND (time_in IS NULL OR time_in = '')
            '''

        where_params = [staff_db_id, school_id, event_date]
        db.execute(
            f"UPDATE attendance SET {', '.join(set_parts)} {where_clause}",
            tuple(params + where_params)
        )

    def _insert_attendance_row(field_values):
        insert_parts = []
        params = []
        for key, value in field_values.items():
            if value is None:
                continue
            if key in {'staff_id', 'school_id', 'date', 'time_in', 'time_out', 'status', 'notes'} or _has_column(key):
                insert_parts.append(key)
                params.append(value)

        if not insert_parts:
            return

        db.execute(
            f"INSERT INTO attendance ({', '.join(insert_parts)}) VALUES ({', '.join(['?'] * len(insert_parts))})",
            tuple(params)
        )

    def _recompute_and_persist_metrics(row_id=None):
        if row_id:
            row = db.execute(
                'SELECT id, time_in, time_out, status FROM attendance WHERE id = ?',
                (row_id,)
            ).fetchone()
        else:
            row = db.execute(
                '''
                SELECT id, time_in, time_out, status
                FROM attendance
                WHERE staff_id = ? AND school_id = ? AND date = ?
                ORDER BY
                    CASE
                        WHEN COALESCE(time_in, '') <> '' AND COALESCE(time_out, '') = '' THEN 0
                        WHEN COALESCE(time_in, '') = '' THEN 1
                        ELSE 2
                    END,
                    COALESCE(time_in, '') DESC
                LIMIT 1
                ''',
                (staff_db_id, school_id, event_date)
            ).fetchone()

        if not row or not row['time_in']:
            return

        check_in_time = _parse_hhmmss(row['time_in'])
        check_out_time = _parse_hhmmss(row['time_out']) if row['time_out'] else None
        if not check_in_time:
            return

        status_result = shift_manager.calculate_attendance_status(staff_shift_type, check_in_time, check_out_time)
        late_minutes = int(status_result.get('late_duration_minutes', 0) or 0)
        early_minutes = int(status_result.get('early_departure_minutes', 0) or 0)

        existing_status = str(row['status'] or '').strip().lower()
        if existing_status in {'approved_regularization', 'leave', 'holiday', 'on_duty', 'on_permission', 'absent'}:
            final_status = row['status']
        elif late_minutes > 0:
            final_status = 'late'
        elif early_minutes > 0:
            final_status = 'left_soon'
        else:
            final_status = 'present'

        db.execute(
            '''
            UPDATE attendance
            SET status = ?,
                late_duration_minutes = ?,
                early_departure_minutes = ?,
                shift_type = ?,
                shift_start_time = ?,
                shift_end_time = ?
            WHERE id = ?
            ''',
            (
                final_status,
                late_minutes,
                early_minutes,
                staff_shift_type,
                status_result.get('shift_start_time').strftime('%H:%M:%S') if status_result.get('shift_start_time') else None,
                status_result.get('shift_end_time').strftime('%H:%M:%S') if status_result.get('shift_end_time') else None,
                row['id']
            )
        )

    if existing and existing['time_in'] and existing['time_out']:
        return {'success': False, 'message': 'Already checked in and checked out for today'}

    try:
        if not existing:
            status_result = shift_manager.calculate_attendance_status(staff_shift_type, timestamp.time())
            _insert_attendance_row({
                'staff_id': staff_db_id,
                'school_id': school_id,
                'date': event_date,
                'time_in': current_time,
                'status': status_result.get('status', 'present'),
                'notes': f'Attendance via {mode_name}',
                'attendance_mode': mode_name,
                'late_duration_minutes': int(status_result.get('late_duration_minutes', 0) or 0),
                'shift_type': staff_shift_type,
                'shift_start_time': status_result.get('shift_start_time').strftime('%H:%M:%S') if status_result.get('shift_start_time') else None,
                'shift_end_time': status_result.get('shift_end_time').strftime('%H:%M:%S') if status_result.get('shift_end_time') else None,
            })
            action = 'check-in'
        elif not existing['time_in']:
            status_result = shift_manager.calculate_attendance_status(staff_shift_type, timestamp.time())
            _update_attendance_row({
                'time_in': current_time,
                'status': status_result.get('status', 'present'),
                'attendance_mode': mode_name,
                'notes': f'Attendance via {mode_name}',
                'late_duration_minutes': int(status_result.get('late_duration_minutes', 0) or 0),
                'shift_type': staff_shift_type,
                'shift_start_time': status_result.get('shift_start_time').strftime('%H:%M:%S') if status_result.get('shift_start_time') else None,
                'shift_end_time': status_result.get('shift_end_time').strftime('%H:%M:%S') if status_result.get('shift_end_time') else None,
            })
            action = 'check-in'
        else:
            parsed_check_in = _parse_hhmmss(existing['time_in'])
            checkout_result = shift_manager.calculate_attendance_status(
                staff_shift_type,
                parsed_check_in or timestamp.time(),
                timestamp.time()
            )

            final_status = existing['status'] or 'present'
            early_minutes = int(checkout_result.get('early_departure_minutes', 0) or 0)
            if early_minutes > 0 and str(final_status).lower() != 'late':
                final_status = 'left_soon'

            _update_attendance_row({
                'time_out': current_time,
                'status': final_status,
                'attendance_mode': mode_name,
                'notes': f'Attendance via {mode_name}',
                'early_departure_minutes': early_minutes,
            }, is_checkout=True)
            action = 'check-out'

        _recompute_and_persist_metrics(existing['id'] if existing else None)

        db.commit()
        return {'success': True, 'message': f'{action.title()} recorded via {mode_name}', 'action': action}
    except Exception as exc:
        db.rollback()
        return {'success': False, 'message': f'Attendance write failed: {exc}'}


@csrf.exempt
@app.route('/api/attendance/qr/issue', methods=['POST'])
def attendance_qr_issue():
    if 'user_id' not in session or session.get('user_type') not in ['admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'error': 'School context missing'}), 400

    mode = get_school_attendance_mode(school_id)
    if mode != 'dynamic_qr':
        return jsonify({'success': False, 'error': f'Dynamic QR is disabled for this school (current mode: {mode}).'}), 400

    _ensure_attendance_selection_runtime_tables()
    qr_config = get_attendance_config(
        f'attendance_qr_config_school_{school_id}',
        {'rotation_seconds': 30}
    )
    ttl_seconds = max(30, min(120, int(qr_config.get('rotation_seconds', 30) or 30)))

    token, payload, expires_at = _build_signed_qr_token(school_id, ttl_seconds)
    db = get_db()
    db.execute(
        '''
        INSERT INTO attendance_qr_tokens (school_id, token_jti, token_hash, expires_at, status)
        VALUES (?, ?, ?, ?, 'active')
        ''',
        (school_id, payload['jti'], _hash_secure_value(token), expires_at)
    )
    db.commit()

    return jsonify({
        'success': True,
        'token': token,
        'expires_at': expires_at.isoformat(),
        'rotation_seconds': ttl_seconds
    })


@csrf.exempt
@app.route('/api/attendance/qr/verify', methods=['POST'])
def attendance_qr_verify():
    if 'user_id' not in session or session.get('user_type') != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'dynamic_qr':
        return jsonify({'success': False, 'error': f'Dynamic QR is disabled for this school (current mode: {mode}).'}), 400

    request_payload = request.get_json(silent=True) or {}
    token = request_payload.get('token') or request_payload.get('qr_token')
    if not token:
        return jsonify({'success': False, 'error': 'Token is required'}), 400

    _ensure_attendance_selection_runtime_tables()
    payload, error = _decode_signed_qr_token(token)
    if error:
        return jsonify({'success': False, 'error': error}), 400
    if int(payload.get('school_id', 0)) != int(school_id):
        return jsonify({'success': False, 'error': 'Token does not belong to this school'}), 400

    db = get_db()
    token_hash = _hash_secure_value(token)
    row = db.execute(
        '''
        SELECT id, status, expires_at
        FROM attendance_qr_tokens
        WHERE school_id = ? AND token_jti = ? AND token_hash = ?
        ORDER BY id DESC
        LIMIT 1
        ''',
        (school_id, payload.get('jti'), token_hash)
    ).fetchone()
    if not row:
        return jsonify({'success': False, 'error': 'Unknown QR token'}), 400
    if row['status'] != 'active':
        return jsonify({'success': False, 'error': 'QR token already used'}), 400

    attendance_result = _record_mode_attendance(session['user_id'], school_id, 'dynamic_qr')
    if not attendance_result['success']:
        return jsonify({'success': False, 'error': attendance_result['message']}), 400

    db.execute(
        'UPDATE attendance_qr_tokens SET status = ?, used_by_staff_id = ?, used_at = ? WHERE id = ?',
        ('used', session['user_id'], datetime.datetime.now(), row['id'])
    )
    db.commit()

    return jsonify({
        'success': True,
        'message': attendance_result['message'],
        'action': attendance_result['action'],
        'mode': 'dynamic_qr',
        'server_time': datetime.datetime.utcnow().isoformat() + 'Z',
        'client': request_payload.get('client', 'web')
    })


@app.route('/api/attendance/qr/render', methods=['GET'])
def attendance_qr_render():
    if 'user_id' not in session or session.get('user_type') not in ['admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'dynamic_qr':
        return jsonify({'success': False, 'error': f'Dynamic QR is disabled for this school (current mode: {mode}).'}), 400

    token = str(request.args.get('token', '')).strip()
    if not token:
        return jsonify({'success': False, 'error': 'token query parameter is required'}), 400
    if len(token) > 4096:
        return jsonify({'success': False, 'error': 'token is too long'}), 400

    svg_bytes = _build_qr_svg(token)
    response = make_response(svg_bytes)
    response.headers['Content-Type'] = 'image/svg+xml'
    response.headers['Cache-Control'] = 'no-store, max-age=0'
    return response


@csrf.exempt
@app.route('/api/attendance/idscan/punch', methods=['POST'])
def attendance_idscan_punch():
    if 'user_id' not in session or session.get('user_type') not in ['admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'id_scan':
        return jsonify({'success': False, 'error': f'ID Scan is disabled for this school (current mode: {mode}).'}), 400

    data = request.get_json(silent=True) or {}
    scan_value = str(data.get('scan_value', '')).strip()
    if not scan_value:
        return jsonify({'success': False, 'error': 'scan_value is required'}), 400

    db = get_db()
    staff_row = db.execute(
        '''
        SELECT id, staff_id, full_name
        FROM staff
        WHERE school_id = ? AND (staff_id = ? OR card_uid = ?)
        ORDER BY CASE WHEN staff_id = ? THEN 0 ELSE 1 END, id ASC
        LIMIT 1
        ''',
        (school_id, scan_value, scan_value, scan_value)
    ).fetchone()

    if not staff_row:
        return jsonify({'success': False, 'error': 'No matching staff for scanned value'}), 404

    attendance_result = _record_mode_attendance(staff_row['id'], school_id, 'id_scan')
    if not attendance_result['success']:
        return jsonify({'success': False, 'error': attendance_result['message']}), 400

    return jsonify({
        'success': True,
        'message': attendance_result['message'],
        'action': attendance_result['action'],
        'staff': {
            'id': staff_row['id'],
            'staff_id': staff_row['staff_id'],
            'full_name': staff_row['full_name']
        }
    })


@csrf.exempt
@app.route('/api/attendance/otp/request', methods=['POST'])
def attendance_otp_request():
    if 'user_id' not in session or session.get('user_type') != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'otp':
        return jsonify({'success': False, 'error': f'OTP mode is disabled for this school (current mode: {mode}).'}), 400

    data = request.get_json(silent=True) or {}
    channel = str(data.get('channel', 'email')).strip().lower()
    office_otp = str(data.get('office_otp', '')).strip()
    if channel not in ['email', 'sms']:
        return jsonify({'success': False, 'error': 'Invalid channel'}), 400
    if len(office_otp) != 4 or not office_otp.isdigit():
        return jsonify({'success': False, 'error': '4-digit office OTP is required'}), 400

    _ensure_attendance_selection_runtime_tables()
    if not _is_valid_office_otp_code(school_id, office_otp):
        return jsonify({'success': False, 'error': 'Invalid or expired office OTP. Please use latest OTP shown in admin panel.'}), 400

    provider_config = get_attendance_config('company_otp_provider_config', {})
    email_ready = _smtp_provider_ready(provider_config)
    sms_ready = _twilio_provider_ready(provider_config)

    db = get_db()
    staff_row = db.execute('SELECT email, phone, full_name FROM staff WHERE id = ?', (session['user_id'],)).fetchone()
    staff_email = str(staff_row['email']).strip() if staff_row and staff_row['email'] else ''
    staff_phone = str(staff_row['phone']).strip() if staff_row and staff_row['phone'] else ''

    requested_channel = channel
    if channel == 'email':
        if not email_ready or not staff_email:
            if sms_ready and staff_phone:
                channel = 'sms'
            else:
                if not email_ready:
                    return jsonify({
                        'success': False,
                        'error': 'Email OTP provider is not configured. Please set SMTP host, port, from email, and app password.'
                    }), 400
                return jsonify({'success': False, 'error': 'Staff email is not configured'}), 400
    else:
        if not sms_ready or not staff_phone:
            if email_ready and staff_email:
                channel = 'email'
            else:
                if not sms_ready:
                    return jsonify({
                        'success': False,
                        'error': 'SMS OTP provider is not configured. Please set Twilio SID, Auth Token, and From Number.'
                    }), 400
                return jsonify({'success': False, 'error': 'Staff phone number is not configured'}), 400

    otp_config = get_attendance_config(
        f'attendance_otp_config_school_{school_id}',
        {'ttl_seconds': 180, 'max_attempts': 3, 'channels': ['email', 'sms']}
    )
    ttl_seconds = max(60, min(600, int(otp_config.get('ttl_seconds', 180) or 180)))
    max_attempts = max(1, min(10, int(otp_config.get('max_attempts', 3) or 3)))

    otp_code = f'{secrets.randbelow(1000000):06d}'
    expires_at = datetime.datetime.now() + datetime.timedelta(seconds=ttl_seconds)

    db.execute(
        '''
        INSERT INTO attendance_otp_challenges
        (school_id, staff_id, channel, otp_hash, expires_at, attempts, max_attempts, status)
        VALUES (?, ?, ?, ?, ?, 0, ?, 'pending')
        ''',
        (school_id, session['user_id'], channel, _hash_secure_value(otp_code), expires_at, max_attempts)
    )
    db.commit()

    # Delivery implementation starts with email; SMS integration placeholder is tracked.
    delivery_message = 'OTP generated.'
    if channel == 'email':
        email_result = _send_smtp_email(
            provider_config,
            staff_email,
            'Attendance OTP',
            f'Hello {staff_row["full_name"]}, your attendance OTP is {otp_code}. It expires in {ttl_seconds} seconds.'
        )
        if not email_result.get('success'):
            return jsonify({'success': False, 'error': email_result.get('error', 'Unable to send OTP email')}), 500
        delivery_message = 'OTP sent to your email.'
    else:
        sms_body = f'Your attendance OTP is {otp_code}. It expires in {ttl_seconds} seconds.'
        sms_result = _send_twilio_sms(provider_config, staff_phone, sms_body)
        if not sms_result.get('success'):
            return jsonify({'success': False, 'error': sms_result.get('error', 'Unable to send SMS OTP')}), 500
        delivery_message = 'OTP sent to your mobile number.'

    if requested_channel != channel:
        delivery_message = f'{delivery_message} (Auto-switched from {requested_channel} to {channel} due to availability.)'

    return jsonify({'success': True, 'message': delivery_message, 'expires_in_seconds': ttl_seconds})


@csrf.exempt
@app.route('/api/attendance/otp/verify', methods=['POST'])
def attendance_otp_verify():
    if 'user_id' not in session or session.get('user_type') != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'otp':
        return jsonify({'success': False, 'error': f'OTP mode is disabled for this school (current mode: {mode}).'}), 400

    data = request.get_json(silent=True) or {}
    otp_code = str(data.get('otp', '')).strip()
    if not otp_code:
        return jsonify({'success': False, 'error': 'OTP is required'}), 400

    _ensure_attendance_selection_runtime_tables()
    db = get_db()
    challenge = db.execute(
        '''
        SELECT *
        FROM attendance_otp_challenges
        WHERE school_id = ? AND staff_id = ? AND status = 'pending'
        ORDER BY id DESC
        LIMIT 1
        ''',
        (school_id, session['user_id'])
    ).fetchone()

    if not challenge:
        return jsonify({'success': False, 'error': 'No active OTP challenge found'}), 404

    expires_at = challenge['expires_at']
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.datetime.fromisoformat(expires_at.replace('Z', ''))
        except ValueError:
            try:
                expires_at = datetime.datetime.strptime(expires_at, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                expires_at = datetime.datetime.now() - datetime.timedelta(seconds=1)

    if expires_at and expires_at < datetime.datetime.now():
        db.execute('UPDATE attendance_otp_challenges SET status = ? WHERE id = ?', ('expired', challenge['id']))
        db.commit()
        return jsonify({'success': False, 'error': 'OTP has expired'}), 400
    if challenge['attempts'] >= challenge['max_attempts']:
        db.execute('UPDATE attendance_otp_challenges SET status = ? WHERE id = ?', ('locked', challenge['id']))
        db.commit()
        return jsonify({'success': False, 'error': 'Maximum OTP attempts exceeded'}), 400

    if _hash_secure_value(otp_code) != challenge['otp_hash']:
        db.execute(
            'UPDATE attendance_otp_challenges SET attempts = attempts + 1 WHERE id = ?',
            (challenge['id'],)
        )
        db.commit()
        return jsonify({'success': False, 'error': 'Invalid OTP'}), 400

    attendance_result = _record_mode_attendance(session['user_id'], school_id, 'otp')
    if not attendance_result['success']:
        return jsonify({'success': False, 'error': attendance_result['message']}), 400

    db.execute(
        'UPDATE attendance_otp_challenges SET status = ?, verified_at = ? WHERE id = ?',
        ('verified', datetime.datetime.now(), challenge['id'])
    )
    db.commit()

    return jsonify({'success': True, 'message': attendance_result['message'], 'action': attendance_result['action']})


@app.route('/admin/attendance/dynamic-qr')
def admin_dynamic_qr_screen():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'dynamic_qr':
        flash(f'Dynamic QR screen is unavailable because attendance mode is {mode}.', 'warning')
        return redirect(url_for('admin_dashboard'))

    qr_config = get_attendance_config(
        f'attendance_qr_config_school_{school_id}',
        {'rotation_seconds': 30, 'screen_name': 'Main Attendance Display'}
    )
    module_enabled = get_module_enabled(school_id) if school_id else {}
    return render_template(
        'admin_dynamic_qr_attendance.html',
        school_id=school_id,
        module_enabled=module_enabled,
        qr_config=qr_config,
        rotation_seconds=max(30, min(120, int(qr_config.get('rotation_seconds', 30) or 30)))
    )


@csrf.exempt
@app.route('/api/attendance/otp/office-issue', methods=['POST'])
def attendance_office_otp_issue():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'otp':
        return jsonify({'success': False, 'error': f'OTP mode is disabled for this school (current mode: {mode}).'}), 400

    _ensure_attendance_selection_runtime_tables()
    issued = _issue_office_otp_code(school_id, 30)
    return jsonify({
        'success': True,
        'office_otp': issued['otp_code'],
        'expires_at': issued['expires_at'].isoformat(),
        'ttl_seconds': issued['ttl_seconds']
    })


@app.route('/api/attendance/otp/office-status', methods=['GET'])
def attendance_office_otp_status():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    if session.get('user_type') not in ['admin', 'staff'] and not session.get('is_sub_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'otp':
        return jsonify({'success': False, 'error': f'OTP mode is disabled for this school (current mode: {mode}).'}), 400

    _ensure_attendance_selection_runtime_tables()
    db = get_db()
    row = db.execute(
        '''
        SELECT id, expires_at, status
        FROM attendance_office_otp_codes
        WHERE school_id = ? AND status = 'active'
        ORDER BY id DESC
        LIMIT 1
        ''',
        (school_id,)
    ).fetchone()

    if not row:
        return jsonify({'success': False, 'error': 'No active office OTP. Ask admin to open OTP display.'}), 404

    expires_at = row['expires_at']
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.datetime.fromisoformat(expires_at.replace('Z', ''))
        except ValueError:
            try:
                expires_at = datetime.datetime.strptime(expires_at, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                expires_at = datetime.datetime.now() - datetime.timedelta(seconds=1)

    now = datetime.datetime.now()
    if expires_at and expires_at < now:
        db.execute('UPDATE attendance_office_otp_codes SET status = ? WHERE id = ?', ('expired', row['id']))
        db.commit()
        return jsonify({'success': False, 'error': 'Office OTP expired. Waiting for next refresh.'}), 404

    remaining_seconds = int(max(0, (expires_at - now).total_seconds())) if expires_at else 0
    return jsonify({
        'success': True,
        'expires_at': expires_at.isoformat() if expires_at else None,
        'remaining_seconds': remaining_seconds
    })


@app.route('/admin/attendance/otp-display')
def admin_otp_display_screen():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'otp':
        flash(f'Office OTP screen is unavailable because attendance mode is {mode}.', 'warning')
        return redirect(url_for('admin_dashboard'))

    module_enabled = get_module_enabled(school_id) if school_id else {}
    return render_template(
        'admin_otp_display.html',
        school_id=school_id,
        module_enabled=module_enabled,
        rotation_seconds=30
    )


@csrf.exempt
@app.route('/api/admin/dynamic-qr/verify-password', methods=['POST'])
def verify_dynamic_qr_exit_password():
    """Require password re-auth before leaving Dynamic QR admin screen."""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or {}
    password = str(payload.get('password', '')).strip()
    if not password:
        return jsonify({'success': False, 'error': 'Password is required'}), 400

    db = get_db()
    user_id = session.get('user_id')
    school_id = session.get('school_id')

    # School admin authentication.
    if session.get('user_type') == 'admin':
        admin = db.execute(
            'SELECT password FROM admins WHERE id = ? AND school_id = ? LIMIT 1',
            (user_id, school_id)
        ).fetchone()
        if not admin or not check_password_hash(admin['password'], password):
            return jsonify({'success': False, 'error': 'Invalid admin password'}), 401
        return jsonify({'success': True})

    # Sub-admin (staff) authentication fallback.
    staff = db.execute(
        'SELECT password_hash, password FROM staff WHERE id = ? AND school_id = ? LIMIT 1',
        (user_id, school_id)
    ).fetchone()
    if not staff:
        return jsonify({'success': False, 'error': 'Staff account not found'}), 404

    password_hash = ''
    if 'password_hash' in staff.keys() and staff['password_hash']:
        password_hash = staff['password_hash']
    elif 'password' in staff.keys() and staff['password']:
        password_hash = staff['password']

    if not password_hash or not check_password_hash(password_hash, password):
        return jsonify({'success': False, 'error': 'Invalid admin password'}), 401

    return jsonify({'success': True})


@app.route('/admin/attendance/id-scan-kiosk')
def admin_id_scan_kiosk():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'id_scan':
        flash(f'ID Scan kiosk is unavailable because attendance mode is {mode}.', 'warning')
        return redirect(url_for('admin_dashboard'))

    id_scan_config = get_attendance_config(
        f'attendance_idscan_config_school_{school_id}',
        {'kiosk_label': 'ID Scan Kiosk', 'allow_continuous_scan': True}
    )
    module_enabled = get_module_enabled(school_id) if school_id else {}
    return render_template(
        'admin_id_scan_kiosk.html',
        school_id=school_id,
        module_enabled=module_enabled,
        id_scan_config=id_scan_config
    )


@app.route('/staff/attendance/qr-scanner')
def staff_qr_scanner_page():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'dynamic_qr':
        flash(f'QR scanner is unavailable because attendance mode is {mode}.', 'warning')
        return redirect(url_for('staff_dashboard'))

    module_enabled = get_module_enabled(school_id) if school_id else {}
    return render_template('staff_qr_scanner.html', module_enabled=module_enabled, school_id=school_id)


@app.route('/staff/attendance/otp')
def staff_otp_attendance_page():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    school_id = session.get('school_id')
    mode = get_school_attendance_mode(school_id)
    if mode != 'otp':
        flash(f'OTP attendance is unavailable because attendance mode is {mode}.', 'warning')
        return redirect(url_for('staff_dashboard'))

    module_enabled = get_module_enabled(school_id) if school_id else {}
    otp_config = get_attendance_config(
        f'attendance_otp_config_school_{school_id}',
        {'ttl_seconds': 180, 'max_attempts': 3, 'channels': ['email', 'sms']}
    )
    provider_config = get_attendance_config('company_otp_provider_config', {})

    available_channels = []
    if _as_bool(provider_config.get('smtp_enabled')):
        available_channels.append('email')
    if _as_bool(provider_config.get('twilio_enabled')):
        available_channels.append('sms')

    # Enforce per-school channel selection as well.
    configured_channels = otp_config.get('channels')
    if not isinstance(configured_channels, list) or not configured_channels:
        configured_channels = ['email', 'sms']

    available_channels = [ch for ch in available_channels if ch in configured_channels]

    if not available_channels:
        # Final fallback: keep UI usable and let request endpoint return precise
        # provider errors (e.g. SMTP not enabled) if channel is unavailable.
        available_channels = [ch for ch in configured_channels if ch in ('email', 'sms')]

    if not available_channels:
        available_channels = ['email', 'sms']

    return render_template(
        'staff_otp_attendance.html',
        module_enabled=module_enabled,
        school_id=school_id,
        otp_config=otp_config,
        available_channels=available_channels
    )

@app.route('/get_attendance_summary')
def get_attendance_summary():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    db = get_db()

    if session['user_type'] == 'staff':
        staff_id = session['user_id']
        today = datetime.date.today()

        # Get accurate attendance summary for current month using the new logic
        attendance_summary = _calculate_accurate_attendance_summary(staff_id, today.year, today.month)

        return jsonify({
            'success': True,
            'present': attendance_summary['present_days'],
            'absent': attendance_summary['absent_days'],
            'late': attendance_summary['late_days'],
            'leave': attendance_summary['leave_days'],
            'on_duty': attendance_summary['on_duty_days'],
            'working_days': attendance_summary['working_days']
        })

    return jsonify({'success': False, 'error': 'Unauthorized'})

@app.route('/get_staff_details')
def get_staff_details():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.args.get('id')
    db = get_db()

    staff = db.execute('SELECT * FROM staff WHERE id = ?', (staff_id,)).fetchone()
    if not staff:
        return jsonify({'success': False, 'error': 'Staff not found'})

    # Get attendance records
    attendance = db.execute('''
        SELECT date, time_in, time_out, status
        FROM attendance
        WHERE staff_id = ?
        ORDER BY date DESC
    ''', (staff_id,)).fetchall()

    return jsonify({
        'success': True,
        'staff': dict(staff),
        'attendance': [dict(a) for a in attendance]
    })


@app.route('/export_staff_data')
def export_staff_data():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

@app.route('/add_admin', methods=['POST'])
def add_admin():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 401

    # Get form data
    school_id = request.form.get('school_id')
    username = request.form.get('username')
    password = request.form.get('password')
    full_name = request.form.get('full_name')
    email = request.form.get('email')

    # Validate required fields
    if not school_id:
        return jsonify({'success': False, 'error': 'School ID is required'}), 400
    
    if not username or not username.strip():
        return jsonify({'success': False, 'error': 'Username is required'}), 400
    
    if not password or not password.strip():
        return jsonify({'success': False, 'error': 'Password is required'}), 400
        
    if not full_name or not full_name.strip():
        return jsonify({'success': False, 'error': 'Full name is required'}), 400

    # Validate school exists
    db = get_db()
    school = db.execute('SELECT id FROM schools WHERE id = ?', (school_id,)).fetchone()
    if not school:
        return jsonify({'success': False, 'error': 'Invalid school ID'}), 400

    # Clean data
    username = username.strip()
    full_name = full_name.strip()
    email = email.strip() if email else None
    
    # Hash password
    password_hash = generate_password_hash(password)

    try:
        db.execute('''
            INSERT INTO admins (school_id, username, password, full_name, email)
            VALUES (?, ?, ?, ?, ?)
        ''', (school_id, username, password_hash, full_name, email))
        db.commit()
        return jsonify({'success': True, 'message': 'Administrator added successfully'})
    except sqlite3.IntegrityError as e:
        if 'username' in str(e).lower():
            return jsonify({'success': False, 'error': 'Username already exists'}), 400
        else:
            return jsonify({'success': False, 'error': 'Database constraint error'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': 'An unexpected error occurred'}), 500

@app.route('/delete_admin', methods=['POST'])
def delete_admin():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 401

    admin_id = request.form.get('admin_id')
    
    if not admin_id:
        return jsonify({'success': False, 'error': 'Admin ID is required'}), 400

    db = get_db()

    try:
        # Check if admin exists
        admin = db.execute('SELECT id FROM admins WHERE id = ?', (admin_id,)).fetchone()
        if not admin:
            return jsonify({'success': False, 'error': 'Administrator not found'}), 404

        db.execute('DELETE FROM admins WHERE id = ?', (admin_id,))
        db.commit()
        return jsonify({'success': True, 'message': 'Administrator deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': 'An error occurred while deleting administrator'}), 500



@app.route('/get_biometric_verifications')
def get_biometric_verifications():
    """Get biometric verification history for staff"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    db = get_db()

    # Get verification history
    query = '''
        SELECT verification_type, verification_time, biometric_method,
               verification_status, notes, device_ip
        FROM biometric_verifications
        WHERE staff_id = ?
    '''
    params = [staff_id]

    if start_date and end_date:
        query += ' AND DATE(verification_time) BETWEEN ? AND ?'
        params.extend([start_date, end_date])

    query += ' ORDER BY verification_time DESC LIMIT 50'

    verifications = db.execute(query, params).fetchall()

    return jsonify({
        'success': True,
        'verifications': [dict(v) for v in verifications]
    })

@app.route('/get_today_attendance_status')
def get_today_attendance_status():
    """Get today's attendance status for staff"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session.get('school_id')
    today = datetime.date.today()

    db = get_db()

    # Get today's attendance
    attendance = db.execute('''
        SELECT
            MIN(time_in) AS time_in,
            MAX(time_out) AS time_out,
            MAX(
                CASE
                    WHEN LOWER(COALESCE(regularization_status, '')) = 'approved' THEN 3
                    WHEN LOWER(COALESCE(status, '')) = 'late' THEN 2
                    WHEN LOWER(COALESCE(status, '')) = 'present' THEN 1
                    ELSE 0
                END
            ) AS status_rank
        FROM attendance
        WHERE staff_id = ? AND school_id = ? AND date = ?
    ''', (staff_id, school_id, today)).fetchone()

    attendance_record = None
    total_hours = '--:--'
    if attendance and (attendance['time_in'] or attendance['time_out'] or (attendance['status_rank'] or 0) > 0):
        if (attendance['status_rank'] or 0) >= 2:
            if (attendance['status_rank'] or 0) >= 3:
                status_value = 'present'
            else:
                status_value = 'late'
        elif (attendance['status_rank'] or 0) == 1:
            status_value = 'present'
        else:
            status_value = 'present' if attendance['time_in'] else 'absent'

        attendance_record = {
            'time_in': attendance['time_in'],
            'time_out': attendance['time_out'],
            'status': status_value
        }

        if attendance['time_in'] and attendance['time_out']:
            try:
                start_dt = datetime.datetime.strptime(attendance['time_in'], '%H:%M:%S')
                end_dt = datetime.datetime.strptime(attendance['time_out'], '%H:%M:%S')
                if end_dt < start_dt:
                    end_dt += datetime.timedelta(days=1)
                diff = end_dt - start_dt
                total_minutes = int(diff.total_seconds() // 60)
                total_hours = f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
            except Exception:
                total_hours = '--:--'

    # Get today's verifications
    verifications = db.execute('''
        SELECT verification_type, verification_time, biometric_method, verification_status
        FROM biometric_verifications
        WHERE staff_id = ? AND DATE(verification_time) = ?
        ORDER BY verification_time DESC
    ''', (staff_id, today)).fetchall()

    # Determine available actions based on current status
    available_actions = []
    # Check-in and check-out are always available (can be updated multiple times)
    available_actions.append('check-in')
    available_actions.append('check-out')

    # Format attendance times to 12-hour format
    formatted_attendance = format_attendance_times_to_12hr(attendance_record) if attendance_record else None

    # Debug information
    debug_info = {
        'staff_id': staff_id,
        'school_id': school_id,
        'today': str(today),
        'attendance_found': attendance_record is not None,
        'raw_attendance': attendance_record
    }

    return jsonify({
        'success': True,
        'attendance': formatted_attendance,
        'total_hours': total_hours,
        'verifications': [dict(v) for v in verifications],
        'available_actions': available_actions,
        'debug': debug_info
    })

def get_staff_status_for_date(staff_id, date, school_id, db):
    """
    Get comprehensive staff status for a given date.
    Checks holidays first, then approved applications, then falls back to attendance records.
    
    Priority order:
    1. Holiday (institution-wide or department-specific) -> "Holiday"
    2. Approved Leave application -> "On Leave"
    3. Approved On Duty application -> "On Duty"  
    4. Approved Permission application -> "On Permission"
    5. Standard attendance status (Present/Absent/Late)
    
    Args:
        staff_id: Staff database ID
        date: Date to check (datetime.date)
        school_id: School ID
        db: Database connection
    
    Returns:
        str: Status string ("Holiday", "On Leave", "On Duty", "On Permission", "present", "absent", "late")
    """
    
    # Import holiday checking function
    from database import is_holiday
    
    # Get staff department for holiday checking
    staff_info = db.execute('''
        SELECT department FROM staff WHERE id = ?
    ''', (staff_id,)).fetchone()
    
    staff_department = staff_info['department'] if staff_info else None
    
    # Check if this date is a holiday
    if is_holiday(date, department=staff_department, school_id=school_id):
        return "Holiday"
    
    # Check for approved Leave application
    leave_app = db.execute('''
        SELECT id FROM leave_applications 
        WHERE staff_id = ? AND school_id = ? AND status = 'approved'
        AND ? BETWEEN start_date AND end_date
    ''', (staff_id, school_id, date.isoformat())).fetchone()
    
    if leave_app:
        return "On Leave"
    
    # Check for approved On Duty application
    on_duty_app = db.execute('''
        SELECT id FROM on_duty_applications 
        WHERE staff_id = ? AND school_id = ? AND status = 'approved'
        AND ? BETWEEN start_date AND end_date
    ''', (staff_id, school_id, date.isoformat())).fetchone()
    
    if on_duty_app:
        return "On Duty"
    
    # Check for approved Permission application (single day)
    permission_app = db.execute('''
        SELECT id FROM permission_applications 
        WHERE staff_id = ? AND school_id = ? AND status = 'approved'
        AND permission_date = ?
    ''', (staff_id, school_id, date.isoformat())).fetchone()
    
    if permission_app:
        return "On Permission"
    
    # No approved applications found, check attendance record
    attendance = db.execute('''
        SELECT
            MIN(time_in) AS time_in,
            MAX(time_out) AS time_out,
            MAX(
                CASE
                    WHEN LOWER(COALESCE(status, '')) = 'late' THEN 2
                    WHEN LOWER(COALESCE(status, '')) = 'present' THEN 1
                    ELSE 0
                END
            ) AS status_rank
        FROM attendance
        WHERE staff_id = ? AND school_id = ? AND date = ?
    ''', (staff_id, school_id, date.isoformat())).fetchone()
    
    if attendance and attendance['time_in']:
        # Get staff shift type for accurate status calculation
        staff_info = db.execute('''
            SELECT COALESCE(shift_type, 'general') as shift_type
            FROM staff WHERE id = ?
        ''', (staff_id,)).fetchone()
        
        shift_type = staff_info['shift_type'] if staff_info else 'general'
        
        # Parse check-in time and calculate real-time status
        try:
            time_in_str = attendance['time_in']
            check_in_time = datetime.datetime.strptime(time_in_str, '%H:%M:%S').time()
            
            # Calculate real-time status using ShiftManager
            from shift_management import ShiftManager
            shift_manager = ShiftManager()
            
            check_out_time = None
            if attendance['time_out']:
                check_out_time = datetime.datetime.strptime(attendance['time_out'], '%H:%M:%S').time()
            
            status_result = shift_manager.calculate_attendance_status(
                shift_type, check_in_time, check_out_time
            )
            
            # Map ShiftManager status to valid admin dashboard status
            calculated_status = status_result['status']
            
            # Ensure only valid statuses are returned for admin dashboard
            valid_statuses = ['present', 'late', 'absent']
            
            if calculated_status in valid_statuses:
                return calculated_status
            elif calculated_status == 'left_soon':
                # Early departure still counts as present for dashboard purposes
                return 'present'
            else:
                # Any other unexpected status defaults to present if they checked in
                print(f"Warning: Unexpected status '{calculated_status}' for staff {staff_id}, defaulting to 'present'")
                return 'present'
            
        except Exception as e:
            # Fall back to stored status if calculation fails
            print(f"Status calculation error for staff {staff_id}: {e}")
            if attendance and (attendance['status_rank'] or 0) >= 2:
                return 'late'
            if attendance and (attendance['status_rank'] or 0) == 1:
                return 'present'
            return "present"
    else:
        return "absent"  # Default status if no attendance record

@app.route('/get_realtime_attendance')
def get_realtime_attendance():
    """Get real-time attendance data for admin dashboard"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session.get('school_id', 1)
    today = datetime.date.today()

    db = get_db()

    staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]
    active_filter = ''
    if 'is_active' in staff_columns:
        active_filter = ' AND COALESCE(s.is_active, 1) = 1'
    elif 'status' in staff_columns:
        active_filter = " AND LOWER(COALESCE(s.status, 'active')) = 'active'"

    # Get all active staff for this school
    staff_query = f'''
         SELECT s.id as staff_id, s.staff_id as staff_number, s.full_name, s.department,
             s.shift_type, s.next_shift_type, s.next_shift_effective_date,
               a.time_in, a.time_out
        FROM staff s
        LEFT JOIN (
            SELECT staff_id,
                   MIN(time_in) AS time_in,
                   MAX(time_out) AS time_out
            FROM attendance
            WHERE date = ? AND school_id = ?
            GROUP BY staff_id
        ) a ON s.id = a.staff_id
        WHERE s.school_id = ?{active_filter}
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    '''
    staff_list = db.execute(staff_query, (today, school_id, school_id)).fetchall()

    # Build attendance data with correct status logic
    attendance_data = []
    status_counts = {
        'total_staff': 0,
        'present': 0,
        'absent': 0,
        'late': 0,
        'on_leave': 0,
        'on_duty': 0,
        'on_permission': 0,
        'holiday': 0
    }
    
    for staff in staff_list:
        # Get comprehensive status using the new logic
        status = get_staff_status_for_date(staff['staff_id'], today, school_id, db)
        
        # Build staff record
        staff_record = {
            'staff_id': staff['staff_id'],
            'staff_number': staff['staff_number'],
            'full_name': staff['full_name'],
            'department': staff['department'],
            'time_in': staff['time_in'],
            'time_out': staff['time_out'],
            'status': status,
            'shift_type': resolve_effective_shift_type(staff, today)
        }
        
        # Format times to 12-hour format
        formatted_staff = format_attendance_times_to_12hr(staff_record)
        attendance_data.append(formatted_staff)
        
        # Count statuses for summary
        status_counts['total_staff'] += 1
        
        if status == 'Holiday':
            status_counts['holiday'] += 1
        elif status == 'On Leave':
            status_counts['on_leave'] += 1
        elif status == 'On Duty':
            status_counts['on_duty'] += 1
        elif status == 'On Permission':
            status_counts['on_permission'] += 1
        elif status == 'present':
            status_counts['present'] += 1
        elif status == 'late':
            status_counts['late'] += 1
        else:  # absent or any other status
            status_counts['absent'] += 1

    return jsonify({
        'success': True,
        'attendance_data': attendance_data,
        'summary': status_counts,
        'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    })

@app.route('/export_company_report')
def export_company_report():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_company_report(start_date, end_date)

@app.route('/export_staff_report')
def export_staff_report():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_individual_staff_report(staff_id, start_date, end_date)

@app.route('/export_monthly_report')
def export_monthly_report():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        return jsonify({'success': False, 'error': 'Year and month are required'})

    if month < 1 or month > 12:
        return jsonify({'success': False, 'error': 'Invalid month. Must be between 1 and 12'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_monthly_report(school_id, year, month)

@app.route('/export_department_report')
def export_department_report():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department = request.args.get('department')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report (using staff attendance report filtered by department)
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

@app.route('/export_yearly_report')
def export_yearly_report():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    year = request.args.get('year', type=int)

    if not year:
        return jsonify({'success': False, 'error': 'Year is required'})

    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

# Enhanced Staff Management Routes
@app.route('/bulk_import_staff', methods=['POST'])
def bulk_import_staff():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})

    # Save uploaded file temporarily
    filename = secure_filename(file.filename)
    temp_path = os.path.join('temp', filename)
    os.makedirs('temp', exist_ok=True)
    file.save(temp_path)

    try:
        staff_manager = StaffManager()
        result = staff_manager.bulk_import_staff(temp_path, school_id)
        return jsonify(result)
    finally:
        # Clean up temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)

@app.route('/advanced_search_staff')
def advanced_search_staff():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']

    filters = {
        'search_term': request.args.get('search_term'),
        'department': request.args.get('department'),
        'position': request.args.get('position'),
        'gender': request.args.get('gender'),
        'date_from': request.args.get('date_from'),
        'date_to': request.args.get('date_to'),
        'limit': request.args.get('limit')  # Remove default limit to match server-side query
    }

    # Remove None values
    filters = {k: v for k, v in filters.items() if v}

    staff_manager = StaffManager()
    staff_list = staff_manager.advanced_search_staff(school_id, filters)

    return jsonify({'success': True, 'staff': staff_list})

@app.route('/upload_staff_photo', methods=['POST'])
def upload_staff_photo():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    if not staff_id:
        return jsonify({'success': False, 'error': 'Staff ID required'})

    if 'photo' not in request.files:
        return jsonify({'success': False, 'error': 'No photo uploaded'})

    photo_file = request.files['photo']
    if photo_file.filename == '':
        return jsonify({'success': False, 'error': 'No photo selected'})

    staff_manager = StaffManager()
    result = staff_manager.manage_staff_photo(staff_id, photo_file)

    return jsonify(result)

@app.route('/get_department_analytics')
def get_department_analytics():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    staff_manager = StaffManager()
    analytics = staff_manager.get_department_analytics(school_id)

    return jsonify({'success': True, 'analytics': analytics})

@app.route('/generate_staff_id')
def generate_staff_id():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.args.get('department')

    staff_manager = StaffManager()
    staff_id = staff_manager.generate_staff_id(school_id, department)

    return jsonify({'success': True, 'staff_id': staff_id})

@app.route('/bulk_update_staff', methods=['POST'])
def bulk_update_staff():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    updates = request.json.get('updates', [])

    if not updates:
        return jsonify({'success': False, 'error': 'No updates provided'})

    staff_manager = StaffManager()
    result = staff_manager.bulk_update_staff(updates, school_id)

    return jsonify(result)

@app.route('/staff_management_enhanced')
def staff_management_enhanced():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    return render_template('staff_management_enhanced.html', today=datetime.datetime.now().strftime('%Y-%m-%d'))

@app.route('/download_staff_template')
def download_staff_template():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Create a sample Excel template
    import pandas as pd
    from io import BytesIO

    # Sample data for template: includes all Add Staff form fields
    template_data = {
        'staff_id': ['EMP001', 'EMP002'],
        'first_name': ['John', 'Jane'],
        'last_name': ['Doe', 'Smith'],
        'date_of_birth': ['1990-01-01', '1988-06-20'],
        'date_of_joining': ['2023-01-01', '2023-02-01'],
        'department': ['IT', 'HR'],
        'destination': ['Developer', 'HR Manager'],
        'gender': ['Male', 'Female'],
        'phone': ['9876543210', '9123456780'],
        'email': ['john.doe@example.com', 'jane.smith@example.com'],
        'shift_type': ['general', 'general'],
        'password': ['password123', 'password123'],
        'basic_salary': [35000, 42000],
        'hra': [8000, 10000],
        'transport_allowance': [2000, 2500],
        'other_allowances': [1000, 1200],
        'pf_deduction': [1800, 2000],
        'esi_deduction': [750, 850],
        'professional_tax': [200, 200],
        'other_deductions': [0, 0],
        'bank_account_name': ['John Doe', 'Jane Smith'],
        'bank_name': ['State Bank of India', 'HDFC Bank'],
        'bank_account_number': ['123456789012', '987654321098'],
        'ifsc_code': ['SBIN0001234', 'HDFC0001234'],
        'pan_number': ['ABCDE1234F', 'PQRSX6789Z']
    }

    df = pd.DataFrame(template_data)

    instructions_data = {
        'Field': [
            'staff_id', 'first_name', 'last_name', 'date_of_birth', 'date_of_joining',
            'department', 'destination', 'gender', 'phone', 'email', 'shift_type',
            'password', 'basic_salary', 'hra', 'transport_allowance', 'other_allowances',
            'pf_deduction', 'esi_deduction', 'professional_tax', 'other_deductions',
            'bank_account_name', 'bank_name', 'bank_account_number', 'ifsc_code', 'pan_number'
        ],
        'Required': [
            'Yes', 'Yes', 'Yes', 'No', 'No',
            'No', 'Yes', 'No', 'No', 'No', 'No',
            'Yes', 'Yes', 'No', 'No', 'No',
            'No', 'No', 'No', 'No',
            'No', 'No', 'No', 'No', 'No'
        ],
        'Format / Notes': [
            'Unique staff code per school',
            'Text',
            'Text',
            'YYYY-MM-DD',
            'YYYY-MM-DD',
            'Department name',
            'Position / role',
            'Male/Female/Other',
            'Phone number',
            'Email address',
            'general/morning/afternoon/evening/night/overtime',
            'Required login password for staff account',
            'Required numeric amount',
            'Numeric amount',
            'Numeric amount',
            'Numeric amount',
            'Numeric amount',
            'Numeric amount',
            'Numeric amount',
            'Numeric amount',
            'Name as per bank account',
            'Bank name',
            'Account number',
            '11-char IFSC code',
            '10-char PAN'
        ]
    }
    instructions_df = pd.DataFrame(instructions_data)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Staff Template', index=False)
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=staff_import_template.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# Advanced Attendance Management Routes
@app.route('/process_attendance_advanced', methods=['POST'])
def process_attendance_advanced():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    verification_type = request.form.get('verification_type')  # check-in, check-out, overtime-in, overtime-out
    timestamp_str = request.form.get('timestamp')

    if not all([staff_id, verification_type, timestamp_str]):
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    try:
        timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
        school_id = session['school_id']

        attendance_manager = AdvancedAttendanceManager()
        result = attendance_manager.process_attendance_with_overtime(
            int(staff_id), school_id, verification_type, timestamp
        )

        return jsonify(result)

    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid timestamp format'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/manage_shifts')
def manage_shifts():
    # Redirect to shift management
    return redirect(url_for('shift_management'))

@app.route('/admin/work_time_assignment')
def work_time_assignment_redirect():
    # Backward-compat redirect — use /admin/shift_management going forward
    return redirect(url_for('shift_management'))

@app.route('/admin/update_shift', methods=['POST'])
def update_shift():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
        
    shift_id = request.form.get('shift_id')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    grace_period = request.form.get('grace_period_minutes')
    description = request.form.get('description')
    
    if not all([shift_id, start_time, end_time]):
        return jsonify({'success': False, 'error': 'Missing required fields'})
        
    try:
        db = get_db()
        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'error': 'School context is required'})
        db.execute('''
            UPDATE shift_definitions 
            SET start_time = ?, end_time = ?, grace_period_minutes = ?, description = ?
            WHERE id = ? AND school_id = ?
        ''', (start_time, end_time, grace_period, description, shift_id, school_id))
        db.commit()
        
        # Reload shift manager
        if hasattr(app, 'shift_manager'):
            app.shift_manager.reload_shift_definitions()
            
        return jsonify({'success': True, 'message': 'Shift updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/add_shift', methods=['POST'])
def add_shift():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    shift_type_raw = (request.form.get('shift_type') or '').strip().lower()
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    grace_period = request.form.get('grace_period_minutes')
    description = (request.form.get('description') or '').strip()

    if not all([shift_type_raw, start_time, end_time]):
        return jsonify({'success': False, 'error': 'Missing required fields'})

    import re
    shift_type = re.sub(r'\s+', '_', shift_type_raw)
    if not re.fullmatch(r'[a-z0-9_\-]{2,30}', shift_type):
        return jsonify({'success': False, 'error': 'Shift name must be 2-30 characters and use only letters, numbers, spaces, hyphen, or underscore'})

    try:
        grace_period_value = int(grace_period) if grace_period not in (None, '') else 10
    except ValueError:
        return jsonify({'success': False, 'error': 'Grace period must be a valid number'})

    if grace_period_value < 0 or grace_period_value > 120:
        return jsonify({'success': False, 'error': 'Grace period must be between 0 and 120 minutes'})

    try:
        db = get_db()
        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'error': 'School context is required'})
        existing = db.execute('''
            SELECT id FROM shift_definitions
            WHERE school_id = ? AND LOWER(shift_type) = LOWER(?)
        ''', (school_id, shift_type)).fetchone()
        if existing:
            return jsonify({'success': False, 'error': 'A shift with this name already exists'})

        db.execute('''
            INSERT INTO shift_definitions (school_id, shift_type, start_time, end_time, grace_period_minutes, description, is_active)
            VALUES (?, ?, ?, ?, ?, ?, 1)
        ''', (school_id, shift_type, start_time, end_time, grace_period_value, description))
        db.commit()

        # Reload shift manager so custom shift can be used immediately
        if hasattr(app, 'shift_manager'):
            app.shift_manager.reload_shift_definitions()

        return jsonify({'success': True, 'message': 'Shift added successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete_shift', methods=['POST'])
def delete_shift():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    if session.get('is_sub_admin'):
        shift_permissions = session.get('permissions', {}).get('Shift Management', {})
        if not shift_permissions.get('delete'):
            return jsonify({'success': False, 'error': 'You do not have permission to delete shifts'})

    shift_id = request.form.get('shift_id')
    if not shift_id:
        return jsonify({'success': False, 'error': 'Shift ID is required'})

    try:
        db = get_db()
        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'error': 'School context is required'})

        shift_row = db.execute('''
            SELECT id, shift_type FROM shift_definitions
            WHERE id = ? AND school_id = ? AND is_active = 1
        ''', (shift_id, school_id)).fetchone()

        if not shift_row:
            return jsonify({'success': False, 'error': 'Shift not found or already deleted'})

        shift_type = (shift_row['shift_type'] or '').strip().lower()
        if shift_type == 'general':
            return jsonify({'success': False, 'error': 'Default General shift cannot be deleted'})

        assigned_staff = db.execute('''
            SELECT COUNT(*) AS cnt
            FROM staff
            WHERE school_id = ? AND LOWER(COALESCE(shift_type, 'general')) = LOWER(?)
        ''', (school_id, shift_type)).fetchone()['cnt']

        if assigned_staff > 0:
            return jsonify({
                'success': False,
                'error': f'Cannot delete shift. {assigned_staff} staff member(s) are currently assigned to this shift.'
            })

        assigned_departments = db.execute('''
            SELECT COUNT(*) AS cnt
            FROM department_shift_mappings
            WHERE school_id = ? AND LOWER(default_shift_type) = LOWER(?)
        ''', (school_id, shift_type)).fetchone()['cnt']

        if assigned_departments > 0:
            return jsonify({
                'success': False,
                'error': f'Cannot delete shift. {assigned_departments} department mapping(s) use this shift.'
            })

        db.execute('''
            UPDATE shift_definitions
            SET is_active = 0
            WHERE id = ? AND school_id = ?
        ''', (shift_id, school_id))
        db.commit()

        if hasattr(app, 'shift_manager'):
            app.shift_manager.reload_shift_definitions()

        return jsonify({'success': True, 'message': 'Shift deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_overtime_summary')
def get_overtime_summary():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Staff can view their own overtime, admin can view any staff's overtime
    if session['user_type'] == 'staff':
        staff_id = session['user_id']
    else:
        staff_id = request.args.get('staff_id')
        if not staff_id:
            return jsonify({'success': False, 'error': 'Staff ID required'})

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range required'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.get_overtime_summary(int(staff_id), start_date, end_date)

    return jsonify(result)

@app.route('/create_regularization_request', methods=['POST'])
def create_regularization_request():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Prefer login staff_id (e.g., ST001) over internal DB id as requested.
    staff_identifier = (session.get('staff_id') or session.get('user_id'))
    school_id = session['school_id']
    attendance_id_raw = (request.form.get('attendance_id') or '').strip()
    request_type = request.form.get('request_type')  # late_arrival, early_departure
    original_time = request.form.get('original_time')
    expected_time = request.form.get('expected_time')
    reason = (request.form.get('reason') or '').strip()

    if not all([attendance_id_raw, request_type, original_time, expected_time, reason]):
        return jsonify({'success': False, 'error': 'All fields are required'})

    if request_type not in ['late_arrival', 'early_departure']:
        return jsonify({'success': False, 'error': 'Invalid request type'})

    try:
        # Be tolerant to accidental non-numeric UI value wrappers.
        attendance_match = re.search(r'\d+', attendance_id_raw)
        if not attendance_match:
            raise ValueError('No numeric attendance id')
        attendance_id = int(attendance_match.group())
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid attendance record'})

    if len(reason) < 5:
        return jsonify({'success': False, 'error': 'Please provide a brief reason (minimum 5 characters)'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.create_attendance_regularization_request(
        staff_identifier, school_id, attendance_id, request_type,
        original_time, expected_time, reason
    )

    return jsonify(result)

@app.route('/process_regularization_request', methods=['POST'])
def process_regularization_request():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    request_id = request.form.get('request_id')
    decision = request.form.get('decision')  # approve, reject
    admin_reason = (request.form.get('admin_reason') or '').strip()
    admin_id = session['user_id']

    if not all([request_id, decision]):
        return jsonify({'success': False, 'error': 'Request ID and decision are required'})

    if decision not in ['approve', 'reject']:
        return jsonify({'success': False, 'error': 'Invalid decision'})

    if not admin_reason:
        return jsonify({'success': False, 'error': 'Admin reason is required'})

    try:
        request_id = int(request_id)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid request ID'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.process_regularization_request(
        request_id, admin_id, decision, admin_reason
    )

    return jsonify(result)

@app.route('/get_regularization_requests')
def get_regularization_requests():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.get_pending_regularization_requests(session['school_id'])
    return jsonify(result)

@app.route('/get_attendance_analytics')
def get_attendance_analytics():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range required'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.get_attendance_analytics(school_id, start_date, end_date)

    return jsonify(result)

@app.route('/auto_mark_leave_attendance', methods=['POST'])
def auto_mark_leave_attendance():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    date = request.form.get('date', datetime.datetime.now().strftime('%Y-%m-%d'))

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.auto_mark_leave_attendance(school_id, date)

    return jsonify(result)

# Reporting Dashboard Routes
@app.route('/reporting_dashboard')
def reporting_dashboard():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    today = datetime.datetime.now()
    return render_template('reporting_dashboard.html',
                         today=today.strftime('%Y-%m-%d'),
                         current_month=today.strftime('%Y-%m'),
                         current_year=today.year)

@app.route('/salary_management')
@requires_permission('salary_management', 'view')
def salary_management():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('staff_dashboard'))
    
    school_id = session.get('school_id')
    today = datetime.datetime.now()
    
    # Get module settings for navigation
    module_enabled = get_module_enabled(school_id) if school_id else {}
    
    return render_template('salary_management.html', current_year=today.year, module_enabled=module_enabled)

@app.route('/payroll_processing_review')
@requires_permission('salary_management', 'view')
def payroll_processing_review():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('staff_dashboard'))

    school_id = session.get('school_id')
    today = datetime.datetime.now()
    today_str = today.strftime('%Y-%m-%d')
    db = get_db()

    module_enabled = get_module_enabled(school_id) if school_id else {}
    last_calc_meta = session.get('last_salary_calc_meta', {}) if isinstance(session.get('last_salary_calc_meta', {}), dict) else {}

    # Use explicit query filters when provided; otherwise reuse latest salary calculation filters.
    has_explicit_filters = any(request.args.get(k) for k in ['year', 'month', 'department', 'shift', 'gender', 'search'])
    last_filters = session.get('last_salary_calc_filters', {}) if isinstance(session.get('last_salary_calc_filters', {}), dict) else {}

    selected_year = request.args.get('year', type=int) if has_explicit_filters else last_filters.get('year', today.year)
    selected_month = request.args.get('month', type=int) if has_explicit_filters else last_filters.get('month', today.month)
    selected_department = (request.args.get('department') or '').strip() if has_explicit_filters else (last_filters.get('department') or '').strip()
    selected_shift = (request.args.get('shift') or '').strip() if has_explicit_filters else (last_filters.get('shift') or '').strip()
    selected_gender = (request.args.get('gender') or '').strip() if has_explicit_filters else (last_filters.get('gender') or '').strip()
    search_term = (request.args.get('search') or '').strip() if has_explicit_filters else (last_filters.get('search') or '').strip()

    if not selected_year:
        selected_year = today.year
    if not selected_month:
        selected_month = today.month

    calculated_by = (last_calc_meta.get('calculated_by') or '').strip()
    calculated_at_display = ''
    calculated_at_raw = last_calc_meta.get('calculated_at')
    if calculated_at_raw:
        try:
            calculated_at_dt = datetime.datetime.fromisoformat(str(calculated_at_raw))
            calculated_at_display = f"{calculated_at_dt.strftime('%d %b %Y, %I:%M:%S')}.{calculated_at_dt.microsecond // 1000:03d} {calculated_at_dt.strftime('%p')}"
        except Exception:
            calculated_at_display = str(calculated_at_raw)

    is_selected_month_calculated = False
    if school_id and selected_year and selected_month:
        try:
            db.execute('''
                CREATE TABLE IF NOT EXISTS payroll_calculation_runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    school_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    calculated_at TIMESTAMP NOT NULL,
                    calculated_by TEXT,
                    total_staff INTEGER NOT NULL DEFAULT 0,
                    total_earnings REAL NOT NULL DEFAULT 0,
                    total_deductions REAL NOT NULL DEFAULT 0,
                    total_net_salary REAL NOT NULL DEFAULT 0,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(school_id, year, month)
                )
            ''')

            month_run = db.execute('''
                SELECT calculated_at, calculated_by
                FROM payroll_calculation_runs
                WHERE school_id = ? AND year = ? AND month = ?
            ''', (school_id, selected_year, selected_month)).fetchone()

            is_selected_month_calculated = bool(month_run)

            if month_run:
                if not calculated_by:
                    calculated_by = (month_run['calculated_by'] or '').strip()
                if not calculated_at_display and month_run['calculated_at']:
                    try:
                        month_calculated_at_dt = datetime.datetime.fromisoformat(str(month_run['calculated_at']))
                        calculated_at_display = f"{month_calculated_at_dt.strftime('%d %b %Y, %I:%M:%S')}.{month_calculated_at_dt.microsecond // 1000:03d} {month_calculated_at_dt.strftime('%p')}"
                    except Exception:
                        calculated_at_display = str(month_run['calculated_at'])
        except Exception:
            is_selected_month_calculated = False

    # Backward compatibility: treat current selected month as calculated if latest session meta matches.
    if not is_selected_month_calculated and calculated_at_raw:
        try:
            if int(last_filters.get('year', 0) or 0) == int(selected_year) and int(last_filters.get('month', 0) or 0) == int(selected_month):
                is_selected_month_calculated = True
        except Exception:
            pass

    # Build shift options from school-specific data used in Shift & Time Settings context.
    available_shifts = []
    shift_keys = set()

    if school_id:
        try:
            configured_shifts = db.execute('''
                SELECT shift_type
                FROM shift_definitions
                WHERE school_id = ? AND is_active = 1
                ORDER BY start_time, shift_type
            ''', (school_id,)).fetchall()
            for row in configured_shifts:
                shift_name = (row['shift_type'] or '').strip()
                if not shift_name:
                    continue
                key = shift_name.lower()
                if key in shift_keys:
                    continue
                shift_keys.add(key)
                available_shifts.append(shift_name)
        except Exception:
            pass

        try:
            school_staff_shifts = db.execute('''
                SELECT DISTINCT TRIM(shift_type) AS shift_type
                FROM staff
                WHERE school_id = ?
                  AND shift_type IS NOT NULL
                  AND TRIM(shift_type) <> ''
                ORDER BY shift_type
            ''', (school_id,)).fetchall()
            for row in school_staff_shifts:
                shift_name = (row['shift_type'] or '').strip()
                if not shift_name:
                    continue
                key = shift_name.lower()
                if key in shift_keys:
                    continue
                shift_keys.add(key)
                available_shifts.append(shift_name)
        except Exception:
            pass

        try:
            mapped_shifts = db.execute('''
                SELECT DISTINCT TRIM(default_shift_type) AS shift_type
                FROM department_shift_mappings
                WHERE school_id = ?
                  AND default_shift_type IS NOT NULL
                  AND TRIM(default_shift_type) <> ''
                ORDER BY shift_type
            ''', (school_id,)).fetchall()
            for row in mapped_shifts:
                shift_name = (row['shift_type'] or '').strip()
                if not shift_name:
                    continue
                key = shift_name.lower()
                if key in shift_keys:
                    continue
                shift_keys.add(key)
                available_shifts.append(shift_name)
        except Exception:
            pass

    # Build payroll records for review page from calculated salaries.
    payroll_records = []
    review_records = []
    reviewed_records = []
    total_net_salary = 0.0
    total_earnings = 0.0
    total_deductions = 0.0

    if school_id and selected_year and selected_month and is_selected_month_calculated:
        try:
            db.execute('''
                CREATE TABLE IF NOT EXISTS payroll_review_status (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    school_id INTEGER NOT NULL,
                    staff_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    status TEXT NOT NULL DEFAULT 'pending',
                    reviewed_by TEXT,
                    reviewed_at TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(school_id, staff_id, year, month)
                )
            ''')

            status_rows = db.execute('''
                SELECT staff_id, status
                FROM payroll_review_status
                WHERE school_id = ? AND year = ? AND month = ?
            ''', (school_id, selected_year, selected_month)).fetchall()
            status_map = {row['staff_id']: (row['status'] or 'pending').lower() for row in status_rows}

            staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]

            staff_query = '''
                SELECT id, staff_id, full_name, department,
                       COALESCE(shift_type, '') AS shift_type,
                       CASE
                           WHEN COALESCE(next_shift_type, '') <> ''
                                AND COALESCE(next_shift_effective_date, '') <> ''
                                AND ? >= next_shift_effective_date
                           THEN next_shift_type
                           ELSE COALESCE(shift_type, '')
                       END AS effective_shift_type,
                       COALESCE(gender, '') AS gender
                FROM staff
                WHERE school_id = ?
            '''
            staff_params = [today_str, school_id]

            if 'is_active' in staff_columns:
                staff_query += ' AND COALESCE(is_active, 1) = 1'
            elif 'status' in staff_columns:
                staff_query += " AND LOWER(COALESCE(status, 'active')) = 'active'"

            if selected_department:
                staff_query += ' AND department = ?'
                staff_params.append(selected_department)

            if selected_shift:
                staff_query += " AND LOWER(CASE WHEN COALESCE(next_shift_type, '') <> '' AND COALESCE(next_shift_effective_date, '') <> '' AND ? >= next_shift_effective_date THEN next_shift_type ELSE COALESCE(shift_type, '') END) = LOWER(?)"
                staff_params.append(today_str)
                staff_params.append(selected_shift)

            if selected_gender:
                staff_query += " AND LOWER(COALESCE(gender, '')) = LOWER(?)"
                staff_params.append(selected_gender)

            if search_term:
                staff_query += " AND (LOWER(COALESCE(full_name, '')) LIKE ? OR LOWER(COALESCE(staff_id, '')) LIKE ? OR LOWER(COALESCE(department, '')) LIKE ?)"
                like_term = f"%{search_term.lower()}%"
                staff_params.extend([like_term, like_term, like_term])

            staff_query += ' ORDER BY CAST(staff_id AS INTEGER) ASC'
            filtered_staff = db.execute(staff_query, staff_params).fetchall()

            salary_calculator = SalaryCalculator(school_id=school_id)

            for staff in filtered_staff:
                salary_result = salary_calculator.calculate_monthly_salary(staff['id'], selected_year, selected_month)
                if not salary_result.get('success'):
                    continue

                breakdown = salary_result.get('salary_breakdown', {})
                earnings = breakdown.get('earnings', {})
                deductions = breakdown.get('deductions', {})
                attendance_summary = breakdown.get('attendance_summary', {})
                attendance_details = salary_calculator._get_monthly_attendance(staff['id'], selected_year, selected_month)
                leave_details = salary_calculator._get_monthly_leaves(staff['id'], selected_year, selected_month)
                permission_details = salary_calculator._get_monthly_permissions(staff['id'], selected_year, selected_month)

                attendance_entries = []
                for attendance_row in attendance_details:
                    attendance_entries.append({
                        'date': attendance_row.get('date').isoformat() if hasattr(attendance_row.get('date'), 'isoformat') else attendance_row.get('date'),
                        'status': attendance_row.get('status') or 'N/A',
                        'time_in': attendance_row.get('time_in') or '--',
                        'time_out': attendance_row.get('time_out') or '--',
                        'late_duration_minutes': attendance_row.get('late_duration_minutes') or 0,
                        'early_departure_minutes': attendance_row.get('early_departure_minutes') or 0,
                        'shift_type': attendance_row.get('shift_type') or 'General',
                        'overtime_in': attendance_row.get('overtime_in') or '--',
                        'overtime_out': attendance_row.get('overtime_out') or '--'
                    })

                leave_entries = []
                for leave_row in leave_details:
                    leave_entries.append({
                        'leave_type': leave_row.get('leave_type') or 'N/A',
                        'start_date': leave_row.get('start_date').isoformat() if hasattr(leave_row.get('start_date'), 'isoformat') else leave_row.get('start_date'),
                        'end_date': leave_row.get('end_date').isoformat() if hasattr(leave_row.get('end_date'), 'isoformat') else leave_row.get('end_date'),
                        'reason': leave_row.get('reason') or 'No reason provided'
                    })

                permission_entries = []
                for permission_row in permission_details:
                    permission_entries.append({
                        'permission_type': permission_row.get('permission_type') or 'N/A',
                        'permission_date': permission_row.get('permission_date').isoformat() if hasattr(permission_row.get('permission_date'), 'isoformat') else permission_row.get('permission_date'),
                        'start_time': permission_row.get('start_time') or '--',
                        'end_time': permission_row.get('end_time') or '--',
                        'duration_hours': permission_row.get('duration_hours') or 0,
                        'reason': permission_row.get('reason') or 'No reason provided'
                    })

                # Load salary adjustments for this staff for this month
                try:
                    db.execute('''
                        CREATE TABLE IF NOT EXISTS salary_adjustments (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            school_id INTEGER NOT NULL,
                            staff_id INTEGER NOT NULL,
                            year INTEGER NOT NULL,
                            month INTEGER NOT NULL,
                            adjustment_type TEXT NOT NULL DEFAULT 'bonus',
                            amount REAL NOT NULL DEFAULT 0.0,
                            reason TEXT,
                            created_by TEXT,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            UNIQUE(school_id, staff_id, year, month, adjustment_type)
                        )
                    ''')
                except Exception:
                    pass

                adjustments = db.execute('''
                    SELECT adjustment_type, amount, reason
                    FROM salary_adjustments
                    WHERE school_id = ? AND staff_id = ? AND year = ? AND month = ?
                    ORDER BY adjustment_type
                ''', (school_id, staff['id'], selected_year, selected_month)).fetchall()

                manual_bonus = 0.0
                manual_deduction = 0.0
                for adj in adjustments:
                    if adj['adjustment_type'] == 'bonus':
                        manual_bonus += float(adj['amount'] or 0)
                    elif adj['adjustment_type'] == 'deduction':
                        manual_deduction += float(adj['amount'] or 0)

                detail_breakdown = {
                    'calculation_period': breakdown.get('calculation_period') or f'{selected_year}-{selected_month:02d}',
                    'working_days': breakdown.get('working_days', 0),
                    'per_day_salary': breakdown.get('per_day_salary', 0),
                    'per_hour_salary': breakdown.get('per_hour_salary', 0),
                    'attendance_summary': attendance_summary,
                    'earnings': earnings,
                    'deductions': deductions,
                    'attendance_entries': attendance_entries,
                    'leave_entries': leave_entries,
                    'permission_entries': permission_entries,
                    'manual_bonus': manual_bonus,
                    'manual_deduction': manual_deduction,
                    'adjustments': [dict(a) for a in adjustments]
                }

                row = {
                    'id': staff['id'],
                    'staff_id': staff.get('staff_id'),
                    'staff_name': staff.get('full_name'),
                    'department': staff.get('department') or 'N/A',
                    'shift_type': staff.get('shift_type') or 'General',
                    'effective_shift_type': staff.get('effective_shift_type') or staff.get('shift_type') or 'General',
                    'gender': staff.get('gender') or 'N/A',
                    'present_days': attendance_summary.get('present_days', 0),
                    'absent_days': attendance_summary.get('absent_days', 0),
                    'total_earnings': float(earnings.get('total_earnings', 0) or 0),
                    'total_deductions': float(deductions.get('total_deductions', 0) or 0),
                    'net_salary': float(
                        (earnings.get('total_earnings', 0) or 0)
                        - (deductions.get('total_deductions', 0) or 0)
                        + manual_bonus
                        - manual_deduction
                    ),
                    'review_status': status_map.get(staff['id'], 'pending'),
                    'detail_breakdown': detail_breakdown
                }
                payroll_records.append(row)

                if row['review_status'] == 'pending':
                    review_records.append(row)
                else:
                    reviewed_records.append(row)

                total_earnings += row['total_earnings']
                total_deductions += row['total_deductions']
                total_net_salary += row['net_salary']
        except Exception:
            payroll_records = []
            review_records = []
            reviewed_records = []
            total_earnings = 0.0
            total_deductions = 0.0
            total_net_salary = 0.0

    return render_template(
        'payroll_processing_review.html',
        current_year=today.year,
        module_enabled=module_enabled,
        available_shifts=available_shifts,
        payroll_records=payroll_records,
        review_records=review_records,
        reviewed_records=reviewed_records,
        selected_year=selected_year,
        selected_month=selected_month,
        selected_department=selected_department,
        selected_shift=selected_shift,
        selected_gender=selected_gender,
        search_term=search_term,
        calculated_by=calculated_by,
        calculated_at_display=calculated_at_display,
        is_selected_month_calculated=is_selected_month_calculated,
        total_earnings=total_earnings,
        total_deductions=total_deductions,
        total_net_salary=total_net_salary,
        total_staff=len(payroll_records),
        review_count=len(review_records),
        reviewed_count=len(reviewed_records)
    )

@app.route('/update_payroll_review_status', methods=['POST'])
@requires_permission('salary_management', 'view')
def update_payroll_review_status():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    school_id = session.get('school_id')
    staff_id = request.form.get('staff_id', type=int)
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)
    status = (request.form.get('status') or 'pending').strip().lower()
    redirect_to = request.form.get('redirect_to') or url_for('payroll_processing_review')

    if not school_id or not staff_id or not year or not month:
        return redirect(redirect_to)

    # Normalize status aliases for compatibility with existing UI labels.
    if status in ('review', 'reviewed'):
        status = 'reviewed'
    elif status in ('complete', 'completed'):
        status = 'completed'
    elif status in ('pay released', 'pay_released', 'released', 'paid'):
        status = 'pay_released'

    if status not in ('pending', 'reviewed', 'completed', 'pay_released'):
        status = 'pending'

    try:
        db = get_db()
        db.execute('''
            CREATE TABLE IF NOT EXISTS payroll_review_status (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                staff_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                reviewed_by TEXT,
                reviewed_at TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(school_id, staff_id, year, month)
            )
        ''')

        reviewed_by = session.get('full_name') or session.get('username') or str(session.get('user_id'))
        reviewed_at = datetime.datetime.now().isoformat(timespec='seconds') if status != 'pending' else None

        existing = db.execute('''
            SELECT id
            FROM payroll_review_status
            WHERE school_id = ? AND staff_id = ? AND year = ? AND month = ?
        ''', (school_id, staff_id, year, month)).fetchone()

        if existing:
            db.execute('''
                UPDATE payroll_review_status
                SET status = ?,
                    reviewed_by = ?,
                    reviewed_at = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (status, reviewed_by, reviewed_at, existing['id']))
        else:
            db.execute('''
                INSERT INTO payroll_review_status
                    (school_id, staff_id, year, month, status, reviewed_by, reviewed_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (school_id, staff_id, year, month, status, reviewed_by, reviewed_at))

        db.commit()
    except Exception as e:
        app.logger.exception('Failed to update payroll review status: %s', e)

    return redirect(redirect_to)

@app.route('/bulk_update_payroll_review_status', methods=['POST'])
@requires_permission('salary_management', 'view')
def bulk_update_payroll_review_status():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'No school selected'}), 400

    try:
        # Get parameters
        selected_staff_ids_str = request.form.get('selected_staff_ids', '[]')
        
        # If it's a JSON string, parse it
        if isinstance(selected_staff_ids_str, str):
            selected_staff_ids = json.loads(selected_staff_ids_str)
        else:
            selected_staff_ids = selected_staff_ids_str
        
        year = request.form.get('year', type=int)
        month = request.form.get('month', type=int)
        status = (request.form.get('status') or 'pending').strip().lower()

        if not selected_staff_ids or not year or not month:
            return jsonify({'success': False, 'message': 'Missing required parameters'}), 400

        # Normalize status aliases for compatibility
        if status in ('review', 'reviewed'):
            status = 'reviewed'
        elif status in ('complete', 'completed'):
            status = 'completed'
        elif status in ('pay released', 'pay_released', 'released', 'paid'):
            status = 'pay_released'

        if status not in ('pending', 'reviewed', 'completed', 'pay_released'):
            status = 'pending'

        db = get_db()
        db.execute('''
            CREATE TABLE IF NOT EXISTS payroll_review_status (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                staff_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                reviewed_by TEXT,
                reviewed_at TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(school_id, staff_id, year, month)
            )
        ''')

        reviewed_by = session.get('full_name') or session.get('username') or str(session.get('user_id'))
        reviewed_at = datetime.datetime.now().isoformat(timespec='seconds') if status != 'pending' else None

        # Process each staff member
        updated_count = 0
        for staff_id in selected_staff_ids:
            staff_id = int(staff_id)
            
            existing = db.execute('''
                SELECT id
                FROM payroll_review_status
                WHERE school_id = ? AND staff_id = ? AND year = ? AND month = ?
            ''', (school_id, staff_id, year, month)).fetchone()

            if existing:
                db.execute('''
                    UPDATE payroll_review_status
                    SET status = ?,
                        reviewed_by = ?,
                        reviewed_at = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (status, reviewed_by, reviewed_at, existing['id']))
            else:
                db.execute('''
                    INSERT INTO payroll_review_status
                        (school_id, staff_id, year, month, status, reviewed_by, reviewed_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ''', (school_id, staff_id, year, month, status, reviewed_by, reviewed_at))
            
            updated_count += 1

        db.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully updated {updated_count} payroll records to {status}!'
        })

    except json.JSONDecodeError as e:
        app.logger.exception('Failed to parse JSON in bulk update: %s', e)
        return jsonify({'success': False, 'message': 'Invalid JSON format'}), 400
    except Exception as e:
        app.logger.exception('Failed to bulk update payroll review status: %s', e)
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/save_salary_adjustment', methods=['POST'])
@requires_permission('salary_management', 'view')
def save_salary_adjustment():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    payload = request.get_json(silent=True) or request.form

    def to_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def to_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    school_id = session.get('school_id')
    staff_id = to_int(payload.get('staff_id'))
    year = to_int(payload.get('year'))
    month = to_int(payload.get('month'))
    adjustment_type = str(payload.get('adjustment_type') or 'bonus').strip().lower()
    amount = to_float(payload.get('amount'))
    reason = str(payload.get('reason') or '').strip()

    if not school_id or not staff_id or not year or not month or amount is None:
        return jsonify({'success': False, 'error': 'Missing required fields'}), 400

    if adjustment_type not in ('bonus', 'deduction'):
        adjustment_type = 'bonus'

    if amount < 0:
        amount = abs(amount)

    try:
        db = get_db()
        db.execute('''
            CREATE TABLE IF NOT EXISTS salary_adjustments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                staff_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                adjustment_type TEXT NOT NULL DEFAULT 'bonus',
                amount REAL NOT NULL DEFAULT 0.0,
                reason TEXT,
                created_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(school_id, staff_id, year, month, adjustment_type)
            )
        ''')

        created_by = session.get('full_name') or session.get('username') or str(session.get('user_id'))

        existing = db.execute('''
            SELECT id
            FROM salary_adjustments
            WHERE school_id = ? AND staff_id = ? AND year = ? AND month = ? AND adjustment_type = ?
        ''', (school_id, staff_id, year, month, adjustment_type)).fetchone()

        if existing:
            db.execute('''
                UPDATE salary_adjustments
                SET amount = ?,
                    reason = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (amount, reason, existing['id']))
        else:
            db.execute('''
                INSERT INTO salary_adjustments
                    (school_id, staff_id, year, month, adjustment_type, amount, reason, created_by, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ''', (school_id, staff_id, year, month, adjustment_type, amount, reason, created_by))

        db.commit()

        return jsonify({
            'success': True,
            'message': f'{adjustment_type.capitalize()} of ₹{amount:,.2f} saved for {year}-{month:02d}',
            'amount': amount,
            'adjustment_type': adjustment_type
        })
    except Exception as e:
        app.logger.exception('Failed to save salary adjustment: %s', e)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get_summary_dashboard')
def get_summary_dashboard():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    reporting_dashboard = ReportingDashboard()
    summary = reporting_dashboard.get_summary_dashboard(school_id)

    return jsonify({'success': True, 'summary': summary})

@app.route('/generate_report')
def generate_report():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    report_type = request.args.get('report_type')

    if not report_type:
        return jsonify({'success': False, 'error': 'Report type is required'})

    reporting_dashboard = ReportingDashboard()

    try:
        if report_type == 'daily':
            date = request.args.get('date')
            if not date:
                return jsonify({'success': False, 'error': 'Date is required for daily report'})
            report = reporting_dashboard.generate_daily_report(school_id, date)

        elif report_type == 'weekly':
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            if not start_date or not end_date:
                return jsonify({'success': False, 'error': 'Date range is required for weekly report'})
            report = reporting_dashboard.generate_weekly_report(school_id, start_date, end_date)

        elif report_type == 'monthly':
            year = request.args.get('year', type=int)
            month = request.args.get('month', type=int)
            if not year or not month:
                return jsonify({'success': False, 'error': 'Year and month are required for monthly report'})
            report = reporting_dashboard.generate_monthly_report(school_id, year, month)

        elif report_type == 'yearly':
            year = request.args.get('year', type=int)
            if not year:
                return jsonify({'success': False, 'error': 'Year is required for yearly report'})
            report = reporting_dashboard.generate_yearly_report(school_id, year)

        elif report_type == 'department':
            department = request.args.get('department')
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            if not all([department, start_date, end_date]):
                return jsonify({'success': False, 'error': 'Department and date range are required'})
            report = reporting_dashboard.generate_department_report(school_id, department, start_date, end_date)

        elif report_type == 'custom':
            filters = {
                'start_date': request.args.get('start_date'),
                'end_date': request.args.get('end_date'),
                'department': request.args.get('department'),
                'position': request.args.get('position'),
                'status': request.args.get('status')
            }
            # Remove None values
            filters = {k: v for k, v in filters.items() if v}
            report = reporting_dashboard.generate_custom_report(school_id, filters)

        elif report_type == 'trends':
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            if not start_date or not end_date:
                return jsonify({'success': False, 'error': 'Date range is required for trends report'})
            report = reporting_dashboard.generate_trends_report(school_id, start_date, end_date)

        elif report_type == 'summary':
            report = reporting_dashboard.get_summary_dashboard(school_id)
            report['report_type'] = 'summary'

        else:
            return jsonify({'success': False, 'error': 'Invalid report type'})

        return jsonify({'success': True, 'report': report})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_departments')
def get_departments():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    db = get_db()

    departments = db.execute('''
        SELECT DISTINCT department as name
        FROM staff
        WHERE school_id = ? AND department IS NOT NULL AND department != ''
        ORDER BY department
    ''', (school_id,)).fetchall()

    return jsonify({
        'success': True,
        'departments': [dept['name'] for dept in departments]
    })

@app.route('/export_report_excel')
def export_report_excel():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    report_type = request.args.get('report_type')

    if not report_type:
        return jsonify({'success': False, 'error': 'Report type is required'})

    # Generate appropriate Excel report based on type
    excel_generator = ExcelReportGenerator()

    if report_type == 'daily':
        date = request.args.get('date', datetime.datetime.now().strftime('%Y-%m-%d'))
        return excel_generator.create_staff_attendance_report(school_id, date, date)
    elif report_type in ['weekly', 'custom', 'trends']:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        if start_date and end_date:
            return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)
    elif report_type == 'monthly':
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        if year and month:
            return excel_generator.create_monthly_report(school_id, year, month)
    elif report_type == 'yearly':
        year = request.args.get('year', type=int)
        if year:
            start_date = f"{year}-01-01"
            end_date = f"{year}-12-31"
            return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

    # Fallback to general report
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    return excel_generator.create_staff_attendance_report(school_id, today, today)

@app.route('/generate_admin_report')
def generate_admin_report():
    """Generate and download reports from admin reports page"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        school_id = session['school_id']
        report_type = request.args.get('report_type')
        format_type = request.args.get('format', 'excel').lower()

        if not report_type:
            return jsonify({'success': False, 'error': 'Report type is required'})

        # Get filter parameters
        year = request.args.get('year', datetime.datetime.now().year, type=int)
        month = request.args.get('month', type=int)
        department = request.args.get('department', '')

        # Create Excel generator for all reports
        excel_generator = ExcelReportGenerator()

        # Route to appropriate report generation based on report_type
        if report_type == 'monthly_salary':
            return generate_monthly_salary_report(school_id, year, month, department, format_type)
        elif report_type == 'payroll_summary':
            return generate_payroll_summary_report(school_id, year, month, format_type)
        elif report_type == 'department_salary':
            return generate_department_salary_report(school_id, year, department, format_type)
        elif report_type == 'staff_directory':
            return generate_staff_directory_report(school_id, format_type)
        elif report_type == 'department_report':
            return generate_department_analysis_report(school_id, year, month, format_type)
        elif report_type == 'performance_report':
            return generate_performance_report(school_id, year, month, department, format_type)
        elif report_type == 'daily_attendance':
            date = request.args.get('date', datetime.datetime.now().strftime('%Y-%m-%d'))
            return generate_daily_attendance_report(school_id, date, department, format_type)
        elif report_type == 'monthly_attendance':
            if month:
                return excel_generator.create_monthly_report(school_id, year, month)
            else:
                # Current month
                return excel_generator.create_monthly_report(school_id, year, datetime.datetime.now().month)
        elif report_type == 'overtime_report':
            return generate_overtime_report(school_id, year, month, format_type)
        else:
            return jsonify({'success': False, 'error': f'Unknown report type: {report_type}'})

    except Exception as e:
        print(f"Report generation error: {str(e)}")
        return jsonify({'success': False, 'error': f'Report generation failed: {str(e)}'})

@app.route('/test_performance_report_json')
def test_performance_report_json():
    """Test endpoint for performance report that returns JSON data for validation"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        school_id = session['school_id']
        department = request.args.get('department', '')
        
        # Temporarily modify the generate_performance_report function to return JSON
        # by setting the format_type to 'json' which we'll handle in the function
        
        # Create a fake request context for the performance report function
        original_args = request.args
        
        # Call the performance report generator directly, but we need to modify it to return JSON
        from flask import g
        import datetime
        
        db = get_db()
        
        # Get from_date and to_date parameters
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        
        # Date range logic (same as in the function)
        if from_date_str and to_date_str:
            try:
                start_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
                end_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
            except ValueError:
                today = datetime.date.today()
                start_date = datetime.date(today.year, today.month, 1)
                end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        else:
            today = datetime.date.today()
            start_date = datetime.date(today.year, today.month, 1) 
            end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        
        start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
        
        # Build department filter
        dept_clause = ''
        staff_params = [school_id]
        if department and str(department).strip():
            dept_clause = ' AND COALESCE(s.department, "") = ?'
            staff_params.append(department)
        
        # Get staff data
        staff_performance_query = f"""
            SELECT 
                s.id as staff_db_id,
                s.staff_id,
                s.full_name,
                COALESCE(NULLIF(TRIM(s.department), ''), 'Unassigned') as department,
                COALESCE(NULLIF(TRIM(s.position), ''), 'Not Specified') as position,
                s.date_of_joining
            FROM staff s
            WHERE s.school_id = ? AND s.is_active = 1 {dept_clause}
            ORDER BY CAST(s.staff_id AS INTEGER) ASC
        """
        
        staff_rows = db.execute(staff_performance_query, tuple(staff_params)).fetchall()
        
        # Calculate working days (moved outside the loop)
        total_days = (end_date - start_date).days + 1
        working_days = 0
        current_dt = start_date
        weekend_days = 0
        
        # Get holidays
        holidays_in_period = db.execute("""
            SELECT start_date, end_date, holiday_name 
            FROM holidays 
            WHERE school_id = ? AND is_active = 1
            AND NOT (end_date < ? OR start_date > ?)
        """, (school_id, start_str, end_str)).fetchall()
        
        holiday_dates = set()
        for holiday in holidays_in_period:
            h_start = datetime.datetime.strptime(holiday['start_date'], '%Y-%m-%d').date()
            h_end = datetime.datetime.strptime(holiday['end_date'], '%Y-%m-%d').date()
            curr_date = h_start
            while curr_date <= h_end:
                holiday_dates.add(curr_date)
                curr_date += datetime.timedelta(days=1)
        
        # Count working days
        current_dt = start_date
        while current_dt <= end_date:
            if current_dt.weekday() < 6:  # Monday-Saturday = 0-5, Sunday = 6
                if current_dt not in holiday_dates:
                    working_days += 1
            else:
                weekend_days += 1
            current_dt += datetime.timedelta(days=1)
        
        # Get attendance data for each staff
        final_performance_rows = []
        
        for staff in staff_rows:
            staff_db_id = staff['staff_db_id']
            
            # Attendance data
            attendance_data = db.execute("""
                SELECT 
                    COUNT(*) as total_records,
                    SUM(CASE WHEN status IN ('present', 'late') THEN 1 ELSE 0 END) as days_present,
                    SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as days_absent,
                    SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as days_late
                FROM attendance 
                WHERE staff_id = ? AND date BETWEEN ? AND ?
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            # Leave data
            leave_data = db.execute("""
                SELECT 
                    COALESCE(COUNT(CASE WHEN status = 'approved' THEN id END), 0) as approved_leave_count,
                    COALESCE(SUM(CASE WHEN status = 'approved' 
                                THEN julianday(end_date) - julianday(start_date) + 1 
                                ELSE 0 END), 0) as days_on_leave
                FROM leave_applications 
                WHERE staff_id = ? AND NOT (end_date < ? OR start_date > ?)
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            # OD data  
            od_data = db.execute("""
                SELECT 
                    COALESCE(COUNT(CASE WHEN status = 'approved' THEN id END), 0) as approved_od_count,
                    COALESCE(SUM(CASE WHEN status = 'approved' 
                                THEN julianday(end_date) - julianday(start_date) + 1 
                                ELSE 0 END), 0) as days_on_od
                FROM on_duty_applications 
                WHERE staff_id = ? AND NOT (end_date < ? OR start_date > ?)
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            # Permission data
            permission_data = db.execute("""
                SELECT 
                    COALESCE(COUNT(CASE WHEN status = 'approved' THEN id END), 0) as approved_permission_count,
                    COALESCE(SUM(CASE WHEN status = 'approved' THEN duration_hours ELSE 0 END), 0) as total_permission_hours
                FROM permission_applications 
                WHERE staff_id = ? AND permission_date BETWEEN ? AND ?
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            permission_days = round(float(permission_data['total_permission_hours'] or 0) / 8, 1)
            
            final_performance_rows.append({
                'staff_id': staff['staff_id'],
                'staff_name': staff['full_name'],
                'department': staff['department'],
                'position': staff['position'],
                'days_present': int(attendance_data['days_present'] or 0),
                'days_absent': int(attendance_data['days_absent'] or 0),
                'days_late': int(attendance_data['days_late'] or 0),
                'days_on_od_applied': int(od_data['days_on_od'] or 0),
                'days_on_leave_applied': int(leave_data['days_on_leave'] or 0),
                'days_with_permission_applied': permission_days,
                'total_working_days': working_days,
                'total_attendance_records': int(attendance_data['total_records'] or 0)
            })
            
            # Only process first 3 staff for quick testing to avoid timeout
            if len(final_performance_rows) >= 3:
                break
        
        return jsonify({
            'success': True,
            'data': final_performance_rows,
            'summary': {
                'total_staff': len(final_performance_rows),
                'date_range': f'{start_str} to {end_str}',
                'department_filter': department if department else 'All Departments',
                'total_working_days': working_days,
                'weekend_days': weekend_days,
                'holiday_days': len(holiday_dates),
                'total_days': total_days
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Test failed: {str(e)}'})

def generate_monthly_salary_report(school_id, year, month, department, format_type):
    """Generate monthly salary report with comprehensive deduction calculation"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import calendar
    import datetime

    db = get_db()
    staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]

    # Build query based on filters
    where_conditions = ['s.school_id = ?']
    params = [school_id]

    if 'is_active' in staff_columns:
        where_conditions.append('COALESCE(s.is_active, 1) = 1')
    elif 'status' in staff_columns:
        where_conditions.append("LOWER(COALESCE(s.status, 'active')) = 'active'")

    if department:
        where_conditions.append('s.department = ?')
        params.append(department)

    # Get staff basic information
    staff_query = f'''
        SELECT s.id, s.staff_id, s.full_name, s.department, s.destination as position,
               s.basic_salary, s.hra, s.transport_allowance, s.other_allowances,
               s.pf_deduction, s.esi_deduction, s.professional_tax, s.other_deductions,
               s.date_of_joining
        FROM staff s
        WHERE {' AND '.join(where_conditions)}
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    '''

    staff_data = db.execute(staff_query, params).fetchall()
    
    # Calculate working days in the month (excluding holidays)
    if year and month:
        total_days = calendar.monthrange(year, month)[1]
        working_days = 0
        current_dt = datetime.date(year, month, 1)
        end_dt = datetime.date(year, month, total_days)
        
        while current_dt <= end_dt:
            if current_dt.weekday() < 6:  # Monday-Saturday = 0-5, Sunday = 6
                working_days += 1
            current_dt += datetime.timedelta(days=1)
    else:
        working_days = 26  # Default working days
    
    # Use SalaryCalculator for consistent deduction calculations
    current_school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=current_school_id)
    
    processed_staff_data = []
    for staff in staff_data:
        staff_db_id = staff['id']
        
        # Calculate salary using SalaryCalculator for consistency
        if year and month:
            salary_result = salary_calculator.calculate_monthly_salary(staff_db_id, year, month)
            if salary_result['success']:
                breakdown = salary_result['salary_breakdown']
                earnings = breakdown['earnings']
                deductions = breakdown['deductions']
                
                processed_staff_data.append({
                    'id': staff['id'],
                    'staff_id': staff['staff_id'],
                    'full_name': staff['full_name'],
                    'department': staff['department'],
                    'position': staff['position'],
                    'basic_salary': earnings['basic_salary'],
                    'allowances': earnings['hra'] + earnings['transport_allowance'] + earnings['other_allowances'],
                    'deductions': deductions['total_deductions'],  # Uses SalaryCalculator's comprehensive calculation
                    'net_salary': breakdown['net_salary'],
                    'date_of_joining': staff['date_of_joining'],
                    'salary_breakdown': breakdown  # Store full breakdown for detailed report generation
                })
            else:
                # Fallback to basic calculation if SalaryCalculator fails
                base_salary = float(staff['basic_salary'] or 0)
                allowances = (float(staff['hra'] or 0) + 
                             float(staff['transport_allowance'] or 0) + 
                             float(staff['other_allowances'] or 0))
                static_deductions = (float(staff['pf_deduction'] or 0) + 
                                   float(staff['esi_deduction'] or 0) + 
                                   float(staff['professional_tax'] or 0) + 
                                   float(staff['other_deductions'] or 0))
                gross_salary = base_salary + allowances
                net_salary = gross_salary - static_deductions
                
                processed_staff_data.append({
                    'id': staff['id'],
                    'staff_id': staff['staff_id'],
                    'full_name': staff['full_name'],
                    'department': staff['department'],
                    'position': staff['position'],
                    'basic_salary': base_salary,
                    'allowances': allowances,
                    'deductions': static_deductions,
                    'net_salary': net_salary,
                    'date_of_joining': staff['date_of_joining']
                })
        else:
            # No year/month specified - use basic calculation
            base_salary = float(staff['basic_salary'] or 0)
            allowances = (float(staff['hra'] or 0) + 
                         float(staff['transport_allowance'] or 0) + 
                         float(staff['other_allowances'] or 0))
            static_deductions = (float(staff['pf_deduction'] or 0) + 
                               float(staff['esi_deduction'] or 0) + 
                               float(staff['professional_tax'] or 0) + 
                               float(staff['other_deductions'] or 0))
            gross_salary = base_salary + allowances
            net_salary = gross_salary - static_deductions
            
            processed_staff_data.append({
                'id': staff['id'],
                'staff_id': staff['staff_id'],
                'full_name': staff['full_name'],
                'department': staff['department'],
                'position': staff['position'],
                'basic_salary': base_salary,
                'allowances': allowances,
                'deductions': static_deductions,
                'net_salary': net_salary,
                'date_of_joining': staff['date_of_joining']
            })

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Salary Report"

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2F5597")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Add title
    month_name = calendar.month_name[month] if month else 'All Months'
    ws.merge_cells('A1:R1')  # Updated to cover all 18 columns
    title_cell = ws['A1']
    title_cell.value = f"Payroll Summary Report - {month_name} {year}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add summary
    ws['A3'] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    if department:
        ws['A4'] = f"Department: {department}"

    # Headers - Updated to include detailed deduction breakdown
    headers = [
        'S.No', 'Staff ID', 'Name', 'Department', 'Position', 'Basic Salary', 
        'HRA', 'Transport', 'Other Allow.', 'Absent Ded.', 'Late Penalty', 'Early Dept.',
        'PF', 'ESI', 'Prof. Tax', 'Other Ded.', 'Total Ded.', 'Net Salary'
    ]
    header_row = 6

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Add data with detailed deduction breakdown
    total_basic = total_hra = total_transport = total_other_allow = 0
    total_absent_ded = total_late_penalty = total_early_dept = 0
    total_pf = total_esi = total_prof_tax = total_other_ded = total_deductions = total_net = 0

    for row_idx, staff in enumerate(processed_staff_data, header_row + 1):
        # Use the already calculated SalaryCalculator data for consistency
        if year and month and 'salary_breakdown' in staff:
            # Use the breakdown data stored during processed_staff_data creation
            breakdown = staff['salary_breakdown']
            earnings = breakdown['earnings']
            deductions = breakdown['deductions']
            
            hra = earnings.get('hra', 0)
            transport = earnings.get('transport_allowance', 0)
            other_allow = earnings.get('other_allowances', 0)
            absent_ded = deductions.get('absent_deduction', 0)
            late_penalty = deductions.get('late_arrival_penalty', 0)
            early_dept = deductions.get('early_departure_penalty', 0)
            pf = deductions.get('pf_deduction', 0)
            esi = deductions.get('esi_deduction', 0)
            prof_tax = deductions.get('professional_tax', 0)
            other_ded = deductions.get('other_deductions', 0)
            total_deductions_calculated = deductions.get('total_deductions', 0)
        else:
            # Calculate detailed breakdown using SalaryCalculator if not already stored
            if year and month:
                salary_result = salary_calculator.calculate_monthly_salary(staff['id'], year, month)
                if salary_result['success']:
                    breakdown = salary_result['salary_breakdown']
                    earnings = breakdown['earnings']
                    deductions = breakdown['deductions']
                    
                    hra = earnings.get('hra', 0)
                    transport = earnings.get('transport_allowance', 0)
                    other_allow = earnings.get('other_allowances', 0)
                    absent_ded = deductions.get('absent_deduction', 0)
                    late_penalty = deductions.get('late_arrival_penalty', 0)
                    early_dept = deductions.get('early_departure_penalty', 0)
                    pf = deductions.get('pf_deduction', 0)
                    esi = deductions.get('esi_deduction', 0)
                    prof_tax = deductions.get('professional_tax', 0)
                    other_ded = deductions.get('other_deductions', 0)
                    total_deductions_calculated = deductions.get('total_deductions', 0)
                else:
                    # Fallback to basic data
                    hra = transport = other_allow = 0
                    absent_ded = late_penalty = early_dept = 0
                    pf = esi = prof_tax = other_ded = 0
                    total_deductions_calculated = staff.get('deductions', 0)
            else:
                # No year/month - use basic data
                hra = transport = other_allow = 0
                absent_ded = late_penalty = early_dept = 0  
                pf = esi = prof_tax = other_ded = 0
                total_deductions_calculated = staff.get('deductions', 0)

        ws.cell(row=row_idx, column=1, value=row_idx - header_row)
        ws.cell(row=row_idx, column=2, value=staff['staff_id'] or 'N/A')
        ws.cell(row=row_idx, column=3, value=staff['full_name'])
        ws.cell(row=row_idx, column=4, value=staff['department'] or 'N/A')
        ws.cell(row=row_idx, column=5, value=staff['position'] or 'N/A')
        ws.cell(row=row_idx, column=6, value=staff['basic_salary'] or 0)
        ws.cell(row=row_idx, column=7, value=hra)
        ws.cell(row=row_idx, column=8, value=transport)
        ws.cell(row=row_idx, column=9, value=other_allow)
        ws.cell(row=row_idx, column=10, value=absent_ded)
        ws.cell(row=row_idx, column=11, value=late_penalty)
        ws.cell(row=row_idx, column=12, value=early_dept)
        ws.cell(row=row_idx, column=13, value=pf)
        ws.cell(row=row_idx, column=14, value=esi)
        ws.cell(row=row_idx, column=15, value=prof_tax)
        ws.cell(row=row_idx, column=16, value=other_ded)
        ws.cell(row=row_idx, column=17, value=total_deductions_calculated)
        ws.cell(row=row_idx, column=18, value=staff['net_salary'] or 0)

        # Add to totals
        total_basic += staff['basic_salary'] or 0
        total_hra += hra
        total_transport += transport
        total_other_allow += other_allow
        total_absent_ded += absent_ded
        total_late_penalty += late_penalty
        total_early_dept += early_dept
        total_pf += pf
        total_esi += esi
        total_prof_tax += prof_tax
        total_other_ded += other_ded
        total_deductions += total_deductions_calculated
        total_net += staff['net_salary'] or 0

        # Apply border to all cells
        for col in range(1, 19):  # Updated to cover all 18 columns
            ws.cell(row=row_idx, column=col).border = border

    # Add totals row
    total_row = len(processed_staff_data) + header_row + 1
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_basic).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_hra).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=total_transport).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=total_other_allow).font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=total_absent_ded).font = Font(bold=True)
    ws.cell(row=total_row, column=11, value=total_late_penalty).font = Font(bold=True)
    ws.cell(row=total_row, column=12, value=total_early_dept).font = Font(bold=True)
    ws.cell(row=total_row, column=13, value=total_pf).font = Font(bold=True)
    ws.cell(row=total_row, column=14, value=total_esi).font = Font(bold=True)
    ws.cell(row=total_row, column=15, value=total_prof_tax).font = Font(bold=True)
    ws.cell(row=total_row, column=16, value=total_other_ded).font = Font(bold=True)
    ws.cell(row=total_row, column=17, value=total_deductions).font = Font(bold=True)
    ws.cell(row=total_row, column=18, value=total_net).font = Font(bold=True)

    # Auto-adjust column widths
    for col in range(1, 19):  # Updated to cover all 18 columns
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'monthly_salary_report_{year}_{month or "all"}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def generate_staff_directory_report(school_id, format_type):
    """Generate comprehensive staff directory report"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    db = get_db()

    # Get comprehensive staff information
    staff_data = db.execute('''
        SELECT s.staff_id, s.full_name, s.first_name, s.last_name,
               s.date_of_birth, s.date_of_joining, s.department, s.destination as position,
               s.gender, s.phone, s.email, s.shift_type, s.basic_salary,
               s.created_at
        FROM staff s
        WHERE s.school_id = ?
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    ''', (school_id,)).fetchall()

    # Common headers and rows for all formats
    headers = [
        'S.No', 'Staff ID', 'Full Name', 'Department', 'Position',
        'Gender', 'Phone', 'Email', 'Date of Joining', 'Date of Birth', 'Shift Type'
    ]
    rows = []
    for idx, staff in enumerate(staff_data, start=1):
        rows.append([
            idx,
            staff['staff_id'] or 'N/A',
            staff['full_name'],
            staff['department'] or 'N/A',
            staff['position'] or 'N/A',
            staff['gender'] or 'N/A',
            staff['phone'] or 'N/A',
            staff['email'] or 'N/A',
            staff['date_of_joining'] or 'N/A',
            staff['date_of_birth'] or 'N/A',
            staff['shift_type'] or 'General'
        ])

    # CSV export
    if format_type == 'csv':
        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Staff Directory Report', datetime.datetime.now().strftime('%Y-%m-%d')])
        writer.writerow([])
        writer.writerow(headers)
        writer.writerows(rows)
        response = make_response(output.getvalue())
        filename = f'staff_directory_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response

    # PDF export
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            import io

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []

            title = Paragraph(f"Staff Directory Report - {datetime.datetime.now().strftime('%Y-%m-%d')}", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))

            table_data = [headers] + rows
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(table)

            doc.build(elements)
            pdf_value = buffer.getvalue()
            buffer.close()

            response = make_response(pdf_value)
            filename = f'staff_directory_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return response
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'PDF generation requires reportlab. Install with: pip install reportlab'
            })
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel export
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Directory"

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2F5597")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Add title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Staff Directory Report - Generated on {datetime.datetime.now().strftime('%Y-%m-%d')}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Headers row
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row_idx = header_row + 1
    for r in rows:
        for col, value in enumerate(r, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
        row_idx += 1

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'staff_directory_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

# Placeholder functions for other report types - can be expanded later
def generate_payroll_summary_report(school_id, year, month, format_type):
    """Generate enhanced Payroll Summary Report with detailed breakdown.
    
    Includes:
    1. Summary Section: Total payroll expenses, department breakdown, staff count
    2. Detailed Staff Records: All salary components, allowances, and deductions
    3. Attendance-based deductions for absent/late days
    4. Professional Excel formatting with totals and styling
    """
    import datetime, calendar, io
    from flask import request

    db = get_db()

    # Handle date range from request parameters (enhanced filtering)
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    department = request.args.get('department', '').strip()
    
    # Determine date range for the payroll report
    if from_date_str and to_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            # Fallback to year/month if date parsing fails
            if not isinstance(year, int) or year <= 0:
                year = datetime.date.today().year
            if not isinstance(month, int) or not (1 <= month <= 12):
                month = datetime.date.today().month
            start_date = datetime.date(year, month, 1)
            end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    elif year and month:
        start_date = datetime.date(year, month, 1)
        end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    else:
        # Default to current month
        today = datetime.date.today()
        start_date = datetime.date(today.year, today.month, 1)
        end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    
    start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
    period_label = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"

    # Build department filter clause
    dept_clause = ''
    staff_params = [school_id]
    if department:
        dept_clause = ' AND COALESCE(s.department, "") = ?'
        staff_params.append(department)

    # Get staff data with salary components
    staff_query = f"""
        SELECT 
            s.id as staff_db_id,
            s.staff_id,
            s.full_name,
            COALESCE(NULLIF(TRIM(s.department), ''), 'Unassigned') as department,
            COALESCE(NULLIF(TRIM(s.destination), ''), 'Not Specified') as position,
            COALESCE(s.basic_salary, 0) as base_salary,
            COALESCE(s.hra, 0) as hra_allowance,
            COALESCE(s.transport_allowance, 0) as transport_allowance,
            COALESCE(s.other_allowances, 0) as other_allowances,
            COALESCE(s.pf_deduction, 0) as pf_deduction,
            COALESCE(s.esi_deduction, 0) as esi_deduction,
            COALESCE(s.professional_tax, 0) as professional_tax,
            COALESCE(s.other_deductions, 0) as other_deductions,
            s.date_of_joining
        FROM staff s
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause}
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    """
    
    staff_rows = db.execute(staff_query, tuple(staff_params)).fetchall()

    # Calculate working days in the period (excluding weekends and holidays)
    total_days = (end_date - start_date).days + 1
    working_days = 0
    current_dt = start_date
    weekend_days = 0
    
    # Get holidays in the period
    holidays_in_period = db.execute("""
        SELECT start_date, end_date, holiday_name 
        FROM holidays 
        WHERE school_id = ? AND is_active = 1
        AND NOT (end_date < ? OR start_date > ?)
    """, (school_id, start_str, end_str)).fetchall()
    
    # Create set of holiday dates
    holiday_dates = set()
    for holiday in holidays_in_period:
        h_start = datetime.datetime.strptime(holiday['start_date'], '%Y-%m-%d').date()
        h_end = datetime.datetime.strptime(holiday['end_date'], '%Y-%m-%d').date()
        curr_date = h_start
        while curr_date <= h_end:
            holiday_dates.add(curr_date)
            curr_date += datetime.timedelta(days=1)
    
    # Count working days (Mon-Sat, excluding holidays and Sundays)
    while current_dt <= end_date:
        if current_dt.weekday() < 6:  # Monday-Saturday = 0-5, Sunday = 6
            if current_dt not in holiday_dates:
                working_days += 1
        else:
            weekend_days += 1
        current_dt += datetime.timedelta(days=1)
    
    holiday_days = len(holiday_dates)

    # Process each staff member for detailed payroll calculation using SalaryCalculator
    current_school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=current_school_id)
    
    payroll_records = []
    total_payroll_expense = 0.0
    department_totals = {}
    
    for staff in staff_rows:
        staff_db_id = staff['staff_db_id']
        staff_id = staff['staff_id']
        
        # Use SalaryCalculator for consistent calculations
        salary_result = salary_calculator.calculate_monthly_salary(staff_db_id, year, month)
        if salary_result['success']:
            breakdown = salary_result['salary_breakdown']
            earnings = breakdown['earnings']
            deductions = breakdown['deductions']
            attendance = breakdown['attendance_summary']
            
            # Extract key values from SalaryCalculator
            base_salary = earnings['basic_salary']
            hra_allowance = earnings['hra']
            transport_allowance = earnings['transport_allowance']
            other_allowances = earnings['other_allowances']
            total_allowances = hra_allowance + transport_allowance + other_allowances
            gross_pay = earnings['total_earnings']
            total_deductions = deductions['total_deductions']  # Comprehensive deductions from SalaryCalculator
            net_payroll = breakdown['net_salary']
            days_present = attendance['present_days']
            days_absent = attendance['absent_days']
            days_late = attendance.get('late_days', 0)  # May not exist in older versions
        else:
            # Fallback to basic calculation if SalaryCalculator fails
            base_salary = float(staff['base_salary'])
            hra_allowance = float(staff['hra_allowance'])
            transport_allowance = float(staff['transport_allowance'])
            other_allowances = float(staff['other_allowances'])
            total_allowances = hra_allowance + transport_allowance + other_allowances
            gross_pay = base_salary + total_allowances
            
            # Get attendance data for fallback calculation
            attendance_data = db.execute("""
                SELECT 
                    SUM(CASE WHEN status IN ('present', 'late') THEN 1 ELSE 0 END) as days_present,
                    SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as days_absent,
                    SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as days_late
                FROM attendance 
                WHERE staff_id = ? AND date BETWEEN ? AND ?
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            days_present = int(attendance_data['days_present'] or 0)
            days_absent = int(attendance_data['days_absent'] or 0)
            days_late = int(attendance_data['days_late'] or 0)
            
            # Basic deductions calculation
            pf_deduction = float(staff['pf_deduction'])
            esi_deduction = float(staff['esi_deduction'])
            professional_tax = float(staff['professional_tax'])
            other_deductions = float(staff['other_deductions'])
            total_deductions = pf_deduction + esi_deduction + professional_tax + other_deductions
            net_payroll = gross_pay - total_deductions
        
        # Add to department totals
        dept = staff['department']
        if dept not in department_totals:
            department_totals[dept] = {'count': 0, 'total': 0.0}
        department_totals[dept]['count'] += 1
        department_totals[dept]['total'] += net_payroll
        
        # Add to total payroll expense
        total_payroll_expense += net_payroll
        
        # Create detailed staff record using appropriate data source
        if salary_result and salary_result['success']:
            # Use SalaryCalculator data
            attendance_summary = attendance
            deduction_details = deductions
            record_deductions = {
                'absent_days': {'count': attendance_summary.get('absent_days', 0), 'amount': deduction_details.get('absent_deduction', 0)},
                'late_arrivals': {'count': attendance_summary.get('late_days', 0), 'amount': deduction_details.get('late_arrival_penalty', 0)},
                'early_departure': {'count': attendance_summary.get('early_departure_days', 0), 'amount': deduction_details.get('early_departure_penalty', 0)},
                'pf': deduction_details.get('pf_deduction', 0),
                'esi': deduction_details.get('esi_deduction', 0),
                'professional_tax': deduction_details.get('professional_tax', 0),
                'other': deduction_details.get('other_deductions', 0),
                'total': total_deductions
            }
            attendance_data_record = {
                'days_present': attendance_summary.get('present_days', 0),
                'days_absent': attendance_summary.get('absent_days', 0),
                'days_late': attendance_summary.get('late_days', 0),
                'working_days': working_days
            }
        else:
            # Use fallback data
            record_deductions = {
                'absent_days': {'count': days_absent, 'amount': 0},
                'late_arrivals': {'count': days_late, 'amount': 0},
                'early_departure': {'count': 0, 'amount': 0},
                'pf': pf_deduction,
                'esi': esi_deduction,
                'professional_tax': professional_tax,
                'other': other_deductions,
                'total': total_deductions
            }
            attendance_data_record = {
                'days_present': days_present,
                'days_absent': days_absent,
                'days_late': days_late,
                'working_days': working_days
            }
        
        payroll_records.append({
            'staff_id': staff_id,
            'staff_name': staff['full_name'],
            'department': staff['department'],
            'position': staff['position'],
            'base_salary': base_salary,
            'allowances': {
                'hra': hra_allowance,
                'transport': transport_allowance,
                'other': other_allowances,
                'total': total_allowances
            },
            'gross_pay': gross_pay,
            'deductions': record_deductions,
            'net_payroll': net_payroll,
            'attendance_summary': attendance_data_record
        })

    # Create summary section
    summary_data = {
        'total_payroll_expense': total_payroll_expense,
        'total_staff_count': len(payroll_records),
        'period_covered': period_label,
        'working_days': working_days,
        'department_breakdown': department_totals
    }

    # Generate Excel report with professional formatting
    if format_type == 'excel':
        return _generate_payroll_summary_excel(payroll_records, summary_data, start_date, end_date, department)
    
    # Return JSON for API requests
    return jsonify({
        'success': True,
        'summary': summary_data,
        'payroll_records': payroll_records
    })


def _generate_payroll_summary_excel(payroll_records, summary_data, start_date, end_date, department=None):
    """Generate professionally formatted Excel report for payroll summary"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll Summary Report'

    # Define professional styles
    title_font = Font(bold=True, size=16, color='FFFFFF')
    title_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
    
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    subheader_font = Font(bold=True, size=11, color='FFFFFF')
    subheader_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    summary_font = Font(bold=True, size=11)
    summary_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Row tracker
    current_row = 1

    # TITLE SECTION
    ws.merge_cells(f'A{current_row}:R{current_row}')  # Updated to R to cover 18 columns
    title_cell = ws[f'A{current_row}']
    title_cell.value = 'PAYROLL SUMMARY REPORT'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    title_cell.border = border
    current_row += 2

    # SUMMARY SECTION
    ws.merge_cells(f'A{current_row}:R{current_row}')  # Updated to R to cover 18 columns
    summary_title = ws[f'A{current_row}']
    summary_title.value = 'PAYROLL SUMMARY'
    summary_title.font = header_font
    summary_title.fill = header_fill
    summary_title.alignment = center_align
    summary_title.border = border
    current_row += 1

    # Summary details
    summary_details = [
        ('Period Covered:', summary_data['period_covered']),
        ('Total Staff Count:', str(summary_data['total_staff_count'])),
        ('Total Working Days:', str(summary_data['working_days'])),
        ('Total Payroll Expense:', f"₹{summary_data['total_payroll_expense']:,.2f}"),
        ('Department Filter:', department if department else 'All Departments')
    ]

    for label, value in summary_details:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        ws[f'A{current_row}'].font = summary_font
        ws[f'B{current_row}'].font = Font(size=11)
        ws[f'A{current_row}'].fill = summary_fill
        ws[f'B{current_row}'].fill = summary_fill
        ws[f'A{current_row}'].border = border
        ws[f'B{current_row}'].border = border
        current_row += 1

    current_row += 1

    # DEPARTMENT BREAKDOWN (if multiple departments)
    if len(summary_data['department_breakdown']) > 1:
        ws.merge_cells(f'A{current_row}:D{current_row}')
        dept_title = ws[f'A{current_row}']
        dept_title.value = 'DEPARTMENT BREAKDOWN'
        dept_title.font = subheader_font
        dept_title.fill = subheader_fill
        dept_title.alignment = center_align
        dept_title.border = border
        current_row += 1

        # Department breakdown headers
        dept_headers = ['Department', 'Staff Count', 'Total Amount', 'Average per Staff']
        for col, header in enumerate(dept_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        current_row += 1

        # Department breakdown data
        for dept, data in summary_data['department_breakdown'].items():
            avg_per_staff = data['total'] / data['count'] if data['count'] > 0 else 0
            row_data = [dept, data['count'], f"₹{data['total']:,.2f}", f"₹{avg_per_staff:,.2f}"]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if col > 1:  # Right align numbers
                    cell.alignment = right_align
            current_row += 1

        current_row += 1

    # DETAILED STAFF RECORDS SECTION
    ws.merge_cells(f'A{current_row}:R{current_row}')  # Updated to R to cover 18 columns
    details_title = ws[f'A{current_row}']
    details_title.value = 'DETAILED STAFF PAYROLL RECORDS'
    details_title.font = header_font
    details_title.fill = header_fill
    details_title.alignment = center_align
    details_title.border = border
    current_row += 1

    # Main table headers - Updated to show comprehensive deduction breakdown
    headers = [
        'Staff ID', 'Staff Name', 'Department', 'Position', 'Base Salary',
        'HRA', 'Transport', 'Other Allow.', 'Gross Pay',
        'Absent Ded.', 'Late Penalty', 'Early Dept.', 'PF', 'ESI', 'Prof. Tax', 'Other Ded.', 'Total Ded.', 'Net Pay'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    current_row += 1

    # Data rows with alternating colors
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    total_gross_pay = 0.0
    total_deductions = 0.0
    total_net_pay = 0.0

    for idx, record in enumerate(payroll_records):
        row_fill = light_fill if idx % 2 == 0 else None
        
        # Prepare row data - Updated to match new header structure
        row_data = [
            record['staff_id'],
            record['staff_name'],
            record['department'],
            record['position'],
            f"₹{record['base_salary']:,.2f}",
            f"₹{record['allowances']['hra']:,.2f}",
            f"₹{record['allowances']['transport']:,.2f}",
            f"₹{record['allowances']['other']:,.2f}",
            f"₹{record['gross_pay']:,.2f}",
            f"₹{record['deductions']['absent_days']['amount']:,.2f}",  # Absent Deduction Amount
            f"₹{record['deductions']['late_arrivals']['amount']:,.2f}",  # Late Penalty Amount
            f"₹{record['deductions'].get('early_departure', {}).get('amount', 0):,.2f}",  # Early Departure Penalty
            f"₹{record['deductions']['pf']:,.2f}",
            f"₹{record['deductions']['esi']:,.2f}",
            f"₹{record['deductions']['professional_tax']:,.2f}",
            f"₹{record['deductions']['other']:,.2f}",
            f"₹{record['deductions']['total']:,.2f}",  # Total Deductions (now includes ALL types)
            f"₹{record['net_payroll']:,.2f}"
        ]
        
        # Write row data
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col >= 5:  # Right align numeric columns (salary and deduction columns)
                cell.alignment = right_align
            else:
                cell.alignment = center_align
        
        # Update totals
        total_gross_pay += record['gross_pay']
        total_deductions += record['deductions']['total']
        total_net_pay += record['net_payroll']
        
        current_row += 1

    # TOTALS ROW - Updated to match new column structure
    totals_data = [
        'TOTALS', '', '', '', '',
        '', '', '', f"₹{total_gross_pay:,.2f}",
        '', '', '', '', '', '', f"₹{total_deductions:,.2f}", f"₹{total_net_pay:,.2f}"
    ]
    
    for col, value in enumerate(totals_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = Font(bold=True, size=12)
        cell.fill = summary_fill
        cell.border = border
        cell.alignment = right_align if col >= 5 else center_align

    # Auto-adjust column widths - Updated for new column structure
    column_widths = [12, 25, 18, 18, 15, 12, 12, 12, 15, 12, 12, 12, 12, 12, 12, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save workbook and return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import make_response
    resp = make_response(output.getvalue())
    fname = f"payroll_summary_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    
    return resp

def generate_department_salary_report(school_id, year, department, format_type):
    """Generate comprehensive department-wise salary report with detailed breakdown including attendance-based deductions"""
    import datetime
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import request
    
    db = get_db()
    
    # Get current month as default or from request parameters
    current_month = datetime.datetime.now().month
    current_year = year or datetime.datetime.now().year
    
    # Check if month is specified in request
    month = request.args.get('month', current_month, type=int)
    
    # Build query with optional department filtering
    query = """
        SELECT 
            id, staff_id, full_name, department, destination as position,
            COALESCE(basic_salary, 0) as basic_salary,
            COALESCE(hra, 0) as hra,
            COALESCE(transport_allowance, 0) as transport_allowance,
            COALESCE(other_allowances, 0) as other_allowances,
            COALESCE(pf_deduction, 0) as pf_deduction,
            COALESCE(esi_deduction, 0) as esi_deduction,
            COALESCE(professional_tax, 0) as professional_tax,
            COALESCE(other_deductions, 0) as other_deductions
        FROM staff 
        WHERE school_id = ? AND is_active = 1
    """
    
    params = [school_id]
    
    # Add department filter if specified
    if department and department.strip():
        query += " AND department = ?"
        params.append(department.strip())
    
    query += " ORDER BY CAST(staff_id AS INTEGER) ASC"
    
    staff_data = db.execute(query, params).fetchall()
    
    # Calculate working days for the specified month
    total_days = calendar.monthrange(current_year, month)[1]
    working_days = 0
    current_dt = datetime.date(current_year, month, 1)
    end_dt = datetime.date(current_year, month, total_days)
    
    while current_dt <= end_dt:
        if current_dt.weekday() < 6:  # Monday-Saturday = 0-5, Sunday = 6
            working_days += 1
        current_dt += datetime.timedelta(days=1)
    
    # Group staff by department using SalaryCalculator for consistent calculations
    current_school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=current_school_id)
    
    departments = {}
    for staff in staff_data:
        dept = staff['department'] or 'Unassigned'
        if dept not in departments:
            departments[dept] = []
        
        staff_db_id = staff['id']
        
        # Use SalaryCalculator for consistent calculations
        salary_result = salary_calculator.calculate_monthly_salary(staff_db_id, current_year, month)
        if salary_result['success']:
            breakdown = salary_result['salary_breakdown']
            earnings = breakdown['earnings']
            deductions = breakdown['deductions']
            
            base_salary = earnings['basic_salary']
            total_allowances = earnings['hra'] + earnings['transport_allowance'] + earnings['other_allowances']
            total_deductions = deductions['total_deductions']  # Comprehensive deductions from SalaryCalculator
            gross_pay = earnings['total_earnings']  # Total earnings before deductions
        else:
            # Fallback to basic calculation if SalaryCalculator fails
            base_salary = float(staff['basic_salary'] or 0)
            hra = float(staff['hra'] or 0)
            transport_allowance = float(staff['transport_allowance'] or 0)
            other_allowances = float(staff['other_allowances'] or 0)
            total_allowances = hra + transport_allowance + other_allowances
            
            pf_deduction = float(staff['pf_deduction'] or 0)
            esi_deduction = float(staff['esi_deduction'] or 0)
            professional_tax = float(staff['professional_tax'] or 0)
            other_deductions = float(staff['other_deductions'] or 0)
            total_deductions = pf_deduction + esi_deduction + professional_tax + other_deductions
            gross_pay = base_salary + total_allowances
        
        # Create staff record with calculated values
        if salary_result and salary_result['success']:
            # Use SalaryCalculator data
            attendance = breakdown['attendance_summary']
            staff_record = {
                'staff_id': staff['staff_id'],
                'name': staff['full_name'],
                'position': staff['position'] or 'N/A',
                'base_salary': base_salary,
                'allowances': {
                    'hra': earnings['hra'],
                    'transport': earnings['transport_allowance'],
                    'other': earnings['other_allowances'],
                    'total': total_allowances
                },
                'deductions': {
                    'absent_days': {'count': attendance.get('absent_days', 0), 'amount': deductions.get('absent_deduction', 0)},
                    'late_arrivals': {'count': attendance.get('late_days', 0), 'amount': deductions.get('late_arrival_penalty', 0)},
                    'early_departure': {'count': attendance.get('early_departure_days', 0), 'amount': deductions.get('early_departure_penalty', 0)},
                    'pf': deductions.get('pf_deduction', 0),
                    'esi': deductions.get('esi_deduction', 0),
                    'professional_tax': deductions.get('professional_tax', 0),
                    'other': deductions.get('other_deductions', 0),
                    'total': total_deductions
                },
                'gross_pay': gross_pay
            }
        else:
            # Use fallback data
            staff_record = {
                'staff_id': staff['staff_id'],
                'name': staff['full_name'],
                'position': staff['position'] or 'N/A',
                'base_salary': base_salary,
                'allowances': {
                    'hra': hra,
                    'transport': transport_allowance,
                    'other': other_allowances,
                    'total': total_allowances
                },
                'deductions': {
                    'absent_days': {'count': 0, 'amount': 0},
                    'late_arrivals': {'count': 0, 'amount': 0},
                    'early_departure': {'count': 0, 'amount': 0},
                    'pf': pf_deduction,
                    'esi': esi_deduction,
                    'professional_tax': professional_tax,
                    'other': other_deductions,
                    'total': total_deductions
                },
                'gross_pay': gross_pay
            }
        departments[dept].append(staff_record)
    
    if format_type == 'json':
        return jsonify({
            'report_type': 'Department Wise Salary Report',
            'school_id': school_id,
            'year': year,
            'generated_at': datetime.datetime.now().isoformat(),
            'departments': departments,
            'summary': {
                'total_departments': len(departments),
                'total_staff': sum(len(dept_staff) for dept_staff in departments.values()),
                'total_gross_pay': sum(
                    sum(staff['gross_pay'] for staff in dept_staff) 
                    for dept_staff in departments.values()
                )
            }
        })
    
    elif format_type == 'excel':
        return _generate_department_salary_excel(departments, school_id, year)
    
    else:
        return jsonify({'error': 'Unsupported format type'}), 400

def _generate_department_salary_excel(departments, school_id, year):
    """Generate professionally formatted Excel file for department-wise salary report"""
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import make_response
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Department Wise Salary Report"
    
    # Define styles
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    dept_header_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    dept_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    data_font = Font(name='Arial', size=11)
    odd_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    current_row = 1
    
    # Title section - Updated to cover 15 columns (A to O)
    ws.merge_cells(f'A{current_row}:P{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"DEPARTMENT WISE SALARY REPORT - {year}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    current_row += 2
    
    # Report info
    ws[f'A{current_row}'] = f"Generated on: {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws[f'A{current_row}'].font = Font(name='Arial', size=10, italic=True)
    current_row += 1
    
    ws[f'A{current_row}'] = f"School ID: {school_id}"
    ws[f'A{current_row}'].font = Font(name='Arial', size=10, italic=True)
    current_row += 2
    
    # Calculate totals for summary
    total_staff = sum(len(dept_staff) for dept_staff in departments.values())
    total_gross_pay = sum(
        sum(staff['gross_pay'] for staff in dept_staff) 
        for dept_staff in departments.values()
    )
    
    # Summary section
    ws[f'A{current_row}'] = "SUMMARY"
    ws[f'A{current_row}'].font = Font(name='Arial', size=12, bold=True)
    current_row += 1
    
    ws[f'A{current_row}'] = f"Total Departments: {len(departments)}"
    ws[f'C{current_row}'] = f"Total Staff: {total_staff}"
    ws[f'E{current_row}'] = f"Total Gross Pay: ₹{total_gross_pay:,.2f}"
    
    for cell in [ws[f'A{current_row}'], ws[f'C{current_row}'], ws[f'E{current_row}']]:
        cell.font = Font(name='Arial', size=10, bold=True)
    current_row += 2
    
    # Main table headers - Updated to include attendance-based deductions
    headers = [
        'Staff ID', 'Staff Name', 'Position/Job Title', 'Base Salary', 
        'HRA', 'Transport', 'Other Allow.', 'Absent Ded.', 'Late Penalty', 'Early Dept.',
        'PF Deduction', 'ESI Deduction', 'Prof. Tax', 'Other Ded.', 'Total Deductions', 'Gross Pay'
    ]
    
    # Process each department
    for dept_name in sorted(departments.keys()):
        dept_staff = departments[dept_name]
        
        # Department header - Updated to cover 16 columns (A to P)
        ws.merge_cells(f'A{current_row}:P{current_row}')
        dept_header_cell = ws[f'A{current_row}']
        dept_header_cell.value = f"{dept_name.upper()} DEPARTMENT ({len(dept_staff)} Staff)"
        dept_header_cell.font = dept_header_font
        dept_header_cell.fill = dept_header_fill
        dept_header_cell.alignment = center_alignment
        dept_header_cell.border = thin_border
        current_row += 1
        
        # Column headers for this department
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # Department staff data
        dept_total_gross = 0
        for i, staff in enumerate(dept_staff):
            row_data = [
                staff['staff_id'],
                staff['name'],
                staff['position'],
                f"₹{staff['base_salary']:,.2f}",
                f"₹{staff['allowances']['hra']:,.2f}",
                f"₹{staff['allowances']['transport']:,.2f}",
                f"₹{staff['allowances']['other']:,.2f}",
                f"₹{staff['deductions']['absent_days']['amount']:,.2f}",  # Absent Deduction
                f"₹{staff['deductions']['late_arrivals']['amount']:,.2f}",  # Late Penalty
                f"₹{staff['deductions'].get('early_departure', {}).get('amount', 0):,.2f}",  # Early Departure Penalty
                f"₹{staff['deductions']['pf']:,.2f}",
                f"₹{staff['deductions']['esi']:,.2f}",
                f"₹{staff['deductions']['professional_tax']:,.2f}",
                f"₹{staff['deductions']['other']:,.2f}",
                f"₹{staff['deductions']['total']:,.2f}",  # Total Deductions (now includes ALL types)
                f"₹{staff['gross_pay']:,.2f}"
            ]
            
            dept_total_gross += staff['gross_pay']
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border
                
                # Alternate row coloring
                if i % 2 == 1:
                    cell.fill = odd_row_fill
                    
                # Alignment
                if col <= 3:  # Text columns
                    cell.alignment = left_alignment
                else:  # Number columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            current_row += 1
        
        # Department total row - Updated to cover columns A to N
        ws.merge_cells(f'A{current_row}:N{current_row}')
        total_cell = ws[f'A{current_row}']
        total_cell.value = f"DEPARTMENT TOTAL"
        total_cell.font = Font(name='Arial', size=11, bold=True)
        total_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        total_cell.alignment = center_alignment
        total_cell.border = thin_border
        
        gross_total_cell = ws[f'O{current_row}']  # Updated to column O (15th column)
        gross_total_cell.value = f"₹{dept_total_gross:,.2f}"
        gross_total_cell.font = Font(name='Arial', size=11, bold=True)
        gross_total_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        gross_total_cell.alignment = Alignment(horizontal='right', vertical='center')
        gross_total_cell.border = thin_border
        
        current_row += 2  # Space between departments
    
    # Adjust column widths - Updated for 15 columns
    column_widths = [12, 25, 20, 15, 12, 12, 12, 12, 12, 12, 12, 12, 12, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=department_salary_report_{year}.xlsx'
    
    return response

def generate_department_analysis_report(school_id, year=None, month=None, format_type='excel'):
    """Generate comprehensive Department Report with multi-format export.
    Includes: total staff per department, gender + position breakdown,
    average tenure, salary min/avg/max, and attendance stats for selected month.
    """
    import datetime, calendar, io

    db = get_db()

    # Resolve period (attendance window + tenure as of end of month)
    today = datetime.date.today()
    if not isinstance(year, int) or year <= 0:
        year = today.year
    if not isinstance(month, int) or not (1 <= month <= 12):
        month = today.month
    start_date = datetime.date(year, month, 1)
    end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')

    # Departments
    departments = [r['dept'] for r in db.execute(
        """
        SELECT DISTINCT department AS dept
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        ORDER BY department
        """, (school_id,)
    ).fetchall()]

    # Totals
    total_map = {r['department']: r['total'] for r in db.execute(
        """
        SELECT department, COUNT(*) AS total
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        GROUP BY department
        """, (school_id,)
    ).fetchall()}

    # Gender breakdown
    gender_map = {}
    for g in db.execute(
        """
        SELECT department, COALESCE(gender, 'Other') AS gender, COUNT(*) AS cnt
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        GROUP BY department, COALESCE(gender, 'Other')
        """, (school_id,)
    ).fetchall():
        dept = g['department']
        gender_map.setdefault(dept, {'Male': 0, 'Female': 0, 'Other': 0})
        key = g['gender'] if g['gender'] in ('Male', 'Female', 'Other') else 'Other'
        gender_map[dept][key] += g['cnt']

    # Position breakdown rows (for sheet)
    position_rows = [
        {'department': p['department'], 'position': p['position'] or 'Unspecified', 'count': p['cnt']}
        for p in db.execute(
            """
            SELECT department, COALESCE(position, 'Unspecified') AS position, COUNT(*) AS cnt
            FROM staff
            WHERE school_id = ? AND COALESCE(department, '') <> ''
            GROUP BY department, COALESCE(position, 'Unspecified')
            ORDER BY department, position
            """, (school_id,)
        ).fetchall()
    ]

    # Salary stats (convert sqlite3.Row -> plain dict)
    salary_rows = db.execute(
        """
        SELECT department,
               MIN(COALESCE(basic_salary, 0)) AS min_salary,
               AVG(COALESCE(basic_salary, 0)) AS avg_salary,
               MAX(COALESCE(basic_salary, 0)) AS max_salary
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        GROUP BY department
        """, (school_id,)
    ).fetchall()
    salary_map = {}
    for s in salary_rows:
        dept = s['department']
        salary_map[dept] = {
            'min_salary': float(s['min_salary'] or 0),
            'avg_salary': float(s['avg_salary'] or 0),
            'max_salary': float(s['max_salary'] or 0),
        }

    # Tenure as of end_date
    from collections import defaultdict
    tenure_sum_days, tenure_count = defaultdict(int), defaultdict(int)
    for r in db.execute(
        """
        SELECT department, date_of_joining
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        """, (school_id,)
    ).fetchall():
        doj = r['date_of_joining']
        if doj:
            try:
                doj_dt = datetime.datetime.strptime(doj, '%Y-%m-%d').date()
                days = (end_date - doj_dt).days
                if days >= 0:
                    tenure_sum_days[r['department']] += days
                    tenure_count[r['department']] += 1
            except Exception:
                pass
    avg_tenure_years = {d: (tenure_sum_days[d]/tenure_count[d]/365.0) if tenure_count[d] else 0.0 for d in departments}

    # Attendance for selected month
    # Build attendance_map as plain dicts to avoid sqlite3.Row .get errors
    attendance_rows = db.execute(
        """
        SELECT s.department AS department,
               SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) AS present_count,
               SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) AS absent_count,
               SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) AS late_count,
               SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) AS leave_count
        FROM staff s
        LEFT JOIN attendance a ON a.staff_id = s.id AND a.date BETWEEN ? AND ?
        WHERE s.school_id = ? AND COALESCE(s.department, '') <> ''
        GROUP BY s.department
        ORDER BY s.department
        """, (start_str, end_str, school_id)
    ).fetchall()
    attendance_map = {}
    for a in attendance_rows:
        attendance_map[a['department']] = {
            'present_count': int(a['present_count'] or 0),
            'absent_count': int(a['absent_count'] or 0),
            'late_count': int(a['late_count'] or 0),
            'leave_count': int(a['leave_count'] or 0),
        }

    # Summary rows
    summary = []
    for dept in departments:
        g = gender_map.get(dept, {'Male': 0, 'Female': 0, 'Other': 0})
        sstat = salary_map.get(dept, {'min_salary': 0, 'avg_salary': 0, 'max_salary': 0})
        att = attendance_map.get(dept, {'present_count': 0, 'absent_count': 0, 'late_count': 0, 'leave_count': 0})
        present = int(att.get('present_count', 0) or 0)
        absent = int(att.get('absent_count', 0) or 0)
        late = int(att.get('late_count', 0) or 0)
        leave = int(att.get('leave_count', 0) or 0)
        denom = present + absent + late + leave
        present_rate = (present/denom*100.0) if denom else 0.0
        summary.append({
            'department': dept,
            'total_staff': int(total_map.get(dept, 0)),
            'male': int(g.get('Male', 0) or 0),
            'female': int(g.get('Female', 0) or 0),
            'other': int(g.get('Other', 0) or 0),
            'avg_tenure_years': round(avg_tenure_years.get(dept, 0.0), 2),
            'min_salary': float(sstat.get('min_salary', 0) or 0),
            'avg_salary': float(sstat.get('avg_salary', 0) or 0),
            'max_salary': float(sstat.get('max_salary', 0) or 0),
            'present': present, 'absent': absent, 'late': late, 'leave': leave,
            'present_rate': round(present_rate, 2)
        })

    # CSV
    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"Department Report - {calendar.month_name[month]} {year}"])
        writer.writerow([])
        headers = ['Department','Total Staff','Male','Female','Other','Avg Tenure (yrs)',
                   'Min Salary','Avg Salary','Max Salary','Present','Absent','Late','Leave','Present Rate (%)']
        writer.writerow(headers)
        for row in summary:
            writer.writerow([
                row['department'], row['total_staff'], row['male'], row['female'], row['other'], row['avg_tenure_years'],
                f"{row['min_salary']:.2f}", f"{row['avg_salary']:.2f}", f"{row['max_salary']:.2f}",
                row['present'], row['absent'], row['late'], row['leave'], row['present_rate']
            ])
        resp = make_response(output.getvalue())
        fname = f"department_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
        resp.headers['Content-Type'] = 'text/csv'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    # PDF
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []
            elements.append(Paragraph(f"Department Report - {calendar.month_name[month]} {year}", styles['Title']))
            elements.append(Spacer(1, 12))
            headers = ['Department','Total','Male','Female','Other','AvgTenure','MinSal','AvgSal','MaxSal','Present','Absent','Late','Leave','%Present']
            data = [headers]
            for r in summary:
                data.append([
                    r['department'], r['total_staff'], r['male'], r['female'], r['other'],
                    f"{r['avg_tenure_years']:.2f}", f"{r['min_salary']:.0f}", f"{r['avg_salary']:.0f}", f"{r['max_salary']:.0f}",
                    r['present'], r['absent'], r['late'], r['leave'], f"{r['present_rate']:.1f}%"
                ])
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(table)
            doc.build(elements)
            pdf = buffer.getvalue(); buffer.close()
            resp = make_response(pdf)
            fname = f"department_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp
        except ImportError:
            return jsonify({'success': False, 'error': 'PDF generation requires reportlab. Install with: pip install reportlab'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel with Summary, Positions, Staff Details
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Summary'

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Department','Total Staff','Male','Female','Other','Avg Tenure (yrs)',
               'Min Salary','Avg Salary','Max Salary','Present','Absent','Late','Leave','Present Rate (%)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h); c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = Alignment(horizontal='center')

    r = 2
    for row in summary:
        vals = [row['department'], row['total_staff'], row['male'], row['female'], row['other'], row['avg_tenure_years'],
                row['min_salary'], row['avg_salary'], row['max_salary'], row['present'], row['absent'], row['late'], row['leave'], row['present_rate']]
        for cidx, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=cidx, value=val); cell.border = border
        r += 1
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Positions sheet
    ws2 = wb.create_sheet(title='Positions')
    pos_headers = ['Department','Position','Count']
    for col, h in enumerate(pos_headers, 1):
        c = ws2.cell(row=1, column=col, value=h); c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = Alignment(horizontal='center')
    r = 2
    for pr in position_rows:
        ws2.cell(row=r, column=1, value=pr['department'])
        ws2.cell(row=r, column=2, value=pr['position'])
        ws2.cell(row=r, column=3, value=pr['count'])
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = border
        r += 1
    for col in range(1, 4):
        ws2.column_dimensions[get_column_letter(col)].width = 24

    # Staff Details sheet
    staff_details = db.execute(
        """
        SELECT department, staff_id, full_name, COALESCE(destination,'Unspecified') AS position,
               COALESCE(gender,'Other') AS gender, date_of_joining, COALESCE(basic_salary,0) AS basic_salary
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        ORDER BY CAST(staff_id AS INTEGER) ASC
        """, (school_id,)
    ).fetchall()
    ws3 = wb.create_sheet(title='Staff Details')
    sd_headers = ['Department','Staff ID','Full Name','Position','Gender','Date of Joining','Basic Salary']
    for col, h in enumerate(sd_headers, 1):
        c = ws3.cell(row=1, column=col, value=h); c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = Alignment(horizontal='center')
    r = 2
    for s in staff_details:
        vals = [s['department'], s['staff_id'], s['full_name'], s['position'], s['gender'], s['date_of_joining'], float(s['basic_salary'] or 0)]
        for cidx, val in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=cidx, value=val); cell.border = border
        r += 1
    widths = [18, 14, 26, 18, 12, 16, 14]
    for col, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    output = BytesIO(); wb.save(output); output.seek(0)
    resp = make_response(output.getvalue())
    fname = f"department_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

def generate_performance_report(school_id, year=None, month=None, department=None, format_type='excel'):
    """Generate enhanced Staff Performance Metrics and Evaluations Report.
    
    Includes the specific fields requested by user:
    - Staff ID, Name, Department, Position
    - Days Present (count)
    - Days Absent (count) 
    - Days on OD (Official Duty) applied
    - Days on Leave applied
    - Days with Permission applied
    
    Supports date range filtering and Excel export.
    """
    import datetime, calendar, io
    from flask import request

    db = get_db()

    # Handle date range from request parameters
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    # Determine date range for the report
    if from_date_str and to_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            # Fallback to month/year if date parsing fails
            today = datetime.date.today()
            if not isinstance(year, int) or year <= 0:
                year = today.year
            if not isinstance(month, int) or not (1 <= month <= 12):
                month = today.month
            start_date = datetime.date(year, month, 1)
            end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    elif year and month:
        start_date = datetime.date(year, month, 1)
        end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    else:
        # Default to current month
        today = datetime.date.today()
        start_date = datetime.date(today.year, today.month, 1)
        end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    
    start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
    period_label = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"

    # Build department filter clause
    dept_clause = ''
    params_base = [start_str, end_str, school_id]
    if department and str(department).strip():
        dept_clause = ' AND COALESCE(s.department, "") = ? '
        params_base.append(department)

    # Main query to get staff performance data with all required fields
    staff_performance_query = f"""
        SELECT 
            s.id as staff_db_id,
            s.staff_id,
            s.full_name,
            COALESCE(NULLIF(TRIM(s.department), ''), 'Unassigned') as department,
            COALESCE(NULLIF(TRIM(s.destination), ''), 'Not Specified') as position,
            s.date_of_joining
            
        FROM staff s
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    """

    # Adjust parameters for staff query (remove date parameters)
    staff_params = [school_id]
    if department and str(department).strip():
        staff_params.append(department)
    
    staff_rows = db.execute(staff_performance_query, tuple(staff_params)).fetchall()

    # Now get attendance data for each staff member separately for accuracy
    performance_rows = []
    
    for staff in staff_rows:
        staff_id = staff['staff_id']
        staff_db_id = staff['staff_db_id']
        
        # Get attendance data for this specific staff member
        attendance_data = db.execute("""
            SELECT 
                COUNT(*) as total_records,
                SUM(CASE WHEN status IN ('present', 'late') THEN 1 ELSE 0 END) as days_present,
                SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as days_absent,
                SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as days_late,
                SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) as days_present_only,
                SUM(CASE WHEN status = 'early_departure' THEN 1 ELSE 0 END) as days_early_departure
            FROM attendance 
            WHERE staff_id = ? AND date BETWEEN ? AND ?
        """, (staff_db_id, start_str, end_str)).fetchone()
        
        days_present = int(attendance_data['days_present'] or 0)
        days_absent = int(attendance_data['days_absent'] or 0)
        days_late = int(attendance_data['days_late'] or 0)
        days_present_only = int(attendance_data['days_present_only'] or 0)
        total_attendance_records = int(attendance_data['total_records'] or 0)
        
        performance_rows.append({
            'staff_id': staff['staff_id'],
            'staff_db_id': staff_db_id,
            'staff_name': staff['full_name'],
            'department': staff['department'],
            'position': staff['position'],
            'days_present': days_present,
            'days_absent': days_absent,
            'days_late': days_late,
            'days_present_only': days_present_only,
            'total_attendance_records': total_attendance_records
        })

    # Get Leave Applications data for all staff
    leave_query = f"""
        SELECT 
            s.staff_id,
            COALESCE(COUNT(CASE WHEN l.status = 'approved' THEN l.id END), 0) as approved_leave_count,
            COALESCE(SUM(CASE WHEN l.status = 'approved' 
                        THEN julianday(l.end_date) - julianday(l.start_date) + 1 
                        ELSE 0 END), 0) as days_on_leave
        FROM staff s
        LEFT JOIN leave_applications l ON s.id = l.staff_id 
            AND NOT (l.end_date < ? OR l.start_date > ?)
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        GROUP BY s.id, s.staff_id
    """
    
    leave_params = [start_str, end_str, school_id]
    if department and str(department).strip():
        leave_params.append(department)
    
    leave_data = db.execute(leave_query, tuple(leave_params)).fetchall()
    leave_map = {row['staff_id']: {
        'applications': int(row['approved_leave_count']), 
        'days': float(row['days_on_leave'])
    } for row in leave_data}

    # Get Official Duty (OD) Applications data
    od_query = f"""
        SELECT 
            s.staff_id,
            COALESCE(COUNT(CASE WHEN od.status = 'approved' THEN od.id END), 0) as approved_od_count,
            COALESCE(SUM(CASE WHEN od.status = 'approved' 
                        THEN julianday(od.end_date) - julianday(od.start_date) + 1 
                        ELSE 0 END), 0) as days_on_od
        FROM staff s
        LEFT JOIN on_duty_applications od ON s.id = od.staff_id 
            AND NOT (od.end_date < ? OR od.start_date > ?)
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        GROUP BY s.id, s.staff_id
    """
    
    od_params = [start_str, end_str, school_id]
    if department and str(department).strip():
        od_params.append(department)
    
    od_data = db.execute(od_query, tuple(od_params)).fetchall()
    od_map = {row['staff_id']: {
        'applications': int(row['approved_od_count']), 
        'days': float(row['days_on_od'])
    } for row in od_data}

    # Get Permission Applications data
    permission_query = f"""
        SELECT 
            s.staff_id,
            COALESCE(COUNT(CASE WHEN p.status = 'approved' THEN p.id END), 0) as approved_permission_count,
            COALESCE(SUM(CASE WHEN p.status = 'approved' THEN p.duration_hours ELSE 0 END), 0) as total_permission_hours
        FROM staff s
        LEFT JOIN permission_applications p ON s.id = p.staff_id 
            AND p.permission_date BETWEEN ? AND ?
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        GROUP BY s.id, s.staff_id
    """
    
    permission_params = [start_str, end_str, school_id]
    if department and str(department).strip():
        permission_params.append(department)
    
    permission_data = db.execute(permission_query, tuple(permission_params)).fetchall()
    permission_map = {row['staff_id']: {
        'applications': int(row['approved_permission_count']), 
        'hours': float(row['total_permission_hours'])
    } for row in permission_data}

    # Calculate working days in the period (excluding weekends and holidays)
    total_days = (end_date - start_date).days + 1
    working_days = 0
    current_dt = start_date
    weekend_days = 0
    
    # Get holidays in the period
    holidays_in_period = db.execute("""
        SELECT start_date, end_date, holiday_name 
        FROM holidays 
        WHERE school_id = ? AND is_active = 1
        AND NOT (end_date < ? OR start_date > ?)
    """, (school_id, start_str, end_str)).fetchall()
    
    # Create set of holiday dates
    holiday_dates = set()
    for holiday in holidays_in_period:
        h_start = datetime.datetime.strptime(holiday['start_date'], '%Y-%m-%d').date()
        h_end = datetime.datetime.strptime(holiday['end_date'], '%Y-%m-%d').date()
        curr_date = h_start
        while curr_date <= h_end:
            holiday_dates.add(curr_date)
            curr_date += datetime.timedelta(days=1)
    
    # Count working days (Mon-Sat, excluding holidays and Sundays)
    while current_dt <= end_date:
        if current_dt.weekday() < 6:  # Monday-Saturday = 0-5, Sunday = 6
            if current_dt not in holiday_dates:
                working_days += 1
        else:
            weekend_days += 1
        current_dt += datetime.timedelta(days=1)
    
    holiday_days = len(holiday_dates)

    # Combine all data into final performance report rows
    final_performance_rows = []
    for perf in performance_rows:
        staff_id = perf['staff_id']
        
        # Get leave data for this staff
        leave_info = leave_map.get(staff_id, {'applications': 0, 'days': 0})
        
        # Get OD data for this staff
        od_info = od_map.get(staff_id, {'applications': 0, 'days': 0})
        
        # Get permission data for this staff
        permission_info = permission_map.get(staff_id, {'applications': 0, 'hours': 0})
        permission_days = round(float(permission_info['hours']) / 8, 1) if permission_info['hours'] else 0
        
        # Calculate correct absent days
        # Absent = Working days - Present days - Approved leave days - Approved OD days
        days_present = perf['days_present']
        days_on_leave = int(leave_info['days'])
        days_on_od = int(od_info['days'])
        
        # Calculate actual absent days (excluding leave and OD)
        calculated_absent_days = max(0, working_days - days_present - days_on_leave - days_on_od)
        
        final_performance_rows.append({
            'staff_id': perf['staff_id'],
            'staff_name': perf['staff_name'],
            'department': perf['department'],
            'position': perf['position'],
            'days_present': days_present,
            'days_absent': calculated_absent_days,  # Use calculated absent days
            'days_on_od_applied': days_on_od,
            'days_on_leave_applied': days_on_leave,
            'days_with_permission_applied': permission_days,
            'total_working_days': working_days,
            'total_attendance_records': perf['total_attendance_records'],
            'days_late': perf['days_late'],
            'days_present_only': perf['days_present_only']
        })

    # CSV export
    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"Staff Performance Metrics Report"])
        writer.writerow([f"Period: {period_label}"])
        if department:
            writer.writerow([f"Department: {department}"])
        writer.writerow([])
        writer.writerow(['Staff ID', 'Staff Name', 'Department', 'Position', 'Days Present', 'Days Absent', 'Days Late', 'Days on OD Applied', 'Days on Leave Applied', 'Days with Permission Applied', 'Total Working Days', 'Attendance Records'])
        
        for row in final_performance_rows:
            writer.writerow([
                row['staff_id'], row['staff_name'], row['department'], row['position'],
                row['days_present'], row['days_absent'], row['days_late'], row['days_on_od_applied'],
                row['days_on_leave_applied'], row['days_with_permission_applied'],
                row['total_working_days'], row['total_attendance_records']
            ])
        
        # Add summary totals
        if final_performance_rows:
            writer.writerow([])
            writer.writerow(['SUMMARY TOTALS'])
            total_staff = len(final_performance_rows)
            total_present = sum(row['days_present'] for row in final_performance_rows)
            total_absent = sum(row['days_absent'] for row in final_performance_rows)
            total_late = sum(row['days_late'] for row in final_performance_rows)
            total_od = sum(row['days_on_od_applied'] for row in final_performance_rows)
            total_leave = sum(row['days_on_leave_applied'] for row in final_performance_rows)
            total_permission = sum(row['days_with_permission_applied'] for row in final_performance_rows)
            total_attendance_records = sum(row['total_attendance_records'] for row in final_performance_rows)
            
            writer.writerow([
                f'Total Staff: {total_staff}', '', '', '',
                total_present, total_absent, total_late, total_od,
                total_leave, total_permission, working_days, total_attendance_records
            ])
        
        resp = make_response(output.getvalue())
        fname = f"staff_performance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
        resp.headers['Content-Type'] = 'text/csv'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    # PDF export
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []
            
            elements.append(Paragraph(f"Staff Performance Metrics Report", styles['Title']))
            elements.append(Paragraph(f"Period: {period_label}", styles['Normal']))
            if department:
                elements.append(Paragraph(f"Department: {department}", styles['Normal']))
            elements.append(Spacer(1, 12))
            
            # Create table data
            headers = ['Staff ID', 'Staff Name', 'Department', 'Position', 'Days Present', 'Days Absent', 'Days Late', 'Days on OD', 'Days on Leave', 'Permission (Days)', 'Working Days', 'Records']
            table_data = [headers]
            
            for row in final_performance_rows:
                table_data.append([
                    row['staff_id'], row['staff_name'], row['department'], row['position'],
                    str(row['days_present']), str(row['days_absent']), str(row['days_late']),
                    str(row['days_on_od_applied']), str(row['days_on_leave_applied']), 
                    str(row['days_with_permission_applied']), str(row['total_working_days']),
                    str(row['total_attendance_records'])
                ])
            
            # Add summary totals row
            if final_performance_rows:
                total_staff = len(final_performance_rows)
                total_present = sum(row['days_present'] for row in final_performance_rows)
                total_absent = sum(row['days_absent'] for row in final_performance_rows)
                total_late = sum(row['days_late'] for row in final_performance_rows)
                total_od = sum(row['days_on_od_applied'] for row in final_performance_rows)
                total_leave = sum(row['days_on_leave_applied'] for row in final_performance_rows)
                total_permission = sum(row['days_with_permission_applied'] for row in final_performance_rows)
                total_attendance_records = sum(row['total_attendance_records'] for row in final_performance_rows)
                
                table_data.append([
                    f'TOTALS ({total_staff} Staff)', '', '', '',
                    str(total_present), str(total_absent), str(total_late), str(total_od),
                    str(total_leave), str(total_permission), str(working_days), str(total_attendance_records)
                ])
            
            table = Table(table_data, repeatRows=1)
            # Style the table
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]
            
            # If there are totals, highlight the last row
            if final_performance_rows and len(table_data) > len(final_performance_rows) + 1:
                last_row = len(table_data) - 1
                table_style.extend([
                    ('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#E6F3FF')),
                    ('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'),
                ])
            
            table.setStyle(TableStyle(table_style))
            elements.append(table)
            
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            
            resp = make_response(pdf)
            fname = f"staff_performance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
            resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp
        except ImportError:
            return jsonify({'success': False, 'error': 'PDF generation requires reportlab. Install with: pip install reportlab'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel export  
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Staff Performance Report'

    # Styling
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Report title and metadata
    ws.cell(row=1, column=1, value=f'Staff Performance Metrics Report').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'Period: {period_label}').font = Font(bold=True, size=11)
    if department:
        ws.cell(row=3, column=1, value=f'Department: {department}').font = Font(bold=True, size=11)
        header_row = 5
    else:
        header_row = 4

    # Column headers
    headers = ['Staff ID', 'Staff Name', 'Department', 'Position', 'Days Present', 'Days Absent', 'Days Late', 'Days on OD', 'Days on Leave', 'Permission (Days)', 'Total Working Days', 'Attendance Records']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    data_start_row = header_row + 1
    for row_idx, data in enumerate(final_performance_rows, data_start_row):
        values = [
            data['staff_id'], data['staff_name'], data['department'], data['position'],
            data['days_present'], data['days_absent'], data['days_late'],
            data['days_on_od_applied'], data['days_on_leave_applied'], 
            data['days_with_permission_applied'], data['total_working_days'],
            data['total_attendance_records']
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col >= 5:  # Numeric columns - center align
                cell.alignment = Alignment(horizontal='center')

    # Add summary totals row
    if final_performance_rows:
        summary_row = data_start_row + len(final_performance_rows) + 1
        
        # Calculate totals
        total_staff = len(final_performance_rows)
        total_present = sum(row['days_present'] for row in final_performance_rows)
        total_absent = sum(row['days_absent'] for row in final_performance_rows)
        total_late = sum(row['days_late'] for row in final_performance_rows)
        total_od = sum(row['days_on_od_applied'] for row in final_performance_rows)
        total_leave = sum(row['days_on_leave_applied'] for row in final_performance_rows)
        total_permission = sum(row['days_with_permission_applied'] for row in final_performance_rows)
        total_attendance_records = sum(row['total_attendance_records'] for row in final_performance_rows)
        
        # Summary header
        summary_header_cell = ws.cell(row=summary_row, column=1, value="SUMMARY TOTALS")
        summary_header_cell.font = Font(bold=True, size=12)
        summary_header_cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        
        # Summary values
        summary_values = [
            f"Total Staff: {total_staff}", "", "", "",
            total_present, total_absent, total_late, total_od, 
            total_leave, total_permission, working_days, total_attendance_records
        ]
        
        for col, value in enumerate(summary_values, 1):
            cell = ws.cell(row=summary_row, column=col, value=value)
            if col >= 5:  # Numeric columns
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            cell.border = border

    # Auto-adjust column widths
    column_widths = [12, 25, 18, 18, 12, 12, 12, 12, 14, 16, 16, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save workbook and return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    fname = f"staff_performance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


def generate_daily_attendance_report(school_id, date_str=None, department=None, format_type='excel'):
    """Generate comprehensive Daily Attendance Report (multi-format).
    Includes per-staff records for selected date, department summary, and overall stats.
    """
    import datetime, io

    db = get_db()

    # Resolve date
    try:
        if not date_str:
            date_obj = datetime.date.today()
        else:
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        date_obj = datetime.date.today()
    date_str = date_obj.strftime('%Y-%m-%d')

    # Check attendance table columns to avoid SQL errors on older schemas
    try:
        cols = {row['name'] for row in db.execute("PRAGMA table_info(attendance)").fetchall()}
    except Exception:
        cols = set()
    has_work_hours = 'work_hours' in cols
    has_overtime = 'overtime_hours' in cols
    has_late_min = 'late_duration_minutes' in cols
    has_early_min = 'early_departure_minutes' in cols

    # Optional department filter
    dept_clause = ''
    params = [date_str, school_id]
    if department and str(department).strip():
        dept_clause = ' AND COALESCE(s.department, "") = ?'
        params.append(department.strip())

    select_fields = [
        's.staff_id',
        's.full_name',
        's.department',
        "COALESCE(s.destination, '') AS position",
        'a.time_in',
        'a.time_out',
        'a.status',
    ]
    if has_work_hours:
        select_fields.append('a.work_hours')
    else:
        select_fields.append('0 AS work_hours')
    if has_overtime:
        select_fields.append('a.overtime_hours')
    else:
        select_fields.append('0 AS overtime_hours')
    if has_late_min:
        select_fields.append('a.late_duration_minutes')
    else:
        select_fields.append('0 AS late_duration_minutes')
    if has_early_min:
        select_fields.append('a.early_departure_minutes')
    else:
        select_fields.append('0 AS early_departure_minutes')

    query = f"""
        SELECT {', '.join(select_fields)}
        FROM staff s
        LEFT JOIN attendance a ON a.staff_id = s.id AND a.date = ?
        WHERE s.school_id = ? {dept_clause}
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    """
    rows = db.execute(query, tuple(params)).fetchall()

    # Normalize rows and build summaries
    def fmt_time(t):
        return t if (t and isinstance(t, str)) else (t.strftime('%H:%M') if t else '')

    staff_records = []
    dept_summary = {}
    overall = {k: 0 for k in ['present','absent','late','leave','on_duty','holiday']}
    total_staff = 0

    for r in rows:
        total_staff += 1
        status = (r['status'] or '').lower() if r['status'] else 'absent'
        if status not in overall:
            # Map unknown to absent by default
            status = 'absent'
        overall[status] += 1

        dept = r['department'] or 'Unassigned'
        ds = dept_summary.setdefault(dept, {k: 0 for k in ['present','absent','late','leave','on_duty','holiday']})
        ds[status] += 1

        staff_records.append({
            'staff_id': r['staff_id'],
            'full_name': r['full_name'],
            'department': dept,
            'position': r['position'],
            'time_in': r['time_in'] or '',
            'time_out': r['time_out'] or '',
            'status': status.capitalize(),
            'late_minutes': int(r['late_duration_minutes'] or 0),
            'early_departure_minutes': int(r['early_departure_minutes'] or 0),
            'work_hours': float(r['work_hours'] or 0),
            'overtime_hours': float(r['overtime_hours'] or 0),
        })

    # Compute department rows with totals and present rate
    dept_rows = []
    for dept, ds in sorted(dept_summary.items()):
        total = sum(ds.values())
        present = ds['present'] + ds['on_duty'] + ds['late']  # treat on_duty/late as attended
        present_rate = (present / total * 100.0) if total else 0.0
        dept_rows.append({
            'department': dept,
            'total': total,
            **ds,
            'present_rate': round(present_rate, 2)
        })

    date_label = date_obj.strftime('%Y-%m-%d')

    # CSV export
    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"Daily Attendance Report - {date_label}"])
        writer.writerow([])
        # Overall summary
        writer.writerow(['Total Staff','Present','Late','Leave','On Duty','Holiday','Absent'])
        writer.writerow([total_staff, overall['present'], overall['late'], overall['leave'], overall['on_duty'], overall['holiday'], overall['absent']])
        writer.writerow([])
        # Department summary
        writer.writerow(['Department','Total','Present','Late','Leave','On Duty','Holiday','Absent','Present Rate (%)'])
        for d in dept_rows:
            writer.writerow([d['department'], d['total'], d['present'], d['late'], d['leave'], d['on_duty'], d['holiday'], d['absent'], d['present_rate']])
        writer.writerow([])
        # Staff records
        writer.writerow(['Staff ID','Full Name','Department','Position','Time In','Time Out','Status','Late (min)','Early Dep (min)','Work Hrs','OT Hrs'])
        for s in staff_records:
            writer.writerow([s['staff_id'], s['full_name'], s['department'], s['position'], s['time_in'], s['time_out'], s['status'], s['late_minutes'], s['early_departure_minutes'], s['work_hours'], s['overtime_hours']])
        resp = make_response(output.getvalue())
        fname = f"daily_attendance_report_{date_obj.strftime('%Y_%m_%d')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
        resp.headers['Content-Type'] = 'text/csv'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    # PDF export
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []
            elements.append(Paragraph(f"Daily Attendance Report - {date_label}", styles['Title']))
            elements.append(Spacer(1, 12))
            # Overall summary table
            ov_headers = ['Total','Present','Late','Leave','On Duty','Holiday','Absent']
            ov_data = [ov_headers, [total_staff, overall['present'], overall['late'], overall['leave'], overall['on_duty'], overall['holiday'], overall['absent']]]
            ov_table = Table(ov_data, repeatRows=1)
            ov_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(ov_table)
            elements.append(Spacer(1, 12))
            # Department summary table
            dep_headers = ['Department','Total','Present','Late','Leave','On Duty','Holiday','Absent','%Present']
            dep_data = [dep_headers]
            for d in dept_rows:
                dep_data.append([d['department'], d['total'], d['present'], d['late'], d['leave'], d['on_duty'], d['holiday'], d['absent'], d['present_rate']])
            dep_table = Table(dep_data, repeatRows=1)
            dep_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(dep_table)
            elements.append(Spacer(1, 12))
            # Staff table (compact)
            st_headers = ['Staff ID','Name','Dept','Pos','In','Out','Status','Late','Early','Hrs','OT']
            st_data = [st_headers]
            for s in staff_records:
                st_data.append([s['staff_id'], s['full_name'], s['department'], s['position'], s['time_in'], s['time_out'], s['status'], s['late_minutes'], s['early_departure_minutes'], s['work_hours'], s['overtime_hours']])
            st_table = Table(st_data, repeatRows=1)
            st_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
            ]))
            elements.append(st_table)
            doc.build(elements)
            pdf = buffer.getvalue(); buffer.close()
            resp = make_response(pdf)
            fname = f"daily_attendance_report_{date_obj.strftime('%Y_%m_%d')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp
        except ImportError:
            return jsonify({'success': False, 'error': 'PDF generation requires reportlab. Install with: pip install reportlab'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel export
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    
    # Create Daily Records sheet FIRST as the active sheet
    ws_daily = wb.active
    ws_daily.title = 'Daily Records'

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # DAILY RECORDS SHEET (MAIN DATA - First sheet for immediate visibility)
    rec_headers = ['Staff ID','Full Name','Department','Position','Time In','Time Out','Status','Late (min)','Early Dep (min)','Work Hrs','OT Hrs']
    
    # Add title
    ws_daily.cell(row=1, column=1, value=f'Daily Attendance Report - Individual Staff Records - {date_label}').font = Font(bold=True, size=12)
    ws_daily.merge_cells('A1:K1')
    ws_daily.append([])  # Empty row
    
    # Add headers
    for col, h in enumerate(rec_headers, 1):
        c = ws_daily.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    
    # Add staff data
    r = 4
    for s in staff_records:
        vals = [s['staff_id'], s['full_name'], s['department'], s['position'], s['time_in'], s['time_out'], s['status'], s['late_minutes'], s['early_departure_minutes'], s['work_hours'], s['overtime_hours']]
        for cidx, val in enumerate(vals, 1):
            cell = ws_daily.cell(row=r, column=cidx, value=val)
            cell.border = border
        r += 1
    
    # Set column widths
    widths = [14, 26, 18, 18, 12, 12, 12, 12, 14, 12, 12]
    for col, w in enumerate(widths, 1):
        ws_daily.column_dimensions[get_column_letter(col)].width = w

    # SUMMARY SHEET (Second sheet)
    ws_summary = wb.create_sheet(title='Summary')
    ws_summary.cell(row=1, column=1, value=f'Daily Attendance Report - Summary - {date_label}').font = Font(bold=True, size=12)
    ws_summary.append([])

    # Overall summary
    sum_headers = ['Total Staff','Present','Late','Leave','On Duty','Holiday','Absent']
    for col, h in enumerate(sum_headers, 1):
        c = ws_summary.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    vals = [total_staff, overall['present'], overall['late'], overall['leave'], overall['on_duty'], overall['holiday'], overall['absent']]
    for cidx, val in enumerate(vals, 1):
        cell = ws_summary.cell(row=4, column=cidx, value=val)
        cell.border = border
    for col in range(1, len(sum_headers)+1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    # Department Summary sheet (Third sheet)
    ws_dept = wb.create_sheet(title='Department Summary')
    dep_headers = ['Department','Total','Present','Late','Leave','On Duty','Holiday','Absent','Present Rate (%)']
    for col, h in enumerate(dep_headers, 1):
        c = ws_dept.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    r = 2
    for d in dept_rows:
        vals = [d['department'], d['total'], d['present'], d['late'], d['leave'], d['on_duty'], d['holiday'], d['absent'], d['present_rate']]
        for cidx, val in enumerate(vals, 1):
            cell = ws_dept.cell(row=r, column=cidx, value=val)
            cell.border = border
        r += 1
    for col in range(1, len(dep_headers)+1):
        ws_dept.column_dimensions[get_column_letter(col)].width = 20

    output = BytesIO(); wb.save(output); output.seek(0)
    resp = make_response(output.getvalue())
    fname = f"daily_attendance_report_{date_obj.strftime('%Y_%m_%d')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


    resp = make_response(output.getvalue())
    fname = f"performance_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

def generate_overtime_report(school_id, year, month, format_type):
    """Generate comprehensive overtime report with individual staff data"""
    excel_generator = ExcelReportGenerator()
    
    if month:
        # Generate overtime-specific report with individual staff overtime data
        return excel_generator.create_overtime_report(school_id, year, month)
    else:
        # For non-monthly requests, generate current month overtime report
        today = datetime.datetime.now()
        return excel_generator.create_overtime_report(school_id, today.year, today.month)



# Analytics Dashboard Routes
@app.route('/analytics_dashboard')
def analytics_dashboard():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    today = datetime.datetime.now()
    thirty_days_ago = today - datetime.timedelta(days=30)

    return render_template('analytics_dashboard.html',
                         start_date=thirty_days_ago.strftime('%Y-%m-%d'),
                         end_date=today.strftime('%Y-%m-%d'))

@app.route('/get_analytics_data')
def get_analytics_data():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department = request.args.get('department')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        data_viz = DataVisualization()

        # Generate all chart data
        analytics_data = {
            'attendance_pie': data_viz.generate_attendance_pie_chart(school_id, start_date, end_date),
            'daily_trends': data_viz.generate_daily_trends_chart(school_id, start_date, end_date),
            'department_comparison': data_viz.generate_department_comparison_chart(school_id, start_date, end_date),
            'weekly_pattern': data_viz.generate_weekly_pattern_chart(school_id, start_date, end_date),
            'overtime_analysis': data_viz.generate_overtime_analysis_chart(school_id, start_date, end_date)
        }

        return jsonify({'success': True, 'data': analytics_data})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_performance_metrics')
def get_performance_metrics():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        data_viz = DataVisualization()
        metrics = data_viz.generate_performance_metrics(school_id, start_date, end_date)

        return jsonify({'success': True, 'data': metrics})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_heatmap_data')
def get_heatmap_data():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        return jsonify({'success': False, 'error': 'Year and month are required'})

    try:
        data_viz = DataVisualization()
        heatmap_data = data_viz.generate_monthly_heatmap_data(school_id, year, month)

        return jsonify({'success': True, 'data': heatmap_data})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export_analytics_data')
def export_analytics_data():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    export_type = request.args.get('export_type', 'analytics')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    # Use the Excel generator to create analytics report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

# Notification System Routes
@app.route('/get_notifications')
def get_notifications():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    user_id = session['user_id']
    user_type = session['user_type']
    limit = request.args.get('limit', 50, type=int)
    unread_only = request.args.get('unread_only', False, type=bool)

    notification_manager = NotificationManager()
    result = notification_manager.get_user_notifications(user_id, user_type, limit, unread_only)

    return jsonify(result)

@app.route('/mark_notification_read', methods=['POST'])
def mark_notification_read():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    notification_id = request.form.get('notification_id', type=int)
    user_id = session['user_id']

    if not notification_id:
        return jsonify({'success': False, 'error': 'Notification ID required'})

    notification_manager = NotificationManager()
    result = notification_manager.mark_notification_read(notification_id, user_id)

    return jsonify(result)

@app.route('/send_system_notification', methods=['POST'])
def send_system_notification():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    title = request.form.get('title')
    message = request.form.get('message')
    user_type = request.form.get('user_type', 'all')
    school_id = session['school_id']

    if not title or not message:
        return jsonify({'success': False, 'error': 'Title and message are required'})

    notification_manager = NotificationManager()
    result = notification_manager.send_system_notification(user_type, title, message, school_id)

    return jsonify(result)

@app.route('/get_notification_count')
def get_notification_count():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    user_id = session['user_id']
    user_type = session['user_type']

    db = get_db()
    unread_count = db.execute('''
        SELECT COUNT(*) as count FROM notifications
        WHERE user_id = ? AND user_type = ? AND is_read = 0
    ''', (user_id, user_type)).fetchone()

    return jsonify({
        'success': True,
        'unread_count': unread_count['count']
    })

# Integrate notifications with existing attendance processing
@app.route('/process_attendance_with_notifications', methods=['POST'])
def process_attendance_with_notifications():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id', type=int)
    verification_type = request.form.get('verification_type')
    timestamp_str = request.form.get('timestamp')

    if not all([staff_id, verification_type, timestamp_str]):
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    try:
        timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
        school_id = session['school_id']

        # Process attendance
        attendance_manager = AdvancedAttendanceManager()
        attendance_result = attendance_manager.process_attendance_with_overtime(
            staff_id, school_id, verification_type, timestamp
        )

        # Send notifications based on attendance result
        if attendance_result['success']:
            notification_manager = NotificationManager()

            if attendance_result.get('status') == 'late':
                # Send late arrival alert
                notification_manager.send_attendance_alert(
                    staff_id, 'late_arrival', {
                        'time_in': attendance_result.get('time_in'),
                        'late_minutes': attendance_result.get('late_minutes', 0)
                    }
                )
            elif attendance_result.get('overtime_hours', 0) > 0:
                # Send overtime alert
                notification_manager.send_attendance_alert(
                    staff_id, 'overtime_alert', {
                        'overtime_hours': attendance_result.get('overtime_hours')
                    }
                )

        return jsonify(attendance_result)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Integrate notifications with leave processing
@app.route('/process_leave_with_notifications', methods=['POST'])
def process_leave_with_notifications():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    leave_id = request.form.get('leave_id', type=int)
    action = request.form.get('action')  # approve or reject
    admin_reason = request.form.get('admin_reason')

    if not leave_id or not action:
        return jsonify({'success': False, 'error': 'Leave ID and action are required'})

    try:
        db = get_db()
        admin_id = session['user_id']

        # Get leave application details before updating
        status = 'approved' if action == 'approve' else 'rejected'
        if status == 'approved':
            leave_details = db.execute('''
                SELECT staff_id, school_id, start_date
                FROM leave_applications
                WHERE id = ?
            ''', (leave_id,)).fetchone()

        # Update leave status
        db.execute('''
            UPDATE leave_applications
            SET status = ?, processed_by = ?, processed_at = ?
            WHERE id = ?
        ''', (status, admin_id, datetime.datetime.now(), leave_id))
        db.commit()

        # Update quota usage if leave is approved
        if status == 'approved' and leave_details:
            try:
                quota_year = datetime.datetime.strptime(leave_details['start_date'], '%Y-%m-%d').year
                update_result = update_quota_usage(leave_details['staff_id'], leave_details['school_id'], quota_year)
                print(f"Quota updated for staff {leave_details['staff_id']}: {update_result}")
            except Exception as e:
                print(f"Error updating quota usage: {e}")

        # Send notification
        notification_manager = NotificationManager()
        notification_result = notification_manager.send_leave_notification(
            leave_id, status, admin_reason
        )

        return jsonify({
            'success': True,
            'message': f'Leave application {status} successfully',
            'notification_sent': notification_result['success']
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/dashboard')
def staff_dashboard():
    """Staff dashboard with module-based navigation"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    db = get_db()
    staff_db_id = session['user_id']
    school_id = session.get('school_id')  # Get school_id from session

    # Get staff information including scheduled shift changes
    staff_info = db.execute('''
        SELECT id, staff_id, full_name, shift_type, next_shift_type, next_shift_effective_date 
        FROM staff WHERE id = ?
    ''', (staff_db_id,)).fetchone()

    if not staff_info:
        return redirect(url_for('index'))
    
    # Resolve current active shift type for TODAY
    today = datetime.date.today()
    today_str = today.strftime('%Y-%m-%d')
    active_shift_type = staff_info['shift_type'] or 'general'
    if (staff_info['next_shift_type'] and 
        staff_info['next_shift_effective_date'] and 
        today_str >= staff_info['next_shift_effective_date']):
        active_shift_type = staff_info['next_shift_type']

    # Get attendance records for the current month
    first_day = today.replace(day=1)
    last_day = (today.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)

    attendance_records = db.execute('''
        SELECT a.date, a.time_in, a.time_out, a.status, 
               COALESCE(a.shift_type, ?) as record_shift_type,
               a.shift_start_time, a.shift_end_time
        FROM attendance a
        WHERE a.staff_id = ? AND a.date BETWEEN ? AND ?
        ORDER BY a.date DESC
    ''', (active_shift_type, staff_db_id, first_day, last_day)).fetchall()

    # For each record, determine display status
    attendance = []
    for record in attendance_records:
        entry_time = record['time_in']
        shift_type_name = record['record_shift_type']
        status = record['status']
        
        # Display logic
        if entry_time:
            # We trust the status already stored in the database by AdvancedAttendanceManager
            status_display = status.capitalize() if status else 'Present'
        else:
            status_display = status.capitalize() if status else 'Absent'
            
        attendance.append({
            'date': record['date'],
            'entry_time': entry_time or '--:--:--',
            'shift_type': shift_type_name.capitalize() if shift_type_name else 'General',
            'status': status_display
        })

    # Get leave applications
    leaves = db.execute('''
         SELECT id, leave_type, start_date, end_date, reason,
             CASE WHEN COALESCE(withdrawn, 0) = 1 THEN 'withdrawn' ELSE status END as status,
               COALESCE(withdrawn, 0) as withdrawn
        FROM leave_applications
        WHERE staff_id = ?
        ORDER BY start_date DESC
    ''', (staff_db_id,)).fetchall()

    # Get on-duty applications
    on_duty_applications = db.execute('''
         SELECT id, duty_type, start_date, end_date, start_time, end_time, location, purpose, reason,
             CASE WHEN COALESCE(withdrawn, 0) = 1 THEN 'withdrawn' ELSE status END as status,
               COALESCE(withdrawn, 0) as withdrawn
        FROM on_duty_applications
        WHERE staff_id = ?
        ORDER BY start_date DESC
    ''', (staff_db_id,)).fetchall()

    # Get permission applications
    permission_applications = db.execute('''
         SELECT id, permission_type, permission_date, start_time, end_time, duration_hours, reason,
             CASE WHEN COALESCE(withdrawn, 0) = 1 THEN 'withdrawn' ELSE status END as status,
               COALESCE(withdrawn, 0) as withdrawn
        FROM permission_applications
        WHERE staff_id = ?
        ORDER BY permission_date DESC
    ''', (staff_db_id,)).fetchall()

    # Eligible attendance records for regularization (last 7 days only)
    lookback_start = today - datetime.timedelta(days=7)
    regularization_candidates = db.execute('''
        SELECT id, date, time_in, time_out, status,
               COALESCE(late_duration_minutes, 0) as late_duration_minutes,
               COALESCE(early_departure_minutes, 0) as early_departure_minutes,
             regularization_requested, regularization_status,
             shift_start_time, shift_end_time
        FROM attendance
        WHERE staff_id = ?
          AND school_id = ?
          AND date BETWEEN ? AND ?
            AND COALESCE(regularization_status, '') != 'pending'
          AND (
                (COALESCE(late_duration_minutes, 0) > 0 AND time_in IS NOT NULL)
                OR
                (COALESCE(early_departure_minutes, 0) > 0 AND time_out IS NOT NULL)
          )
        ORDER BY date DESC
    ''', (staff_db_id, school_id, lookback_start, today)).fetchall()

    # Recent regularization requests for staff visibility
    regularization_requests = db.execute('''
        SELECT r.id, r.request_type, r.original_time, r.expected_time,
               r.duration_minutes, r.staff_reason, r.admin_reason,
               r.status, r.requested_at, r.processed_at,
               a.date as attendance_date
        FROM attendance_regularization_requests r
        JOIN attendance a ON a.id = r.attendance_id
        WHERE r.staff_id = ? AND r.school_id = ?
        ORDER BY r.requested_at DESC
        LIMIT 20
    ''', (staff_db_id, school_id)).fetchall()

    # Get module settings for navigation (staff can see limited nav based on admin settings)
    module_enabled = get_module_enabled(school_id) if school_id else {}
    attendance_mode = get_school_attendance_mode(school_id) if school_id else 'biometric'

    return render_template('staff_dashboard.html',
                         attendance=attendance,
                         leaves=leaves,
                         on_duty_applications=on_duty_applications,
                         permission_applications=permission_applications,
                         regularization_candidates=regularization_candidates,
                         regularization_requests=regularization_requests,
                         module_enabled=module_enabled,
                         attendance_mode=attendance_mode,
                         today=today,
                         staff_info=staff_info)

@app.route('/admin/department_shifts')
def department_shifts():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']
    today_str = datetime.date.today().strftime('%Y-%m-%d')

    # Get current department shift mappings
    mappings = db.execute('''
        SELECT department, default_shift_type, created_at, updated_at
        FROM department_shift_mappings
        WHERE school_id = ?
        ORDER BY department
    ''', (school_id,)).fetchall()

    # Get all departments currently in use
    departments = db.execute('''
        SELECT DISTINCT department
        FROM staff
        WHERE school_id = ? AND department IS NOT NULL AND department != ''
        ORDER BY department
    ''', (school_id,)).fetchall()

    # Get all active shift definitions with times for dropdowns
    shift_defs = db.execute('''
        SELECT shift_type, start_time, end_time, grace_period_minutes, description
        FROM shift_definitions WHERE school_id = ? AND is_active = 1 ORDER BY id
    ''', (school_id,)).fetchall()

    # Build a quick lookup dict: shift_type -> {start, end}
    shift_time_map = {s['shift_type']: {'start': s['start_time'][:5], 'end': s['end_time'][:5]}
                     for s in shift_defs}

    # Get institution info for branding
    institution = db.execute(
        'SELECT * FROM schools WHERE id = ?',
        (school_id,)
    ).fetchone()

    # Get module settings for navigation
    module_enabled = get_module_enabled(school_id)

    return render_template('department_shifts.html', mappings=mappings, departments=departments,
                           shift_defs=shift_defs, shift_time_map=shift_time_map,
                           institution=institution, module_enabled=module_enabled)

@app.route('/api/department_shifts', methods=['GET', 'POST', 'DELETE'])
def api_department_shifts():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'message': 'Unauthorized access'}), 401

    db = get_db()
    school_id = session['school_id']

    if request.method == 'GET':
        try:
            # Get all department shift mappings with shift times
            mappings = db.execute('''
                SELECT dsm.department, dsm.default_shift_type, dsm.created_at, dsm.updated_at,
                       sd.start_time, sd.end_time
                FROM department_shift_mappings dsm
                LEFT JOIN shift_definitions sd
                    ON sd.shift_type = dsm.default_shift_type
                    AND sd.school_id = dsm.school_id
                    AND sd.is_active = 1
                WHERE dsm.school_id = ?
                ORDER BY dsm.department
            ''', [school_id]).fetchall()

            mappings_list = []
            for mapping in mappings:
                start = (mapping['start_time'] or '')[:5]
                end   = (mapping['end_time']   or '')[:5]
                mappings_list.append({
                    'department': mapping['department'],
                    'default_shift_type': mapping['default_shift_type'],
                    'start_time': start,
                    'end_time': end,
                    'created_at': mapping['created_at'],
                    'updated_at': mapping['updated_at']
                })

            return jsonify({
                'success': True,
                'mappings': mappings_list
            })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error loading department shifts: {str(e)}'
            }), 500

    elif request.method == 'POST':
        try:
            data = request.get_json()
            department = data.get('department', '').strip()
            shift_type = data.get('shift_type', '').strip()

            if not department or not shift_type:
                return jsonify({
                    'success': False,
                    'message': 'Department and shift type are required'
                }), 400

            # Check if mapping already exists
            existing = db.execute('''
                SELECT id FROM department_shift_mappings
                WHERE school_id = ? AND department = ?
            ''', [school_id, department]).fetchone()

            if existing:
                return jsonify({
                    'success': False,
                    'message': f'Department mapping for "{department}" already exists'
                }), 400

            # Create new mapping
            current_time = datetime.datetime.now().isoformat()
            db.execute('''
                INSERT INTO department_shift_mappings
                (school_id, department, default_shift_type, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', [school_id, department, shift_type, current_time, current_time])

            db.commit()

            return jsonify({
                'success': True,
                'message': f'Department shift mapping for "{department}" created successfully'
            })

        except Exception as e:
            db.rollback()
            return jsonify({
                'success': False,
                'message': f'Error creating department mapping: {str(e)}'
            }), 500

    elif request.method == 'DELETE':
        try:
            data = request.get_json()
            department = data.get('department', '').strip()

            if not department:
                return jsonify({
                    'success': False,
                    'message': 'Department is required'
                }), 400

            # Delete the mapping
            result = db.execute('''
                DELETE FROM department_shift_mappings
                WHERE school_id = ? AND department = ?
            ''', [school_id, department])

            if result.rowcount == 0:
                return jsonify({
                    'success': False,
                    'message': f'Department mapping for "{department}" not found'
                }), 404

            db.commit()

            return jsonify({
                'success': True,
                'message': f'Department shift mapping for "{department}" deleted successfully'
            })

        except Exception as e:
            db.rollback()
            return jsonify({
                'success': False,
                'message': f'Error deleting department mapping: {str(e)}'
            }), 500

# Test route for department shifts debugging
@app.route('/test/department_shifts')
def test_department_shifts():
    return render_template('test_department_shifts.html')

# Debug route to check database tables
@app.route('/api/debug/tables')
def debug_tables():
    try:
        db = get_db()

        # Get all tables
        tables = db.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        table_list = [row['name'] for row in tables]

        result = f"Available tables: {table_list}\n\n"

        # Check if department_shift_mappings table exists
        if 'department_shift_mappings' in table_list:
            result += "department_shift_mappings table EXISTS\n"

            # Get table structure
            schema = db.execute("PRAGMA table_info(department_shift_mappings)").fetchall()
            result += "Table structure:\n"
            for col in schema:
                result += f"  {col['name']} ({col['type']})\n"

            # Get record count
            count = db.execute("SELECT COUNT(*) as count FROM department_shift_mappings").fetchone()
            result += f"\nRecord count: {count['count']}\n"

            # Get sample records
            if count['count'] > 0:
                records = db.execute("SELECT * FROM department_shift_mappings LIMIT 5").fetchall()
                result += "\nSample records:\n"
                for record in records:
                    result += f"  {dict(record)}\n"
        else:
            result += "department_shift_mappings table does NOT exist\n"

        return result, 200, {'Content-Type': 'text/plain'}

    except Exception as e:
        return f"Error checking database: {str(e)}", 500, {'Content-Type': 'text/plain'}

@csrf.exempt
@app.route('/api/module-settings/toggle', methods=['POST'])
def toggle_module_settings():
    """Toggle module enabled/disabled status for a school"""
    if 'user_id' not in session or session.get('user_type') != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        school_id = data.get('school_id')
        module_name = data.get('module_name')
        is_enabled = data.get('is_enabled', True)
        
        if not school_id or not module_name:
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        # Map module name to database column
        module_column_map = {
            'staff_management': 'staff_management_enabled',
            'shift_management': 'shift_management_enabled',
            'salary_management': 'salary_management_enabled',
            'payroll_processing_review': 'salary_management_enabled',
            'timetable_management': 'timetable_management_enabled',
            'reports': 'reports_enabled',
            'biometric_devices': 'biometric_devices_enabled',
            'department_shift_assignments': 'department_shift_assignments_enabled',
            'holiday_management': 'holiday_management_enabled',
            'quota_management': 'quota_management_enabled',
            'sub_admin_management': 'sub_admin_management_enabled',
            'student_management': 'student_management_enabled',
        }
        
        column_name = module_column_map.get(module_name)
        if not column_name:
            return jsonify({'success': False, 'error': 'Invalid module name'}), 400
        
        db = get_db()
        
        # Update the module status
        db.execute(
            f'UPDATE schools SET {column_name} = ? WHERE id = ?',
            (1 if is_enabled else 0, school_id)
        )
        db.commit()
        
        return jsonify({
            'success': True,
            'message': f'Module {module_name} {"enabled" if is_enabled else "disabled"} successfully',
            'module_name': module_name,
            'is_enabled': is_enabled
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/staff_management')
@requires_permission('staff_management', 'view')
def staff_management():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    from database import get_all_departments, initialize_default_departments, get_all_positions, initialize_default_positions
    
    db = get_db()
    school_id = session['school_id']
    today_str = datetime.date.today().strftime('%Y-%m-%d')
    staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]
    has_next_shift = 'next_shift_type' in staff_columns and 'next_shift_effective_date' in staff_columns

    # Initialize default departments and positions if needed
    initialize_default_departments(school_id)
    initialize_default_positions(school_id)

    # Get all staff with comprehensive details
    if has_next_shift:
        staff = db.execute('''
            SELECT id, staff_id, full_name, first_name, last_name,
                   date_of_birth, date_of_joining, department, destination,
                   position, gender, phone, email, shift_type,
                   CASE
                       WHEN COALESCE(next_shift_type, '') <> ''
                            AND COALESCE(next_shift_effective_date, '') <> ''
                            AND ? >= next_shift_effective_date
                       THEN next_shift_type
                       ELSE COALESCE(shift_type, 'general')
                   END AS effective_shift_type,
                   photo_url
            FROM staff
            WHERE school_id = ?
            ORDER BY CAST(staff_id AS INTEGER) ASC
        ''', (today_str, school_id)).fetchall()
    else:
        staff = db.execute('''
            SELECT id, staff_id, full_name, first_name, last_name,
                   date_of_birth, date_of_joining, department, destination,
                   position, gender, phone, email, shift_type,
                   COALESCE(shift_type, 'general') AS effective_shift_type,
                   photo_url
            FROM staff
            WHERE school_id = ?
            ORDER BY CAST(staff_id AS INTEGER) ASC
        ''', (school_id,)).fetchall()

    # Get department shift mappings for reference
    dept_mappings = db.execute('''
        SELECT department, default_shift_type
        FROM department_shift_mappings
        WHERE school_id = ?
    ''', (school_id,)).fetchall()

    dept_shift_map = {mapping['department']: mapping['default_shift_type'] for mapping in dept_mappings}
    
    # Get all available departments and positions for dropdowns
    departments = get_all_departments(school_id)
    positions = get_all_positions(school_id)

    # Get all active shift definitions for dropdowns
    shift_defs = db.execute('''
        SELECT shift_type, start_time, end_time, description
        FROM shift_definitions WHERE school_id = ? AND is_active = 1 ORDER BY id
    ''', (school_id,)).fetchall()

    # Auto-fix any staff records with 'User' suffix (run once when page loads)
    try:
        fix_user_suffix_in_staff_names()
    except Exception as e:
        print(f"Warning: Could not auto-fix user suffix: {e}")

    # Get module settings for navigation
    module_enabled = get_module_enabled(school_id)

    return render_template('staff_management.html', staff=staff, dept_shift_map=dept_shift_map,
                           departments=departments, positions=positions, shift_defs=shift_defs,
                           module_enabled=module_enabled)

@app.route('/admin/add_department', methods=['POST'])
@csrf.exempt
def add_department_route():
    """Add a new custom department"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    from database import add_department
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        department_name = data.get('name', '').strip()
        department_description = data.get('description', '').strip()
        
        if not department_name:
            return jsonify({'success': False, 'error': 'Department name is required'}), 400
        
        school_id = session['school_id']
        result = add_department(department_name, department_description, school_id)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify({'success': False, 'error': result['message']}), 400
            
    except Exception as e:
        print(f"Error in add_department route: {e}")
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/admin/get_departments_list', methods=['GET'])
@csrf.exempt
def get_departments_list():
    """Get all departments for a school"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    from database import get_all_departments
    
    try:
        school_id = session['school_id']
        departments = get_all_departments(school_id)
        return jsonify({'success': True, 'departments': departments}), 200
    except Exception as e:
        print(f"Error in get_departments_list route: {e}")
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/admin/delete_department', methods=['POST'])
@csrf.exempt
def delete_department_route():
    """Delete a department"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    from database import delete_department
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        department_id = data.get('id')
        
        if not department_id:
            return jsonify({'success': False, 'error': 'Department ID is required'}), 400
        
        school_id = session['school_id']
        result = delete_department(department_id, school_id)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify(result), 400
            
    except Exception as e:
        print(f"Error in delete_department route: {e}")
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/admin/add_position', methods=['POST'])
@csrf.exempt
def add_position_route():
    """Add a new custom position"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    from database import add_position
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        position_name = data.get('name', '').strip()
        position_description = data.get('description', '').strip()
        
        if not position_name:
            return jsonify({'success': False, 'error': 'Position name is required'}), 400
        
        school_id = session['school_id']
        result = add_position(position_name, position_description, school_id)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify({'success': False, 'error': result['message']}), 400
            
    except Exception as e:
        print(f"Error in add_position route: {e}")
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/admin/get_positions_list', methods=['GET'])
@csrf.exempt
def get_positions_list():
    """Get all positions for a school"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    from database import get_all_positions
    
    try:
        school_id = session['school_id']
        positions = get_all_positions(school_id)
        return jsonify({'success': True, 'positions': positions}), 200
    except Exception as e:
        print(f"Error in get_positions_list route: {e}")
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/admin/delete_position', methods=['POST'])
@csrf.exempt
def delete_position_route():
    """Delete a position (soft delete)"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    from database import delete_position
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        position_id = data.get('id')
        
        if not position_id:
            return jsonify({'success': False, 'error': 'Position ID is required'}), 400
        
        school_id = session['school_id']
        result = delete_position(position_id, school_id)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify({'success': False, 'message': result['message']}), 400
            
    except Exception as e:
        print(f"Error in delete_position route: {e}")
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/admin/shift_management')
@requires_permission('shift_management', 'view')
def shift_management():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']

    # Get total staff count for sidebar
    total_staff_count = db.execute('''
        SELECT COUNT(*) as count FROM staff WHERE school_id = ?
    ''', (school_id,)).fetchone()['count']

    # Get defined shifts
    shifts = db.execute('''
        SELECT * FROM shift_definitions 
        WHERE school_id = ? AND is_active = 1
        ORDER BY id ASC
    ''', (school_id,)).fetchall()
    
    # Get institution info for branding
    institution = db.execute(
        'SELECT * FROM schools WHERE id = ?',
        (school_id,)
    ).fetchone()

    # Get module settings for navigation
    module_enabled = get_module_enabled(school_id)

    return render_template('shift_management.html', 
                          total_staff_count=total_staff_count, 
                          shifts=shifts,
                          institution=institution,
                          module_enabled=module_enabled)

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']
    use_modern_ui = session.get('use_modern_ui', False)

    # Get all staff
    staff = db.execute('''
        SELECT id, staff_id, full_name, department, position
        FROM staff
        WHERE school_id = ?
        ORDER BY CAST(staff_id AS INTEGER) ASC
    ''', (school_id,)).fetchall()

    # Get pending leave applications
    pending_leaves = db.execute('''
        SELECT l.id, s.full_name, s.photo_url, l.leave_type, l.start_date, l.end_date, l.reason
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ? AND l.status = 'pending'
        ORDER BY l.applied_at
    ''', (school_id,)).fetchall()

    # Get pending on-duty applications
    pending_on_duty = db.execute('''
        SELECT od.id, s.full_name, s.photo_url, od.duty_type, od.start_date, od.end_date, od.start_time, od.end_time, od.location, od.purpose, od.reason
        FROM on_duty_applications od
        JOIN staff s ON od.staff_id = s.id
        WHERE od.school_id = ? AND od.status = 'pending'
        ORDER BY od.applied_at
    ''', (school_id,)).fetchall()

    # Get pending permission applications
    pending_permissions = db.execute('''
        SELECT p.id, s.full_name, s.photo_url, p.permission_type, p.permission_date, p.start_time, p.end_time, p.duration_hours, p.reason
        FROM permission_applications p
        JOIN staff s ON p.staff_id = s.id
        WHERE p.school_id = ? AND p.status = 'pending'
        ORDER BY p.applied_at
    ''', (school_id,)).fetchall()

    # Get pending attendance regularization requests
    pending_regularizations = db.execute('''
        SELECT r.id, r.request_type, r.original_time, r.expected_time,
               r.duration_minutes, r.staff_reason, r.requested_at,
               s.full_name, s.photo_url,
               a.date as attendance_date
        FROM attendance_regularization_requests r
        JOIN staff s ON r.staff_id = s.id
        JOIN attendance a ON r.attendance_id = a.id
        WHERE r.school_id = ? AND r.status = 'pending'
        ORDER BY r.requested_at
    ''', (school_id,)).fetchall()

    # Get today's attendance summary using comprehensive status logic
    today = datetime.date.today()
    
    # Get all staff for attendance calculation
    all_staff = db.execute('''
        SELECT s.id as staff_id, s.staff_id as staff_number, s.full_name, s.department,
               a.time_in, a.time_out, s.photo_url
        FROM staff s
        LEFT JOIN (
            SELECT staff_id,
                   MIN(time_in) AS time_in,
                   MAX(time_out) AS time_out
            FROM attendance
            WHERE date = ? AND school_id = ?
            GROUP BY staff_id
        ) a ON s.id = a.staff_id
        WHERE s.school_id = ?
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    ''', (today, school_id, school_id)).fetchall()
    
    # Calculate attendance summary with correct status logic
    status_counts = {
        'total_staff': 0,
        'present': 0,
        'absent': 0,
        'late': 0,
        'on_leave': 0,
        'on_duty': 0,
        'on_permission': 0,
        'holiday': 0
    }
    
    today_attendance = []
    
    for staff in all_staff:
        # Get comprehensive status using the new logic
        status = get_staff_status_for_date(staff['staff_id'], today, school_id, db)
        
        # Build staff record for today_attendance
        staff_record = {
            'staff_id': staff['staff_id'],
            'staff_number': staff['staff_number'],
            'full_name': staff['full_name'],
            'department': staff['department'],
            'time_in': staff['time_in'],
            'time_out': staff['time_out'],
            'status': status,
            'photo_url': staff['photo_url']
        }
        
        today_attendance.append(staff_record)
        
        # Count statuses for summary
        status_counts['total_staff'] += 1
        
        if status == 'Holiday':
            status_counts['holiday'] += 1
        elif status == 'On Leave':
            status_counts['on_leave'] += 1
        elif status == 'On Duty':
            status_counts['on_duty'] += 1
        elif status == 'On Permission':
            status_counts['on_permission'] += 1
        elif status == 'present':
            status_counts['present'] += 1
        elif status == 'late':
            status_counts['late'] += 1
        else:  # absent or any other status
            status_counts['absent'] += 1
    
    # Convert to match expected format
    attendance_summary = status_counts

    # Get timetable module status for this school
    timetable_settings = db.execute('SELECT is_enabled FROM timetable_settings WHERE school_id = ?', (school_id,)).fetchone()
    timetable_enabled = bool(timetable_settings['is_enabled']) if timetable_settings else False

    # Get module enabled settings for navigation
    module_enabled = get_module_enabled(school_id)
    attendance_mode = get_school_attendance_mode(school_id)

    # Get all shift definitions for filter
    all_shifts = db.execute('''
        SELECT id, shift_type, start_time, end_time
        FROM shift_definitions
        WHERE school_id = ? AND is_active = 1
        ORDER BY start_time
    ''', (school_id,)).fetchall()
    
    # Determine current shift based on current time
    current_time = datetime.datetime.now().time()
    current_shift = None
    for shift in all_shifts:
        shift_start = datetime.datetime.strptime(shift['start_time'], '%H:%M:%S').time() if isinstance(shift['start_time'], str) else shift['start_time']
        shift_end = datetime.datetime.strptime(shift['end_time'], '%H:%M:%S').time() if isinstance(shift['end_time'], str) else shift['end_time']
        
        # Handle shifts that cross midnight
        if shift_start <= shift_end:
            if shift_start <= current_time <= shift_end:
                current_shift = shift['shift_type']
                break
        else:  # shift crosses midnight
            if current_time >= shift_start or current_time <= shift_end:
                current_shift = shift['shift_type']
                break
    
    # Filter today_attendance to include only active staff by default
    today_attendance_with_status = []
    for record in today_attendance:
        # Get staff active status
        staff_info = db.execute('''
            SELECT COALESCE(is_active, 1) as is_active, shift_type, next_shift_type, next_shift_effective_date
            FROM staff
            WHERE id = ? AND school_id = ?
        ''', (record['staff_id'], school_id)).fetchone()
        
        if staff_info:
            record['is_active'] = bool(staff_info['is_active'])
            record['shift_type'] = resolve_effective_shift_type(staff_info, today)
            today_attendance_with_status.append(record)
    
    today_attendance = today_attendance_with_status

    if use_modern_ui:
        return render_template('admin_dashboard_modern.html',
                             staff=staff,
                             pending_leaves=pending_leaves,
                             pending_on_duty=pending_on_duty,
                             pending_permissions=pending_permissions,
                             pending_regularizations=pending_regularizations,
                             attendance_summary=attendance_summary,
                             today_attendance=today_attendance,
                             today=today,
                             timetable_enabled=timetable_enabled,
                             attendance_mode=attendance_mode,
                             module_enabled=module_enabled,
                             recent_activities=[],  # Add recent activities data
                             performance={},  # Add performance metrics
                             biometric_status={},  # Add biometric status
                             last_backup='Today',  # Add backup info
                             all_shifts=all_shifts,
                             current_shift=current_shift)
    else:
        return render_template('admin_dashboard.html',
                             staff=staff,
                             pending_leaves=pending_leaves,
                             pending_on_duty=pending_on_duty,
                             pending_permissions=pending_permissions,
                             pending_regularizations=pending_regularizations,
                             attendance_summary=attendance_summary,
                             today_attendance=today_attendance,
                             today=today,
                             timetable_enabled=timetable_enabled,
                             attendance_mode=attendance_mode,
                             module_enabled=module_enabled,
                             all_shifts=all_shifts,
                             current_shift=current_shift)


@app.route('/admin/timetable')
@requires_permission('timetable_management', 'view')
def admin_timetable():
    """Admin Timetable Management page"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))
    
    school_id = session.get('school_id')
    module_enabled = get_module_enabled(school_id)
    
    return render_template('timetable_management.html', module_enabled=module_enabled)


@app.route('/admin/student-management', methods=['GET', 'POST'])
def admin_student_management():
    """Admin Student Management page"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))
    
    db = get_db()
    school_id = session['school_id']

    def _ensure_exam_assignments_table():
        """Create exam assignments table if it does not exist (MySQL/SQLite compatible fallback)."""
        try:
            db.execute('''
                CREATE TABLE IF NOT EXISTS exam_assignments (
                    id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    school_id INTEGER NOT NULL,
                    class_name VARCHAR(100) NOT NULL,
                    exam_title VARCHAR(255) NOT NULL,
                    exam_date DATE NOT NULL,
                    description TEXT,
                    created_by INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            db.commit()
            return
        except Exception:
            db.rollback()

        # SQLite fallback (dev/test environments)
        db.execute('''
            CREATE TABLE IF NOT EXISTS exam_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                class_name TEXT NOT NULL,
                exam_title TEXT NOT NULL,
                exam_date TEXT NOT NULL,
                description TEXT,
                created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        db.commit()

    def _ensure_student_holidays_table():
        """Create student holiday planner table if it does not exist (MySQL/SQLite compatible fallback)."""
        try:
            db.execute('''
                CREATE TABLE IF NOT EXISTS student_holidays (
                    id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    school_id INTEGER NOT NULL,
                    holiday_title VARCHAR(255) NOT NULL,
                    holiday_start_date DATE NOT NULL,
                    holiday_end_date DATE NOT NULL,
                    description TEXT,
                    target_mode VARCHAR(20) NOT NULL DEFAULT 'bulk',
                    target_class VARCHAR(100),
                    target_section VARCHAR(100),
                    created_by INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            db.commit()
            return
        except Exception:
            db.rollback()

        db.execute('''
            CREATE TABLE IF NOT EXISTS student_holidays (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                holiday_title TEXT NOT NULL,
                holiday_start_date TEXT NOT NULL,
                holiday_end_date TEXT NOT NULL,
                description TEXT,
                target_mode TEXT NOT NULL DEFAULT 'bulk',
                target_class TEXT,
                target_section TEXT,
                created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        db.commit()
    
    # Check if student management module is enabled
    module_enabled = get_module_enabled(school_id)
    if not module_enabled.get('student_management', True):
        flash('Student Management module is disabled for this school.', 'warning')
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        try:
            if action == 'create':
                # Create new student - Extended fields
                admission_number = request.form.get('admission_number')
                roll_number = request.form.get('roll_number')
                student_type = request.form.get('student_type', 'Day Scholar')
                first_name = request.form.get('first_name')
                last_name = request.form.get('last_name')
                date_of_birth = request.form.get('date_of_birth')
                age = request.form.get('age')
                gender = request.form.get('gender')
                student_mobile = request.form.get('student_mobile')
                address = request.form.get('address')
                class_name = request.form.get('class')
                section = request.form.get('section')
                academic_year = request.form.get('academic_year')
                
                # Parent details
                parent_name = request.form.get('parent_name')
                parent_mobile = request.form.get('parent_mobile')
                parent_email = request.form.get('parent_email')
                mother_name = request.form.get('mother_name')
                mother_phone = request.form.get('mother_phone')
                parent_occupation = request.form.get('parent_occupation')
                
                # Academic details
                tenth_marks = request.form.get('tenth_marks')
                tenth_percentage = request.form.get('tenth_percentage')
                twelfth_marks = request.form.get('twelfth_marks')
                twelfth_percentage = request.form.get('twelfth_percentage')
                skills = request.form.get('skills')
                
                # Documents
                tc_number = request.form.get('tc_number')
                aadhar_number = request.form.get('aadhar_number')
                
                # Photo upload handling
                photo_data = None
                if 'student_photo' in request.files:
                    photo_file = request.files['student_photo']
                    if photo_file and photo_file.filename:
                        import base64
                        photo_bytes = photo_file.read()
                        photo_data = f"data:image/{photo_file.filename.rsplit('.', 1)[1].lower()};base64,{base64.b64encode(photo_bytes).decode('utf-8')}"
                
                # Custom fields
                import json
                custom_fields = {}
                for key in request.form:
                    if key.startswith('custom_field_'):
                        field_name = key.replace('custom_field_', '')
                        custom_fields[field_name] = request.form.get(key)
                custom_fields_json = json.dumps(custom_fields) if custom_fields else None
                
                # Validate: Either admission_number or roll_number must be provided
                if not admission_number and not roll_number:
                    flash('Either Admission Number or Roll Number is required!', 'error')
                    return redirect(url_for('admin_student_management'))
                
                # Generate student_id
                if admission_number:
                    student_id = f"STU{admission_number}"
                else:
                    student_id = f"STU{roll_number}"
                
                # Generate default password (can be changed later)
                default_password = generate_password_hash('student123')
                
                # Create full_name
                full_name = f"{first_name} {last_name}"
                
                # Check if student_id already exists
                existing = db.execute('''
                    SELECT id FROM students 
                    WHERE school_id = ? AND (student_id = ? OR admission_number = ? OR roll_number = ?)
                ''', (school_id, student_id, admission_number, roll_number)).fetchone()
                
                if existing:
                    flash('Student with this admission/roll number already exists!', 'error')
                else:
                    db.execute('''
                        INSERT INTO students (
                            school_id, student_id, password, full_name, first_name, last_name,
                            admission_number, roll_number, student_type, date_of_birth, age, gender,
                            student_mobile, address, `class`, section, academic_year,
                            parent_name, parent_phone, parent_email, mother_name, mother_phone,
                            parent_occupation, tenth_marks, tenth_percentage, twelfth_marks,
                            twelfth_percentage, skills, tc_number, aadhar_number, custom_fields, photo_data
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (school_id, student_id, default_password, full_name, first_name, last_name,
                          admission_number, roll_number, student_type, date_of_birth, age, gender,
                          student_mobile, address, class_name, section, academic_year,
                          parent_name, parent_mobile, parent_email, mother_name, mother_phone,
                          parent_occupation, tenth_marks, tenth_percentage, twelfth_marks,
                          twelfth_percentage, skills, tc_number, aadhar_number, custom_fields_json, photo_data))
                    db.commit()
                    flash(f'Student {full_name} created successfully! Default password: student123', 'success')
                
            elif action == 'edit':
                # Update existing student
                student_id = request.form.get('student_id')
                roll_number = request.form.get('roll_number')
                first_name = request.form.get('first_name')
                last_name = request.form.get('last_name')
                date_of_birth = request.form.get('date_of_birth')
                age = request.form.get('age')
                gender = request.form.get('gender')
                student_mobile = request.form.get('student_mobile')
                parent_mobile = request.form.get('parent_mobile')
                address = request.form.get('address')
                class_name = request.form.get('class')
                section = request.form.get('section')
                academic_year = request.form.get('academic_year')
                
                # Create full_name
                full_name = f"{first_name} {last_name}"
                
                db.execute('''
                    UPDATE students SET
                        roll_number = ?, first_name = ?, last_name = ?, full_name = ?,
                        date_of_birth = ?, age = ?, gender = ?, phone = ?, parent_phone = ?,
                        address = ?, `class` = ?, section = ?, academic_year = ?
                    WHERE id = ? AND school_id = ?
                ''', (roll_number, first_name, last_name, full_name, date_of_birth, age,
                      gender, student_mobile, parent_mobile, address, class_name, section,
                      academic_year, student_id, school_id))
                db.commit()
                flash('Student updated successfully!', 'success')
                
            elif action == 'delete':
                # Delete student
                student_id = request.form.get('student_id')
                db.execute('DELETE FROM students WHERE id = ? AND school_id = ?', 
                          (student_id, school_id))
                db.commit()
                flash('Student deleted successfully!', 'success')

            elif action == 'assign_exam':
                _ensure_exam_assignments_table()

                class_name = (request.form.get('exam_class') or '').strip()
                subject_list = request.form.getlist('exam_subject[]')
                date_list = request.form.getlist('exam_date[]')
                description = (request.form.get('exam_description') or '').strip()

                if not class_name:
                    flash('Class is required for exam assignment.', 'error')
                    return redirect(url_for('admin_student_management'))

                if not isinstance(subject_list, list) or not isinstance(date_list, list) or len(subject_list) != len(date_list):
                    flash('Invalid exam rows. Please add subject and date correctly.', 'error')
                    return redirect(url_for('admin_student_management'))

                valid_rows = []
                for idx in range(len(subject_list)):
                    subject_name = (subject_list[idx] or '').strip()
                    date_value = (date_list[idx] or '').strip()
                    if not subject_name and not date_value:
                        continue
                    if not subject_name or not date_value:
                        flash('Each exam row needs both subject and date.', 'error')
                        return redirect(url_for('admin_student_management'))
                    try:
                        exam_date = datetime.date.fromisoformat(date_value).strftime('%Y-%m-%d')
                    except Exception:
                        flash(f'Invalid exam date format in row {idx + 1}.', 'error')
                        return redirect(url_for('admin_student_management'))
                    valid_rows.append((subject_name, exam_date))

                if not valid_rows:
                    flash('Add at least one subject and date to assign exam.', 'error')
                    return redirect(url_for('admin_student_management'))

                for subject_name, exam_date in valid_rows:
                    db.execute('''
                        INSERT INTO exam_assignments (
                            school_id, class_name, exam_title, exam_date, description, created_by
                        ) VALUES (?, ?, ?, ?, ?, ?)
                    ''', (school_id, class_name, subject_name, exam_date, description, session.get('user_id')))
                db.commit()
                flash(f'{len(valid_rows)} exam assignment(s) saved for class {class_name}.', 'success')

            elif action == 'update_exam':
                _ensure_exam_assignments_table()

                exam_id = (request.form.get('exam_id') or '').strip()
                class_name = (request.form.get('exam_class') or '').strip()
                subject_list = request.form.getlist('exam_subject[]')
                date_list = request.form.getlist('exam_date[]')
                description = (request.form.get('exam_description') or '').strip()

                if not exam_id:
                    flash('Exam ID is required for update.', 'error')
                    return redirect(url_for('admin_student_management'))

                if not class_name:
                    flash('Class is required for exam update.', 'error')
                    return redirect(url_for('admin_student_management'))

                if not isinstance(subject_list, list) or not isinstance(date_list, list) or len(subject_list) != len(date_list):
                    flash('Invalid exam rows. Please provide one subject and date for update.', 'error')
                    return redirect(url_for('admin_student_management'))

                valid_rows = []
                for idx in range(len(subject_list)):
                    subject_name = (subject_list[idx] or '').strip()
                    date_value = (date_list[idx] or '').strip()
                    if not subject_name and not date_value:
                        continue
                    if not subject_name or not date_value:
                        flash('Each exam row needs both subject and date.', 'error')
                        return redirect(url_for('admin_student_management'))
                    try:
                        exam_date = datetime.date.fromisoformat(date_value).strftime('%Y-%m-%d')
                    except Exception:
                        flash(f'Invalid exam date format in row {idx + 1}.', 'error')
                        return redirect(url_for('admin_student_management'))
                    valid_rows.append((subject_name, exam_date))

                if len(valid_rows) != 1:
                    flash('Update mode supports exactly one subject and one date.', 'error')
                    return redirect(url_for('admin_student_management'))

                subject_name, exam_date = valid_rows[0]
                result = db.execute('''
                    UPDATE exam_assignments
                    SET class_name = ?, exam_title = ?, exam_date = ?, description = ?
                    WHERE id = ? AND school_id = ?
                ''', (class_name, subject_name, exam_date, description, exam_id, school_id))
                db.commit()

                if result.rowcount:
                    flash('Exam assignment updated successfully.', 'success')
                else:
                    flash('Exam assignment not found or already removed.', 'warning')

            elif action == 'delete_exam':
                _ensure_exam_assignments_table()

                exam_id = (request.form.get('exam_id') or '').strip()
                if not exam_id:
                    flash('Exam ID is required for delete.', 'error')
                    return redirect(url_for('admin_student_management'))

                result = db.execute('''
                    DELETE FROM exam_assignments
                    WHERE id = ? AND school_id = ?
                ''', (exam_id, school_id))
                db.commit()

                if result.rowcount:
                    flash('Exam assignment deleted successfully.', 'success')
                else:
                    flash('Exam assignment not found or already removed.', 'warning')

            elif action == 'assign_student_holiday':
                _ensure_student_holidays_table()

                holiday_title = (request.form.get('holiday_title') or '').strip()
                holiday_start_date = (request.form.get('holiday_start_date') or '').strip()
                holiday_end_date = (request.form.get('holiday_end_date') or '').strip()
                holiday_description = (request.form.get('holiday_description') or '').strip()
                holiday_mode = (request.form.get('holiday_mode') or 'bulk').strip().lower()
                holiday_class = (request.form.get('holiday_class') or '').strip()
                holiday_section = (request.form.get('holiday_section') or '').strip()

                if not holiday_title:
                    flash('Holiday title is required.', 'error')
                    return redirect(url_for('admin_student_management'))

                try:
                    start_date = datetime.date.fromisoformat(holiday_start_date).strftime('%Y-%m-%d')
                    end_date = datetime.date.fromisoformat(holiday_end_date).strftime('%Y-%m-%d')
                except Exception:
                    flash('Please provide valid holiday start and end dates.', 'error')
                    return redirect(url_for('admin_student_management'))

                if end_date < start_date:
                    flash('Holiday end date cannot be earlier than start date.', 'error')
                    return redirect(url_for('admin_student_management'))

                if holiday_mode not in ('bulk', 'class_section'):
                    holiday_mode = 'bulk'

                if holiday_mode == 'class_section' and not holiday_class:
                    flash('Please select class for class/section holiday mode.', 'error')
                    return redirect(url_for('admin_student_management'))

                target_class_value = holiday_class if holiday_mode == 'class_section' else None
                target_section_value = holiday_section if holiday_mode == 'class_section' and holiday_section else None

                db.execute('''
                    INSERT INTO student_holidays (
                        school_id, holiday_title, holiday_start_date, holiday_end_date, description,
                        target_mode, target_class, target_section, created_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    school_id,
                    holiday_title,
                    start_date,
                    end_date,
                    holiday_description,
                    holiday_mode,
                    target_class_value,
                    target_section_value,
                    session.get('user_id')
                ))
                db.commit()

                if holiday_mode == 'bulk':
                    flash('Student holiday assigned in bulk successfully.', 'success')
                else:
                    section_suffix = f" - {target_section_value}" if target_section_value else ' (All Sections)'
                    flash(f'Student holiday assigned for {target_class_value}{section_suffix}.', 'success')

            elif action == 'update_student_holiday':
                _ensure_student_holidays_table()

                holiday_id = (request.form.get('holiday_id') or '').strip()
                holiday_title = (request.form.get('holiday_title') or '').strip()
                holiday_start_date = (request.form.get('holiday_start_date') or '').strip()
                holiday_end_date = (request.form.get('holiday_end_date') or '').strip()
                holiday_description = (request.form.get('holiday_description') or '').strip()
                holiday_mode = (request.form.get('holiday_mode') or 'bulk').strip().lower()
                holiday_class = (request.form.get('holiday_class') or '').strip()
                holiday_section = (request.form.get('holiday_section') or '').strip()

                if not holiday_id:
                    flash('Holiday ID is required for update.', 'error')
                    return redirect(url_for('admin_student_management'))

                if not holiday_title:
                    flash('Holiday title is required.', 'error')
                    return redirect(url_for('admin_student_management'))

                try:
                    start_date = datetime.date.fromisoformat(holiday_start_date).strftime('%Y-%m-%d')
                    end_date = datetime.date.fromisoformat(holiday_end_date).strftime('%Y-%m-%d')
                except Exception:
                    flash('Please provide valid holiday start and end dates.', 'error')
                    return redirect(url_for('admin_student_management'))

                if end_date < start_date:
                    flash('Holiday end date cannot be earlier than start date.', 'error')
                    return redirect(url_for('admin_student_management'))

                if holiday_mode not in ('bulk', 'class_section'):
                    holiday_mode = 'bulk'

                if holiday_mode == 'class_section' and not holiday_class:
                    flash('Please select class for class/section holiday mode.', 'error')
                    return redirect(url_for('admin_student_management'))

                target_class_value = holiday_class if holiday_mode == 'class_section' else None
                target_section_value = holiday_section if holiday_mode == 'class_section' and holiday_section else None

                result = db.execute('''
                    UPDATE student_holidays
                    SET holiday_title = ?,
                        holiday_start_date = ?,
                        holiday_end_date = ?,
                        description = ?,
                        target_mode = ?,
                        target_class = ?,
                        target_section = ?
                    WHERE id = ? AND school_id = ?
                ''', (
                    holiday_title,
                    start_date,
                    end_date,
                    holiday_description,
                    holiday_mode,
                    target_class_value,
                    target_section_value,
                    holiday_id,
                    school_id
                ))
                db.commit()

                if result.rowcount:
                    flash('Student holiday updated successfully.', 'success')
                else:
                    flash('Student holiday not found or already removed.', 'warning')

            elif action == 'delete_student_holiday':
                _ensure_student_holidays_table()

                holiday_id = (request.form.get('holiday_id') or '').strip()
                if not holiday_id:
                    flash('Holiday ID is required for delete.', 'error')
                    return redirect(url_for('admin_student_management'))

                result = db.execute('''
                    DELETE FROM student_holidays
                    WHERE id = ? AND school_id = ?
                ''', (holiday_id, school_id))
                db.commit()

                if result.rowcount:
                    flash('Student holiday deleted successfully.', 'success')
                else:
                    flash('Student holiday not found or already removed.', 'warning')
                
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('admin_student_management'))
    
    # GET request - display page
    # Get all students - use backticks for MySQL reserved keyword 'class'
    students = db.execute('''
        SELECT * FROM students
        WHERE school_id = ?
        ORDER BY `class`, section, roll_number
    ''', (school_id,)).fetchall()
    
    # Get statistics
    total_students = len(students)
    
    # Students added this month
    current_month = datetime.date.today().strftime('%Y-%m')
    students_this_month = db.execute('''
        SELECT COUNT(*) as count FROM students
        WHERE school_id = ? AND strftime('%Y-%m', created_at) = ?
    ''', (school_id, current_month)).fetchone()['count']
    
    # Fetch classes and sections from timetable academic hierarchy
    # Get active academic levels (grades/classes)
    academic_levels = db.execute('''
        SELECT id, level_name, level_number 
        FROM timetable_academic_levels 
        WHERE school_id = ? AND is_active = 1 
        ORDER BY level_number ASC
    ''', (school_id,)).fetchall()
    
    # Get active sections with their associated levels
    timetable_sections = db.execute('''
        SELECT ts.id, ts.section_name, ts.section_code, ts.level_id, tal.level_name
        FROM timetable_sections ts
        JOIN timetable_academic_levels tal ON ts.level_id = tal.id
        WHERE ts.school_id = ? AND ts.is_active = 1
        ORDER BY tal.level_number, ts.section_name
    ''', (school_id,)).fetchall()
    
    # Build class and section lists
    unique_classes = [level['level_name'] for level in academic_levels]
    
    # Build sections mapping by class for dynamic filtering
    import json
    sections_by_class = {}
    # Also map (class, section) -> section_id for timetable lookups
    section_id_map = {}  # { 'ClassName': { 'SectionName': section_id } }
    level_id_map = {}    # { 'ClassName': level_id }
    for level in academic_levels:
        level_id_map[level['level_name']] = level['id']
    for section in timetable_sections:
        class_name = section['level_name']
        if class_name not in sections_by_class:
            sections_by_class[class_name] = []
            section_id_map[class_name] = {}
        sections_by_class[class_name].append(section['section_name'])
        section_id_map[class_name][section['section_name']] = section['id']
    
    # Get all unique section names (across all levels) for filter dropdown
    unique_sections_set = set()
    for section in timetable_sections:
        unique_sections_set.add(section['section_name'])
    unique_sections = sorted(list(unique_sections_set))
    
    # Calculate total classes from academic hierarchy
    total_classes = len(academic_levels)

    # Build class -> subjects map from timetable assignments
    class_subjects_map = {}
    subject_rows = db.execute('''
        SELECT DISTINCT tal.level_name AS class_name, ha.subject_name
        FROM timetable_hierarchical_assignments ha
        JOIN timetable_academic_levels tal ON ha.level_id = tal.id
        WHERE ha.school_id = ?
          AND ha.subject_name IS NOT NULL
          AND TRIM(ha.subject_name) <> ''
        ORDER BY tal.level_name, ha.subject_name
    ''', (school_id,)).fetchall()

    for row in subject_rows:
        cls = row['class_name']
        sub = row['subject_name']
        if cls not in class_subjects_map:
            class_subjects_map[cls] = []
        if sub not in class_subjects_map[cls]:
            class_subjects_map[cls].append(sub)

    _ensure_exam_assignments_table()
    exam_assignments = db.execute('''
        SELECT id, class_name, exam_title, exam_date, description, created_at
        FROM exam_assignments
        WHERE school_id = ?
        ORDER BY exam_date DESC, id DESC
        LIMIT 25
    ''', (school_id,)).fetchall()

    _ensure_student_holidays_table()
    student_holidays = db.execute('''
        SELECT id, holiday_title, holiday_start_date, holiday_end_date, description,
               target_mode, target_class, target_section, created_at
        FROM student_holidays
        WHERE school_id = ?
        ORDER BY holiday_start_date DESC, id DESC
        LIMIT 25
    ''', (school_id,)).fetchall()
    
    # Get module enabled status
    module_enabled = get_module_enabled(school_id)

    school_row = db.execute('SELECT name FROM schools WHERE id = ?', (school_id,)).fetchone()
    school_name = school_row['name'] if school_row else 'VishnoRex School'

    return render_template('admin_student_management.html',
                         students=students,
                         total_students=total_students,
                         students_this_month=students_this_month,
                         total_classes=total_classes,
                         student_holidays=student_holidays,
                         exam_assignments=exam_assignments,
                         class_subjects_map=json.dumps(class_subjects_map),
                         unique_classes=unique_classes,
                         unique_sections=unique_sections,
                         sections_by_class=json.dumps(sections_by_class),
                         section_id_map=json.dumps(section_id_map),
                         level_id_map=json.dumps(level_id_map),
                         module_enabled=module_enabled,
                         school_name=school_name)


@app.route('/admin/students/download-template')
def download_student_template():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    import pandas as pd
    from io import BytesIO

    # Includes all Add Student form details except direct photo file upload.
    template_data = {
        'admission_number': ['ADM1001', 'ADM1002'],
        'roll_number': ['1', '2'],
        'student_type': ['Day Scholar', 'Hostel'],
        'first_name': ['Arun', 'Meena'],
        'last_name': ['Kumar', 'Ravi'],
        'date_of_birth': ['2010-06-12', '2011-09-03'],
        'age': [15, 14],
        'gender': ['Male', 'Female'],
        'class': ['Grade 10', 'Grade 9'],
        'section': ['A', 'B'],
        'academic_year': ['2025-2026', '2025-2026'],
        'student_mobile': ['9876543210', '9123456780'],
        'address': ['12, Lake Road, City', '45, Park Street, City'],
        'tenth_is_diploma': ['No', 'No'],
        'tenth_marks': ['450/500', '430/500'],
        'tenth_percentage': [90.0, 86.0],
        'twelfth_is_diploma': ['No', 'No'],
        'twelfth_marks': ['', ''],
        'twelfth_percentage': ['', ''],
        'skills': ['Coding, Chess', 'Drawing, Music'],
        'parent_name': ['Kumar', 'Ravi'],
        'parent_mobile': ['9876500001', '9876500002'],
        'parent_email': ['kumar.parent@example.com', 'ravi.parent@example.com'],
        'mother_name': ['Lakshmi', 'Saroja'],
        'mother_phone': ['9876501001', '9876501002'],
        'parent_occupation': ['Engineer', 'Teacher'],
        'tc_number': ['TC-2025-10', 'TC-2025-11'],
        'aadhar_number': ['123456789012', '234567890123'],
        'custom_fields_json': ['{"blood_group":"B+"}', '{"blood_group":"O+"}']
    }

    instructions_data = {
        'Field': [
            'admission_number', 'roll_number', 'student_type', 'first_name', 'last_name',
            'date_of_birth', 'age', 'gender', 'class', 'section', 'academic_year',
            'student_mobile', 'address', 'tenth_is_diploma', 'tenth_marks', 'tenth_percentage',
            'twelfth_is_diploma', 'twelfth_marks', 'twelfth_percentage', 'skills', 'parent_name',
            'parent_mobile', 'parent_email', 'mother_name', 'mother_phone', 'parent_occupation',
            'tc_number', 'aadhar_number', 'custom_fields_json'
        ],
        'Required': [
            'Either admission_number or roll_number',
            'Either admission_number or roll_number',
            'Yes', 'Yes', 'Yes',
            'Yes', 'No', 'Yes', 'Yes', 'Yes', 'Yes',
            'No', 'Yes', 'No', 'No', 'No',
            'No', 'No', 'No', 'No', 'Yes',
            'Yes', 'No', 'No', 'No', 'No',
            'No', 'No', 'No'
        ],
        'Format / Notes': [
            'Text; used to generate student login ID',
            'Text; used when admission_number is empty',
            'Day Scholar or Hostel',
            'Text',
            'Text',
            'YYYY-MM-DD',
            'Optional; auto-calculated if blank',
            'Male/Female/Other',
            'Must match configured class name in system',
            'Must match configured section for the selected class',
            'Example: 2025-2026',
            '10-digit mobile',
            'Complete address',
            'Yes/No',
            'Example: 450/500',
            'Numeric',
            'Yes/No',
            'Example: 480/500',
            'Numeric',
            'Comma-separated free text',
            'Father/Guardian name',
            '10-digit mobile',
            'Email',
            'Text',
            '10-digit mobile',
            'Text',
            'Text',
            '12-digit Aadhaar',
            'JSON object for custom fields (optional)'
        ]
    }

    template_df = pd.DataFrame(template_data)
    instructions_df = pd.DataFrame(instructions_data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        template_df.to_excel(writer, sheet_name='Student Template', index=False)
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=student_import_template.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/admin/students/bulk-upload', methods=['POST'])
def bulk_upload_students():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})

    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'No file selected'})

    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({'success': False, 'error': 'Unsupported file format. Use .xlsx, .xls, or .csv'})

    import pandas as pd
    import json

    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Unable to read file: {e}'})

    if df.empty:
        return jsonify({'success': False, 'error': 'Uploaded file is empty'})

    df.columns = [str(c).strip().lower() for c in df.columns]

    school_id = session['school_id']
    db = get_db()
    columns = db.execute('PRAGMA table_info(students)').fetchall()
    column_names = {col['name'] for col in columns}

    def clean_text(value):
        if pd.isna(value):
            return ''
        text = str(value).strip()
        if text.lower() in {'nan', 'none', 'null'}:
            return ''
        return text

    def clean_mobile(value):
        raw = clean_text(value)
        if not raw:
            return ''
        digits = ''.join(ch for ch in raw if ch.isdigit())
        return digits[:10] if digits else ''

    def clean_date(value):
        if pd.isna(value):
            return ''
        if isinstance(value, (datetime.datetime, datetime.date, pd.Timestamp)):
            return pd.to_datetime(value).strftime('%Y-%m-%d')
        text = clean_text(value)
        if not text:
            return ''
        try:
            return pd.to_datetime(text).strftime('%Y-%m-%d')
        except Exception:
            return ''

    def clean_int(value):
        if pd.isna(value):
            return None
        text = clean_text(value)
        if not text:
            return None
        try:
            return int(float(text))
        except Exception:
            return None

    def clean_float(value):
        if pd.isna(value):
            return None
        text = clean_text(value)
        if not text:
            return None
        try:
            return float(text)
        except Exception:
            return None

    def quote_col(name):
        return f'`{name}`'

    imported_count = 0
    errors = []

    for idx, row in df.iterrows():
        row_no = idx + 2
        try:
            admission_number = clean_text(row.get('admission_number'))
            roll_number = clean_text(row.get('roll_number'))
            student_type = clean_text(row.get('student_type')) or 'Day Scholar'
            first_name = clean_text(row.get('first_name'))
            last_name = clean_text(row.get('last_name'))
            date_of_birth = clean_date(row.get('date_of_birth'))
            age = clean_int(row.get('age'))
            gender = clean_text(row.get('gender'))
            class_name = clean_text(row.get('class'))
            section = clean_text(row.get('section'))
            academic_year = clean_text(row.get('academic_year'))
            student_mobile = clean_mobile(row.get('student_mobile'))
            address = clean_text(row.get('address'))

            tenth_marks = clean_text(row.get('tenth_marks'))
            tenth_percentage = clean_float(row.get('tenth_percentage'))
            twelfth_marks = clean_text(row.get('twelfth_marks'))
            twelfth_percentage = clean_float(row.get('twelfth_percentage'))
            skills = clean_text(row.get('skills'))

            parent_name = clean_text(row.get('parent_name'))
            parent_mobile = clean_mobile(row.get('parent_mobile'))
            parent_email = clean_text(row.get('parent_email'))
            mother_name = clean_text(row.get('mother_name'))
            mother_phone = clean_mobile(row.get('mother_phone'))
            parent_occupation = clean_text(row.get('parent_occupation'))

            tc_number = clean_text(row.get('tc_number'))
            aadhar_number = ''.join(ch for ch in clean_text(row.get('aadhar_number')) if ch.isdigit())

            custom_fields_json = clean_text(row.get('custom_fields_json'))
            if custom_fields_json:
                try:
                    json.loads(custom_fields_json)
                except Exception:
                    errors.append(f'Row {row_no}: custom_fields_json must be valid JSON')
                    continue

            if not admission_number and not roll_number:
                errors.append(f'Row {row_no}: either admission_number or roll_number is required')
                continue
            if not first_name or not last_name:
                errors.append(f'Row {row_no}: first_name and last_name are required')
                continue
            if not date_of_birth:
                errors.append(f'Row {row_no}: date_of_birth is required in YYYY-MM-DD')
                continue
            if not gender:
                errors.append(f'Row {row_no}: gender is required')
                continue
            if not class_name or not section:
                errors.append(f'Row {row_no}: class and section are required')
                continue
            if not academic_year:
                errors.append(f'Row {row_no}: academic_year is required')
                continue
            if not address:
                errors.append(f'Row {row_no}: address is required')
                continue
            if not parent_name or not parent_mobile:
                errors.append(f'Row {row_no}: parent_name and parent_mobile are required')
                continue

            student_id = f"STU{admission_number}" if admission_number else f"STU{roll_number}"
            full_name = f"{first_name} {last_name}".strip()

            existing = db.execute('''
                SELECT id FROM students
                WHERE school_id = ? AND (student_id = ? OR admission_number = ? OR roll_number = ?)
            ''', (school_id, student_id, admission_number, roll_number)).fetchone()

            if existing:
                errors.append(f'Row {row_no}: student with same student_id/admission_number/roll_number already exists')
                continue

            insert_map = {
                'school_id': school_id,
                'student_id': student_id,
                'password': generate_password_hash('student123'),
                'full_name': full_name,
                'first_name': first_name,
                'last_name': last_name,
                'admission_number': admission_number,
                'roll_number': roll_number,
                'student_type': student_type,
                'date_of_birth': date_of_birth,
                'age': age,
                'gender': gender,
                'address': address,
                'class': class_name,
                'section': section,
                'academic_year': academic_year,
                'parent_name': parent_name,
                'parent_phone': parent_mobile,
                'parent_email': parent_email,
                'mother_name': mother_name,
                'mother_phone': mother_phone,
                'parent_occupation': parent_occupation,
                'tenth_marks': tenth_marks,
                'tenth_percentage': tenth_percentage,
                'twelfth_marks': twelfth_marks,
                'twelfth_percentage': twelfth_percentage,
                'skills': skills,
                'tc_number': tc_number,
                'aadhar_number': aadhar_number,
                'custom_fields': custom_fields_json,
                'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            # Support either legacy phone or explicit student_mobile column.
            if 'student_mobile' in column_names:
                insert_map['student_mobile'] = student_mobile
            elif 'phone' in column_names:
                insert_map['phone'] = student_mobile

            filtered = {k: v for k, v in insert_map.items() if k in column_names and v is not None}
            columns_sql = ', '.join(quote_col(c) for c in filtered.keys())
            values = list(filtered.values())
            placeholders = ', '.join(['?' for _ in values])
            query = f'INSERT INTO students ({columns_sql}) VALUES ({placeholders})'
            db.execute(query, values)
            imported_count += 1

        except Exception as e:
            errors.append(f'Row {row_no}: {e}')

    db.commit()

    return jsonify({
        'success': True,
        'imported_count': imported_count,
        'errors': errors,
        'total_rows': len(df)
    })


@app.route('/api/student-dashboard-data')
def api_student_dashboard_data():
    """API endpoint to get student attendance dashboard data for today"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    db = get_db()
    school_id = session['school_id']
    today = datetime.date.today().strftime('%Y-%m-%d')

    def _serialize_time_value(value):
        """Convert DB time-like values to a JSON-safe string."""
        if value in (None, ''):
            return '--:--:--'
        if isinstance(value, str):
            return value
        try:
            return value.strftime('%H:%M:%S')
        except Exception:
            try:
                return value.isoformat()
            except Exception:
                return str(value)

    # Get all students with today's attendance
    students = db.execute('''
        SELECT s.id, s.student_id, s.full_name, s.first_name, s.last_name,
               s.`class`, s.section, s.roll_number, s.photo_data, s.gender,
               sa.morning_status, sa.morning_time, sa.afternoon_status, sa.afternoon_time
        FROM students s
        LEFT JOIN student_attendance sa ON s.id = sa.student_id AND sa.attendance_date = ?
        WHERE s.school_id = ?
        ORDER BY s.`class`, s.section, s.roll_number
    ''', (today, school_id)).fetchall()

    student_list = []
    summary = {'total': 0, 'present': 0, 'absent': 0, 'late': 0, 'leave': 0, 'not_marked': 0}

    for s in students:
        summary['total'] += 1
        morning = s['morning_status'] or ''
        afternoon = s['afternoon_status'] or ''

        # Determine overall status
        if morning == 'present' or afternoon == 'present':
            status = 'Present'
            summary['present'] += 1
        elif morning == 'late' or afternoon == 'late':
            status = 'Late'
            summary['late'] += 1
        elif morning == 'leave' or afternoon == 'leave':
            status = 'On Leave'
            summary['leave'] += 1
        elif morning == 'absent' or afternoon == 'absent':
            status = 'Absent'
            summary['absent'] += 1
        else:
            status = 'Not Marked'
            summary['not_marked'] += 1

        student_list.append({
            'id': s['id'],
            'student_id': s['student_id'],
            'full_name': s['full_name'],
            'class': s['class'],
            'section': s['section'],
            'roll_number': s['roll_number'],
            'gender': s['gender'] or '',
            'morning_status': morning,
            'morning_time': _serialize_time_value(s['morning_time']),
            'afternoon_status': afternoon,
            'afternoon_time': _serialize_time_value(s['afternoon_time']),
            'status': status,
            'has_photo': bool(s['photo_data'])
        })

    return jsonify({
        'success': True,
        'date': today,
        'summary': summary,
        'students': student_list
    })


@app.route('/api/student-attendance-rules', methods=['GET', 'POST'])
def api_student_attendance_rules():
    """Get and update admin-configurable student attendance time-slot rules."""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    try:
        db = get_db()
        school_id = session['school_id']
        setting_key = f"student_attendance_rule_school_{school_id}"

        from database import get_system_setting, set_system_setting

        def _serialize_time(value):
            """Convert DB time values to JSON-safe string representation."""
            if value is None:
                return ''
            if isinstance(value, str):
                return value
            try:
                return value.strftime('%H:%M:%S')
            except Exception:
                try:
                    return value.isoformat()
                except Exception:
                    return str(value)

        available_time_slots = []

        # Preferred source: explicit timetable time slots.
        try:
            slot_rows = db.execute('''
                SELECT id, slot_label, period_sequence, start_time, end_time
                FROM timetable_period_timings
                WHERE school_id = ? AND is_active = 1
                ORDER BY COALESCE(period_sequence, 9999) ASC, start_time ASC, id ASC
            ''', (school_id,)).fetchall()

            for row in slot_rows:
                slot_id = int(row['id'])
                available_time_slots.append({
                    'time_slot_id': slot_id,
                    'slot_label': row['slot_label'] or f"Time Slot {slot_id}",
                    'period_sequence': row['period_sequence'],
                    'start_time': _serialize_time(row['start_time']),
                    'end_time': _serialize_time(row['end_time'])
                })
        except Exception:
            # Backward-compatible fallback for older databases: infer slots from timetable_periods.
            period_rows = db.execute('''
                SELECT DISTINCT period_number, period_name, start_time, end_time
                FROM timetable_periods
                WHERE school_id = ?
                ORDER BY period_number ASC
            ''', (school_id,)).fetchall()

            for row in period_rows:
                slot_id = int(row['period_number'])
                available_time_slots.append({
                    'time_slot_id': slot_id,
                    'slot_label': row['period_name'] or f"Period {slot_id}",
                    'period_sequence': slot_id,
                    'start_time': _serialize_time(row['start_time']),
                    'end_time': _serialize_time(row['end_time'])
                })

        raw_setting = get_system_setting(setting_key, None)
        rule = {'mode': 'all', 'time_slot_ids': []}
        if raw_setting:
            try:
                parsed = json.loads(raw_setting)
                mode = str(parsed.get('mode', 'all')).lower()
                slot_ids = parsed.get('time_slot_ids', None)
                # Backward compatibility for older data saved with period_numbers.
                if slot_ids is None:
                    slot_ids = parsed.get('period_numbers', [])
                if mode in ['all', 'selected'] and isinstance(slot_ids, list):
                    rule['mode'] = mode
                    unique_slot_ids = []
                    for item in slot_ids:
                        try:
                            slot_id = int(item)
                        except (TypeError, ValueError):
                            continue
                        if slot_id not in unique_slot_ids:
                            unique_slot_ids.append(slot_id)
                    rule['time_slot_ids'] = unique_slot_ids
            except (TypeError, ValueError, json.JSONDecodeError):
                pass

        valid_time_slot_ids = {slot['time_slot_id'] for slot in available_time_slots}
        rule['time_slot_ids'] = [slot_id for slot_id in rule['time_slot_ids'] if slot_id in valid_time_slot_ids]

        if request.method == 'POST':
            data = request.get_json(silent=True) or {}
            mode = str(data.get('mode', 'all')).lower()
            time_slot_ids = data.get('time_slot_ids', [])

            if mode not in ['all', 'selected']:
                return jsonify({'success': False, 'error': 'Invalid attendance mode'}), 400

            if not isinstance(time_slot_ids, list):
                return jsonify({'success': False, 'error': 'Invalid time slot selection'}), 400

            clean_time_slot_ids = []
            for item in time_slot_ids:
                try:
                    slot_id = int(item)
                except (TypeError, ValueError):
                    continue
                if slot_id in valid_time_slot_ids and slot_id not in clean_time_slot_ids:
                    clean_time_slot_ids.append(slot_id)

            if mode == 'selected' and not clean_time_slot_ids:
                return jsonify({'success': False, 'error': 'Select at least one time slot'}), 400

            if mode == 'all':
                clean_time_slot_ids = []

            rule = {
                'mode': mode,
                'time_slot_ids': clean_time_slot_ids
            }

            description = 'Student attendance rule per school: all time slots or selected time slots'
            serialized_rule = json.dumps(rule)

            save_ok = set_system_setting(setting_key, serialized_rule, description)

            # Fallback path: direct update/insert for environments where helper upsert can fail.
            if not save_ok:
                try:
                    existing_setting = db.execute('''
                        SELECT id FROM system_settings WHERE setting_key = ?
                    ''', (setting_key,)).fetchone()

                    current_time = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

                    if existing_setting:
                        db.execute('''
                            UPDATE system_settings
                            SET setting_value = ?,
                                description = COALESCE(?, description),
                                updated_at = ?
                            WHERE setting_key = ?
                        ''', (serialized_rule, description, current_time, setting_key))
                    else:
                        db.execute('''
                            INSERT INTO system_settings (setting_key, setting_value, description, updated_at)
                            VALUES (?, ?, ?, ?)
                        ''', (setting_key, serialized_rule, description, current_time))

                    db.commit()
                    save_ok = True
                except Exception as save_error:
                    db.rollback()
                    return jsonify({'success': False, 'error': f'Could not save attendance rules: {str(save_error)}'}), 500

            if not save_ok:
                return jsonify({'success': False, 'error': 'Could not save attendance rules'}), 500

            return jsonify({
                'success': True,
                'message': 'Attendance rules updated successfully',
                'rule': rule,
                'available_time_slots': available_time_slots
            })

        return jsonify({
            'success': True,
            'rule': rule,
            'available_time_slots': available_time_slots
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Could not load attendance rules: {str(e)}'}), 500


@app.route('/api/student-attendance-override-data')
def api_student_attendance_override_data():
    """Return class/section/date scoped student attendance for admin manual override."""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    try:
        school_id = session['school_id']
        class_name = (request.args.get('class') or '').strip()
        section_name = (request.args.get('section') or '').strip()
        attendance_date = (request.args.get('date') or '').strip()

        if not class_name or not section_name or not attendance_date:
            return jsonify({'success': False, 'error': 'Class, section and date are required'}), 400

        try:
            datetime.datetime.strptime(attendance_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date format'}), 400

        db = get_db()
        rows = db.execute('''
            SELECT s.id, s.student_id, s.full_name, s.roll_number, s.`class`, s.section,
                   sa.morning_status, sa.afternoon_status, sa.morning_time, sa.afternoon_time
            FROM students s
            LEFT JOIN student_attendance sa
                ON sa.student_id = s.id AND sa.attendance_date = ?
            WHERE s.school_id = ?
              AND s.`class` = ?
              AND s.section = ?
            ORDER BY COALESCE(NULLIF(TRIM(s.roll_number), ''), s.student_id), s.full_name
        ''', (attendance_date, school_id, class_name, section_name)).fetchall()

        def _serialize_time_value(value):
            if value is None:
                return ''
            if isinstance(value, str):
                return value
            try:
                return value.strftime('%H:%M:%S')
            except Exception:
                try:
                    return value.isoformat()
                except Exception:
                    return str(value)

        students = []
        for row in rows:
            students.append({
                'id': row['id'],
                'student_id': row['student_id'],
                'full_name': row['full_name'],
                'roll_number': row['roll_number'] or '',
                'class': row['class'],
                'section': row['section'],
                'morning_status': (row['morning_status'] or '').lower(),
                'afternoon_status': (row['afternoon_status'] or '').lower(),
                'morning_time': _serialize_time_value(row['morning_time']),
                'afternoon_time': _serialize_time_value(row['afternoon_time'])
            })

        return jsonify({
            'success': True,
            'class': class_name,
            'section': section_name,
            'date': attendance_date,
            'students': students
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to load override data: {str(e)}'}), 500


@app.route('/api/student-attendance-override-save', methods=['POST'])
def api_student_attendance_override_save():
    """Persist admin manual override attendance statuses for morning and afternoon sessions."""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    try:
        payload = request.get_json(silent=True) or {}
        attendance_date = (payload.get('date') or '').strip()
        updates = payload.get('updates', [])

        if not attendance_date:
            return jsonify({'success': False, 'error': 'Date is required'}), 400
        if not isinstance(updates, list) or not updates:
            return jsonify({'success': False, 'error': 'No attendance updates provided'}), 400

        try:
            datetime.datetime.strptime(attendance_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date format'}), 400

        valid_statuses = {'', 'present', 'absent', 'late', 'leave'}
        school_id = session['school_id']
        session_user_id = session['user_id']
        now_time = datetime.datetime.now().strftime('%H:%M:%S')

        db = get_db()

        # morning_marked_by/afternoon_marked_by reference staff.id.
        # For admin accounts that are not present in staff, store NULL to avoid FK violations.
        marker_row = db.execute('''
            SELECT id
            FROM staff
            WHERE id = ? AND school_id = ?
        ''', (session_user_id, school_id)).fetchone()
        marked_by_staff_id = marker_row['id'] if marker_row else None

        for item in updates:
            try:
                student_id = int(item.get('student_id'))
            except (TypeError, ValueError):
                return jsonify({'success': False, 'error': 'Invalid student id in update payload'}), 400

            belongs = db.execute('''
                SELECT id
                FROM students
                WHERE id = ? AND school_id = ?
            ''', (student_id, school_id)).fetchone()
            if not belongs:
                return jsonify({'success': False, 'error': f'Invalid student id: {student_id}'}), 400

            morning_status = (item.get('morning_status') or '').strip().lower()
            afternoon_status = (item.get('afternoon_status') or '').strip().lower()

            if morning_status not in valid_statuses or afternoon_status not in valid_statuses:
                return jsonify({'success': False, 'error': 'Invalid attendance status supplied'}), 400

            existing = db.execute('''
                SELECT id, morning_time, afternoon_time
                FROM student_attendance
                WHERE student_id = ? AND attendance_date = ?
            ''', (student_id, attendance_date)).fetchone()

            morning_time = now_time if morning_status else None
            afternoon_time = now_time if afternoon_status else None

            if existing:
                db.execute('''
                    UPDATE student_attendance
                    SET morning_status = ?,
                        morning_time = ?,
                        morning_marked_by = ?,
                        morning_notes = ?,
                        afternoon_status = ?,
                        afternoon_time = ?,
                        afternoon_marked_by = ?,
                        afternoon_notes = ?
                    WHERE student_id = ? AND attendance_date = ?
                ''', (
                    morning_status if morning_status else None,
                    morning_time,
                    marked_by_staff_id if morning_status else None,
                    'Admin manual override' if morning_status else None,
                    afternoon_status if afternoon_status else None,
                    afternoon_time,
                    marked_by_staff_id if afternoon_status else None,
                    'Admin manual override' if afternoon_status else None,
                    student_id,
                    attendance_date
                ))
            else:
                db.execute('''
                    INSERT INTO student_attendance (
                        student_id, school_id, attendance_date,
                        morning_status, morning_time, morning_marked_by, morning_notes,
                        afternoon_status, afternoon_time, afternoon_marked_by, afternoon_notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    student_id,
                    school_id,
                    attendance_date,
                    morning_status if morning_status else None,
                    morning_time,
                    marked_by_staff_id if morning_status else None,
                    'Admin manual override' if morning_status else None,
                    afternoon_status if afternoon_status else None,
                    afternoon_time,
                    marked_by_staff_id if afternoon_status else None,
                    'Admin manual override' if afternoon_status else None
                ))

        db.commit()
        return jsonify({'success': True, 'message': 'Attendance override saved successfully'})
    except Exception as e:
        try:
            get_db().rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'error': f'Failed to save attendance override: {str(e)}'}), 500


# ========================= FEE MANAGEMENT APIs =========================

def _fee_print_user_authorized():
    """Allow any logged-in user (admin, sub-admin, staff, student) to request fee receipt printing."""
    has_admin_session  = 'user_id' in session
    has_student_session = 'student_id' in session
    return has_admin_session or has_student_session


def _detect_windows_printers():
    """Return installed printers and best thermal/default candidates on Windows."""
    if os.name != 'nt':
        return {'supported': False, 'reason': 'Auto printer routing is supported on Windows only'}
    try:
        import win32print  # type: ignore
    except Exception:
        return {'supported': False, 'reason': 'pywin32 is not available'}

    flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
    enum_rows = win32print.EnumPrinters(flags)
    names = []
    for row in enum_rows:
        if isinstance(row, tuple) and len(row) >= 3 and row[2]:
            names.append(str(row[2]))
    names = list(dict.fromkeys(names))

    try:
        default_printer = win32print.GetDefaultPrinter()
    except Exception:
        default_printer = None

    thermal_keywords = (
        'thermal', 'pos', 'receipt', '58mm', '80mm', 'xp-80', 'xprinter',
        'epson tm', 'tm-t', 'tsp100', 'sunmi', 'zjiang', 'posbox'
    )
    thermal_candidates = [p for p in names if any(k in p.lower() for k in thermal_keywords)]

    # Strict mode: thermal is enabled only when printer is both READY and USB connected.
    thermal_printer = None
    thermal_ready = False
    for candidate in thermal_candidates:
        ready, _ = _is_windows_printer_ready(candidate)
        usb_connected, _ = _is_windows_printer_usb_connected(candidate)
        present, _ = _is_windows_printer_physically_present(candidate)
        if ready and usb_connected and present:
            thermal_printer = candidate
            thermal_ready = True
            break

    return {
        'supported': True,
        'all': names,
        'default': default_printer,
        'thermal': thermal_printer,
        'thermal_ready': thermal_ready,
        'thermal_candidates': thermal_candidates,
    }


def _is_windows_printer_usb_connected(printer_name):
    """Check whether a Windows printer is attached via a USB port."""
    if os.name != 'nt':
        return False, 'Not Windows'
    try:
        import win32print  # type: ignore
        handle = win32print.OpenPrinter(printer_name)
        try:
            info = win32print.GetPrinter(handle, 2)
            port_name = str(info.get('pPortName') or '')
        finally:
            win32print.ClosePrinter(handle)
    except Exception as exc:
        return False, f'Unable to query printer port: {exc}'

    # Typical direct USB ports are USB001/USB002.
    # Many receipt drivers map USB devices to virtual local ports like CP001/DOT4_001.
    port_upper = port_name.upper()
    is_usb = (
        'USB' in port_upper
        or port_upper.startswith('CP')
        or port_upper.startswith('DOT4')
    )
    if not is_usb:
        return False, f'Not USB port ({port_name})'
    return True, f'USB connected ({port_name})'


def _is_windows_printer_physically_present(printer_name):
    """Check whether Windows reports the printer queue device as physically present."""
    if os.name != 'nt':
        return False, 'Not Windows'

    # Queue status can remain "Normal" while cable is unplugged.
    # PnP PrintQueue Present flag is a stricter hardware-availability signal.
    safe_name = str(printer_name or '').replace("'", "''")
    ps_script = (
        "$d = Get-PnpDevice | Where-Object { $_.Class -eq 'PrintQueue' -and $_.FriendlyName -eq '" + safe_name + "' } "
        "| Select-Object -First 1 FriendlyName,Status,Present; "
        "if ($null -eq $d) { '' } else { $d | ConvertTo-Json -Compress }"
    )

    try:
        result = subprocess.run(
            ['powershell', '-NoProfile', '-Command', ps_script],
            capture_output=True,
            text=True,
            timeout=4,
            check=False,
        )
    except Exception as exc:
        return False, f'Unable to query PnP presence: {exc}'

    out = (result.stdout or '').strip()
    if not out:
        return False, 'PrintQueue device not found in PnP'

    try:
        payload = json.loads(out)
    except Exception:
        return False, f'Unexpected PnP response: {out[:120]}'

    present = bool(payload.get('Present'))
    status = str(payload.get('Status') or '').strip().lower()
    if present and status == 'ok':
        return True, 'PnP present'
    return False, f'PnP not present (status={payload.get("Status")}, present={payload.get("Present")})'


def _is_windows_printer_ready(printer_name):
    """Check whether a Windows printer queue appears online/ready for printing."""
    if os.name != 'nt':
        return False, 'Not Windows'
    try:
        import win32print  # type: ignore
        handle = win32print.OpenPrinter(printer_name)
        try:
            info = win32print.GetPrinter(handle, 2)
            status = int(info.get('Status', 0) or 0)
        finally:
            win32print.ClosePrinter(handle)
    except Exception as exc:
        return False, f'Unable to query printer status: {exc}'

    bad_bits = (
        getattr(win32print, 'PRINTER_STATUS_OFFLINE', 0x80)
        | getattr(win32print, 'PRINTER_STATUS_ERROR', 0x2)
        | getattr(win32print, 'PRINTER_STATUS_NOT_AVAILABLE', 0x1000)
        | getattr(win32print, 'PRINTER_STATUS_PAPER_JAM', 0x8)
        | getattr(win32print, 'PRINTER_STATUS_PAPER_OUT', 0x10)
        | getattr(win32print, 'PRINTER_STATUS_USER_INTERVENTION', 0x100000)
        | getattr(win32print, 'PRINTER_STATUS_DOOR_OPEN', 0x400000)
    )

    if status & bad_bits:
        return False, f'Printer not ready (status={status})'
    return True, 'Ready'


def _build_fee_receipt_text(receipt):
    """Build printer-friendly plain-text receipt for thermal printers."""
    width = 48

    def esc(v):
        return str(v or '').strip()

    def money(v):
        try:
            return f"{float(v or 0):.2f}"
        except Exception:
            return "0.00"

    def center(text):
        text = esc(text)
        if len(text) >= width:
            return text[:width]
        pad = (width - len(text)) // 2
        return (' ' * pad) + text

    fees = receipt.get('fees') or []
    total = sum(float((f or {}).get('total') or 0) for f in fees)
    paid = sum(float((f or {}).get('paid') or 0) for f in fees)
    balance = sum(float((f or {}).get('balance') or 0) for f in fees)

    lines = [
        center(receipt.get('schoolName') or 'School'),
        center('FEE PAYMENT RECEIPT'),
        '-' * width,
        f"Receipt : {esc(receipt.get('receiptNo'))}",
        f"Paid On : {esc(receipt.get('paidDate') or '-')}",
        f"Printed : {esc(receipt.get('printedAt') or '-')}",
        f"Student : {esc(receipt.get('studentName'))}",
        f"ID      : {esc(receipt.get('studentId'))}",
        f"Class   : {esc(receipt.get('classSection'))}",
        '-' * width,
    ]

    for idx, fee in enumerate(fees, start=1):
        fee_type = esc((fee or {}).get('feeType'))
        fee_note = esc((fee or {}).get('notes'))
        fee_label = f"{fee_type} ({fee_note})" if fee_note else fee_type
        lines.append(f"{idx}. {fee_label}"[:width])
        lines.append(f"   Total:{money((fee or {}).get('total'))}  Paid:{money((fee or {}).get('paid'))}  Bal:{money((fee or {}).get('balance'))}")

    lines.extend([
        '-' * width,
        f"TOTAL   : {money(total)}",
        f"PAID    : {money(paid)}",
        f"BALANCE : {money(balance)}",
        f"Mode    : {esc(receipt.get('paymentMode') or 'Cash')}",
    ])

    notes = esc(receipt.get('notes'))
    if notes:
        lines.append(f"Notes   : {notes}"[:width])

    lines.extend([
        '-' * width,
        center('Thank you!'),
        '\n\n'
    ])
    return '\n'.join(lines)


def _send_raw_to_windows_printer(printer_name, payload_text, cut_paper=False):
    """Send plain text directly to a Windows printer using RAW mode."""
    import win32print  # type: ignore

    data = payload_text.encode('utf-8', errors='replace')
    if cut_paper:
        data += b'\n\n\x1dV\x00'

    handle = win32print.OpenPrinter(printer_name)
    try:
        win32print.StartDocPrinter(handle, 1, ('Fee Receipt', None, 'RAW'))
        try:
            win32print.StartPagePrinter(handle)
            win32print.WritePrinter(handle, data)
            win32print.EndPagePrinter(handle)
        finally:
            win32print.EndDocPrinter(handle)
    finally:
        win32print.ClosePrinter(handle)


@app.route('/api/fees/printer-status', methods=['GET'])
def api_fees_printer_status():
    """Return whether a ready thermal printer is connected – used by frontend to decide print path."""
    # No strict auth needed – this is non-sensitive hardware metadata.
    # Any page that shows the print button is already behind a login wall.
    if not ('user_id' in session or 'student_id' in session):
        return jsonify({'success': True, 'thermal_ready': False, 'thermal_name': None})
    detected = _detect_windows_printers()
    thermal = detected.get('thermal')
    return jsonify({
        'success': True,
        'thermal_ready': bool(detected.get('thermal_ready')),
        'thermal_name': thermal or None,
    })



@app.route('/api/fees/auto-print', methods=['POST'])
@csrf.exempt
def api_fees_auto_print():
    """Thermal print endpoint: print only when a connected USB thermal printer is ready."""
    if not _fee_print_user_authorized():
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    data = request.get_json() or {}
    receipt = data.get('receipt') or {}
    if not isinstance(receipt, dict) or not receipt.get('fees'):
        return jsonify({'success': False, 'error': 'Invalid receipt payload'}), 400

    detected = _detect_windows_printers()
    if not detected.get('supported'):
        return jsonify({
            'success': False,
            'error': detected.get('reason') or 'Thermal print not available'
        })

    thermal_printer = detected.get('thermal')
    if not thermal_printer:
        return jsonify({
            'success': False,
            'error': 'Thermal printer is not connected or not ready'
        })

    try:
        txt = _build_fee_receipt_text(receipt)
        _send_raw_to_windows_printer(thermal_printer, txt, cut_paper=True)
        return jsonify({
            'success': True,
            'printed': True,
            'printer_type': 'thermal',
            'printer_name': thermal_printer
        })
    except Exception as exc:
        logger.error(f"Auto thermal print failed: {exc}")
        return jsonify({
            'success': False,
            'error': 'Thermal print failed'
        })

@app.route('/api/fees/types', methods=['GET'])
def api_fee_types_get():
    """Get all fee types for the school"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    db = get_db()
    school_id = session['school_id']
    rows = db.execute(
        'SELECT id, name, description FROM fee_types WHERE school_id = ? ORDER BY name',
        (school_id,)
    ).fetchall()
    return jsonify({'success': True, 'fee_types': [dict(r) for r in rows]})


@app.route('/api/fees/types', methods=['POST'])
@csrf.exempt
def api_fee_types_create():
    """Create a new fee type"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Fee type name is required'}), 400
    db = get_db()
    school_id = session['school_id']
    try:
        db.execute(
            'INSERT INTO fee_types (school_id, name, description) VALUES (?, ?, ?)',
            (school_id, name, description or None)
        )
        db.commit()
        row = db.execute(
            'SELECT id, name, description FROM fee_types WHERE school_id = ? AND name = ?',
            (school_id, name)
        ).fetchone()
        return jsonify({'success': True, 'fee_type': dict(row)})
    except Exception as e:
        if 'UNIQUE' in str(e) or 'Duplicate' in str(e):
            return jsonify({'success': False, 'error': 'A fee type with this name already exists'}), 409
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fees/types/<int:ft_id>', methods=['DELETE'])
@csrf.exempt
def api_fee_types_delete(ft_id):
    """Delete a fee type"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    db = get_db()
    school_id = session['school_id']
    existing = db.execute(
        'SELECT id FROM fee_types WHERE id = ? AND school_id = ?', (ft_id, school_id)
    ).fetchone()
    if not existing:
        return jsonify({'success': False, 'error': 'Fee type not found'}), 404
    db.execute('DELETE FROM student_fees WHERE fee_type_id = ? AND school_id = ?', (ft_id, school_id))
    db.execute('DELETE FROM fee_types WHERE id = ? AND school_id = ?', (ft_id, school_id))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/fees/students')
def api_fees_students():
    """Get students filtered by class/section for fee assignment"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    db = get_db()
    school_id = session['school_id']
    cls = request.args.get('class', '').strip()
    section = request.args.get('section', '').strip()
    query = '''
        SELECT id, student_id, full_name, `class`, section, roll_number
        FROM students WHERE school_id = ?
    '''
    params = [school_id]
    if cls:
        query += ' AND `class` = ?'
        params.append(cls)
    if section:
        query += ' AND section = ?'
        params.append(section)
    query += ' ORDER BY `class`, section, roll_number, full_name'
    rows = db.execute(query, params).fetchall()
    return jsonify({'success': True, 'students': [dict(r) for r in rows]})


@app.route('/api/fees/assign', methods=['POST'])
@csrf.exempt
def api_fees_assign():
    """Assign a fee to one or more students"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    student_ids = data.get('student_ids', [])  # list of students.id (int)
    fee_type_id = data.get('fee_type_id')
    amount = data.get('amount')
    due_date = data.get('due_date') or None
    notes = (data.get('notes') or '').strip() or None
    if not student_ids or not fee_type_id or amount is None:
        return jsonify({'success': False, 'error': 'student_ids, fee_type_id and amount are required'}), 400
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid amount'}), 400
    db = get_db()
    school_id = session['school_id']
    admin_id = session['user_id']
    ft = db.execute(
        'SELECT id FROM fee_types WHERE id = ? AND school_id = ?', (fee_type_id, school_id)
    ).fetchone()
    if not ft:
        return jsonify({'success': False, 'error': 'Fee type not found'}), 404
    created = 0
    for sid in student_ids:
        st = db.execute(
            'SELECT id FROM students WHERE id = ? AND school_id = ?', (sid, school_id)
        ).fetchone()
        if not st:
            continue
        db.execute('''
            INSERT INTO student_fees (school_id, student_db_id, fee_type_id, amount, due_date, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (school_id, sid, fee_type_id, amount, due_date, notes, admin_id))
        created += 1
    db.commit()
    return jsonify({'success': True, 'created': created})


@app.route('/api/fees/pay/<int:fee_id>', methods=['POST'])
@csrf.exempt
def api_fees_pay(fee_id):
    """Mark a fee as paid"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    payment_mode = (data.get('payment_mode') or 'Cash').strip()
    db = get_db()
    school_id = session['school_id']
    fee = db.execute(
        'SELECT id, student_db_id, fee_type_id, amount, paid_amount FROM student_fees WHERE id = ? AND school_id = ?',
        (fee_id, school_id)
    ).fetchone()
    if not fee:
        return jsonify({'success': False, 'error': 'Fee record not found'}), 404
    paid_ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    paid_delta = max(0.0, float(fee['amount'] or 0) - float(fee['paid_amount'] or 0))
    db.execute(
        "UPDATE student_fees SET paid_amount=amount, status='paid', paid_date=?, payment_mode=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (paid_ts, payment_mode, fee_id)
    )
    if paid_delta > 0:
        db.execute('''
            INSERT INTO fee_payment_history
            (school_id, student_db_id, student_fee_id, fee_type_id, paid_amount, payment_mode, notes, paid_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (school_id, fee['student_db_id'], fee_id, fee['fee_type_id'], paid_delta, payment_mode,
              'Full payment recorded', session.get('user_id')))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/fees/<int:fee_id>', methods=['PUT'])
@csrf.exempt
def api_fees_update(fee_id):
    """Update a fee record (amount, due_date, fee_type_id, notes, status)"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    db = get_db()
    school_id = session['school_id']
    fee = db.execute(
        'SELECT id FROM student_fees WHERE id = ? AND school_id = ?', (fee_id, school_id)
    ).fetchone()
    if not fee:
        return jsonify({'success': False, 'error': 'Fee record not found'}), 404
    updates = []
    params = []
    if 'amount' in data:
        try:
            updates.append('amount=?'); params.append(float(data['amount']))
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Invalid amount'}), 400
    if 'due_date' in data:
        updates.append('due_date=?'); params.append(data['due_date'] or None)
    if 'fee_type_id' in data and data['fee_type_id']:
        updates.append('fee_type_id=?'); params.append(int(data['fee_type_id']))
    if 'notes' in data:
        updates.append('notes=?'); params.append((data['notes'] or '').strip() or None)
    if 'status' in data and data['status'] in ('pending', 'paid', 'overdue', 'waived', 'partial'):
        updates.append('status=?'); params.append(data['status'])
    if not updates:
        return jsonify({'success': False, 'error': 'Nothing to update'}), 400
    updates.append('updated_at=CURRENT_TIMESTAMP')
    params.append(fee_id)
    db.execute(f"UPDATE student_fees SET {', '.join(updates)} WHERE id=?", params)
    db.commit()
    return jsonify({'success': True})


@app.route('/api/fees/pay-partial/<int:fee_id>', methods=['POST'])
@csrf.exempt
def api_fees_pay_partial(fee_id):
    """Record a partial (or full) payment against a fee record"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    data = request.get_json() or {}
    db = get_db()
    school_id = session['school_id']
    fee = db.execute(
        'SELECT id, student_db_id, fee_type_id, amount, paid_amount FROM student_fees WHERE id = ? AND school_id = ?',
        (fee_id, school_id)
    ).fetchone()
    if not fee:
        return jsonify({'success': False, 'error': 'Fee record not found'}), 404
    try:
        paid_now = float(data.get('paid_now', 0))
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid amount'}), 400
    if paid_now <= 0:
        return jsonify({'success': False, 'error': 'Amount must be greater than 0'}), 400
    total_amount = float(fee['amount'])
    already_paid = float(fee['paid_amount'] or 0)
    new_paid = min(already_paid + paid_now, total_amount)
    payment_mode = (data.get('payment_mode') or 'Cash').strip()
    paid_ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_status = 'paid' if new_paid >= total_amount else 'partial'
    db.execute(
        "UPDATE student_fees SET paid_amount=?, status=?, paid_date=?, payment_mode=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (new_paid, new_status, paid_ts, payment_mode, fee_id)
    )
    credited = max(0.0, new_paid - already_paid)
    if credited > 0:
        db.execute('''
            INSERT INTO fee_payment_history
            (school_id, student_db_id, student_fee_id, fee_type_id, paid_amount, payment_mode, notes, paid_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (school_id, fee['student_db_id'], fee_id, fee['fee_type_id'], credited, payment_mode,
              'Partial payment recorded', session.get('user_id')))
    db.commit()
    return jsonify({'success': True, 'new_paid': new_paid, 'status': new_status, 'total': total_amount})


@app.route('/api/fees/pay-selected', methods=['POST'])
@csrf.exempt
def api_fees_pay_selected():
    """Record payments for multiple selected fee rows of a student"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    data = request.get_json() or {}
    school_id = session['school_id']
    payment_mode = (data.get('payment_mode') or 'Cash').strip()

    try:
        student_db_id = int(data.get('student_db_id'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid student'}), 400

    payments = data.get('payments') or []
    if not isinstance(payments, list) or not payments:
        return jsonify({'success': False, 'error': 'No fee items selected'}), 400

    parsed = []
    seen_fee_ids = set()
    for item in payments:
        try:
            fee_id = int(item.get('fee_id'))
            amount = float(item.get('amount'))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'Invalid selected fee amount'}), 400
        if fee_id in seen_fee_ids:
            return jsonify({'success': False, 'error': 'Duplicate fee item selected'}), 400
        if amount <= 0:
            return jsonify({'success': False, 'error': 'Amount must be greater than 0'}), 400
        seen_fee_ids.add(fee_id)
        parsed.append({'fee_id': fee_id, 'amount': amount})

    db = get_db()
    placeholders = ','.join(['?'] * len(parsed))
    rows = db.execute(f'''
        SELECT id, student_db_id, fee_type_id, amount, paid_amount, status
        FROM student_fees
        WHERE school_id = ? AND id IN ({placeholders})
    ''', [school_id] + [p['fee_id'] for p in parsed]).fetchall()

    if len(rows) != len(parsed):
        return jsonify({'success': False, 'error': 'One or more selected fee records were not found'}), 404

    row_by_id = {r['id']: r for r in rows}
    paid_ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    total_deducted = 0.0
    updated_count = 0

    try:
        for item in parsed:
            fee = row_by_id[item['fee_id']]
            if fee['student_db_id'] != student_db_id:
                raise ValueError('Selected fee does not belong to the chosen student')

            total_amount = float(fee['amount'] or 0)
            already_paid = float(fee['paid_amount'] or 0)
            remaining = max(0.0, total_amount - already_paid)
            pay_now = float(item['amount'])

            if remaining <= 0 or fee['status'] in ('paid', 'waived'):
                raise ValueError('One or more selected fees are not payable')
            if pay_now > remaining:
                raise ValueError('Entered amount cannot exceed balance for selected fee')

            new_paid = already_paid + pay_now
            new_status = 'paid' if new_paid >= total_amount else 'partial'

            db.execute(
                "UPDATE student_fees SET paid_amount=?, status=?, paid_date=?, payment_mode=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (new_paid, new_status, paid_ts, payment_mode, fee['id'])
            )
            db.execute('''
                INSERT INTO fee_payment_history
                (school_id, student_db_id, student_fee_id, fee_type_id, paid_amount, payment_mode, notes, paid_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (school_id, student_db_id, fee['id'], fee['fee_type_id'], pay_now, payment_mode,
                  'Batch selected payment', session.get('user_id')))

            total_deducted += pay_now
            updated_count += 1

        db.commit()
    except ValueError as exc:
        db.rollback()
        return jsonify({'success': False, 'error': str(exc)}), 400
    except Exception:
        db.rollback()
        return jsonify({'success': False, 'error': 'Failed to process selected payments'}), 500

    return jsonify({
        'success': True,
        'updated_count': updated_count,
        'total_deducted': total_deducted
    })


@app.route('/api/fees/history/student/<int:student_db_id>')
def api_fees_history_for_student(student_db_id):
    """Get latest fee payment history for a student"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    db = get_db()
    school_id = session['school_id']
    rows = db.execute('''
        SELECT fph.id, fph.student_fee_id, fph.paid_amount, fph.payment_mode, fph.notes,
               fph.paid_at, ft.name AS fee_type_name
        FROM fee_payment_history fph
        LEFT JOIN fee_types ft ON ft.id = fph.fee_type_id
        WHERE fph.school_id = ? AND fph.student_db_id = ?
        ORDER BY fph.paid_at DESC, fph.id DESC
        LIMIT 50
    ''', (school_id, student_db_id)).fetchall()
    return jsonify({'success': True, 'history': [dict(r) for r in rows]})


@app.route('/api/fees/<int:fee_id>', methods=['DELETE'])
@csrf.exempt
def api_fees_delete(fee_id):
    """Delete a fee record"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    db = get_db()
    school_id = session['school_id']
    fee = db.execute(
        'SELECT id FROM student_fees WHERE id = ? AND school_id = ?', (fee_id, school_id)
    ).fetchone()
    if not fee:
        return jsonify({'success': False, 'error': 'Fee record not found'}), 404
    db.execute('DELETE FROM student_fees WHERE id = ?', (fee_id,))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/fees/student-list')
def api_fees_student_list():
    """Get all fee records for students (admin view, filterable by class/section)"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    db = get_db()
    school_id = session['school_id']
    cls = request.args.get('class', '').strip()
    section = request.args.get('section', '').strip()
    query = '''
        SELECT sf.id, sf.student_db_id, s.student_id, s.full_name, s.`class`, s.section,
               s.roll_number, ft.name as fee_type_name, ft.id as fee_type_id,
               sf.amount, sf.paid_amount, sf.due_date, sf.status, sf.paid_date, sf.payment_mode, sf.notes,
               (
                   SELECT fph.paid_at
                   FROM fee_payment_history fph
                   WHERE fph.school_id = sf.school_id AND fph.student_fee_id = sf.id
                   ORDER BY fph.paid_at DESC, fph.id DESC
                   LIMIT 1
               ) AS paid_at
        FROM student_fees sf
        JOIN students s ON sf.student_db_id = s.id
        JOIN fee_types ft ON sf.fee_type_id = ft.id
        WHERE sf.school_id = ?
    '''
    params = [school_id]
    if cls:
        query += ' AND s.`class` = ?'
        params.append(cls)
    if section:
        query += ' AND s.section = ?'
        params.append(section)
    query += ' ORDER BY s.`class`, s.section, s.roll_number, ft.name'
    rows = db.execute(query, params).fetchall()
    return jsonify({'success': True, 'fees': [dict(r) for r in rows]})


@app.route('/api/fees/student/<int:student_db_id>')
def api_fees_for_student(student_db_id):
    """Get all fee records for a specific student (used by student panel)"""
    # Allow either admin or the student themselves
    is_admin = 'user_id' in session and (session.get('user_type') == 'admin' or session.get('is_sub_admin'))
    is_student = 'student_id' in session and session.get('user_type') == 'student'
    if not is_admin and not is_student:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    db = get_db()
    # Validate student belongs to the current school
    school_id = session.get('school_id')
    student = db.execute('SELECT id, school_id FROM students WHERE id = ?', (student_db_id,)).fetchone()
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404
    if school_id and student['school_id'] != school_id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    rows = db.execute('''
        SELECT sf.id, ft.name as fee_type_name, sf.amount, sf.due_date,
               sf.status, sf.paid_date, sf.payment_mode, sf.notes, sf.created_at
        FROM student_fees sf
        JOIN fee_types ft ON sf.fee_type_id = ft.id
        WHERE sf.student_db_id = ?
        ORDER BY sf.created_at DESC
    ''', (student_db_id,)).fetchall()
    fees = [dict(r) for r in rows]
    total = sum(f['amount'] for f in fees)
    paid = sum(f['amount'] for f in fees if f['status'] == 'paid')
    pending = sum(f['amount'] for f in fees if f['status'] == 'pending')
    overdue = sum(f['amount'] for f in fees if f['status'] == 'overdue')
    return jsonify({'success': True, 'fees': fees,
                    'summary': {'total': total, 'paid': paid, 'pending': pending, 'overdue': overdue}})

# ========================= END FEE MANAGEMENT APIs =========================


@app.route('/api/get-student/<int:student_id>')
def api_get_student(student_id):
    """API endpoint to get student data for editing"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    db = get_db()
    school_id = session['school_id']
    
    student = db.execute('''
        SELECT * FROM students
        WHERE id = ? AND school_id = ?
    ''', (student_id, school_id)).fetchone()
    
    if not student:
        return jsonify({'success': False, 'error': 'Student not found'}), 404
    
    return jsonify({
        'success': True,
        'student': dict(student)
    })


@app.route('/staff/timetable')
def staff_timetable():
    """Staff Timetable Self-Service page"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))
    return render_template('staff_timetable.html')

@app.route('/staff/homework-assigner', methods=['GET', 'POST'])
def staff_homework_assigner():
    """Staff homework assigner page."""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    db = get_db()
    staff_id = session.get('user_id')
    school_id = session.get('school_id')
    today_date = datetime.date.today()
    day_of_week = today_date.weekday() + 1  # Mon=1 ... Sun=7
    if day_of_week == 7:
        day_of_week = 0  # Keep app convention: Sunday=0

    db.execute('''
        CREATE TABLE IF NOT EXISTS staff_homework_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            school_id INTEGER NOT NULL,
            staff_id INTEGER NOT NULL,
            level_id INTEGER,
            section_id INTEGER,
            class_name TEXT NOT NULL,
            section_name TEXT NOT NULL,
            subject_name TEXT NOT NULL,
            period_number INTEGER,
            homework_text TEXT NOT NULL,
            due_date TEXT,
            assigned_date TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    db.execute('''
        CREATE TABLE IF NOT EXISTS student_homework_completion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            homework_id INTEGER NOT NULL,
            student_id INTEGER NOT NULL,
            school_id INTEGER NOT NULL,
            staff_id INTEGER NOT NULL,
            completed_at TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(homework_id, student_id)
        )
    ''')

    # Backward-compatible schema updates for existing databases.
    existing_columns = set()
    try:
        column_rows = db.execute("PRAGMA table_info(staff_homework_assignments)").fetchall()
        existing_columns = {row['name'] for row in column_rows}
    except Exception:
        existing_columns = set()

    if 'attachment_path' not in existing_columns:
        try:
            db.execute('ALTER TABLE staff_homework_assignments ADD COLUMN attachment_path TEXT')
        except Exception:
            pass

    if 'updated_at' not in existing_columns:
        try:
            db.execute('ALTER TABLE staff_homework_assignments ADD COLUMN updated_at TIMESTAMP')
        except Exception:
            pass

    def _delete_attachment_file(path_value):
        if not path_value:
            return
        try:
            full_path = os.path.join(os.getcwd(), path_value)
            if os.path.isfile(full_path):
                os.remove(full_path)
        except Exception:
            # Attachment cleanup failure should not block the main operation.
            pass

    def _save_attachment(file_obj):
        if not file_obj or not getattr(file_obj, 'filename', ''):
            return None

        allowed_ext = {'pdf', 'doc', 'docx', 'ppt', 'pptx', 'xls', 'xlsx', 'txt', 'jpg', 'jpeg', 'png', 'zip'}
        safe_original = secure_filename(file_obj.filename)
        if not safe_original:
            raise ValueError('Invalid file name.')

        ext = safe_original.rsplit('.', 1)[-1].lower() if '.' in safe_original else ''
        if ext not in allowed_ext:
            raise ValueError('Unsupported file type. Use PDF, DOCX, PPT, XLSX, TXT, image, or ZIP.')

        upload_dir = os.path.join('static', 'uploads', 'homework', f'school_{school_id}')
        os.makedirs(upload_dir, exist_ok=True)

        stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_name = f"hw_{staff_id}_{stamp}_{int(time.time() * 1000) % 100000}_{safe_original}"
        relative_path = os.path.join(upload_dir, unique_name).replace('\\', '/')

        file_obj.save(relative_path)
        return relative_path

    if request.method == 'POST':
        action = (request.form.get('action') or 'assign_homework').strip()

        if action == 'mark_homework_completed':
            homework_id_raw = (request.form.get('homework_id') or '').strip()
            selected_student_ids = request.form.getlist('completed_student_ids')

            try:
                homework_id = int(homework_id_raw)
            except (TypeError, ValueError):
                flash('Invalid homework record selected for completion.', 'warning')
                return redirect(url_for('staff_homework_assigner'))

            assignment = db.execute('''
                SELECT id, class_name, section_name
                FROM staff_homework_assignments
                WHERE id = ? AND school_id = ? AND staff_id = ?
            ''', (homework_id, school_id, staff_id)).fetchone()

            if not assignment:
                flash('Homework record not found or access denied.', 'warning')
                return redirect(url_for('staff_homework_assigner'))

            valid_students = db.execute('''
                SELECT id
                FROM students
                WHERE school_id = ?
                  AND LOWER(TRIM(`class`)) = LOWER(TRIM(?))
                  AND LOWER(TRIM(section)) = LOWER(TRIM(?))
            ''', (school_id, assignment['class_name'], assignment['section_name'])).fetchall()
            valid_student_ids = {str(row['id']) for row in valid_students}

            clean_selected_ids = []
            for student_id_value in selected_student_ids:
                student_id_value = str(student_id_value).strip()
                if student_id_value in valid_student_ids and student_id_value not in clean_selected_ids:
                    clean_selected_ids.append(student_id_value)

            if not clean_selected_ids:
                flash('Select at least one student to mark homework as completed.', 'warning')
                return redirect(url_for('staff_homework_assigner'))

            try:
                completion_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for student_id_value in clean_selected_ids:
                    student_id_int = int(student_id_value)

                    existing_completion = db.execute('''
                        SELECT id
                        FROM student_homework_completion
                        WHERE homework_id = ? AND student_id = ?
                    ''', (homework_id, student_id_int)).fetchone()

                    if existing_completion:
                        db.execute('''
                            UPDATE student_homework_completion
                            SET school_id = ?,
                                staff_id = ?,
                                completed_at = ?
                            WHERE id = ?
                        ''', (school_id, staff_id, completion_date, existing_completion['id']))
                    else:
                        db.execute('''
                            INSERT INTO student_homework_completion (
                                homework_id, student_id, school_id, staff_id, completed_at
                            ) VALUES (?, ?, ?, ?, ?)
                        ''', (homework_id, student_id_int, school_id, staff_id, completion_date))

                db.commit()
                flash(f'Marked homework as completed for {len(clean_selected_ids)} student(s).', 'success')
            except Exception as exc:
                db.rollback()
                flash(f'Unable to save completion status: {str(exc)}', 'danger')

            return redirect(url_for('staff_homework_assigner'))

        if action == 'delete_homework':
            homework_id_raw = (request.form.get('homework_id') or '').strip()
            try:
                homework_id = int(homework_id_raw)
            except (TypeError, ValueError):
                flash('Invalid homework record selected.', 'warning')
                return redirect(url_for('staff_homework_assigner'))

            record = db.execute('''
                SELECT id, attachment_path
                FROM staff_homework_assignments
                WHERE id = ? AND school_id = ? AND staff_id = ?
            ''', (homework_id, school_id, staff_id)).fetchone()

            if not record:
                flash('Homework record not found or access denied.', 'warning')
                return redirect(url_for('staff_homework_assigner'))

            try:
                db.execute('''
                    DELETE FROM staff_homework_assignments
                    WHERE id = ? AND school_id = ? AND staff_id = ?
                ''', (homework_id, school_id, staff_id))
                db.commit()
                _delete_attachment_file(record['attachment_path'])
                flash('Homework deleted successfully.', 'success')
            except Exception as exc:
                db.rollback()
                flash(f'Unable to delete homework: {str(exc)}', 'danger')

            return redirect(url_for('staff_homework_assigner'))

        homework_id_raw = (request.form.get('homework_id') or '').strip()
        is_edit = action == 'update_homework' and homework_id_raw

        class_name = (request.form.get('class_name') or '').strip()
        section_name = (request.form.get('section_name') or '').strip()
        subject_name = (request.form.get('subject_name') or '').strip()
        homework_text = (request.form.get('homework_text') or '').strip()
        due_date = (request.form.get('due_date') or '').strip()
        remove_attachment = (request.form.get('remove_attachment') or '').strip() == '1'
        homework_file = request.files.get('homework_file')

        level_id_raw = (request.form.get('level_id') or '').strip()
        section_id_raw = (request.form.get('section_id') or '').strip()
        period_number_raw = (request.form.get('period_number') or '').strip()

        try:
            level_id = int(level_id_raw) if level_id_raw else None
            section_id = int(section_id_raw) if section_id_raw else None
            period_number = int(period_number_raw) if period_number_raw else None
        except (TypeError, ValueError):
            flash('Invalid period or class details. Please choose a period card again.', 'danger')
            return redirect(url_for('staff_homework_assigner'))

        if not class_name or not section_name or not subject_name or not homework_text:
            flash('Please select a class period and enter homework details.', 'warning')
            return redirect(url_for('staff_homework_assigner'))

        if due_date:
            try:
                datetime.date.fromisoformat(due_date)
            except ValueError:
                flash('Invalid due date format.', 'danger')
                return redirect(url_for('staff_homework_assigner'))

        edit_record = None
        homework_id = None
        if is_edit:
            try:
                homework_id = int(homework_id_raw)
            except (TypeError, ValueError):
                flash('Invalid homework ID for edit.', 'danger')
                return redirect(url_for('staff_homework_assigner'))

            edit_record = db.execute('''
                SELECT id, attachment_path
                FROM staff_homework_assignments
                WHERE id = ? AND school_id = ? AND staff_id = ?
            ''', (homework_id, school_id, staff_id)).fetchone()

            if not edit_record:
                flash('Homework record not found for edit.', 'warning')
                return redirect(url_for('staff_homework_assigner'))

        try:
            new_attachment_path = _save_attachment(homework_file)
        except ValueError as ve:
            flash(str(ve), 'warning')
            return redirect(url_for('staff_homework_assigner'))
        except Exception as exc:
            flash(f'Unable to upload file: {str(exc)}', 'danger')
            return redirect(url_for('staff_homework_assigner'))

        try:
            if is_edit:
                final_attachment_path = edit_record['attachment_path']
                if remove_attachment:
                    final_attachment_path = None
                if new_attachment_path:
                    final_attachment_path = new_attachment_path

                db.execute('''
                    UPDATE staff_homework_assignments
                    SET level_id = ?,
                        section_id = ?,
                        class_name = ?,
                        section_name = ?,
                        subject_name = ?,
                        period_number = ?,
                        homework_text = ?,
                        due_date = ?,
                        attachment_path = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                      AND school_id = ?
                      AND staff_id = ?
                ''', (
                    level_id,
                    section_id,
                    class_name,
                    section_name,
                    subject_name,
                    period_number,
                    homework_text,
                    due_date if due_date else None,
                    final_attachment_path,
                    homework_id,
                    school_id,
                    staff_id
                ))
                db.commit()

                if (remove_attachment or new_attachment_path) and edit_record['attachment_path']:
                    _delete_attachment_file(edit_record['attachment_path'])

                flash('Homework updated successfully.', 'success')
            else:
                db.execute('''
                    INSERT INTO staff_homework_assignments (
                        school_id, staff_id, level_id, section_id,
                        class_name, section_name, subject_name, period_number,
                        homework_text, due_date, assigned_date, attachment_path
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    school_id,
                    staff_id,
                    level_id,
                    section_id,
                    class_name,
                    section_name,
                    subject_name,
                    period_number,
                    homework_text,
                    due_date if due_date else None,
                    today_date.strftime('%Y-%m-%d'),
                    new_attachment_path
                ))
                db.commit()
                flash(f'Homework assigned for {class_name}-{section_name} ({subject_name}) successfully.', 'success')
        except Exception as exc:
            db.rollback()
            if new_attachment_path:
                _delete_attachment_file(new_attachment_path)
            flash(f'Unable to assign homework right now: {str(exc)}', 'danger')

        return redirect(url_for('staff_homework_assigner'))

    day_names = {
        0: 'Sunday',
        1: 'Monday',
        2: 'Tuesday',
        3: 'Wednesday',
        4: 'Thursday',
        5: 'Friday',
        6: 'Saturday'
    }

    today_periods = []
    if day_of_week != 0:
        today_periods = db.execute('''
            SELECT
                ha.level_id,
                ha.section_id,
                tal.level_name AS class_name,
                ts.section_name,
                ha.subject_name,
                ha.period_number,
                COALESCE(
                    (
                        SELECT tp.period_name
                        FROM timetable_periods tp
                        WHERE tp.school_id = ha.school_id
                          AND tp.period_number = ha.period_number
                          AND tp.is_active = 1
                          AND (tp.level_id IS NULL OR tp.level_id = ha.level_id)
                          AND (tp.section_id IS NULL OR tp.section_id = ha.section_id)
                        ORDER BY
                            CASE WHEN tp.level_id = ha.level_id THEN 0 ELSE 1 END,
                            CASE WHEN tp.section_id = ha.section_id THEN 0 ELSE 1 END
                        LIMIT 1
                    ),
                    'Period ' || ha.period_number
                ) AS period_name,
                (
                    SELECT tp.start_time
                    FROM timetable_periods tp
                    WHERE tp.school_id = ha.school_id
                      AND tp.period_number = ha.period_number
                      AND tp.is_active = 1
                      AND (tp.level_id IS NULL OR tp.level_id = ha.level_id)
                      AND (tp.section_id IS NULL OR tp.section_id = ha.section_id)
                    ORDER BY
                        CASE WHEN tp.level_id = ha.level_id THEN 0 ELSE 1 END,
                        CASE WHEN tp.section_id = ha.section_id THEN 0 ELSE 1 END
                    LIMIT 1
                ) AS start_time,
                (
                    SELECT tp.end_time
                    FROM timetable_periods tp
                    WHERE tp.school_id = ha.school_id
                      AND tp.period_number = ha.period_number
                      AND tp.is_active = 1
                      AND (tp.level_id IS NULL OR tp.level_id = ha.level_id)
                      AND (tp.section_id IS NULL OR tp.section_id = ha.section_id)
                    ORDER BY
                        CASE WHEN tp.level_id = ha.level_id THEN 0 ELSE 1 END,
                        CASE WHEN tp.section_id = ha.section_id THEN 0 ELSE 1 END
                    LIMIT 1
                ) AS end_time
            FROM timetable_hierarchical_assignments ha
            JOIN timetable_academic_levels tal
              ON tal.id = ha.level_id
             AND tal.school_id = ha.school_id
             AND tal.is_active = 1
            JOIN timetable_sections ts
              ON ts.id = ha.section_id
             AND ts.school_id = ha.school_id
             AND ts.is_active = 1
            WHERE ha.school_id = ?
              AND ha.staff_id = ?
              AND ha.day_of_week = ?
            ORDER BY ha.period_number, tal.level_name, ts.section_name
        ''', (school_id, staff_id, day_of_week)).fetchall()

    recent_assignments = db.execute('''
        SELECT hwa.id, hwa.class_name, hwa.section_name, hwa.subject_name, hwa.period_number,
               hwa.homework_text, hwa.due_date, hwa.assigned_date, hwa.created_at,
               hwa.attachment_path,
               COALESCE((
                   SELECT COUNT(*)
                   FROM student_homework_completion shc
                   WHERE shc.homework_id = hwa.id
               ), 0) AS completed_count
        FROM staff_homework_assignments hwa
        WHERE hwa.school_id = ?
          AND hwa.staff_id = ?
        ORDER BY hwa.assigned_date DESC, hwa.created_at DESC
        LIMIT 20
    ''', (school_id, staff_id)).fetchall()

    assignment_students_map = {}
    for assignment in recent_assignments:
        students = db.execute('''
            SELECT id, student_id, full_name, roll_number
            FROM students
            WHERE school_id = ?
              AND LOWER(TRIM(`class`)) = LOWER(TRIM(?))
              AND LOWER(TRIM(section)) = LOWER(TRIM(?))
            ORDER BY roll_number ASC, full_name ASC
        ''', (school_id, assignment['class_name'], assignment['section_name'])).fetchall()

        completed_students = db.execute('''
            SELECT student_id
            FROM student_homework_completion
            WHERE homework_id = ?
        ''', (assignment['id'],)).fetchall()
        completed_student_ids = {row['student_id'] for row in completed_students}

        assignment_students_map[str(assignment['id'])] = [
            {
                'id': row['id'],
                'student_id': row['student_id'],
                'full_name': row['full_name'],
                'roll_number': row['roll_number'],
                'completed': row['id'] in completed_student_ids
            }
            for row in students
        ]

    return render_template(
        'staff_homework_assigner.html',
        today_periods=today_periods,
        today_name=day_names.get(day_of_week, 'Today'),
        today_date=today_date.strftime('%Y-%m-%d'),
        recent_assignments=recent_assignments,
        assignment_students_map=assignment_students_map
    )

@app.route('/admin/staff-period-assignment')
def staff_period_assignment():
    """Admin Staff Period Assignment page"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))
    return render_template('staff_period_assignment.html')

@app.route('/staff/my-period-assignment')
def staff_view_period_assignment():
    """Staff view of their period assignments"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))
    return render_template('staff_period_assignment.html', view_mode='staff')

@app.route('/api/dashboard/attendance_stats')
def api_dashboard_attendance_stats():
    """Get real-time attendance statistics for dashboard auto-refresh"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        db = get_db()
        school_id = session['school_id']
        today = datetime.date.today()
        
        # Get all staff for attendance calculation
        all_staff = db.execute('''
            SELECT s.id as staff_id
            FROM staff s
            WHERE s.school_id = ?
        ''', (school_id,)).fetchall()
        
        # Calculate attendance summary
        status_counts = {
            'total_staff': 0,
            'present': 0,
            'absent': 0,
            'late': 0,
            'on_leave': 0,
            'on_duty': 0,
            'on_permission': 0,
            'holiday': 0
        }
        
        for staff in all_staff:
            status = get_staff_status_for_date(staff['staff_id'], today, school_id, db)
            status_counts['total_staff'] += 1
            
            if status == 'Holiday':
                status_counts['holiday'] += 1
            elif status == 'On Leave':
                status_counts['on_leave'] += 1
            elif status == 'On Duty':
                status_counts['on_duty'] += 1
            elif status == 'On Permission':
                status_counts['on_permission'] += 1
            elif status == 'present':
                status_counts['present'] += 1
            elif status == 'late':
                status_counts['late'] += 1
            else:
                status_counts['absent'] += 1
        
        return jsonify({
            'success': True,
            'present': status_counts['present'],
            'absent': status_counts['absent'],
            'late': status_counts['late'],
            'on_leave': status_counts['on_leave'],
            'on_duty': status_counts['on_duty'],
            'on_permission': status_counts['on_permission'],
            'holiday': status_counts['holiday'],
            'total_staff': status_counts['total_staff']
        })
        
    except Exception as e:
        print(f"Error getting attendance stats: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/export_dashboard_data')
def export_dashboard_data():
    """Comprehensive admin dashboard data export with multiple format support"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        school_id = session['school_id']
        export_format = request.args.get('format', 'excel').lower()
        export_type = request.args.get('type', 'all').lower()  # all, staff, attendance, applications

        # Use the ExcelReportGenerator for comprehensive reports
        excel_generator = ExcelReportGenerator()

        if export_type == 'staff':
            # Export only staff data using the existing method
            today = datetime.datetime.now().strftime('%Y-%m-%d')
            return excel_generator.create_staff_profile_report(school_id)
        elif export_type == 'attendance':
            # Export attendance data for current month
            today = datetime.date.today()
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            end_date = today.strftime('%Y-%m-%d')
            return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)
        elif export_type == 'applications':
            # Export applications data
            return export_applications_data(school_id, export_format)
        else:
            # Export comprehensive dashboard data
            return export_comprehensive_dashboard_data(school_id, export_format)

    except Exception as e:
        print(f"Dashboard export error: {str(e)}")
        return jsonify({'success': False, 'error': f'Export failed: {str(e)}'})

def export_applications_data(school_id, export_format='excel'):
    """Export applications data (leave, on-duty, permission)"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    db = get_db()

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 1. Leave Applications Sheet
    ws_leave = wb.create_sheet("Leave Applications")
    leave_headers = ['S.No', 'Staff Name', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Reason', 'Status', 'Applied Date']

    # Add headers
    for col, header in enumerate(leave_headers, 1):
        cell = ws_leave.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Get leave applications data
    leave_apps = db.execute('''
        SELECT l.*, s.full_name
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ?
        ORDER BY l.applied_at DESC
    ''', (school_id,)).fetchall()

    for row_idx, app in enumerate(leave_apps, 2):
        ws_leave.cell(row=row_idx, column=1, value=row_idx-1)
        ws_leave.cell(row=row_idx, column=2, value=app['full_name'])
        ws_leave.cell(row=row_idx, column=3, value=app['leave_type'])
        ws_leave.cell(row=row_idx, column=4, value=app['start_date'])
        ws_leave.cell(row=row_idx, column=5, value=app['end_date'])
        ws_leave.cell(row=row_idx, column=6, value=app['total_days'] if 'total_days' in app.keys() else 'N/A')
        ws_leave.cell(row=row_idx, column=7, value=app['reason'] or 'N/A')
        ws_leave.cell(row=row_idx, column=8, value=app['status'])
        ws_leave.cell(row=row_idx, column=9, value=app['applied_at'])

    # 2. On-Duty Applications Sheet
    ws_duty = wb.create_sheet("On-Duty Applications")
    duty_headers = ['S.No', 'Staff Name', 'Duty Type', 'Start Date', 'End Date', 'Start Time', 'End Time', 'Location', 'Purpose', 'Status']

    for col, header in enumerate(duty_headers, 1):
        cell = ws_duty.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    duty_apps = db.execute('''
        SELECT od.*, s.full_name
        FROM on_duty_applications od
        JOIN staff s ON od.staff_id = s.id
        WHERE od.school_id = ?
        ORDER BY od.applied_at DESC
    ''', (school_id,)).fetchall()

    for row_idx, app in enumerate(duty_apps, 2):
        ws_duty.cell(row=row_idx, column=1, value=row_idx-1)
        ws_duty.cell(row=row_idx, column=2, value=app['full_name'])
        ws_duty.cell(row=row_idx, column=3, value=app['duty_type'])
        ws_duty.cell(row=row_idx, column=4, value=app['start_date'])
        ws_duty.cell(row=row_idx, column=5, value=app['end_date'])
        ws_duty.cell(row=row_idx, column=6, value=app['start_time'])
        ws_duty.cell(row=row_idx, column=7, value=app['end_time'])
        ws_duty.cell(row=row_idx, column=8, value=app['location'])
        ws_duty.cell(row=row_idx, column=9, value=app['purpose'])
        ws_duty.cell(row=row_idx, column=10, value=app['status'])

    # 3. Permission Applications Sheet
    ws_perm = wb.create_sheet("Permission Applications")
    perm_headers = ['S.No', 'Staff Name', 'Permission Type', 'Date', 'Start Time', 'End Time', 'Duration (Hours)', 'Reason', 'Status']

    for col, header in enumerate(perm_headers, 1):
        cell = ws_perm.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    perm_apps = db.execute('''
        SELECT p.*, s.full_name
        FROM permission_applications p
        JOIN staff s ON p.staff_id = s.id
        WHERE p.school_id = ?
        ORDER BY p.applied_at DESC
    ''', (school_id,)).fetchall()

    for row_idx, app in enumerate(perm_apps, 2):
        ws_perm.cell(row=row_idx, column=1, value=row_idx-1)
        ws_perm.cell(row=row_idx, column=2, value=app['full_name'])
        ws_perm.cell(row=row_idx, column=3, value=app['permission_type'])
        ws_perm.cell(row=row_idx, column=4, value=app['permission_date'])
        ws_perm.cell(row=row_idx, column=5, value=app['start_time'])
        ws_perm.cell(row=row_idx, column=6, value=app['end_time'])
        ws_perm.cell(row=row_idx, column=7, value=app['duration_hours'])
        ws_perm.cell(row=row_idx, column=8, value=app['reason'])
        ws_perm.cell(row=row_idx, column=9, value=app['status'])

    # Auto-adjust column widths for all sheets
    for ws in [ws_leave, ws_duty, ws_perm]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'applications_report_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def export_comprehensive_dashboard_data(school_id, export_format='excel'):
    """Export comprehensive dashboard data including all sections"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    db = get_db()
    today = datetime.date.today()

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2F5597")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 1. Dashboard Summary Sheet
    ws_summary = wb.create_sheet("Dashboard Summary")

    # Title
    ws_summary.merge_cells('A1:F1')
    title_cell = ws_summary['A1']
    title_cell.value = f"Admin Dashboard Summary - {today.strftime('%Y-%m-%d')}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Attendance Summary
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave,
            SUM(CASE WHEN a.status = 'on_duty' THEN 1 ELSE 0 END) as on_duty
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, school_id)).fetchone()

    # Add attendance summary
    ws_summary['A3'] = "Today's Attendance Summary"
    ws_summary['A3'].font = Font(bold=True, size=14)

    summary_data = [
        ['Metric', 'Count'],
        ['Total Staff', attendance_summary['total_staff']],
        ['Present', attendance_summary['present']],
        ['Absent', attendance_summary['absent']],
        ['Late', attendance_summary['late']],
        ['On Leave', attendance_summary['on_leave']],
        ['On Duty', attendance_summary['on_duty']]
    ]

    for row_idx, row_data in enumerate(summary_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 4:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            cell.border = border

    # 2. Today's Attendance Sheet
    ws_attendance = wb.create_sheet("Today's Attendance")

    attendance_headers = ['S.No', 'Staff ID', 'Staff Name', 'Department', 'Status', 'Time In', 'Time Out']

    for col, header in enumerate(attendance_headers, 1):
        cell = ws_attendance.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    today_attendance = db.execute('''
        SELECT s.staff_id, s.full_name, s.department,
               COALESCE(a.status, 'absent') as status,
               a.time_in, a.time_out
        FROM staff s
        LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
        WHERE s.school_id = ?
        ORDER BY CAST(s.staff_id AS INTEGER) ASC
    ''', (today, school_id)).fetchall()

    for row_idx, record in enumerate(today_attendance, 2):
        ws_attendance.cell(row=row_idx, column=1, value=row_idx-1)
        ws_attendance.cell(row=row_idx, column=2, value=record['staff_id'] or 'N/A')
        ws_attendance.cell(row=row_idx, column=3, value=record['full_name'])
        ws_attendance.cell(row=row_idx, column=4, value=record['department'] or 'N/A')
        ws_attendance.cell(row=row_idx, column=5, value=record['status'])
        ws_attendance.cell(row=row_idx, column=6, value=record['time_in'] or 'N/A')
        ws_attendance.cell(row=row_idx, column=7, value=record['time_out'] or 'N/A')

    # Auto-adjust column widths
    for ws in [ws_summary, ws_attendance]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'dashboard_comprehensive_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response


@app.route('/toggle_modern_ui')
def toggle_modern_ui():
    """Toggle between modern and legacy UI"""
    if 'user_id' not in session:
        return redirect(url_for('index'))

    # Toggle the modern UI setting
    session['use_modern_ui'] = not session.get('use_modern_ui', False)

    # Redirect back to the referring page or dashboard
    return redirect(request.referrer or url_for('admin_dashboard'))


@app.route('/company/dashboard')
def company_dashboard():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return redirect(url_for('index'))

    db = get_db()

    # Get all schools with timetable status
    schools = db.execute('''
        SELECT s.id, s.name, s.address, s.contact_email, s.contact_phone, s.logo_path,
               COUNT(DISTINCT a.id) as admin_count,
               COUNT(DISTINCT st.id) as staff_count,
               COALESCE(ts.is_enabled, 0) as timetable_enabled
        FROM schools s
        LEFT JOIN admins a ON s.id = a.school_id
        LEFT JOIN staff st ON s.id = st.school_id
        LEFT JOIN timetable_settings ts ON s.id = ts.school_id
        GROUP BY s.id
        ORDER BY s.name
    ''').fetchall()

    return render_template('company_dashboard.html', schools=schools)

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    """
    LEGACY ROUTE - DEPRECATED
    This route is kept for backward compatibility but should not be used.
    Use /biometric_attendance instead for proper user-controlled verification.
    """
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Return error to prevent automatic updates
    return jsonify({
        'success': False,
        'error': 'This route is deprecated. Please use biometric verification system for attendance marking.'
    })

def validate_verification_rules(verification_type, existing_attendance, current_time, staff_shift_type='general'):
    """Validate business rules for biometric verification using shift management"""

    if verification_type == 'check-in':
        # Check if user has already checked in today
        if existing_attendance and existing_attendance['time_in']:
            return 'You have already checked in today. Multiple check-ins are not allowed.'
        return None

    elif verification_type == 'check-out':
        # Check if user has already checked out today
        if existing_attendance and existing_attendance['time_out']:
            return 'You have already checked out today. Multiple check-outs are not allowed.'

        # Check if user has checked in first
        if not existing_attendance or not existing_attendance['time_in']:
            return 'You must check in first before checking out.'
        return None

    return 'Invalid verification type'

@app.route('/biometric_attendance', methods=['POST'])
def biometric_attendance():
    """
    DEPRECATED: Handle biometric attendance verification with manual type selection

    This route is now deprecated. Staff should use the biometric device directly,
    and the system will automatically poll for new verifications.
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})

    # Return error message directing users to use the device directly
    return jsonify({
        'success': False,
        'error': 'Please use the biometric device directly to select your verification type and verify.'
    })

@app.route('/check_device_verification', methods=['POST'])
def check_device_verification():
    """Check for recent biometric verification from the device for a specific staff member"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})

    # Allow both staff and admin to use this
    if session['user_type'] not in ['staff', 'admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Handle both staff and admin users
    if session['user_type'] == 'staff':
        staff_id = session['user_id']
    else:
        # For admin, get staff_id from form
        staff_id = request.form.get('staff_id')
        if not staff_id:
            return jsonify({'success': False, 'error': 'Staff ID required for admin verification'})

    school_id = session['school_id']
    active_mode = get_school_attendance_mode(school_id)
    if active_mode != 'biometric':
        return jsonify({
            'success': False,
            'error': f'Device biometric verification is disabled for this school (current mode: {active_mode}).'
        })

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured for your institution'})

    current_datetime = datetime.datetime.now()
    db = get_db()

    # Get staff information
    staff_info = db.execute('''
        SELECT staff_id, full_name FROM staff WHERE id = ?
    ''', (staff_id,)).fetchone()

    if not staff_info:
        return jsonify({'success': False, 'error': 'Staff not found'})

    try:
        # Check for recent verification from the device (within last 30 seconds)
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({
                'success': False,
                'error': 'Failed to connect to biometric device'
            })

        # Look for recent attendance records for this staff member
        recent_cutoff = current_datetime - datetime.timedelta(seconds=30)
        recent_records = zk_device.get_new_attendance_records(recent_cutoff)

        staff_recent_record = None
        for record in recent_records:
            if str(record['user_id']) == str(staff_info['staff_id']):
                staff_recent_record = record
                break

        zk_device.disconnect()

        if not staff_recent_record:
            return jsonify({
                'success': False,
                'error': 'No recent biometric verification found. Please use the biometric device to verify your attendance.'
            })

        # Process the verification from the device
        verification_type = staff_recent_record['verification_type']
        verification_time = staff_recent_record['timestamp']

        # Validate business rules
        today = verification_time.date()
        existing_attendance = db.execute('''
            SELECT * FROM attendance WHERE staff_id = ? AND date = ?
        ''', (staff_id, today)).fetchone()

        # Get staff shift type for validation
        staff_info = db.execute('SELECT shift_type FROM staff WHERE id = ?', (staff_id,)).fetchone()
        staff_shift_type = staff_info['shift_type'] if staff_info and staff_info['shift_type'] else 'general'

        validation_error = validate_verification_rules(verification_type, existing_attendance, verification_time.time(), staff_shift_type)
        if validation_error:
            return jsonify({'success': False, 'error': validation_error})

        # Log successful verification
        db.execute('''
            INSERT INTO biometric_verifications
            (staff_id, school_id, verification_type, verification_time, device_ip, biometric_method, verification_status)
            VALUES (?, ?, ?, ?, ?, 'fingerprint', 'success')
        ''', (staff_id, school_id, verification_type, verification_time, device_ip))

        # Get staff shift type
        staff_info = db.execute('SELECT shift_type FROM staff WHERE id = ?', (staff_id,)).fetchone()
        staff_shift_type = staff_info['shift_type'] if staff_info and staff_info['shift_type'] else 'general'

        # Update attendance based on verification type using enhanced shift management
        current_time = verification_time.strftime('%H:%M:%S')
        shift_manager = ShiftManager()

        if verification_type == 'check-in':
            # Calculate attendance status using shift management
            attendance_result = shift_manager.calculate_attendance_status(
                staff_shift_type, verification_time.time()
            )

            # Only create new record if no check-in exists (validation already checked this)
            if existing_attendance:
                return jsonify({
                    'success': False,
                    'error': 'You have already checked in today.'
                })
            else:
                db.execute('''
                    INSERT INTO attendance (staff_id, school_id, date, time_in, status,
                                          late_duration_minutes, shift_type, shift_start_time, shift_end_time)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (staff_id, school_id, today, current_time, attendance_result['status'],
                      attendance_result['late_duration_minutes'], staff_shift_type,
                      attendance_result['shift_start_time'].strftime('%H:%M:%S'),
                      attendance_result['shift_end_time'].strftime('%H:%M:%S')))

        elif verification_type == 'check-out':
            # Only update check-out if no check-out exists (validation already checked this)
            if existing_attendance and existing_attendance['time_out']:
                return jsonify({
                    'success': False,
                    'error': 'You have already checked out today.'
                })
            elif existing_attendance:
                # Calculate if this is early departure
                attendance_result = shift_manager.calculate_attendance_status(
                    staff_shift_type,
                    datetime.datetime.strptime(existing_attendance['time_in'], '%H:%M:%S').time() if existing_attendance['time_in'] else verification_time.time(),
                    verification_time.time()
                )

                # Update status if early departure detected
                final_status = existing_attendance['status']
                if attendance_result['early_departure_minutes'] and attendance_result['early_departure_minutes'] > 0:
                    if final_status != 'late':  # Don't override late status
                        final_status = 'left_soon'

                db.execute('''
                    UPDATE attendance SET time_out = ?, status = ?, early_departure_minutes = ?
                    WHERE staff_id = ? AND date = ?
                ''', (current_time, final_status, attendance_result['early_departure_minutes'], staff_id, today))
            else:
                return jsonify({
                    'success': False,
                    'error': 'You must check in first before checking out.'
                })

        db.commit()

        # Format times for display in 12-hour format
        current_time_12hr = format_time_to_12hr(current_time)
        verification_time_12hr = verification_time.strftime('%Y-%m-%d %I:%M %p')

        return jsonify({
            'success': True,
            'message': f'{verification_type.title()} recorded successfully at {current_time_12hr}',
            'verification_time': verification_time_12hr,
            'verification_type': verification_type,
            'biometric_method': 'fingerprint',
            'time_recorded': current_time_12hr
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Verification error: {str(e)}'
        })

@app.route('/apply_leave', methods=['POST'])
def apply_leave():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    leave_type = request.form.get('leave_type')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    reason = request.form.get('reason')

    # Validate required fields
    if not all([leave_type, start_date, end_date, reason]):
        return jsonify({'success': False, 'error': 'Please fill all required fields'})

    # Calculate requested days
    try:
        start = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        requested_days = (end - start).days + 1
        
        if requested_days <= 0:
            return jsonify({'success': False, 'error': 'End date must be after or equal to start date'})
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'})

    db = get_db()
    current_year = datetime.datetime.now().year
    
    # Initialize quotas if not exists
    initialize_staff_quotas(staff_id, school_id, current_year)

    # Ensure validation uses the latest usage (including recent withdrawals).
    update_quota_usage(staff_id, school_id, current_year)
    
    # CRITICAL: Check quota availability (Server-Side Validation)
    quota = db.execute("""
        SELECT allocated_days, used_days, (allocated_days - used_days) as remaining_days
        FROM staff_leave_quotas 
        WHERE staff_id = ? AND quota_year = ? AND leave_type = ?
    """, (staff_id, current_year, leave_type)).fetchone()
    
    if not quota:
        return jsonify({
            'success': False, 
            'error': f'No quota found for {leave_type}. Please contact administrator.'
        })
    
    # Check A: Zero Balance
    if quota['remaining_days'] <= 0:
        return jsonify({
            'success': False,
            'error': f'Application Failed: You have used all your available {leave_type}.'
        })
    
    # Check B: Insufficient Balance
    if requested_days > quota['remaining_days']:
        return jsonify({
            'success': False,
            'error': f'Application Failed: Your request for {requested_days} {leave_type} days exceeds your remaining balance of {quota["remaining_days"]} day{"s" if quota["remaining_days"] != 1 else ""}.'
        })

    # If quota checks pass, submit the application
    try:
        # Insert with explicit pending status to ensure no auto-approval
        cursor = db.execute('''
            INSERT INTO leave_applications
            (staff_id, school_id, leave_type, start_date, end_date, reason, status)
            VALUES (?, ?, ?, ?, ?, ?, 'pending')
        ''', (staff_id, school_id, leave_type, start_date, end_date, reason))
        
        new_leave_id = cursor.lastrowid
        db.commit()

        try:
            update_quota_usage(staff_id, school_id, current_year)
        except Exception as quota_err:
            print(f"⚠️ Quota refresh after leave apply failed: {quota_err}")

        # Log the submission for audit trail
        print(f"📝 LEAVE SUBMITTED: Staff {staff_id} submitted {leave_type} leave for {start_date} to {end_date} (ID: {new_leave_id})")

        # Verify the status is actually pending (safety check)
        verification = db.execute('SELECT status FROM leave_applications WHERE id = ?', (new_leave_id,)).fetchone()
        if verification['status'] != 'pending':
            print(f"🚨 WARNING: Leave {new_leave_id} status is '{verification['status']}' instead of 'pending'!")

        return jsonify({
            'success': True, 
            'message': 'Leave application submitted successfully and is pending admin approval',
            'leave_id': new_leave_id,
            'status': 'pending'
        })
    except Exception as e:
        db.rollback()
        print(f"❌ LEAVE SUBMISSION ERROR: {str(e)}")
        return jsonify({'success': False, 'error': f'Failed to submit application: {str(e)}'})

@app.route('/apply_on_duty', methods=['POST'])
def apply_on_duty():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    duty_type = request.form.get('duty_type')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    location = request.form.get('location')
    purpose = request.form.get('purpose')
    reason = request.form.get('reason')

    # Validate required fields
    if not all([duty_type, start_date, end_date, purpose]):
        return jsonify({'success': False, 'error': 'Please fill all required fields'})

    # Calculate requested days
    try:
        start = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        requested_days = (end - start).days + 1
        
        if requested_days <= 0:
            return jsonify({'success': False, 'error': 'End date must be after or equal to start date'})
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'})

    db = get_db()
    current_year = datetime.datetime.now().year
    
    # Initialize quotas if not exists
    initialize_staff_quotas(staff_id, school_id, current_year)

    # Ensure validation uses the latest usage (including recent withdrawals).
    update_quota_usage(staff_id, school_id, current_year)
    
    # CRITICAL: Check quota availability (Server-Side Validation)
    quota = db.execute("""
        SELECT allocated_days, used_days, (allocated_days - used_days) as remaining_days
        FROM staff_od_quotas 
        WHERE staff_id = ? AND quota_year = ?
    """, (staff_id, current_year)).fetchone()
    
    if not quota:
        return jsonify({
            'success': False, 
            'error': 'No On Duty quota found. Please contact administrator.'
        })
    
    # Check A: Zero Balance
    if quota['remaining_days'] <= 0:
        return jsonify({
            'success': False,
            'error': 'Application Failed: You have used all your available On Duty days.'
        })
    
    # Check B: Insufficient Balance
    if requested_days > quota['remaining_days']:
        return jsonify({
            'success': False,
            'error': f'Application Failed: Your request for {requested_days} On Duty days exceeds your remaining balance of {quota["remaining_days"]} day{"s" if quota["remaining_days"] != 1 else ""}.'
        })

    # If quota checks pass, submit the application
    try:
        db.execute('''
            INSERT INTO on_duty_applications
            (staff_id, school_id, duty_type, start_date, end_date, start_time, end_time, location, purpose, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (staff_id, school_id, duty_type, start_date, end_date, start_time, end_time, location, purpose, reason))
        db.commit()

        try:
            update_quota_usage(staff_id, school_id, current_year)
        except Exception as quota_err:
            print(f"⚠️ Quota refresh after on-duty apply failed: {quota_err}")

        return jsonify({'success': True, 'message': 'On-duty application submitted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to submit application: {str(e)}'})

@app.route('/apply_permission', methods=['POST'])
def apply_permission():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    permission_type = request.form.get('permission_type')
    permission_date = request.form.get('permission_date')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    reason = request.form.get('reason')

    # Validate required fields
    if not all([permission_type, permission_date, start_time, end_time, reason]):
        return jsonify({'success': False, 'error': 'Please fill all required fields'})

    # Calculate duration in hours
    try:
        start_dt = datetime.datetime.strptime(start_time, '%H:%M')
        end_dt = datetime.datetime.strptime(end_time, '%H:%M')
        duration = (end_dt - start_dt).total_seconds() / 3600

        if duration <= 0:
            return jsonify({'success': False, 'error': 'End time must be after start time'})

    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid time format'})

    db = get_db()
    current_year = datetime.datetime.now().year
    
    # Initialize quotas if not exists
    initialize_staff_quotas(staff_id, school_id, current_year)

    # Ensure validation uses the latest usage (including recent withdrawals).
    update_quota_usage(staff_id, school_id, current_year)
    
    # CRITICAL: Check quota availability (Server-Side Validation)
    quota = db.execute("""
        SELECT allocated_hours, used_hours, (allocated_hours - used_hours) as remaining_hours
        FROM staff_permission_quotas 
        WHERE staff_id = ? AND quota_year = ?
    """, (staff_id, current_year)).fetchone()
    
    if not quota:
        return jsonify({
            'success': False, 
            'error': 'No Permission quota found. Please contact administrator.'
        })
    
    # Check A: Zero Balance
    if quota['remaining_hours'] <= 0:
        return jsonify({
            'success': False,
            'error': 'Application Failed: You have used all your available Permission hours.'
        })
    
    # Check B: Insufficient Balance
    if duration > quota['remaining_hours']:
        return jsonify({
            'success': False,
            'error': f'Application Failed: Your request for {duration} hours exceeds your remaining balance of {quota["remaining_hours"]} hour{"s" if quota["remaining_hours"] != 1 else ""}.'
        })

    # If quota checks pass, submit the application
    try:
        db.execute('''
            INSERT INTO permission_applications
            (staff_id, school_id, permission_type, permission_date, start_time, end_time, duration_hours, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (staff_id, school_id, permission_type, permission_date, start_time, end_time, duration, reason))
        db.commit()

        try:
            update_quota_usage(staff_id, school_id, current_year)
        except Exception as quota_err:
            print(f"⚠️ Quota refresh after permission apply failed: {quota_err}")

        return jsonify({'success': True, 'message': 'Permission application submitted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to submit application: {str(e)}'})


def _withdraw_staff_application(table_name, application_id, staff_id):
    """Withdraw a pending staff application atomically while preserving history."""
    db = get_db()
    withdrawn_at = datetime.datetime.now()

    cursor = db.execute(f'''
        UPDATE {table_name}
        SET withdrawn = 1,
            withdrawn_at = ?,
            status = 'rejected',
            processed_by = NULL,
            processed_at = ?
        WHERE id = ? AND staff_id = ? AND status = 'pending' AND COALESCE(withdrawn, 0) = 0
    ''', (withdrawn_at, withdrawn_at, application_id, staff_id))

    if cursor.rowcount == 0:
        existing = db.execute(
            f"SELECT status, COALESCE(withdrawn, 0) as withdrawn, staff_id FROM {table_name} WHERE id = ?",
            (application_id,)
        ).fetchone()
        if not existing:
            return False, 'Application not found'
        if str(existing['staff_id']) != str(staff_id):
            return False, 'Unauthorized'
        if existing['withdrawn']:
            return False, 'Application already withdrawn'
        return False, f"Cannot withdraw application in '{existing['status']}' status"

    db.commit()
    return True, 'Application withdrawn successfully'


@app.route('/withdraw_leave/<int:application_id>', methods=['POST'])
def withdraw_leave(application_id):
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    db = get_db()
    success, message = _withdraw_staff_application('leave_applications', application_id, session['user_id'])
    if not success:
        return jsonify({'success': False, 'error': message})
    try:
        row = db.execute('SELECT school_id, start_date FROM leave_applications WHERE id = ?', (application_id,)).fetchone()
        if row:
            quota_year = datetime.datetime.strptime(row['start_date'], '%Y-%m-%d').year
            update_quota_usage(session['user_id'], row['school_id'], quota_year)
    except Exception as quota_err:
        print(f"⚠️ Quota refresh after leave withdrawal failed: {quota_err}")
    return jsonify({'success': True, 'message': message})


@app.route('/withdraw_on_duty/<int:application_id>', methods=['POST'])
def withdraw_on_duty(application_id):
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    db = get_db()
    success, message = _withdraw_staff_application('on_duty_applications', application_id, session['user_id'])
    if not success:
        return jsonify({'success': False, 'error': message})
    try:
        row = db.execute('SELECT school_id, start_date FROM on_duty_applications WHERE id = ?', (application_id,)).fetchone()
        if row:
            quota_year = datetime.datetime.strptime(row['start_date'], '%Y-%m-%d').year
            update_quota_usage(session['user_id'], row['school_id'], quota_year)
    except Exception as quota_err:
        print(f"⚠️ Quota refresh after on-duty withdrawal failed: {quota_err}")
    return jsonify({'success': True, 'message': message})


@app.route('/withdraw_permission/<int:application_id>', methods=['POST'])
def withdraw_permission(application_id):
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    db = get_db()
    success, message = _withdraw_staff_application('permission_applications', application_id, session['user_id'])
    if not success:
        return jsonify({'success': False, 'error': message})
    try:
        row = db.execute('SELECT school_id, permission_date FROM permission_applications WHERE id = ?', (application_id,)).fetchone()
        if row:
            quota_year = datetime.datetime.strptime(row['permission_date'], '%Y-%m-%d').year
            update_quota_usage(session['user_id'], row['school_id'], quota_year)
    except Exception as quota_err:
        print(f"⚠️ Quota refresh after permission withdrawal failed: {quota_err}")
    return jsonify({'success': True, 'message': message})

@app.route('/process_leave', methods=['POST'])
def process_leave():
    """Process leave application with enhanced validation and logging"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized access - Admin required'})

    leave_id = request.form.get('leave_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    # Enhanced validation
    if not leave_id:
        return jsonify({'success': False, 'error': 'Leave ID is required'})
    
    if not decision or decision not in ['approve', 'reject']:
        return jsonify({'success': False, 'error': 'Valid decision (approve/reject) is required'})
    
    try:
        leave_id = int(leave_id)
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid leave ID format'})

    db = get_db()

    # Get leave application details for validation and logging
    leave_app = db.execute('''
        SELECT l.*, COALESCE(l.withdrawn, 0) as withdrawn, s.staff_id, s.full_name
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.id = ?
    ''', (leave_id,)).fetchone()

    if not leave_app:
        return jsonify({'success': False, 'error': 'Leave application not found'})
    
    # Check if already processed
    if leave_app['status'] != 'pending' or leave_app['withdrawn']:
        return jsonify({
            'success': False, 
            'error': 'Leave application is not pending. Cannot process again.'
        })

    status = 'approved' if decision == 'approve' else 'rejected'

    # Log the approval/rejection action
    print(f"🔐 LEAVE PROCESSING: Admin {admin_id} {decision}d leave ID {leave_id} "
          f"for staff {leave_app['staff_id']} ({leave_app['full_name']}) "
          f"- Type: {leave_app['leave_type']}, Dates: {leave_app['start_date']} to {leave_app['end_date']}")

    # Update leave status with admin tracking
    cursor = db.execute('''
        UPDATE leave_applications
        SET status = ?, processed_by = ?, processed_at = ?
        WHERE id = ? AND status = 'pending' AND COALESCE(withdrawn, 0) = 0
    ''', (status, admin_id, processed_at, leave_id))
    
    # Verify update was successful
    updated_rows = cursor.rowcount
    if updated_rows == 0:
        db.rollback()
        return jsonify({'success': False, 'error': 'Failed to update leave status - may already be processed'})
    
    db.commit()

    # Refresh quota usage for both approve/reject to keep reserved balance accurate.
    try:
        quota_year = datetime.datetime.strptime(leave_app['start_date'], '%Y-%m-%d').year
        update_result = update_quota_usage(leave_app['staff_id'], leave_app['school_id'], quota_year)
        print(f"✅ Quota updated for staff {leave_app['staff_id']}: {update_result}")
    except Exception as e:
        print(f"⚠️ Error updating quota usage: {e}")
        # Don't fail the whole operation for quota update errors

    return jsonify({
        'success': True, 
        'message': f'Leave application {status} successfully',
        'leave_id': leave_id,
        'staff_name': leave_app['full_name'],
        'action': decision
    })

@app.route('/process_on_duty', methods=['POST'])
def process_on_duty():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    application_id = request.form.get('application_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_remarks = request.form.get('admin_remarks', '')
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    if not application_id or not decision:
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    db = get_db()

    status = 'approved' if decision == 'approve' else 'rejected'

    try:
        # Get the on-duty application details before updating
        on_duty_app = db.execute('''
            SELECT staff_id, school_id, start_date, end_date, duty_type, location, purpose,
                   status, COALESCE(withdrawn, 0) as withdrawn
            FROM on_duty_applications
            WHERE id = ?
        ''', (application_id,)).fetchone()

        if not on_duty_app:
            return jsonify({'success': False, 'error': 'On-duty application not found'})

        if on_duty_app['status'] != 'pending' or on_duty_app['withdrawn']:
            return jsonify({'success': False, 'error': 'On-duty application is not pending'})

        # Update the application status
        cursor = db.execute('''
            UPDATE on_duty_applications
            SET status = ?, processed_by = ?, processed_at = ?, admin_remarks = ?
            WHERE id = ? AND status = 'pending' AND COALESCE(withdrawn, 0) = 0
        ''', (status, admin_id, processed_at, admin_remarks, application_id))

        if cursor.rowcount == 0:
            db.rollback()
            return jsonify({'success': False, 'error': 'On-duty application is no longer pending'})

        # If approved, mark attendance as "On Duty" for the date range
        if status == 'approved':
            staff_id = on_duty_app['staff_id']
            school_id = on_duty_app['school_id']
            start_date = datetime.datetime.strptime(on_duty_app['start_date'], '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(on_duty_app['end_date'], '%Y-%m-%d').date()

            # Mark attendance for each day in the date range
            current_date = start_date
            while current_date <= end_date:
                date_str = current_date.strftime('%Y-%m-%d')

                # Check if attendance record already exists
                existing_attendance = db.execute('''
                    SELECT id, status FROM attendance
                    WHERE staff_id = ? AND date = ?
                ''', (staff_id, date_str)).fetchone()

                if existing_attendance:
                    # Update existing record to "On Duty" status
                    db.execute('''
                        UPDATE attendance
                        SET status = 'on_duty',
                            on_duty_type = ?,
                            on_duty_location = ?,
                            on_duty_purpose = ?
                        WHERE staff_id = ? AND date = ?
                    ''', (on_duty_app['duty_type'], on_duty_app['location'],
                          on_duty_app['purpose'], staff_id, date_str))
                else:
                    # Create new attendance record with "On Duty" status
                    db.execute('''
                        INSERT INTO attendance
                        (staff_id, school_id, date, status, on_duty_type, on_duty_location, on_duty_purpose)
                        VALUES (?, ?, ?, 'on_duty', ?, ?, ?)
                    ''', (staff_id, school_id, date_str, on_duty_app['duty_type'],
                          on_duty_app['location'], on_duty_app['purpose']))

                current_date += datetime.timedelta(days=1)

        try:
            quota_year = datetime.datetime.strptime(on_duty_app['start_date'], '%Y-%m-%d').year
            update_result = update_quota_usage(on_duty_app['staff_id'], on_duty_app['school_id'], quota_year)
            print(f"OD Quota updated for staff {on_duty_app['staff_id']}: {update_result}")
        except Exception as e:
            print(f"Error updating OD quota usage: {e}")

        db.commit()

        return jsonify({'success': True, 'message': f'On-duty application {status} successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to process application: {str(e)}'})

@app.route('/process_permission', methods=['POST'])
def process_permission():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    application_id = request.form.get('application_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_remarks = request.form.get('admin_remarks', '')
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    if not application_id or not decision:
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    db = get_db()

    status = 'approved' if decision == 'approve' else 'rejected'

    try:
        permission_details = db.execute('''
            SELECT staff_id, school_id, permission_date, status, COALESCE(withdrawn, 0) as withdrawn
            FROM permission_applications
            WHERE id = ?
        ''', (application_id,)).fetchone()

        if not permission_details:
            return jsonify({'success': False, 'error': 'Permission application not found'})

        if permission_details['status'] != 'pending' or permission_details['withdrawn']:
            return jsonify({'success': False, 'error': 'Permission application is not pending'})

        cursor = db.execute('''
            UPDATE permission_applications
            SET status = ?, processed_by = ?, processed_at = ?, admin_remarks = ?
            WHERE id = ? AND status = 'pending' AND COALESCE(withdrawn, 0) = 0
        ''', (status, admin_id, processed_at, admin_remarks, application_id))

        if cursor.rowcount == 0:
            db.rollback()
            return jsonify({'success': False, 'error': 'Permission application is no longer pending'})

        try:
            quota_year = datetime.datetime.strptime(permission_details['permission_date'], '%Y-%m-%d').year
            update_result = update_quota_usage(permission_details['staff_id'], permission_details['school_id'], quota_year)
            print(f"Permission Quota updated for staff {permission_details['staff_id']}: {update_result}")
        except Exception as e:
            print(f"Error updating permission quota usage: {e}")

        db.commit()

        return jsonify({'success': True, 'message': f'Permission application {status} successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to process application: {str(e)}'})














# Weekly Calendar API Endpoints
@app.route('/get_weekly_attendance')
def get_weekly_attendance():
    """Get weekly attendance data for staff profile calendar"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get parameters
    staff_id = request.args.get('staff_id')  # For admin viewing other staff
    week_start = request.args.get('week_start')  # YYYY-MM-DD format

    # Determine which staff to get data for
    if session['user_type'] == 'staff':
        target_staff_id = session['user_id']
    elif session['user_type'] == 'admin' and staff_id:
        target_staff_id = staff_id
    else:
        return jsonify({'success': False, 'error': 'Invalid request'})

    if not week_start:
        return jsonify({'success': False, 'error': 'Week start date required'})

    try:
        # Parse week start date
        week_start_date = datetime.datetime.strptime(week_start, '%Y-%m-%d').date()
        week_end_date = week_start_date + datetime.timedelta(days=6)

        db = get_db()

        # Get staff information including shift type and scheduled changes
        staff_info = db.execute('''
            SELECT id, school_id, staff_id, full_name, shift_type, next_shift_type, next_shift_effective_date
            FROM staff
            WHERE id = ?
        ''', (target_staff_id,)).fetchone()

        if not staff_info:
            return jsonify({'success': False, 'error': 'Staff not found'})

        # Get all active shift definitions for easier lookup
        shift_defs_raw = db.execute('''
            SELECT shift_type, start_time, end_time, grace_period_minutes
            FROM shift_definitions
            WHERE school_id = ? AND is_active = 1
        ''', (staff_info['school_id'],)).fetchall()
        
        all_shift_defs = {sd['shift_type']: sd for sd in shift_defs_raw}

        # Load full shift history for this staff member (for past-date lookups)
        shift_history = db.execute('''
            SELECT shift_type, effective_from, effective_to
            FROM staff_shift_history
            WHERE staff_id = ? AND school_id = ?
            ORDER BY effective_from ASC
        ''', (target_staff_id, staff_info['school_id'])).fetchall()

        def get_shift_for_date(date_str):
            """Return the shift_type that was in effect on a given date per history."""
            for row in reversed(shift_history):  # newest first
                if row['effective_from'] <= date_str:
                    if row['effective_to'] is None or row['effective_to'] >= date_str:
                        return row['shift_type']
            return None  # no history found, caller falls back to current

        # Get attendance records for the week
        attendance_records = db.execute('''
            SELECT date, time_in, time_out, status,
                   late_duration_minutes, early_departure_minutes,
                   on_duty_type, on_duty_location, on_duty_purpose
            FROM attendance
            WHERE staff_id = ? AND date BETWEEN ? AND ?
            ORDER BY date
        ''', (target_staff_id, week_start_date, week_end_date)).fetchall()

        # Get leave applications for the week
        leave_applications = db.execute('''
            SELECT id, leave_type, start_date, end_date, reason, status
            FROM leave_applications
            WHERE staff_id = ? 
              AND status IN ('approved', 'pending')
              AND ((start_date BETWEEN ? AND ?) 
                   OR (end_date BETWEEN ? AND ?)
                   OR (start_date <= ? AND end_date >= ?))
            ORDER BY start_date
        ''', (target_staff_id, week_start_date, week_end_date, 
              week_start_date, week_end_date,
              week_start_date, week_end_date)).fetchall()

        # Get on-duty applications for the week
        on_duty_applications = db.execute('''
            SELECT id, duty_type, start_date, end_date, location, purpose, status
            FROM on_duty_applications
            WHERE staff_id = ? 
              AND status IN ('approved', 'pending')
              AND ((start_date BETWEEN ? AND ?) 
                   OR (end_date BETWEEN ? AND ?)
                   OR (start_date <= ? AND end_date >= ?))
            ORDER BY start_date
        ''', (target_staff_id, week_start_date, week_end_date,
              week_start_date, week_end_date,
              week_start_date, week_end_date)).fetchall()

        # Get permission applications for the week
        permission_applications = db.execute('''
            SELECT id, permission_type, permission_date, start_time, end_time, 
                   duration_hours, reason, status
            FROM permission_applications
            WHERE staff_id = ? 
              AND status IN ('approved', 'pending')
              AND permission_date BETWEEN ? AND ?
            ORDER BY permission_date
        ''', (target_staff_id, week_start_date, week_end_date)).fetchall()

        today_str = datetime.datetime.now().strftime('%Y-%m-%d')

        # Create weekly data structure
        weekly_data = []
        current_date = week_start_date

        for i in range(7):  # 7 days in a week
            date_str = current_date.strftime('%Y-%m-%d')
            day_name = current_date.strftime('%A')

            # Resolve actual shift type for this specific date
            if date_str < today_str:
                # Past day: use historical shift record so previous days never
                # reflect a future/current shift change.
                actual_shift_type = get_shift_for_date(date_str) or staff_info['shift_type'] or 'general'
            else:
                # Today or future: honour any scheduled (next) shift change
                actual_shift_type = staff_info['shift_type'] or 'general'
                if (staff_info['next_shift_type'] and 
                    staff_info['next_shift_effective_date'] and 
                    date_str >= staff_info['next_shift_effective_date']):
                    actual_shift_type = staff_info['next_shift_type']
                
            # Use defined shift or fallback to general
            current_shift_def = all_shift_defs.get(actual_shift_type)
            if not current_shift_def:
                current_shift_def = all_shift_defs.get('general')

            # Find attendance record for this date
            attendance_record = None
            for record in attendance_records:
                if record['date'] == date_str:
                    attendance_record = record
                    break

            # Calculate day data
            day_data = calculate_daily_attendance_data(
                current_date, attendance_record, current_shift_def, actual_shift_type
            )
            day_data['day_name'] = day_name
            day_data['date'] = date_str

            # Check for leave applications on this date
            day_data['leave_applications'] = []
            for leave in leave_applications:
                # Handle both string and date object formats from SQLite
                leave_start = leave['start_date'] if isinstance(leave['start_date'], datetime.date) else datetime.datetime.strptime(leave['start_date'], '%Y-%m-%d').date()
                leave_end = leave['end_date'] if isinstance(leave['end_date'], datetime.date) else datetime.datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
                if leave_start <= current_date <= leave_end:
                    day_data['leave_applications'].append({
                        'id': leave['id'],
                        'type': leave['leave_type'],
                        'reason': leave['reason'],
                        'status': leave['status']
                    })

            # Check for on-duty applications on this date
            day_data['on_duty_applications'] = []
            for od in on_duty_applications:
                # Handle both string and date object formats from SQLite
                od_start = od['start_date'] if isinstance(od['start_date'], datetime.date) else datetime.datetime.strptime(od['start_date'], '%Y-%m-%d').date()
                od_end = od['end_date'] if isinstance(od['end_date'], datetime.date) else datetime.datetime.strptime(od['end_date'], '%Y-%m-%d').date()
                if od_start <= current_date <= od_end:
                    day_data['on_duty_applications'].append({
                        'id': od['id'],
                        'type': od['duty_type'],
                        'location': od['location'],
                        'purpose': od['purpose'],
                        'status': od['status']
                    })

            # Check for permission applications on this date
            day_data['permission_applications'] = []
            for perm in permission_applications:
                # Handle both string and date object formats from SQLite
                perm_date = perm['permission_date'] if isinstance(perm['permission_date'], datetime.date) else datetime.datetime.strptime(perm['permission_date'], '%Y-%m-%d').date()
                if perm_date == current_date:
                    day_data['permission_applications'].append({
                        'id': perm['id'],
                        'type': perm['permission_type'],
                        'start_time': perm['start_time'],
                        'end_time': perm['end_time'],
                        'duration': perm['duration_hours'],
                        'reason': perm['reason'],
                        'status': perm['status']
                    })

            weekly_data.append(day_data)
            current_date += datetime.timedelta(days=1)

        return jsonify({
            'success': True,
            'staff_info': {
                'id': staff_info['id'],
                'staff_id': staff_info['staff_id'],
                'full_name': staff_info['full_name'],
                'shift_type': staff_info['shift_type'] or 'general',
                'next_shift_type': staff_info['next_shift_type'],
                'next_shift_effective_date': staff_info['next_shift_effective_date']
            },
            'week_start': week_start,
            'week_end': week_end_date.strftime('%Y-%m-%d'),
            'weekly_data': weekly_data
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


def calculate_daily_attendance_data(date, attendance_record, shift_def, shift_type):
    """Calculate daily attendance data with delays and early departures"""

    # Default data structure with shift timing information
    day_data = {
        'present_status': 'Absent',
        'shift_type_display': shift_type.replace('_', ' ').title() + ' Shift',
        'shift_start_time': None,
        'shift_end_time': None,
        'morning_thumb': None,
        'evening_thumb': None,
        'delay_info': None,
        'delay_duration': None,
        'arrived_soon_info': None,
        'arrived_soon_duration': None,
        'left_soon_info': None,
        'left_soon_duration': None
    }

    # Add shift timing information in 12-hour format
    if shift_def:
        day_data['shift_start_time'] = format_time_to_12hr(shift_def['start_time'])
        day_data['shift_end_time'] = format_time_to_12hr(shift_def['end_time'])

    if not attendance_record:
        return day_data

    # Check attendance status
    attendance_status = attendance_record['status'] if attendance_record['status'] else 'absent'

    # Handle on-duty status
    if attendance_status == 'on_duty':
        day_data['present_status'] = 'On Duty'
        # Safely access on-duty fields with fallbacks
        day_data['on_duty_type'] = attendance_record['on_duty_type'] if attendance_record['on_duty_type'] else 'Official Work'
        day_data['on_duty_location'] = attendance_record['on_duty_location'] if attendance_record['on_duty_location'] else 'Not specified'
        day_data['on_duty_purpose'] = attendance_record['on_duty_purpose'] if attendance_record['on_duty_purpose'] else 'Official duty'
    
    # Handle Holiday status
    elif attendance_status == 'holiday':
        day_data['present_status'] = 'Holiday'
        day_data['holiday_info'] = attendance_record['notes'] if attendance_record['notes'] else 'General Holiday'

    # Handle Present status with Shift Prefix
    elif attendance_status in ['present', 'late', 'left_soon']:
        shift_prefix = shift_type.replace('_', ' ').title() + ' Shift'
        
        if attendance_status == 'late':
            day_data['present_status'] = f"{shift_prefix} - Late"
        elif attendance_status == 'left_soon':
            day_data['present_status'] = f"{shift_prefix} - Left Early"
        else:
            day_data['present_status'] = f"{shift_prefix} - Present"
    
    else:
        day_data['present_status'] = attendance_status.title() if attendance_status else 'Absent'
        return day_data  # Return early for on-duty status

    # Staff was present if there's a time_in record
    if attendance_record['time_in']:
        # Only set to 'Present' if it wasn't already set with shift details
        if day_data['present_status'] in ['Absent', 'On Duty', 'Holiday']:
             day_data['present_status'] = 'Present'
             
        day_data['morning_thumb'] = format_time_to_12hr(attendance_record['time_in'])

        # Calculate arrival timing relative to shift start
        if shift_def:
            actual_time = datetime.datetime.strptime(attendance_record['time_in'], '%H:%M:%S').time()
            shift_start_time = datetime.datetime.strptime(shift_def['start_time'], '%H:%M:%S').time()
            grace_period = shift_def['grace_period_minutes'] if shift_def['grace_period_minutes'] else 10

            # Convert times to datetime objects for comparison
            actual_datetime = datetime.datetime.combine(datetime.date.today(), actual_time)
            shift_start_datetime = datetime.datetime.combine(datetime.date.today(), shift_start_time)
            grace_cutoff_datetime = shift_start_datetime + datetime.timedelta(minutes=grace_period)

            # Check if arrived early (before shift start time)
            if actual_datetime < shift_start_datetime:
                early_minutes = int((shift_start_datetime - actual_datetime).total_seconds() / 60)
                shift_start_12hr = format_time_to_12hr(shift_def['start_time'])
                actual_time_12hr = day_data['morning_thumb']

                day_data['arrived_soon_info'] = f"Arrived Soon: {format_duration_minutes(early_minutes)} (Expected: {shift_start_12hr}, Actual: {actual_time_12hr})"
                day_data['arrived_soon_duration'] = early_minutes

            # Check if arrived late (after grace period)
            elif actual_datetime > grace_cutoff_datetime:
                if attendance_record['late_duration_minutes'] and attendance_record['late_duration_minutes'] > 0:
                    delay_minutes = attendance_record['late_duration_minutes']
                    grace_cutoff_12hr = format_time_to_12hr(grace_cutoff_datetime.time().strftime('%H:%M:%S'))
                    actual_time_12hr = day_data['morning_thumb']

                    day_data['delay_info'] = f"Delay: {format_duration_minutes(delay_minutes)} (Expected: {grace_cutoff_12hr}, Actual: {actual_time_12hr})"
                    day_data['delay_duration'] = delay_minutes

    # Evening thumb out
    if attendance_record['time_out']:
        day_data['evening_thumb'] = format_time_to_12hr(attendance_record['time_out'])

        # Calculate departure timing relative to shift end
        if shift_def:
            actual_time = datetime.datetime.strptime(attendance_record['time_out'], '%H:%M:%S').time()
            shift_end_time = datetime.datetime.strptime(shift_def['end_time'], '%H:%M:%S').time()

            # Convert times to datetime objects for comparison
            actual_datetime = datetime.datetime.combine(datetime.date.today(), actual_time)
            shift_end_datetime = datetime.datetime.combine(datetime.date.today(), shift_end_time)

            # Check if left early (before shift end time)
            if actual_datetime < shift_end_datetime:
                early_minutes = int((shift_end_datetime - actual_datetime).total_seconds() / 60)
                shift_end_12hr = format_time_to_12hr(shift_def['end_time'])
                actual_time_12hr = day_data['evening_thumb']

                day_data['left_soon_info'] = f"Left Soon: {format_duration_minutes(early_minutes)} (Expected: {shift_end_12hr}, Actual: {actual_time_12hr})"
                day_data['left_soon_duration'] = early_minutes

    # No overtime functionality - removed

    return day_data


def format_time_to_12hr(time_str):
    """Convert 24-hour time string to 12-hour format"""
    if not time_str:
        return None

    try:
        # Handle both HH:MM:SS and HH:MM formats
        if len(time_str.split(':')) == 3:
            time_obj = datetime.datetime.strptime(time_str, '%H:%M:%S').time()
        else:
            time_obj = datetime.datetime.strptime(time_str, '%H:%M').time()

        # Convert to 12-hour format
        return datetime.datetime.combine(datetime.date.today(), time_obj).strftime('%I:%M %p')
    except:
        return time_str  # Return original if conversion fails


def format_duration_minutes(minutes):
    """Format duration in minutes to readable format without prefix"""
    if minutes <= 0:
        return ""

    hours = minutes // 60
    mins = minutes % 60

    if hours > 0:
        if mins > 0:
            return f"{hours}h {mins}m"
        else:
            return f"{hours}h"
    else:
        return f"{mins}m"


def format_duration_for_display(minutes, prefix):
    """Format duration in minutes to display format"""
    if minutes <= 0:
        return None

    duration_str = format_duration_minutes(minutes)
    return f"{prefix}: {duration_str}"


def format_attendance_times_to_12hr(attendance_record):
    """Convert all time fields in an attendance record to 12-hour format"""
    if not attendance_record:
        return attendance_record

    # Create a copy to avoid modifying the original
    formatted_record = dict(attendance_record)

    # Time fields to convert
    time_fields = ['time_in', 'time_out', 'shift_start_time', 'shift_end_time']

    for field in time_fields:
        if field in formatted_record and formatted_record[field]:
            formatted_record[field] = format_time_to_12hr(formatted_record[field])

    return formatted_record


def resolve_effective_shift_type(staff_row, target_date=None):
    """Resolve a staff member's effective shift for a given date."""
    if not staff_row:
        return 'general'

    date_value = target_date or datetime.date.today()
    if isinstance(date_value, datetime.datetime):
        date_str = date_value.date().strftime('%Y-%m-%d')
    elif isinstance(date_value, datetime.date):
        date_str = date_value.strftime('%Y-%m-%d')
    else:
        date_str = str(date_value)

    current_shift = staff_row['shift_type'] if 'shift_type' in staff_row.keys() else None
    next_shift = staff_row['next_shift_type'] if 'next_shift_type' in staff_row.keys() else None
    next_effective = staff_row['next_shift_effective_date'] if 'next_shift_effective_date' in staff_row.keys() else None

    effective_shift = current_shift or 'general'
    if next_shift and next_effective and date_str >= str(next_effective):
        effective_shift = next_shift

    return effective_shift


@app.route('/add_staff_enhanced', methods=['POST'])
def add_staff_enhanced():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    staff_id = request.form.get('staff_id')
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    date_of_birth = request.form.get('date_of_birth')
    date_of_joining = request.form.get('date_of_joining')
    department = request.form.get('department')
    destination = request.form.get('destination')
    gender = request.form.get('gender')
    phone = request.form.get('phone')
    email = request.form.get('email')
    bank_account_name = (request.form.get('bank_account_name') or '').strip()
    bank_name = (request.form.get('bank_name') or '').strip()
    bank_account_number = (request.form.get('bank_account_number') or '').strip()
    ifsc_code = (request.form.get('ifsc_code') or '').strip().upper()
    pan_number = (request.form.get('pan_number') or '').strip().upper()
    shift_type = request.form.get('shift_type', 'general')
    is_active_raw = request.form.get('is_active')
    status_raw = request.form.get('status')
    is_active = 1
    if is_active_raw is not None:
        is_active = 1 if str(is_active_raw).lower() in ['1', 'true', 'on', 'yes', 'active'] else 0
    elif status_raw is not None:
        is_active = 1 if str(status_raw).lower() == 'active' else 0
    password = generate_password_hash(request.form.get('password'))

    # Salary fields
    basic_salary = float(request.form.get('basic_salary', 0) or 0)
    hra = float(request.form.get('hra', 0) or 0)
    transport_allowance = float(request.form.get('transport_allowance', 0) or 0)
    other_allowances = float(request.form.get('other_allowances', 0) or 0)
    dearness_allowance = float(request.form.get('dearness_allowance', 0) or 0)
    esi_deduction = float(request.form.get('esi_deduction', 0) or 0)
    professional_tax = float(request.form.get('professional_tax', 0) or 0)
    other_deductions = float(request.form.get('other_deductions', 0) or 0)
    pf_toggle_raw = request.form.get('pf_enabled')
    pf_opt_in = 1 if str(pf_toggle_raw).strip().lower() in ['1', 'true', 'on', 'yes'] else 0
    pf_manual_override_raw = request.form.get('pf_manual_override')
    pf_manual_override = 1 if str(pf_manual_override_raw).strip().lower() in ['1', 'true', 'on', 'yes'] else 0
    pf_deduction_input = float(request.form.get('pf_deduction', 0) or 0)

    # Create full_name from first_name and last_name
    full_name = f"{first_name} {last_name}".strip() if first_name or last_name else ""

    # Validate age (must be 18 or older)
    if date_of_birth:
        try:
            birth_date = datetime.datetime.strptime(date_of_birth, '%Y-%m-%d').date()
            today = datetime.date.today()
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
            if age < 18:
                return jsonify({'success': False, 'error': 'Staff member must be at least 18 years old'})
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date of birth format'})

    db = get_db()

    # PF auto-calculation per EPFO-style rule set
    active_staff_count = db.execute('''
        SELECT COUNT(*) AS total
        FROM staff
        WHERE school_id = ? AND COALESCE(is_active, 1) = 1
    ''', (school_id,)).fetchone()
    company_employee_count = int((active_staff_count['total'] if active_staff_count else 0) or 0)
    pf_components = calculate_pf_components(
        basic_salary=basic_salary,
        dearness_allowance=dearness_allowance,
        pf_opt_in=bool(pf_opt_in),
        company_employee_count=company_employee_count,
    )
    auto_pf_deduction = float(pf_components.get('employee_pf', 0))
    if pf_opt_in == 0:
        pf_deduction = 0.0
    elif pf_manual_override == 1:
        pf_deduction = max(0.0, pf_deduction_input)
    else:
        pf_deduction = auto_pf_deduction

    # Auto-assign shift type based on department if not manually specified
    if not shift_type or shift_type == 'general':
        if department:
            dept_mapping = db.execute('''
                SELECT default_shift_type
                FROM department_shift_mappings
                WHERE school_id = ? AND department = ?
            ''', (school_id, department)).fetchone()

            if dept_mapping:
                shift_type = dept_mapping['default_shift_type']
            else:
                shift_type = 'general'  # fallback

    # Always create staff in both app and device, do not check for staff ID existence
    device_ip, device_port = get_institution_device()
    biometric_created = False
    biometric_error = None
    if not device_ip:
        biometric_error = 'No biometric device configured for your institution'
    else:
        try:
            zk_device = ZKBiometricDevice(device_ip)
            if zk_device.connect():
                # Always overwrite user on device, do not check existence
                biometric_created = zk_device.create_user(user_id=staff_id, name=full_name, overwrite=True)
                zk_device.disconnect()
                if not biometric_created:
                    biometric_error = 'Failed to create staff on biometric device. Please check device connection.'
            else:
                biometric_error = 'Cannot connect to biometric device. Staff not added.'
        except Exception as e:
            biometric_error = f'Biometric device error: {e}'

    # Handle photo upload (if any)
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo and photo.filename and allowed_file(photo.filename):
            # Ensure uploads directory exists
            upload_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'static/uploads'), 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Generate filename with staff name
            ext = os.path.splitext(photo.filename)[1]
            safe_name = full_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            filename = f"{safe_name}_{staff_id}{ext}"
            photo_path = os.path.join(upload_dir, filename)
            photo.save(photo_path)
            photo_url = f"uploads/{filename}"

    try:
        # Ensure all required columns exist in the table
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        column_names = [col['name'] for col in columns]

        # Build the insert query dynamically based on available columns
        insert_columns = ['school_id', 'staff_id', 'password_hash', 'full_name']
        insert_values = [school_id, staff_id, password, full_name]

        if 'first_name' in column_names and first_name:
            insert_columns.append('first_name')
            insert_values.append(first_name)
        if 'last_name' in column_names and last_name:
            insert_columns.append('last_name')
            insert_values.append(last_name)
        if 'date_of_birth' in column_names and date_of_birth:
            insert_columns.append('date_of_birth')
            insert_values.append(date_of_birth)
        if 'date_of_joining' in column_names and date_of_joining:
            insert_columns.append('date_of_joining')
            insert_values.append(date_of_joining)
        if 'department' in column_names and department:
            insert_columns.append('department')
            insert_values.append(department)
        if 'destination' in column_names and destination:
            insert_columns.append('destination')
            insert_values.append(destination)
        if 'position' in column_names and destination:
            insert_columns.append('position')
            insert_values.append(destination)  # Use the same value for position
        if 'gender' in column_names and gender:
            insert_columns.append('gender')
            insert_values.append(gender)
        if 'phone' in column_names and phone:
            insert_columns.append('phone')
            insert_values.append(phone)
        if 'email' in column_names and email:
            insert_columns.append('email')
            insert_values.append(email)
        if 'bank_account_name' in column_names and bank_account_name:
            insert_columns.append('bank_account_name')
            insert_values.append(bank_account_name)
        if 'bank_name' in column_names and bank_name:
            insert_columns.append('bank_name')
            insert_values.append(bank_name)
        if 'bank_account_number' in column_names and bank_account_number:
            insert_columns.append('bank_account_number')
            insert_values.append(bank_account_number)
        if 'ifsc_code' in column_names and ifsc_code:
            insert_columns.append('ifsc_code')
            insert_values.append(ifsc_code)
        if 'pan_number' in column_names and pan_number:
            insert_columns.append('pan_number')
            insert_values.append(pan_number)
        if 'shift_type' in column_names:
            insert_columns.append('shift_type')
            insert_values.append(shift_type)
        if 'photo_url' in column_names and photo_url:
            insert_columns.append('photo_url')
            insert_values.append(photo_url)
        if 'basic_salary' in column_names:
            insert_columns.append('basic_salary')
            insert_values.append(basic_salary)
        if 'hra' in column_names:
            insert_columns.append('hra')
            insert_values.append(hra)
        if 'transport_allowance' in column_names:
            insert_columns.append('transport_allowance')
            insert_values.append(transport_allowance)
        if 'other_allowances' in column_names:
            insert_columns.append('other_allowances')
            insert_values.append(other_allowances)
        if 'dearness_allowance' in column_names:
            insert_columns.append('dearness_allowance')
            insert_values.append(dearness_allowance)
        if 'pf_opt_in' in column_names:
            insert_columns.append('pf_opt_in')
            insert_values.append(pf_opt_in)
        if 'pf_deduction' in column_names:
            insert_columns.append('pf_deduction')
            insert_values.append(pf_deduction)
        if 'esi_deduction' in column_names:
            insert_columns.append('esi_deduction')
            insert_values.append(esi_deduction)
        if 'professional_tax' in column_names:
            insert_columns.append('professional_tax')
            insert_values.append(professional_tax)
        if 'other_deductions' in column_names:
            insert_columns.append('other_deductions')
            insert_values.append(other_deductions)
        if 'is_active' in column_names:
            insert_columns.append('is_active')
            insert_values.append(is_active)
        if 'status' in column_names:
            insert_columns.append('status')
            insert_values.append('active' if is_active else 'inactive')

        # Add created_at timestamp
        if 'created_at' in column_names:
            insert_columns.append('created_at')
            insert_values.append(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Build and execute the query
        placeholders = ', '.join(['?' for _ in insert_values])
        query = f"INSERT INTO staff ({', '.join(insert_columns)}) VALUES ({placeholders})"

        db.execute(query, insert_values)
        db.commit()

        return jsonify({'success': True, 'biometric_created': biometric_created, 'biometric_error': biometric_error})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add_staff', methods=['POST'])
def add_staff():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    staff_id = request.form.get('staff_id')
    full_name = request.form.get('full_name')
    password = generate_password_hash(request.form.get('password'))
    email = request.form.get('email')
    phone = request.form.get('phone')
    department = request.form.get('department')
    position = request.form.get('position')
    is_active_raw = request.form.get('is_active')
    status_raw = request.form.get('status')
    is_active = 1
    if is_active_raw is not None:
        is_active = 1 if str(is_active_raw).lower() in ['1', 'true', 'on', 'yes', 'active'] else 0
    elif status_raw is not None:
        is_active = 1 if str(status_raw).lower() == 'active' else 0

    # Get salary information
    basic_salary = float(request.form.get('basic_salary', 0))
    hra = float(request.form.get('hra', 0))
    transport_allowance = float(request.form.get('transport_allowance', 0))
    other_allowances = float(request.form.get('other_allowances', 0))
    dearness_allowance = float(request.form.get('dearness_allowance', 0) or 0)
    pf_toggle_raw = request.form.get('pf_enabled')
    pf_opt_in = 1 if str(pf_toggle_raw).strip().lower() in ['1', 'true', 'on', 'yes'] else 0
    pf_manual_override_raw = request.form.get('pf_manual_override')
    pf_manual_override = 1 if str(pf_manual_override_raw).strip().lower() in ['1', 'true', 'on', 'yes'] else 0
    pf_deduction_input = float(request.form.get('pf_deduction', 0) or 0)
    esi_deduction = float(request.form.get('esi_deduction', 0))
    professional_tax = float(request.form.get('professional_tax', 0))
    other_deductions = float(request.form.get('other_deductions', 0))
    bank_account_name = (request.form.get('bank_account_name') or '').strip()
    bank_name = (request.form.get('bank_name') or '').strip()
    bank_account_number = (request.form.get('bank_account_number') or '').strip()
    ifsc_code = (request.form.get('ifsc_code') or '').strip().upper()
    pan_number = (request.form.get('pan_number') or '').strip().upper()

    # biometric_enrolled = request.form.get('biometric_enrolled', 'false').lower() == 'true'  # Not used currently

    db = get_db()

    active_staff_count = db.execute('''
        SELECT COUNT(*) AS total
        FROM staff
        WHERE school_id = ? AND COALESCE(is_active, 1) = 1
    ''', (school_id,)).fetchone()
    company_employee_count = int((active_staff_count['total'] if active_staff_count else 0) or 0)
    pf_components = calculate_pf_components(
        basic_salary=basic_salary,
        dearness_allowance=dearness_allowance,
        pf_opt_in=bool(pf_opt_in),
        company_employee_count=company_employee_count,
    )
    auto_pf_deduction = float(pf_components.get('employee_pf', 0))
    if pf_opt_in == 0:
        pf_deduction = 0.0
    elif pf_manual_override == 1:
        pf_deduction = max(0.0, pf_deduction_input)
    else:
        pf_deduction = auto_pf_deduction

    # Always create staff in both app and device, do not check for staff ID existence
    device_ip, device_port = get_institution_device()
    biometric_created = False
    biometric_error = None
    if not device_ip:
        biometric_error = 'No biometric device configured for your institution'
    else:
        try:
            zk_device = ZKBiometricDevice(device_ip)
            if zk_device.connect():
                # Always overwrite user on device, do not check existence
                biometric_created = zk_device.create_user(user_id=staff_id, name=full_name, overwrite=True)
                zk_device.disconnect()
                if not biometric_created:
                    biometric_error = 'Failed to create staff on biometric device. Please check device connection.'
            else:
                biometric_error = 'Cannot connect to biometric device. Staff not added.'
        except Exception as e:
            biometric_error = f'Biometric device error: {e}'

    # Handle photo upload
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo.filename != '' and allowed_file(photo.filename):
            try:
                # Ensure uploads directory exists
                upload_dir = os.path.join(app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)

                # Generate filename with staff name
                ext = os.path.splitext(photo.filename)[1]
                safe_name = full_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                filename = f"{safe_name}_{staff_id}{ext}"
                # Save only the relative path for static serving
                photo_url = f"uploads/{filename}"
                photo.save(os.path.join(upload_dir, filename))
            except Exception as e:
                print(f"Error saving photo: {e}")
                return jsonify({'success': False, 'error': 'Error saving photo'})
        elif photo.filename != '':
            return jsonify({'success': False, 'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'})
    # Always set a default photo if none uploaded
    if not photo_url:
        photo_url = 'images/applogo.png'  # Fallback to app logo

    try:
        # Ensure 'photo_url' column exists in the table
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        has_photo_url = any(col['name'] == 'photo_url' for col in columns)

        if not has_photo_url:
            db.execute("ALTER TABLE staff ADD COLUMN photo_url TEXT")

        created_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        insert_cursor = db.execute('''
            INSERT INTO staff
            (school_id, staff_id, password_hash, full_name, email, phone, department, position, destination, photo_url, created_at,
                         basic_salary, hra, transport_allowance, other_allowances, dearness_allowance, pf_opt_in, pf_deduction, esi_deduction, professional_tax, other_deductions,
                         bank_account_name, bank_name, bank_account_number, ifsc_code, pan_number)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (school_id, staff_id, password, full_name, email, phone, department, position, position, photo_url, created_at,
                            basic_salary, hra, transport_allowance, other_allowances, dearness_allowance, pf_opt_in, pf_deduction, esi_deduction, professional_tax, other_deductions,
                            bank_account_name, bank_name, bank_account_number, ifsc_code, pan_number))

        staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]
        if 'is_active' in staff_columns or 'status' in staff_columns:
            update_parts = []
            update_values = []
            if 'is_active' in staff_columns:
                update_parts.append('is_active = ?')
                update_values.append(is_active)
            if 'status' in staff_columns:
                update_parts.append('status = ?')
                update_values.append('active' if is_active else 'inactive')

            if update_parts:
                update_values.append(insert_cursor.lastrowid)
                db.execute(f"UPDATE staff SET {', '.join(update_parts)} WHERE id = ?", update_values)

        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/check_staff_id_availability', methods=['POST'])
def check_staff_id_availability():
    """Check if a staff ID is available in the current school"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    school_id = session.get('school_id')

    if not staff_id:
        return jsonify({'success': False, 'error': 'Staff ID is required'})

    if not school_id:
        return jsonify({'success': False, 'error': 'School ID not found in session'})

    try:
        db = get_db()
        existing_staff = db.execute('''
            SELECT full_name, created_at FROM staff
            WHERE school_id = ? AND staff_id = ?
        ''', (school_id, staff_id)).fetchone()

        if existing_staff:
            return jsonify({
                'success': True,
                'available': False,
                'message': f'Staff ID "{staff_id}" is already used by {existing_staff["full_name"]}',
                'suggestion': 'Please choose a different Staff ID'
            })
        else:
            return jsonify({
                'success': True,
                'available': True,
                'message': f'Staff ID "{staff_id}" is available'
            })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_school', methods=['POST'])
def delete_school():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = request.form.get('school_id')

    db = get_db()

    try:
        # Delete all related records first
        db.execute('DELETE FROM admins WHERE school_id = ?', (school_id,))
        db.execute('DELETE FROM staff WHERE school_id = ?', (school_id,))
        db.execute('DELETE FROM schools WHERE id = ?', (school_id,))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_staff_details_enhanced')
def get_staff_details_enhanced():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.args.get('id')
    db = get_db()

    staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]
    da_select = 'COALESCE(dearness_allowance, 0) AS dearness_allowance' if 'dearness_allowance' in staff_columns else '0 AS dearness_allowance'
    pf_opt_in_select = 'COALESCE(pf_opt_in, 0) AS pf_opt_in' if 'pf_opt_in' in staff_columns else '0 AS pf_opt_in'
    if 'is_active' in staff_columns and 'status' in staff_columns:
        active_expr = "COALESCE(is_active, CASE WHEN LOWER(COALESCE(status, 'active')) = 'active' THEN 1 ELSE 0 END) AS is_active"
    elif 'is_active' in staff_columns:
        active_expr = "COALESCE(is_active, 1) AS is_active"
    elif 'status' in staff_columns:
        active_expr = "CASE WHEN LOWER(COALESCE(status, 'active')) = 'active' THEN 1 ELSE 0 END AS is_active"
    else:
        active_expr = "1 AS is_active"

    staff = db.execute(f'''
        SELECT id, staff_id, full_name, first_name, last_name,
               date_of_birth, date_of_joining, department, destination,
               position, gender, phone, email, shift_type, photo_url,
               basic_salary, hra, transport_allowance, other_allowances,
               {da_select},
               {pf_opt_in_select},
             pf_deduction, esi_deduction, professional_tax, other_deductions,
               bank_account_name, bank_name, bank_account_number, ifsc_code, pan_number,
               {active_expr}
        FROM staff
        WHERE id = ?
    ''', (staff_id,)).fetchone()

    if not staff:
        return jsonify({'success': False, 'error': 'Staff not found'})

    return jsonify({
        'success': True,
        'staff': dict(staff)
    })

@app.route('/update_staff_enhanced', methods=['POST'])
def update_staff_enhanced():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_db_id = request.form.get('staff_db_id')
    staff_id = request.form.get('staff_id')
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    date_of_birth = request.form.get('date_of_birth')
    date_of_joining = request.form.get('date_of_joining')
    department = request.form.get('department')
    destination = request.form.get('destination') or request.form.get('position')  # Support both field names
    gender = request.form.get('gender')
    phone = request.form.get('phone')
    email = request.form.get('email')
    bank_account_name = (request.form.get('bank_account_name') or '').strip()
    bank_name = (request.form.get('bank_name') or '').strip()
    bank_account_number = (request.form.get('bank_account_number') or '').strip()
    ifsc_code = (request.form.get('ifsc_code') or '').strip().upper()
    pan_number = (request.form.get('pan_number') or '').strip().upper()
    shift_type = request.form.get('shift_type', 'general')
    is_active_raw = request.form.get('is_active')
    status_raw = request.form.get('status')
    is_active = 1
    if is_active_raw is not None:
        is_active = 1 if str(is_active_raw).lower() in ['1', 'true', 'on', 'yes', 'active'] else 0
    elif status_raw is not None:
        is_active = 1 if str(status_raw).lower() == 'active' else 0

    # Salary fields
    basic_salary = request.form.get('basic_salary', type=float)
    hra = request.form.get('hra', type=float)
    transport_allowance = request.form.get('transport_allowance', type=float)
    other_allowances = request.form.get('other_allowances', type=float)
    dearness_allowance = request.form.get('dearness_allowance', type=float)
    if dearness_allowance is None:
        dearness_allowance = 0.0
    pf_toggle_raw = request.form.get('pf_enabled')
    pf_opt_in = 1 if str(pf_toggle_raw).strip().lower() in ['1', 'true', 'on', 'yes'] else 0
    pf_manual_override_raw = request.form.get('pf_manual_override')
    pf_manual_override = 1 if str(pf_manual_override_raw).strip().lower() in ['1', 'true', 'on', 'yes'] else 0
    pf_deduction_input = request.form.get('pf_deduction', type=float)
    if pf_deduction_input is None:
        pf_deduction_input = 0.0
    esi_deduction = request.form.get('esi_deduction', type=float)
    professional_tax = request.form.get('professional_tax', type=float)
    other_deductions = request.form.get('other_deductions', type=float)

    school_id = session['school_id']

    # Create full_name from first_name and last_name
    full_name = f"{first_name} {last_name}".strip() if first_name or last_name else ""

    if not staff_db_id or not staff_id or not first_name or not last_name:
        return jsonify({'success': False, 'error': 'Staff ID, first name, and last name are required'})

    db = get_db()

    active_staff_count = db.execute('''
        SELECT COUNT(*) AS total
        FROM staff
        WHERE school_id = ? AND COALESCE(is_active, 1) = 1
    ''', (school_id,)).fetchone()
    company_employee_count = int((active_staff_count['total'] if active_staff_count else 0) or 0)
    pf_components = calculate_pf_components(
        basic_salary=float(basic_salary or 0),
        dearness_allowance=float(dearness_allowance or 0),
        pf_opt_in=bool(pf_opt_in),
        company_employee_count=company_employee_count,
    )
    auto_pf_deduction = float(pf_components.get('employee_pf', 0))
    if pf_opt_in == 0:
        pf_deduction = 0.0
    elif pf_manual_override == 1:
        pf_deduction = max(0.0, float(pf_deduction_input or 0))
    else:
        pf_deduction = auto_pf_deduction

    # Handle photo upload (if any)
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo and photo.filename and allowed_file(photo.filename):
            # Ensure uploads directory exists
            upload_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'static/uploads'), 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Generate filename with staff name
            ext = os.path.splitext(photo.filename)[1]
            safe_name = full_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            filename = f"{safe_name}_{staff_id}{ext}"
            photo_path = os.path.join(upload_dir, filename)
            photo.save(photo_path)
            photo_url = f"uploads/{filename}"

    try:
        # Build update query dynamically based on available columns
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        column_names = [col['name'] for col in columns]

        # Special handling for shift_type to apply from next day
        current_staff = db.execute("SELECT shift_type, next_shift_type FROM staff WHERE id = ?", (staff_db_id,)).fetchone()
        
        update_parts = ['staff_id = ?', 'full_name = ?']
        update_values = [staff_id, full_name]

        if 'first_name' in column_names:
            update_parts.append('first_name = ?')
            update_values.append(first_name)
        if 'last_name' in column_names:
            update_parts.append('last_name = ?')
            update_values.append(last_name)
        if 'date_of_birth' in column_names:
            update_parts.append('date_of_birth = ?')
            update_values.append(date_of_birth)
        if 'date_of_joining' in column_names:
            update_parts.append('date_of_joining = ?')
            update_values.append(date_of_joining)
        if 'department' in column_names:
            update_parts.append('department = ?')
            update_values.append(department)
        if 'destination' in column_names:
            update_parts.append('destination = ?')
            update_values.append(destination)
        if 'position' in column_names:
            update_parts.append('position = ?')
            update_values.append(destination)  # Use the same value for position
        if 'gender' in column_names:
            update_parts.append('gender = ?')
            update_values.append(gender)
        if 'phone' in column_names:
            update_parts.append('phone = ?')
            update_values.append(phone)
        if 'email' in column_names:
            update_parts.append('email = ?')
            update_values.append(email)
        if 'is_active' in column_names:
            update_parts.append('is_active = ?')
            update_values.append(is_active)
        if 'status' in column_names:
            update_parts.append('status = ?')
            update_values.append('active' if is_active else 'inactive')
        if 'bank_account_name' in column_names:
            update_parts.append('bank_account_name = ?')
            update_values.append(bank_account_name)
        if 'bank_name' in column_names:
            update_parts.append('bank_name = ?')
            update_values.append(bank_name)
        if 'bank_account_number' in column_names:
            update_parts.append('bank_account_number = ?')
            update_values.append(bank_account_number)
        if 'ifsc_code' in column_names:
            update_parts.append('ifsc_code = ?')
            update_values.append(ifsc_code)
        if 'pan_number' in column_names:
            update_parts.append('pan_number = ?')
            update_values.append(pan_number)
        
        # Shift type delayed logic
        if 'shift_type' in column_names:
            if current_staff and current_staff['shift_type'] != shift_type:
                # If shift is being changed, don't update current shift_type immediately
                # Instead, set next_shift_type and next_shift_effective_date
                if 'next_shift_type' in column_names:
                    update_parts.append('next_shift_type = ?')
                    update_values.append(shift_type)
                if 'next_shift_effective_date' in column_names:
                    tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                    update_parts.append('next_shift_effective_date = ?')
                    update_values.append(tomorrow)

                print(f"Shift change detected for staff {staff_db_id}: Scheduled {shift_type} for tomorrow")
            else:
                # If it's the same shift, or we don't have current data, update normally (or clear next shift)
                update_parts.append('shift_type = ?')
                update_values.append(shift_type)
                if 'next_shift_type' in column_names:
                    update_parts.append('next_shift_type = NULL')
                if 'next_shift_effective_date' in column_names:
                    update_parts.append('next_shift_effective_date = NULL')

        if 'photo_url' in column_names and photo_url:
            update_parts.append('photo_url = ?')
            update_values.append(photo_url)

        # Add salary fields
        if 'basic_salary' in column_names and basic_salary is not None:
            update_parts.append('basic_salary = ?')
            update_values.append(basic_salary)
        if 'hra' in column_names and hra is not None:
            update_parts.append('hra = ?')
            update_values.append(hra)
        if 'transport_allowance' in column_names and transport_allowance is not None:
            update_parts.append('transport_allowance = ?')
            update_values.append(transport_allowance)
        if 'other_allowances' in column_names and other_allowances is not None:
            update_parts.append('other_allowances = ?')
            update_values.append(other_allowances)
        if 'dearness_allowance' in column_names:
            update_parts.append('dearness_allowance = ?')
            update_values.append(dearness_allowance)
        if 'pf_opt_in' in column_names:
            update_parts.append('pf_opt_in = ?')
            update_values.append(pf_opt_in)
        if 'pf_deduction' in column_names:
            update_parts.append('pf_deduction = ?')
            update_values.append(pf_deduction)
        if 'esi_deduction' in column_names and esi_deduction is not None:
            update_parts.append('esi_deduction = ?')
            update_values.append(esi_deduction)
        if 'professional_tax' in column_names and professional_tax is not None:
            update_parts.append('professional_tax = ?')
            update_values.append(professional_tax)
        if 'other_deductions' in column_names and other_deductions is not None:
            update_parts.append('other_deductions = ?')
            update_values.append(other_deductions)

        # Add WHERE clause values
        update_values.extend([staff_db_id, school_id])

        # Build and execute the query
        query = f"UPDATE staff SET {', '.join(update_parts)} WHERE id = ? AND school_id = ?"

        db.execute(query, update_values)

        # Record shift history so the calendar can preserve past-day shift times
        if 'shift_type' in column_names and current_staff and current_staff['shift_type'] != shift_type:
            # Change is scheduled for tomorrow
            _tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            record_shift_history(db, staff_db_id, school_id, shift_type, _tomorrow)

        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export_staff_excel')
def export_staff_excel():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        school_id = session['school_id']
        db = get_db()

        # Get all staff with comprehensive details
        staff = db.execute('''
            SELECT staff_id, full_name, first_name, last_name,
                   date_of_birth, date_of_joining, department, destination,
                   position, gender, phone, email, shift_type, created_at
            FROM staff
            WHERE school_id = ?
            ORDER BY CAST(staff_id AS INTEGER) ASC
        ''', (school_id,)).fetchall()

        # Create proper Excel file using openpyxl
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff Details"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add title
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"Staff Details Report - Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=16, color="2F5597")
        title_cell.alignment = Alignment(horizontal='center')

        # Add headers
        headers = [
            'S.No', 'Staff ID', 'First Name', 'Last Name', 'Full Name',
            'Date of Birth', 'Date of Joining', 'Department', 'Destination/Position',
            'Gender', 'Phone Number', 'Email ID', 'Shift Type', 'Created Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Add data
        for row_idx, staff_member in enumerate(staff, 4):
            row_data = [
                row_idx - 3,  # S.No
                staff_member['staff_id'] or 'N/A',
                staff_member['first_name'] or 'N/A',
                staff_member['last_name'] or 'N/A',
                staff_member['full_name'] or 'N/A',
                staff_member['date_of_birth'] or 'N/A',
                staff_member['date_of_joining'] or 'N/A',
                staff_member['department'] or 'N/A',
                staff_member['destination'] or staff_member['position'] or 'N/A',
                staff_member['gender'] or 'N/A',
                staff_member['phone'] or 'N/A',
                staff_member['email'] or 'N/A',
                staff_member['shift_type'] or 'General',
                staff_member['created_at'] or 'N/A'
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15

        # Freeze header row
        ws.freeze_panes = 'A4'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response with proper Excel headers
        response = make_response(output.getvalue())
        filename = f'staff_details_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

        return response

    except Exception as e:
        # Log the error and return JSON error response
        print(f"Excel export error: {str(e)}")
        return jsonify({'success': False, 'error': f'Excel export failed: {str(e)}'})

    return jsonify({'success': False, 'error': 'Unknown error occurred'})

@app.route('/admin/add_department_shift', methods=['POST'])
def add_department_shift():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.form.get('department')
    shift_type = request.form.get('shift_type')

    if not department or not shift_type:
        return jsonify({'success': False, 'error': 'Department and shift type are required'})

    db = get_db()

    try:
        db.execute('''
            INSERT OR REPLACE INTO department_shift_mappings
            (school_id, department, default_shift_type, updated_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        ''', (school_id, department, shift_type))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/update_department_shift', methods=['POST'])
def update_department_shift():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.form.get('department')
    shift_type = request.form.get('shift_type')

    if not department or not shift_type:
        return jsonify({'success': False, 'error': 'Department and shift type are required'})

    db = get_db()

    try:
        # Check if mapping already exists
        existing = db.execute('''
            SELECT id FROM department_shift_mappings
            WHERE school_id = ? AND department = ?
        ''', (school_id, department)).fetchone()

        if existing:
            # Update existing mapping
            db.execute('''
                UPDATE department_shift_mappings
                SET default_shift_type = ?, updated_at = CURRENT_TIMESTAMP
                WHERE school_id = ? AND department = ?
            ''', (shift_type, school_id, department))
        else:
            # Insert new mapping
            db.execute('''
                INSERT INTO department_shift_mappings (school_id, department, default_shift_type, created_at, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ''', (school_id, department, shift_type))

        staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]
        has_next_shift = 'next_shift_type' in staff_columns and 'next_shift_effective_date' in staff_columns

        if has_next_shift:
            # Keep department shift updates consistent with enhanced staff update behavior.
            _tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')

            affected_staff_rows = db.execute('''
                SELECT id, shift_type
                FROM staff
                WHERE school_id = ? AND department = ?
                AND (
                    COALESCE(shift_type, 'general') != ?
                    OR COALESCE(next_shift_type, '') != ?
                    OR COALESCE(next_shift_effective_date, '') != ?
                )
            ''', (school_id, department, shift_type, shift_type, _tomorrow)).fetchall()

            staff_updated = db.execute('''
                UPDATE staff
                SET next_shift_type = ?,
                    next_shift_effective_date = ?
                WHERE school_id = ? AND department = ?
                AND (
                    COALESCE(shift_type, 'general') != ?
                    OR COALESCE(next_shift_type, '') != ?
                    OR COALESCE(next_shift_effective_date, '') != ?
                )
            ''', (shift_type, _tomorrow, school_id, department, shift_type, shift_type, _tomorrow))

            affected_rows = staff_updated.rowcount

            # Record history only when actual current shift differs from requested shift.
            for _row in affected_staff_rows:
                if (_row['shift_type'] or 'general') != shift_type:
                    record_shift_history(db, _row['id'], school_id, shift_type, _tomorrow)

            message = f'Department {department} scheduled to {shift_type} shift for tomorrow. {affected_rows} staff members updated.'
        else:
            # Legacy schema fallback: apply immediately if next-shift columns are unavailable.
            affected_staff_rows = db.execute('''
                SELECT id FROM staff
                WHERE school_id = ? AND department = ? AND COALESCE(shift_type, 'general') != ?
            ''', (school_id, department, shift_type)).fetchall()

            staff_updated = db.execute('''
                UPDATE staff
                SET shift_type = ?
                WHERE school_id = ? AND department = ? AND COALESCE(shift_type, 'general') != ?
            ''', (shift_type, school_id, department, shift_type))

            affected_rows = staff_updated.rowcount

            _today = datetime.datetime.now().strftime('%Y-%m-%d')
            for _row in affected_staff_rows:
                record_shift_history(db, _row['id'], school_id, shift_type, _today)

            message = f'Department {department} assigned to {shift_type} shift successfully. {affected_rows} staff members updated.'

        db.commit()
        return jsonify({'success': True, 'message': message, 'affected_staff': affected_rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete_department_shift', methods=['POST'])
def delete_department_shift():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.form.get('department')

    if not department:
        return jsonify({'success': False, 'error': 'Department is required'})

    db = get_db()

    try:
        db.execute('''
            DELETE FROM department_shift_mappings
            WHERE school_id = ? AND department = ?
        ''', (school_id, department))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/bulk_update_staff_shifts', methods=['POST'])
def bulk_update_staff_shifts():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    db = get_db()

    try:
        staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]
        has_next_shift = 'next_shift_type' in staff_columns and 'next_shift_effective_date' in staff_columns

        updated_count = 0

        if has_next_shift:
            _tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')

            # Find staff needing a new schedule and avoid re-scheduling identical tomorrow entries.
            staff_to_update = db.execute('''
                SELECT s.id, s.full_name, s.department, s.shift_type, dsm.default_shift_type,
                       s.next_shift_type, s.next_shift_effective_date
                FROM staff s
                JOIN department_shift_mappings dsm ON s.department = dsm.department AND s.school_id = dsm.school_id
                WHERE s.school_id = ? AND s.department IS NOT NULL AND s.department != ''
                AND COALESCE(s.shift_type, 'general') != dsm.default_shift_type
                AND (
                    COALESCE(s.next_shift_type, '') != dsm.default_shift_type
                    OR COALESCE(s.next_shift_effective_date, '') != ?
                )
            ''', (school_id, _tomorrow)).fetchall()

            for staff in staff_to_update:
                db.execute('''
                    UPDATE staff
                    SET next_shift_type = ?, next_shift_effective_date = ?
                    WHERE id = ?
                ''', (staff['default_shift_type'], _tomorrow, staff['id']))
                record_shift_history(db, staff['id'], school_id, staff['default_shift_type'], _tomorrow)
                updated_count += 1
        else:
            # Legacy schema fallback: immediate update and history logging.
            staff_to_update = db.execute('''
                SELECT s.id, s.full_name, s.department, s.shift_type, dsm.default_shift_type
                FROM staff s
                JOIN department_shift_mappings dsm ON s.department = dsm.department AND s.school_id = dsm.school_id
                WHERE s.school_id = ? AND s.department IS NOT NULL AND s.department != ''
                AND COALESCE(s.shift_type, 'general') != dsm.default_shift_type
            ''', (school_id,)).fetchall()

            _today = datetime.datetime.now().strftime('%Y-%m-%d')
            for staff in staff_to_update:
                db.execute('''
                    UPDATE staff
                    SET shift_type = ?
                    WHERE id = ?
                ''', (staff['default_shift_type'], staff['id']))
                record_shift_history(db, staff['id'], school_id, staff['default_shift_type'], _today)
                updated_count += 1

        db.commit()
        return jsonify({'success': True, 'updated_count': updated_count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/preview_staff_shift_changes')
def preview_staff_shift_changes():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    db = get_db()

    try:
        # Get all staff with departments that have shift mappings and would be changed
        changes = db.execute('''
            SELECT s.full_name as staff_name, s.department,
                   COALESCE(s.shift_type, 'general') as current_shift,
                   dsm.default_shift_type as new_shift
            FROM staff s
            JOIN department_shift_mappings dsm ON s.department = dsm.department AND s.school_id = dsm.school_id
            WHERE s.school_id = ? AND s.department IS NOT NULL AND s.department != ''
            AND (s.shift_type IS NULL OR s.shift_type != dsm.default_shift_type)
            ORDER BY CAST(s.staff_id AS INTEGER) ASC
        ''', (school_id,)).fetchall()

        changes_list = [dict(change) for change in changes]
        return jsonify({'success': True, 'changes': changes_list})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update_staff', methods=['POST'])
def update_staff():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    full_name = request.form.get('full_name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    department = request.form.get('department')
    position = request.form.get('position')
    shift_type = request.form.get('shift_type', 'general')
    status = request.form.get('status')
    is_active_raw = request.form.get('is_active')
    is_active = 1
    if is_active_raw is not None:
        is_active = 1 if str(is_active_raw).lower() in ['1', 'true', 'on', 'yes', 'active'] else 0
    elif status is not None:
        is_active = 1 if str(status).lower() == 'active' else 0
    school_id = session['school_id']

    if not staff_id or not full_name:
        return jsonify({'success': False, 'error': 'Staff ID and full name are required'})

    db = get_db()

    # Handle photo upload
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo.filename != '' and allowed_file(photo.filename):
            try:
                # Ensure uploads directory exists
                upload_dir = os.path.join(app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)

                # Generate filename with staff name
                ext = os.path.splitext(photo.filename)[1]
                safe_name = full_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
                filename = f"{safe_name}_{staff_id}{ext}"
                photo_url = os.path.join('uploads', filename)
                photo.save(os.path.join(app.static_folder, photo_url))
            except Exception as e:
                print(f"Error saving photo: {e}")
                return jsonify({'success': False, 'error': 'Error saving photo'})
        elif photo.filename != '':
            return jsonify({'success': False, 'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'})

    try:
        staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]

        update_parts = [
            'full_name = ?',
            'email = ?',
            'phone = ?',
            'department = ?',
            'position = ?',
            'destination = ?',
            'shift_type = ?'
        ]
        update_values = [full_name, email, phone, department, position, position, shift_type]

        if photo_url:
            update_parts.append('photo_url = ?')
            update_values.append(photo_url)
        if 'is_active' in staff_columns:
            update_parts.append('is_active = ?')
            update_values.append(is_active)
        if 'status' in staff_columns:
            update_parts.append('status = ?')
            update_values.append('active' if is_active else 'inactive')

        update_values.extend([staff_id, school_id])
        db.execute(
            f"UPDATE staff SET {', '.join(update_parts)} WHERE id = ? AND school_id = ?",
            update_values
        )

        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_staff', methods=['POST'])
def delete_staff():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    school_id = session['school_id']

    db = get_db()

    # Get staff record to retrieve biometric user ID
    staff = db.execute('SELECT staff_id FROM staff WHERE id = ? AND school_id = ?', (staff_id, school_id)).fetchone()
    biometric_deleted = False
    biometric_error = None
    if staff:
        device_ip, device_port = get_institution_device()
        if not device_ip:
            biometric_error = 'No biometric device configured'
        else:
            try:
                zk_device = ZKBiometricDevice(device_ip)
                if zk_device.connect():
                    biometric_deleted = zk_device.delete_user(staff['staff_id'])
                    zk_device.disconnect()
                else:
                    biometric_error = 'Failed to connect to biometric device'
            except Exception as e:
                biometric_error = str(e)

    db.execute('DELETE FROM staff WHERE id = ? AND school_id = ?', (staff_id, school_id))
    db.commit()

    return jsonify({'success': True, 'biometric_deleted': biometric_deleted, 'biometric_error': biometric_error})

@app.route('/reset_staff_password', methods=['POST'])
def reset_staff_password():
    """Reset staff password (Admin only)"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_db_id = request.form.get('staff_id')  # This is the database ID
    new_password = request.form.get('new_password')
    school_id = session['school_id']

    if not staff_db_id:
        return jsonify({'success': False, 'error': 'Staff ID is required'})

    if not new_password:
        return jsonify({'success': False, 'error': 'New password is required'})

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters long'})

    db = get_db()

    try:
        # Verify staff exists in the same school
        staff = db.execute('''
            SELECT staff_id, full_name FROM staff
            WHERE id = ? AND school_id = ?
        ''', (staff_db_id, school_id)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff not found'})

        # Generate new password hash
        password_hash = generate_password_hash(new_password)

        # Update staff password
        db.execute('''
            UPDATE staff
            SET password_hash = ?
            WHERE id = ? AND school_id = ?
        ''', (password_hash, staff_db_id, school_id))

        db.commit()

        return jsonify({
            'success': True,
            'message': f'Password reset successfully for {staff["full_name"]} (Staff ID: {staff["staff_id"]})'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/set_default_passwords', methods=['POST'])
def set_default_passwords():
    """Set default passwords for all staff without passwords (for testing)"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    default_password = 'password123'

    db = get_db()

    try:
        # Find staff without passwords
        staff_without_passwords = db.execute('''
            SELECT id, staff_id, full_name FROM staff
            WHERE school_id = ? AND (password_hash IS NULL OR password_hash = '')
        ''', (school_id,)).fetchall()

        if not staff_without_passwords:
            return jsonify({'success': True, 'message': 'All staff already have passwords set'})

        # Generate password hash
        password_hash = generate_password_hash(default_password)

        # Update all staff without passwords
        updated_count = 0
        for staff in staff_without_passwords:
            db.execute('''
                UPDATE staff SET password_hash = ? WHERE id = ?
            ''', (password_hash, staff['id']))
            updated_count += 1

        db.commit()

        return jsonify({
            'success': True,
            'message': f'Set default password for {updated_count} staff members. Default password: {default_password}',
            'updated_count': updated_count
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/change_password', methods=['GET', 'POST'])
def staff_change_password():
    """Allow staff to change their own password"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    if request.method == 'GET':
        return render_template('staff_change_password.html')

    # Handle POST request
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if not all([current_password, new_password, confirm_password]):
        return jsonify({'success': False, 'error': 'All fields are required'})

    if new_password != confirm_password:
        return jsonify({'success': False, 'error': 'New passwords do not match'})

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters long'})

    db = get_db()
    staff_id = session['user_id']

    try:
        # Get current staff record
        staff = db.execute('SELECT * FROM staff WHERE id = ?', (staff_id,)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff record not found'})

        # Verify current password
        if not check_password_hash(staff['password_hash'], current_password):
            return jsonify({'success': False, 'error': 'Current password is incorrect'})

        # Update password
        new_password_hash = generate_password_hash(new_password)
        db.execute('UPDATE staff SET password_hash = ? WHERE id = ?', (new_password_hash, staff_id))
        db.commit()

        return jsonify({'success': True, 'message': 'Password changed successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add_school', methods=['POST'])
def add_school():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    name = request.form.get('name')
    address = request.form.get('address')
    contact_email = request.form.get('contact_email')
    contact_phone = request.form.get('contact_phone')

    admin_username = request.form.get('admin_username')
    admin_password = generate_password_hash(request.form.get('admin_password'))
    admin_full_name = request.form.get('admin_full_name')
    admin_email = request.form.get('admin_email')

    db = get_db()

    # Handle logo upload
    logo_path = None
    if 'logo' in request.files:
        logo = request.files['logo']
        if logo.filename != '' and allowed_file(logo.filename):
            try:
                upload_dir = os.path.join(app.static_folder, 'uploads', 'logos')
                os.makedirs(upload_dir, exist_ok=True)

                ext = os.path.splitext(logo.filename)[1]
                safe_name = secure_filename(name).replace(' ', '_').lower()
                filename = f"{safe_name}_logo_{int(time.time())}{ext}"
                logo_path = f"static/uploads/logos/{filename}"
                logo.save(os.path.join(upload_dir, filename))
            except Exception as e:
                print(f"Error saving logo: {e}")
                return jsonify({'success': False, 'error': 'Failed to save school logo'})
        elif logo.filename != '':
            return jsonify({'success': False, 'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'})

    try:
        cursor = db.cursor()

        # Reserve the next school id as MAX(id) + 1 so the new row never gets 0.
        cursor.execute('SELECT COALESCE(MAX(id), 0) + 1 AS next_school_id FROM schools')
        next_school_id = cursor.fetchone()[0]
        try:
            next_school_id = int(next_school_id)
        except (TypeError, ValueError):
            next_school_id = 1
        if next_school_id <= 0:
            next_school_id = 1

        # Add school
        cursor.execute('''
            INSERT INTO schools (id, name, address, contact_email, contact_phone, logo_path, branding_enabled)
            VALUES (?, ?, ?, ?, ?, ?, 1)
        ''', (next_school_id, name, address, contact_email, contact_phone, logo_path))
        school_id = next_school_id

        # Add initial admin for the school
        cursor.execute('''
            INSERT INTO admins (school_id, username, password, full_name, email)
            VALUES (?, ?, ?, ?, ?)
        ''', (school_id, admin_username, admin_password, admin_full_name, admin_email))

        db.commit()
        return jsonify({'success': True})

    except sqlite3.IntegrityError:
        db.rollback()
        return jsonify({'success': False, 'error': 'School or admin username already exists'})

    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_school/<int:school_id>', methods=['GET'])
def get_school(school_id):
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    db = get_db()
    school = db.execute('SELECT * FROM schools WHERE id = ?', (school_id,)).fetchone()
    
    if not school:
        return jsonify({'success': False, 'error': 'School not found'})
    
    return jsonify({
        'success': True,
        'school': {
            'id': school['id'],
            'name': school['name'],
            'address': school['address'],
            'contact_email': school['contact_email'],
            'contact_phone': school['contact_phone'],
            'logo_path': school['logo_path'],
            'branding_enabled': school['branding_enabled']
        }
    })

@app.route('/edit_school/<int:school_id>', methods=['POST'])
def edit_school(school_id):
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    name = request.form.get('name')
    address = request.form.get('address')
    contact_email = request.form.get('contact_email')
    contact_phone = request.form.get('contact_phone')
    branding_enabled = request.form.get('branding_enabled', '1')

    db = get_db()
    
    # Get current school data
    school = db.execute('SELECT logo_path FROM schools WHERE id = ?', (school_id,)).fetchone()
    if not school:
        return jsonify({'success': False, 'error': 'School not found'})
    
    logo_path = school['logo_path']

    # Handle logo upload
    if 'logo' in request.files:
        logo = request.files['logo']
        if logo.filename != '' and allowed_file(logo.filename):
            try:
                upload_dir = os.path.join(app.static_folder, 'uploads', 'logos')
                os.makedirs(upload_dir, exist_ok=True)

                # Delete old logo if exists
                if logo_path:
                    old_logo_full_path = os.path.join(os.getcwd(), logo_path)
                    if os.path.exists(old_logo_full_path):
                        os.remove(old_logo_full_path)

                ext = os.path.splitext(logo.filename)[1]
                safe_name = secure_filename(name).replace(' ', '_').lower()
                filename = f"{safe_name}_logo_{int(time.time())}{ext}"
                logo_path = f"static/uploads/logos/{filename}"
                logo.save(os.path.join(upload_dir, filename))
            except Exception as e:
                print(f"Error saving logo: {e}")
                return jsonify({'success': False, 'error': 'Failed to save school logo'})
        elif logo.filename != '':
            return jsonify({'success': False, 'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'})

    try:
        cursor = db.cursor()
        cursor.execute('''
            UPDATE schools 
            SET name = ?, address = ?, contact_email = ?, contact_phone = ?, 
                logo_path = ?, branding_enabled = ?
            WHERE id = ?
        ''', (name, address, contact_email, contact_phone, logo_path, branding_enabled, school_id))
        
        db.commit()
        return jsonify({'success': True, 'message': 'School updated successfully'})

    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/staff/<int:id>')
def staff_profile(id):
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    db = get_db()
    staff = db.execute('SELECT * FROM staff WHERE id = ? AND school_id = ?',
                      (id, session['school_id'])).fetchone()

    if not staff:
        return redirect(url_for('admin_dashboard'))

    # Get attendance summary for this staff member
    attendance_summary = db.execute('''
        SELECT
            COUNT(CASE WHEN status = 'present' THEN 1 END) as present_count,
            COUNT(CASE WHEN status = 'late' THEN 1 END) as late_count,
            COUNT(CASE WHEN status = 'absent' THEN 1 END) as absent_count,
            COUNT(CASE WHEN status = 'leave' THEN 1 END) as leave_count
        FROM attendance
        WHERE staff_id = ? AND date >= date('now', '-30 days')
    ''', (id,)).fetchone()



    # Get recent biometric verifications (latest per day per type)
    recent_verifications = db.execute('''
        SELECT verification_type, verification_time, biometric_method, verification_status
        FROM biometric_verifications bv1
        WHERE staff_id = ?
          AND verification_time = (
            SELECT MAX(verification_time)
            FROM biometric_verifications bv2
            WHERE bv2.staff_id = bv1.staff_id
              AND bv2.verification_type = bv1.verification_type
              AND DATE(bv2.verification_time) = DATE(bv1.verification_time)
          )
        ORDER BY verification_time DESC
        LIMIT 20
    ''', (id,)).fetchall()

    return render_template('staff_profile.html',
                         staff=staff,
                         attendance_summary=attendance_summary,
                         recent_verifications=recent_verifications)

@app.route('/admin/search_staff')
def search_staff():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    search_term = request.args.get('q', '')
    db = get_db()

    staff = db.execute('''
        SELECT id, staff_id, full_name, department, position
        FROM staff
        WHERE school_id = ? AND full_name LIKE ?
        ORDER BY CAST(staff_id AS INTEGER) ASC
    ''', (session['school_id'], f"%{search_term}%")).fetchall()

    # Get pending leave applications
    pending_leaves = db.execute('''
        SELECT l.id, s.full_name, l.leave_type, l.start_date, l.end_date, l.reason
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ? AND l.status = 'pending'
        ORDER BY l.applied_at
    ''', (session['school_id'],)).fetchall()

    # Get today's attendance summary
    today = datetime.date.today()
    attendance_mode = get_school_attendance_mode(session['school_id'])
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, session['school_id'])).fetchone()

    return render_template('admin_dashboard.html',
                         staff=staff,
                         pending_leaves=pending_leaves,
                         attendance_summary=attendance_summary,
                         attendance_mode=attendance_mode,
                         today=today)

# Update in app.py
@app.route('/toggle_school_visibility', methods=['POST'])
def toggle_school_visibility():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = request.form.get('school_id')
    db = get_db()

    # Ensure column exists
    columns = db.execute("PRAGMA table_info(schools)").fetchall()
    has_is_hidden = any(col['name'] == 'is_hidden' for col in columns)

    if not has_is_hidden:
        db.execute('ALTER TABLE schools ADD COLUMN is_hidden BOOLEAN DEFAULT 0')
        db.commit()

    # Toggle visibility
    db.execute('''
        UPDATE schools
        SET is_hidden = CASE WHEN is_hidden = 1 THEN 0 ELSE 1 END
        WHERE id = ?
    ''', (school_id,))
    db.commit()

    return jsonify({'success': True})




# ZK Biometric Device Integration Routes
@app.route('/sync_biometric_attendance', methods=['POST'])
def sync_biometric_attendance():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    
    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured for your institution'})
    
    auto_create_missing_staff = request.form.get('auto_create_staff', 'true').lower() == 'true'

    try:
        # Use the updated automatic processing function that includes check-out sync
        result = process_device_attendance_automatically(device_ip, school_id)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'synced_count': result.get('processed_count', 0),
                'auto_checkout_enabled': True
            })
        else:
            return jsonify({'success': False, 'error': result['message']})
            
        # Fallback to old method if needed
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        # Use the correct method name
        attendance_records = zk_device.get_attendance_records()
        if not attendance_records:
            zk_device.disconnect()
            return jsonify({'success': False, 'error': 'No attendance records found'})

        db = get_db()
        synced_count = 0
        created_staff_count = 0
        
        # Get all device users for auto-creation if needed
        device_users = {}
        if auto_create_missing_staff:
            all_users = zk_device.get_users()
            for user in all_users:
                # Handle both dict and object formats
                if isinstance(user, dict):
                    user_id = str(user.get('user_id', ''))
                    name = user.get('name', '')
                else:
                    user_id = str(getattr(user, 'user_id', ''))
                    name = getattr(user, 'name', '')
                
                if user_id:
                    device_users[user_id] = name

        for record in attendance_records:
            try:
                record_user_id = str(record['user_id'])
                
                # Get staff database ID from staff_id (biometric user_id)
                staff_record = db.execute('''
                    SELECT id FROM staff WHERE staff_id = ? AND school_id = ?
                ''', (record_user_id, school_id)).fetchone()

                if not staff_record:
                    if auto_create_missing_staff and record_user_id in device_users:
                        # Auto-create staff member
                        device_name = device_users[record_user_id]
                        display_name = device_name.strip() if device_name else f"Device User {record_user_id}"
                        
                        # Parse name into first/last components
                        first_name, last_name = parse_full_name(display_name)
                        
                        # Create staff record
                        db.execute('''
                            INSERT INTO staff 
                            (school_id, staff_id, password_hash, full_name, first_name, last_name, 
                             department, position, email, phone)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (school_id, record_user_id, generate_password_hash('password123'), 
                              display_name, first_name, last_name, 'General', 'Staff Member', '', ''))
                        
                        # Get the newly created staff ID
                        staff_record = db.execute('''
                            SELECT id FROM staff WHERE staff_id = ? AND school_id = ?
                        ''', (record_user_id, school_id)).fetchone()
                        
                        created_staff_count += 1
                        print(f"Auto-created staff: {record_user_id} ({display_name})")
                    else:
                        print(f"Staff with ID {record_user_id} not found in database - skipping attendance record")
                        continue

                staff_db_id = staff_record['id']
                attendance_date = record['timestamp'].date()
                attendance_time = record['timestamp'].strftime('%H:%M:%S')
                verification_type = record.get('verification_type', 'check-in')

                # Check if attendance record exists for this date
                existing_record = db.execute('''
                    SELECT * FROM attendance WHERE staff_id = ? AND date = ?
                ''', (staff_db_id, attendance_date)).fetchone()

                if existing_record:
                    # Update existing record based on verification type
                    if verification_type == 'check-in' and not existing_record['time_in']:
                        db.execute('''
                            UPDATE attendance SET time_in = ?, status = 'present'
                            WHERE staff_id = ? AND date = ?
                        ''', (attendance_time, staff_db_id, attendance_date))
                        synced_count += 1
                    elif verification_type == 'check-out' and not existing_record['time_out']:
                        db.execute('''
                            UPDATE attendance SET time_out = ?
                            WHERE staff_id = ? AND date = ?
                        ''', (attendance_time, staff_db_id, attendance_date))
                        synced_count += 1
                else:
                    # Create new attendance record
                    if verification_type == 'check-in':
                        db.execute('''
                            INSERT INTO attendance (staff_id, school_id, date, time_in, status)
                            VALUES (?, ?, ?, ?, 'present')
                        ''', (staff_db_id, school_id, attendance_date, attendance_time))
                        synced_count += 1

            except Exception as record_error:
                print(f"Error processing record {record}: {record_error}")
                continue

        db.commit()
        zk_device.disconnect()

        message = f'Successfully synced {synced_count} attendance records'
        if created_staff_count > 0:
            message += f' and created {created_staff_count} new staff members'

        return jsonify({
            'success': True,
            'message': message,
            'synced_count': synced_count,
            'created_staff_count': created_staff_count,
            'total_records': len(attendance_records)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'Sync failed: {str(e)}'})

@app.route('/process_device_attendance', methods=['POST'])
def process_device_attendance():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    
    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        process_device_attendance_automatically(zk_device, school_id, db=get_db())
        return jsonify({'success': True, 'message': 'Device attendance processed successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/verify_staff_biometric', methods=['POST'])
def verify_staff_biometric_route():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    
    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        is_valid = verify_staff_biometric(zk_device, staff_id)
        return jsonify({'success': True, 'is_valid': is_valid})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Duplicate route removed - already defined above

@app.route('/apply_on_duty_permission', methods=['POST'])
def apply_on_duty_permission():
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    permission_type = request.form.get('permission_type', 'On Duty')
    start_datetime = request.form.get('start_datetime')
    end_datetime = request.form.get('end_datetime')
    reason = request.form.get('reason')

    if not start_datetime or not end_datetime:
        return jsonify({'success': False, 'error': 'Start and end datetime required'})

    db = get_db()
    try:
        db.execute('''
            INSERT INTO on_duty_permissions
            (staff_id, school_id, permission_type, start_datetime, end_datetime, reason)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (staff_id, school_id, permission_type, start_datetime, end_datetime, reason))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Duplicate route removed - already defined above

@app.route('/test_biometric_connection', methods=['POST'])
def test_biometric_connection():
    """Test connection to ZK biometric device (Ethernet or Cloud)"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution or allow override for testing
    device_ip_override = request.form.get('device_ip')
    if device_ip_override:
        device_ip = device_ip_override
    else:
        device_ip, device_port = get_institution_device()
        if not device_ip:
            return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    device_id = request.form.get('device_id', f"ZK_{device_ip.replace('.', '_')}")
    use_cloud = request.form.get('use_cloud', 'false').lower() == 'true'

    try:
        # Determine connection mode
        if CLOUD_ENABLED and use_cloud:
            # Test cloud connection
            device_config = get_device_config(device_id)
            if not device_config:
                return jsonify({
                    'success': False,
                    'message': f'Device {device_id} not configured for cloud connectivity',
                    'device_id': device_id
                })

            connector = get_cloud_connector()
            if not connector.running:
                return jsonify({
                    'success': False,
                    'message': 'Cloud connector is not running',
                    'device_id': device_id
                })

            status = connector.get_device_status(device_id)
            if status['status'] == 'connected':
                return jsonify({
                    'success': True,
                    'message': 'Cloud connection successful',
                    'device_id': device_id,
                    'device_ip': device_config.local_ip,
                    'total_users': status.get('user_count', 0),
                    'connection_type': 'cloud'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': f'Device status: {status["status"]}',
                    'device_id': device_id,
                    'connection_type': 'cloud'
                })
        else:
            # Test Ethernet connection
            port = 4370 if device_ip == '192.168.1.201' else 32150
            zk_device = ZKBiometricDevice(device_ip, port=port, timeout=15, device_id='1', use_cloud=False)
            if zk_device.connect():
                # Get device info
                users = zk_device.get_users()
                zk_device.disconnect()

                return jsonify({
                    'success': True,
                    'message': 'Ethernet connection successful',
                    'device_ip': device_ip,
                    'total_users': len(users),
                    'connection_type': 'ethernet'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to connect to device via Ethernet',
                    'device_ip': device_ip,
                    'connection_type': 'ethernet'
                })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Connection test failed: {str(e)}',
            'device_ip': device_ip,
            'device_id': device_id
        })

@app.route('/cloud_config', methods=['GET', 'POST'])
def cloud_config():
    """Manage cloud configuration"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    if not CLOUD_ENABLED:
        return jsonify({'success': False, 'error': 'Cloud functionality not available'})

    if request.method == 'GET':
        try:
            config = get_cloud_config()
            connector = get_cloud_connector()

            # Return safe configuration (no sensitive data)
            safe_config = {
                'cloud_provider': config.cloud_provider,
                'api_base_url': config.api_base_url,
                'websocket_url': config.websocket_url,
                'mqtt_broker': config.mqtt_broker,
                'mqtt_port': config.mqtt_port,
                'organization_id': config.organization_id,
                'connection_timeout': config.connection_timeout,
                'retry_attempts': config.retry_attempts,
                'heartbeat_interval': config.heartbeat_interval,
                'use_ssl': config.use_ssl,
                'verify_ssl': config.verify_ssl,
                'encryption_enabled': config.encryption_enabled,
                'auto_sync': config.auto_sync,
                'sync_interval': config.sync_interval,
                'batch_size': config.batch_size,
                'local_backup': config.local_backup,
                'backup_retention_days': config.backup_retention_days,
                'connector_running': connector.running if connector else False,
                'has_api_key': bool(config.api_key),
                'has_secret_key': bool(config.secret_key)
            }

            return jsonify({
                'success': True,
                'config': safe_config
            })

        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    elif request.method == 'POST':
        try:
            # Update configuration
            data = request.get_json()
            if not data:
                return jsonify({'success': False, 'error': 'No data provided'})

            from cloud_config import config_manager

            # Update configuration fields
            config = config_manager.config

            # Update safe fields only
            safe_fields = [
                'cloud_provider', 'api_base_url', 'websocket_url', 'mqtt_broker', 'mqtt_port',
                'organization_id', 'connection_timeout', 'retry_attempts', 'heartbeat_interval',
                'use_ssl', 'verify_ssl', 'encryption_enabled', 'auto_sync', 'sync_interval',
                'batch_size', 'local_backup', 'backup_retention_days'
            ]

            for field in safe_fields:
                if field in data:
                    setattr(config, field, data[field])

            # Handle sensitive fields separately
            if 'api_key' in data and data['api_key']:
                config.api_key = data['api_key']

            if 'secret_key' in data and data['secret_key']:
                config.secret_key = data['secret_key']

            # Save configuration
            config_manager.save_config()

            return jsonify({
                'success': True,
                'message': 'Configuration updated successfully'
            })

        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

@app.route('/cloud_status', methods=['GET'])
def cloud_status():
    """Get cloud connector status"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    if not CLOUD_ENABLED:
        return jsonify({
            'success': True,
            'cloud_enabled': False,
            'message': 'Cloud functionality not available'
        })

    try:
        connector = get_cloud_connector()
        config = get_cloud_config()

        from cloud_config import get_all_devices
        devices = get_all_devices()

        # Get device statuses
        device_statuses = []
        for device in devices:
            if device.cloud_enabled:
                status = connector.get_device_status(device.device_id)
                device_statuses.append({
                    'device_id': device.device_id,
                    'device_name': device.device_name,
                    'local_ip': device.local_ip,
                    'status': status['status'],
                    'last_sync': device.last_sync,
                    'user_count': status.get('user_count', 0)
                })

        return jsonify({
            'success': True,
            'cloud_enabled': True,
            'connector_running': connector.running,
            'websocket_connected': connector.websocket is not None and
                                 connector.websocket.sock and
                                 connector.websocket.sock.connected,
            'last_heartbeat': connector.last_heartbeat.isoformat() if connector.last_heartbeat else None,
            'message_queue_size': len(connector.message_queue),
            'device_count': len(device_statuses),
            'devices': device_statuses,
            'config_valid': bool(config.api_key and config.api_base_url and config.organization_id)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/cloud_dashboard')
def cloud_dashboard():
    """Cloud monitoring dashboard"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('login'))

    if not CLOUD_ENABLED:
        return render_template('error.html',
                             error_message="Cloud functionality not available",
                             error_details="Please install cloud dependencies to access this feature.")

    return render_template('cloud_dashboard.html')

@app.route('/get_biometric_users', methods=['GET'])
def get_biometric_users():
    """Get users from ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured for your institution'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            users = zk_device.get_users()
            zk_device.disconnect()

            return jsonify({
                'success': True,
                'users': users,
                'total_users': len(users)
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get users: {str(e)}'
        })

@app.route('/enroll_biometric_user', methods=['POST'])
def enroll_biometric_user():
    """Enroll a user in the ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    user_id = request.form.get('user_id')
    name = request.form.get('name')
    privilege = int(request.form.get('privilege', 0))
    overwrite = request.form.get('overwrite', 'false').lower() == 'true'

    if not user_id or not name:
        return jsonify({'success': False, 'message': 'User ID and name are required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            # First check if user already exists
            users = zk_device.get_users()
            existing_user = None
            for user in users:
                # Handle both dict and object formats
                user_id_value = user.get('user_id') if isinstance(user, dict) else getattr(user, 'user_id', None)
                if str(user_id_value) == str(user_id):
                    existing_user = user
                    break

            if existing_user and not overwrite:
                zk_device.disconnect()
                return jsonify({
                    'success': False,
                    'message': f'User ID {user_id} already exists on biometric device',
                    'user_exists': True,
                    'existing_user': {
                        'user_id': existing_user.get('user_id') if isinstance(existing_user, dict) else getattr(existing_user, 'user_id', 'Unknown'),
                        'name': existing_user.get('name') if isinstance(existing_user, dict) else getattr(existing_user, 'name', 'Unknown'),
                        'privilege': existing_user.get('privilege') if isinstance(existing_user, dict) else getattr(existing_user, 'privilege', 0)
                    },
                    'suggestion': 'You can either:\n1. Use a different User ID\n2. Enable "Overwrite existing user" option\n3. Create staff account for the existing biometric user'
                })

            result = zk_device.enroll_user(user_id, name, privilege, overwrite=overwrite)
            zk_device.disconnect()

            if result['success']:
                return jsonify({
                    'success': True,
                    'message': result['message'],
                    'action': result.get('action', 'enrolled'),
                    'user_exists': result.get('user_exists', False)
                })
            else:
                return jsonify({
                    'success': False,
                    'message': result['message'],
                    'user_exists': result.get('user_exists', False),
                    'existing_user': result.get('existing_user')
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Enrollment failed: {str(e)}'
        })

@app.route('/check_biometric_user', methods=['POST'])
def check_biometric_user():
    """Check if a user already exists on the ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    user_id = request.form.get('user_id')

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            users = zk_device.get_users()
            zk_device.disconnect()

            # Check if user exists
            for user in users:
                if user['user_id'] == user_id:
                    return jsonify({
                        'success': True,
                        'user_exists': True,
                        'user_data': {
                            'user_id': user['user_id'],
                            'name': user['name'],
                            'privilege': user['privilege']
                        }
                    })

            return jsonify({
                'success': True,
                'user_exists': False,
                'message': f'User {user_id} not found on device'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Check failed: {str(e)}'
        })

@app.route('/start_biometric_enrollment', methods=['POST'])
def start_biometric_enrollment():
    """Start biometric enrollment mode on device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured for your institution'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            success = zk_device.start_enrollment_mode()
            # Don't disconnect here - keep connection for enrollment

            if success:
                return jsonify({
                    'success': True,
                    'message': 'Device ready for biometric enrollment'
                })
            else:
                zk_device.disconnect()
                return jsonify({
                    'success': False,
                    'message': 'Failed to start enrollment mode'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to start enrollment: {str(e)}'
        })

@app.route('/end_biometric_enrollment', methods=['POST'])
def end_biometric_enrollment():
    """End biometric enrollment mode on device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            success = zk_device.end_enrollment_mode()
            zk_device.disconnect()

            if success:
                return jsonify({
                    'success': True,
                    'message': 'Enrollment mode ended'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to end enrollment mode'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to end enrollment: {str(e)}'
        })

@app.route('/verify_biometric_enrollment', methods=['POST'])
def verify_biometric_enrollment():
    """Verify that biometric data was captured for a user"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    user_id = request.form.get('user_id')
    trigger_enrollment = request.form.get('trigger_enrollment', 'false').lower() == 'true'

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            # First check if user exists
            users = zk_device.get_users()
            user_exists = False
            user_data = None

            for user in users:
                if user['user_id'] == user_id:
                    user_exists = True
                    user_data = user
                    break

            # If user doesn't exist or we need to trigger enrollment
            if trigger_enrollment and user_exists:
                # Trigger biometric enrollment for the user
                result = zk_device.trigger_biometric_enrollment(user_id)

                if result['success']:
                    return jsonify({
                        'success': True,
                        'enrolled': False,
                        'enrollment_started': True,
                        'manual_mode': result.get('manual_mode', True),
                        'message': result['message']
                    })
                else:
                    zk_device.disconnect()
                    return jsonify({
                        'success': False,
                        'message': result['message']
                    })

            # Check if user exists and has biometric data
            if user_exists:
                zk_device.disconnect()
                return jsonify({
                    'success': True,
                    'enrolled': True,
                    'user_data': user_data,
                    'message': f'User {user_id} biometric data verified'
                })
            else:
                # User not found, try to create and enroll
                if trigger_enrollment:
                    # First create the user
                    name = request.form.get('name', f'User {user_id}')
                    privilege = int(request.form.get('privilege', 0))

                    # Create user first
                    enroll_result = zk_device.enroll_user(user_id, name, privilege)

                    if enroll_result['success']:
                        # Now trigger biometric enrollment
                        result = zk_device.trigger_biometric_enrollment(user_id)

                        if result['success']:
                            return jsonify({
                                'success': True,
                                'enrolled': False,
                                'user_created': True,
                                'enrollment_started': True,
                                'manual_mode': result.get('manual_mode', True),
                                'message': f'User created and {result["message"]}'
                            })

                    zk_device.disconnect()
                    return jsonify({
                        'success': False,
                        'message': f'Failed to create user: {enroll_result.get("message", "Unknown error")}'
                    })

                zk_device.disconnect()
                return jsonify({
                    'success': True,
                    'enrolled': False,
                    'message': f'User {user_id} not found or biometric data not captured'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Verification failed: {str(e)}'
        })

@app.route('/delete_biometric_user', methods=['POST'])
def delete_biometric_user():
    """Delete a user from the ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    user_id = request.form.get('user_id')

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            success = zk_device.delete_user(user_id)
            zk_device.disconnect()

            if success:
                return jsonify({
                    'success': True,
                    'message': f'User {user_id} deleted successfully from biometric device'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': f'User {user_id} not found on device or deletion failed'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to biometric device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Deletion failed: {str(e)}'
        })

@app.route('/resolve_biometric_conflict', methods=['POST'])
def resolve_biometric_conflict():
    """Resolve biometric user conflicts with multiple resolution options"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    user_id = request.form.get('user_id')
    action = request.form.get('action')  # 'overwrite', 'delete', 'check'
    new_name = request.form.get('new_name', '')

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'message': 'Failed to connect to biometric device'})

        if action == 'check':
            # Just check if user exists and return details
            users = zk_device.get_users()
            for user in users:
                if str(user['user_id']) == str(user_id):
                    zk_device.disconnect()
                    return jsonify({
                        'success': True,
                        'user_exists': True,
                        'user_data': {
                            'user_id': user['user_id'],
                            'name': user['name'],
                            'privilege': user['privilege'],
                            'uid': user['uid']
                        }
                    })

            zk_device.disconnect()
            return jsonify({
                'success': True,
                'user_exists': False,
                'message': f'User {user_id} not found on device'
            })

        elif action == 'delete':
            # Delete the existing user
            success = zk_device.delete_user(user_id)
            zk_device.disconnect()

            return jsonify({
                'success': success,
                'message': f'User {user_id} {"deleted successfully" if success else "deletion failed"}'
            })

        elif action == 'overwrite':
            # Overwrite the existing user
            if not new_name:
                zk_device.disconnect()
                return jsonify({'success': False, 'message': 'New name is required for overwrite'})

            result = zk_device.enroll_user(user_id, new_name, overwrite=True)
            zk_device.disconnect()

            return jsonify(result)

        else:
            zk_device.disconnect()
            return jsonify({'success': False, 'message': 'Invalid action specified'})

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Conflict resolution failed: {str(e)}'
        })

@app.route('/poll_device_attendance', methods=['POST'])
def poll_device_attendance():
    """Poll ZK device for new attendance records and process them automatically"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    school_id = session.get('school_id', 1)

    try:
        result = process_device_attendance_automatically(device_ip, school_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Polling failed: {str(e)}',
            'processed_count': 0
        })

@app.route('/get_latest_device_verifications')
def get_latest_device_verifications():
    """Get the latest biometric verifications from the device for real-time updates"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured for your institution'})
    since_minutes = int(request.args.get('since_minutes', 5))  # Default to last 5 minutes

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({
                'success': False,
                'error': 'Failed to connect to biometric device',
                'verifications': []
            })

        # Get records from the last few minutes
        since_timestamp = datetime.datetime.now() - datetime.timedelta(minutes=since_minutes)
        recent_records = zk_device.get_new_attendance_records(since_timestamp)

        zk_device.disconnect()

        # Format the records for the frontend
        verifications = []
        for record in recent_records:
            verifications.append({
                'user_id': record['user_id'],
                'verification_type': record['verification_type'],
                'timestamp': record['timestamp'].strftime('%Y-%m-%d %I:%M %p'),
                'time_only': record['timestamp'].strftime('%I:%M %p')
            })

        return jsonify({
            'success': True,
            'verifications': verifications,
            'count': len(verifications)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to get latest verifications: {str(e)}',
            'verifications': []
        })

def serialize_row(row):
    """Convert SQLite row to dict and handle time objects for JSON serialization"""
    if not row:
        return None
    
    result = dict(row)
    for key, value in result.items():
        # Convert time objects to string
        if isinstance(value, datetime.time):
            result[key] = value.strftime('%H:%M:%S')
        # Convert datetime objects to string
        elif isinstance(value, datetime.datetime):
            result[key] = value.strftime('%Y-%m-%d %H:%M:%S')
        # Convert date objects to string
        elif isinstance(value, datetime.date):
            result[key] = value.strftime('%Y-%m-%d')
    return result

@app.route('/get_comprehensive_staff_profile')
def get_comprehensive_staff_profile():
    """Get comprehensive staff profile data for admin dashboard modal"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.args.get('id')
    if not staff_id:
        return jsonify({'success': False, 'error': 'Staff ID required'})

    # Get year and month parameters for viewing past records
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    # Default to current month if not specified
    today = datetime.date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    db = get_db()

    try:
        # Get staff information
        staff = db.execute('''
            SELECT s.*, sc.name as school_name
            FROM staff s
            LEFT JOIN schools sc ON s.school_id = sc.id
            WHERE s.id = ?
        ''', (staff_id,)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff not found'})

        # Calculate month boundaries for the requested month
        first_day_of_month = datetime.date(year, month, 1)
        if month == 12:
            last_day_of_month = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            last_day_of_month = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
        
        # For current month, limit to today's date
        if year == today.year and month == today.month:
            last_day_of_month = min(last_day_of_month, today)
        
        existing_attendance = db.execute('''
            SELECT date, time_in, time_out, status,
                   COALESCE(regularization_status, '') AS regularization_status,
                   on_duty_type, on_duty_location, on_duty_purpose
            FROM attendance
            WHERE staff_id = ? AND date >= ? AND date <= ?
            ORDER BY date DESC
        ''', (staff_id, first_day_of_month, last_day_of_month)).fetchall()

        # Get approved leaves for the requested month
        approved_leaves = db.execute('''
            SELECT start_date, end_date
            FROM leave_applications
            WHERE staff_id = ? 
              AND status = 'approved'
              AND ((start_date BETWEEN ? AND ?) 
                   OR (end_date BETWEEN ? AND ?)
                   OR (start_date <= ? AND end_date >= ?))
        ''', (staff_id, first_day_of_month, last_day_of_month, 
              first_day_of_month, last_day_of_month, 
              first_day_of_month, last_day_of_month)).fetchall()

        # Create a set of leave dates
        leave_dates = set()
        for leave in approved_leaves:
            start_date = datetime.datetime.strptime(leave['start_date'], '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
            current_date = max(start_date, first_day_of_month)
            end_date_capped = min(end_date, today)
            
            while current_date <= end_date_capped:
                if current_date.weekday() < 6:  # Monday to Saturday
                    leave_dates.add(current_date)
                current_date += datetime.timedelta(days=1)

        # Create comprehensive attendance records including missing days
        attendance_dict = {}
        
        # Add existing records
        for record in existing_attendance:
            date_str = record['date']
            attendance_dict[date_str] = dict(record)
        
        # Import holiday checking function
        from database import is_holiday
        
        # Get staff department for holiday checking  
        staff_dict = dict(staff) if staff else {}
        staff_department = staff_dict.get('department')
        
        # Generate complete attendance records for all working days in requested month
        current_date = first_day_of_month
        
        while current_date <= last_day_of_month:
            # Skip Sundays (weekday 6)
            if current_date.weekday() < 6:  # Monday to Saturday
                date_str = current_date.strftime('%Y-%m-%d')
                
                if date_str not in attendance_dict:
                    # Check if this date is a holiday
                    staff_school_id = staff_dict.get('school_id')
                    is_date_holiday = is_holiday(current_date, department=staff_department, school_id=staff_school_id)
                    
                    # Determine status for missing days
                    if is_date_holiday:
                        status = 'holiday'
                    elif current_date in leave_dates:
                        status = 'leave'
                    elif current_date == today and year == today.year and month == today.month:
                        # Don't create absent record for today if it's still ongoing (current month only)
                        current_date += datetime.timedelta(days=1)
                        continue
                    else:
                        status = 'absent'
                    
                    # Create missing attendance record
                    attendance_dict[date_str] = {
                        'date': date_str,
                        'time_in': None,
                        'time_out': None,
                        'status': status,
                        'regularization_status': '',
                        'on_duty_type': None,
                        'on_duty_location': None,
                        'on_duty_purpose': None
                    }
                else:
                    # Check existing records for holiday status override
                    existing_record = attendance_dict[date_str]
                    staff_school_id = staff_dict.get('school_id')
                    is_date_holiday = is_holiday(current_date, department=staff_department, school_id=staff_school_id)
                    
                    # If it's a holiday but marked as absent, update to holiday
                    if is_date_holiday and existing_record['status'] == 'absent':
                        existing_record['status'] = 'holiday'
            
            current_date += datetime.timedelta(days=1)
        
        # Convert back to list, normalize regularized statuses, then sort by date (newest first).
        attendance = list(attendance_dict.values())
        for record in attendance:
            record_status = str(record.get('status') or '').strip().lower()
            regularization_state = str(record.get('regularization_status') or '').strip().lower()
            if regularization_state == 'approved' and record_status == 'late':
                record['status'] = 'approved_regularization'

        attendance.sort(key=lambda x: x['date'], reverse=True)

        # Get biometric verifications for requested month
        verifications = db.execute('''
            SELECT verification_type, verification_time, verification_status, device_ip
            FROM biometric_verifications
            WHERE staff_id = ? AND DATE(verification_time) BETWEEN ? AND ?
            ORDER BY verification_time DESC
        ''', (staff_id, first_day_of_month, last_day_of_month)).fetchall()

        # Get leave applications
        leaves = db.execute('''
             SELECT leave_type, start_date, end_date, reason,
                 CASE WHEN COALESCE(withdrawn, 0) = 1 THEN 'withdrawn' ELSE status END as status,
                 applied_at,
                   COALESCE(withdrawn, 0) as withdrawn
            FROM leave_applications
            WHERE staff_id = ?
            ORDER BY applied_at DESC
            LIMIT 20
        ''', (staff_id,)).fetchall()

        # Get on-duty applications
        on_duty_applications = db.execute('''
             SELECT duty_type, start_date, end_date, start_time, end_time, location, purpose, reason,
                 CASE WHEN COALESCE(withdrawn, 0) = 1 THEN 'withdrawn' ELSE status END as status,
                 applied_at, admin_remarks,
                   COALESCE(withdrawn, 0) as withdrawn
            FROM on_duty_applications
            WHERE staff_id = ?
            ORDER BY applied_at DESC
            LIMIT 20
        ''', (staff_id,)).fetchall()

        # Get permission applications
        permission_applications = db.execute('''
             SELECT permission_type, permission_date, start_time, end_time, duration_hours, reason,
                 CASE WHEN COALESCE(withdrawn, 0) = 1 THEN 'withdrawn' ELSE status END as status,
                 applied_at, admin_remarks,
                   COALESCE(withdrawn, 0) as withdrawn
            FROM permission_applications
            WHERE staff_id = ?
            ORDER BY applied_at DESC
            LIMIT 20
        ''', (staff_id,)).fetchall()

        # Calculate total working days in current month (Monday to Saturday)
        def calculate_working_days_in_month(year, month):
            """Calculate total working days (Monday-Saturday) in a given month"""
            import calendar
            # Get the last day of the month
            last_day = calendar.monthrange(year, month)[1]
            working_days = 0
            
            for day in range(1, last_day + 1):
                date_obj = datetime.date(year, month, day)
                # Monday=0, Sunday=6, so weekday < 6 means Monday-Saturday
                if date_obj.weekday() < 6:
                    working_days += 1
            return working_days
        
        # Calculate actual working days in requested month
        current_month_working_days = calculate_working_days_in_month(year, month)
        
        # Calculate accurate attendance statistics
        total_recorded_days = len(attendance)
        present_days = len([a for a in attendance if a['status'] in ['present', 'late', 'on_duty', 'approved_regularization']])
        absent_days = len([a for a in attendance if a['status'] == 'absent'])
        late_days = len([a for a in attendance if a['status'] == 'late'])
        on_duty_days = len([a for a in attendance if a['status'] == 'on_duty'])
        leave_days = len([a for a in attendance if a['status'] == 'leave'])
        holiday_days = len([a for a in attendance if a['status'] == 'holiday'])
        regularized_days = len([a for a in attendance if a['status'] == 'approved_regularization'])
        
        # Calculate actual working days (Monday-Saturday minus holidays)
        actual_working_days = current_month_working_days - holiday_days

        attendance_stats = {
            'total_recorded_days': total_recorded_days,
            'working_days': actual_working_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'regularized_days': regularized_days,
            'on_duty_days': on_duty_days,
            'leave_days': leave_days,
            'holiday_days': holiday_days,
            'attendance_rate': round((present_days / actual_working_days * 100) if actual_working_days > 0 else 0, 1)
        }

        # Format attendance times to 12-hour format
        formatted_attendance = [format_attendance_times_to_12hr(dict(a)) for a in attendance]

        # Get module settings for this school
        school_id = staff_dict.get('school_id')
        module_enabled = get_module_enabled(school_id) if school_id else {}

        return jsonify({
            'success': True,
            'staff': serialize_row(staff),
            'attendance': formatted_attendance,
            'verifications': [serialize_row(v) for v in verifications],
            'leaves': [serialize_row(l) for l in leaves],
            'on_duty_applications': [serialize_row(od) for od in on_duty_applications],
            'permission_applications': [serialize_row(p) for p in permission_applications],
            'attendance_stats': attendance_stats,
            'month_info': {
                'year': year,
                'month': month,
                'month_name': datetime.date(year, month, 1).strftime('%B %Y')
            },
            'timetable_enabled': module_enabled.get('timetable_management', False)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to get staff profile: {str(e)}'
        })

@app.route('/download_staff_attendance_report')
def download_staff_attendance_report():
    """Download staff attendance report for a specific month"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.args.get('id')
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not staff_id:
        return jsonify({'success': False, 'error': 'Staff ID required'})
    
    # Default to current month if not specified
    today = datetime.date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    try:
        # Import ExcelReportGenerator
        from excel_reports import ExcelReportGenerator
        
        excel_generator = ExcelReportGenerator()
        
        # Convert year/month to start_date and end_date
        start_date = datetime.date(year, month, 1)
        if month == 12:
            end_date = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            end_date = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
        
        # Get staff name for filename and validate staff exists
        db = get_db()
        staff = db.execute('SELECT full_name FROM staff WHERE id = ?', (staff_id,)).fetchone()
        
        if not staff:
            return jsonify({'success': False, 'error': 'Staff member not found'})
        
        staff_name = staff['full_name'] if staff and staff['full_name'] else 'Staff'
        
        # Clean staff name for filename
        clean_name = ''.join(c for c in staff_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        if not clean_name:
            clean_name = f"Staff_{staff_id}"
        
        month_name = datetime.date(year, month, 1).strftime('%B_%Y')
        filename = f"{clean_name}_Attendance_Report_{month_name}.xlsx"
        
        # Create Excel workbook directly using openpyxl
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import io
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Add title and metadata
        ws['A1'] = f"Attendance Report - {staff_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Period: {month_name}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f"Generated: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=10, italic=True)
        
        # Get existing attendance records for the month
        existing_attendance = db.execute('''
            SELECT date, time_in, time_out, status, on_duty_type, on_duty_location, on_duty_purpose
            FROM attendance
            WHERE staff_id = ? AND date >= ? AND date <= ?
            ORDER BY date DESC
        ''', (staff_id, start_date, end_date)).fetchall()

        # Get approved leaves for the month
        approved_leaves = db.execute('''
            SELECT start_date, end_date
            FROM leave_applications
            WHERE staff_id = ? 
              AND status = 'approved'
              AND ((start_date BETWEEN ? AND ?) 
                   OR (end_date BETWEEN ? AND ?)
                   OR (start_date <= ? AND end_date >= ?))
        ''', (staff_id, start_date, end_date, 
              start_date, end_date, 
              start_date, end_date)).fetchall()

        # Create a set of leave dates
        leave_dates = set()
        for leave in approved_leaves:
            leave_start = datetime.datetime.strptime(leave['start_date'], '%Y-%m-%d').date()
            leave_end = datetime.datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
            current_date = max(leave_start, start_date)
            end_date_capped = min(leave_end, end_date)
            
            while current_date <= end_date_capped:
                if current_date.weekday() < 6:  # Monday to Saturday
                    leave_dates.add(current_date)
                current_date += datetime.timedelta(days=1)

        # Create comprehensive attendance records including missing days
        attendance_dict = {}
        
        # Add existing records
        for record in existing_attendance:
            date_str = record['date']
            attendance_dict[date_str] = dict(record)
        
        # Import holiday checking function
        from database import is_holiday
        
        # Get staff department for holiday checking
        staff_dept = db.execute('SELECT department, school_id FROM staff WHERE id = ?', (staff_id,)).fetchone()
        staff_department = staff_dept['department'] if staff_dept else None
        staff_school_id = staff_dept['school_id'] if staff_dept else None
        
        # Generate complete attendance records for all working days
        current_date = start_date
        today = datetime.date.today()
        
        while current_date <= end_date:
            # Skip Sundays (weekday 6)
            if current_date.weekday() < 6:  # Monday to Saturday
                date_str = current_date.strftime('%Y-%m-%d')
                
                if date_str not in attendance_dict:
                    # Check if this date is a holiday
                    is_date_holiday = is_holiday(current_date, department=staff_department, school_id=staff_school_id)
                    
                    # Determine status for missing days
                    if is_date_holiday:
                        status = 'Holiday'
                    elif current_date in leave_dates:
                        status = 'Leave'
                    elif current_date == today and year == today.year and month == today.month:
                        # Don't create absent record for today if it's still ongoing (current month only)
                        current_date += datetime.timedelta(days=1)
                        continue
                    else:
                        status = 'Absent'
                    
                    # Create missing attendance record
                    attendance_dict[date_str] = {
                        'date': date_str,
                        'time_in': None,
                        'time_out': None,
                        'status': status,
                        'on_duty_type': None,
                        'on_duty_location': None,
                        'on_duty_purpose': None
                    }
                else:
                    # Check existing records for holiday status override
                    existing_record = attendance_dict[date_str]
                    is_date_holiday = is_holiday(current_date, department=staff_department, school_id=staff_school_id)
                    
                    # If it's a holiday but marked as absent, update to holiday
                    if is_date_holiday and existing_record['status'] == 'absent':
                        existing_record['status'] = 'Holiday'
            
            current_date += datetime.timedelta(days=1)
        
        # Convert back to list and sort by date (newest first)
        all_attendance = sorted(attendance_dict.values(), key=lambda x: x['date'], reverse=True)
        
        # Calculate statistics for summary
        present_count = len([r for r in all_attendance if r['status'] and r['status'].lower() in ['present', 'late']])
        absent_count = len([r for r in all_attendance if r['status'] and r['status'].lower() == 'absent'])
        leave_count = len([r for r in all_attendance if r['status'] and r['status'].lower() == 'leave'])
        holiday_count = len([r for r in all_attendance if r['status'] and r['status'].lower() == 'holiday'])
        on_duty_count = len([r for r in all_attendance if r['status'] and r['status'].lower() == 'on_duty'])
        # Working days should be total calendar working days (Mon-Sat) minus holidays
        total_working_days = _get_working_days_in_month(year, month) - holiday_count
        
        # Add summary statistics
        ws['A5'] = "Summary Statistics:"
        ws['A5'].font = Font(bold=True, size=12)
        ws['A6'] = f"Total Working Days: {total_working_days}"
        ws['A7'] = f"Present: {present_count}"
        ws['A8'] = f"Absent: {absent_count}"
        ws['A9'] = f"Leave: {leave_count}"
        ws['A10'] = f"Holiday: {holiday_count}"
        ws['A11'] = f"On Duty: {on_duty_count}"
        
        # Add headers
        headers = ['Date', 'Day', 'Time In', 'Time Out', 'Status', 'On Duty Type', 'Location', 'Purpose']
        header_row = 13
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Add data rows
        row = 14
        if all_attendance:
            for record in all_attendance:
                date_obj = datetime.datetime.strptime(record['date'], '%Y-%m-%d').date()
                day_name = date_obj.strftime('%A')
                formatted_date = date_obj.strftime('%d/%m/%Y')
                
                # Set row background color based on status
                status = record['status'] or 'Unknown'
                
                ws.cell(row=row, column=1, value=formatted_date)
                ws.cell(row=row, column=2, value=day_name)
                ws.cell(row=row, column=3, value=record['time_in'] or '--:--')
                ws.cell(row=row, column=4, value=record['time_out'] or '--:--')
                ws.cell(row=row, column=5, value=status.title())
                ws.cell(row=row, column=6, value=record['on_duty_type'] or '')
                ws.cell(row=row, column=7, value=record['on_duty_location'] or '')
                ws.cell(row=row, column=8, value=record['on_duty_purpose'] or '')
                
                # Color coding for different statuses
                if status.lower() == 'absent':
                    fill_color = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
                elif status.lower() == 'leave':
                    fill_color = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue
                elif status.lower() == 'holiday':
                    fill_color = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light gray
                elif status.lower() in ['present', 'late']:
                    fill_color = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Light green
                else:
                    fill_color = None
                
                if fill_color:
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).fill = fill_color
                
                row += 1
        else:
            # Add "No records found" message if no data
            ws.cell(row=row, column=1, value="No attendance records found for this period")
            ws.merge_cells(f'A{row}:H{row}')
            ws.cell(row=row, column=1).font = Font(italic=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Cache-Control'] = 'no-cache'
        response.headers['Content-Length'] = str(len(output.getvalue()))
        
        return response
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to generate report: {str(e)}'})

def _get_working_days_in_month(year: int, month: int) -> int:
    """Calculate working days in a month (excluding only Sundays)"""
    from datetime import date, timedelta
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    working_days = 0
    current_date = start_date
    
    while current_date <= end_date:
        # Monday = 0, Sunday = 6
        if current_date.weekday() < 6:  # Monday to Saturday
            working_days += 1
        current_date += timedelta(days=1)
    
    return working_days

# 🚀 PERFORMANCE OPTIMIZATION: Cached attendance summary
_attendance_summary_cache = {}

def _get_cached_attendance_summary(staff_id: int, year: int, month: int):
    """Get attendance summary with caching for improved performance"""
    cache_key = f"{staff_id}_{year}_{month}"
    current_time = datetime.datetime.now()
    
    # Check if we have a cached result that's less than 5 minutes old
    if (cache_key in _attendance_summary_cache and 
        current_time - _attendance_summary_cache[cache_key]['timestamp'] < datetime.timedelta(minutes=5)):
        return _attendance_summary_cache[cache_key]['data']
    
    # Calculate fresh summary
    summary = _calculate_accurate_attendance_summary(staff_id, year, month)
    
    # Cache the result
    _attendance_summary_cache[cache_key] = {
        'data': summary,
        'timestamp': current_time
    }
    
    # Clean old cache entries (keep only last 50 entries)
    if len(_attendance_summary_cache) > 50:
        oldest_keys = sorted(_attendance_summary_cache.keys(), 
                           key=lambda k: _attendance_summary_cache[k]['timestamp'])[:10]
        for old_key in oldest_keys:
            del _attendance_summary_cache[old_key]
    
    return summary

def _calculate_accurate_attendance_summary(staff_id: int, year: int, month: int):
    """Calculate accurate attendance summary with proper absent day calculation"""
    db = get_db()

    def _as_date(value):
        if value is None:
            return None
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        return datetime.datetime.strptime(str(value), '%Y-%m-%d').date()
    
    # Calculate month boundaries
    first_day = datetime.date(year, month, 1)
    if month == 12:
        last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        last_day = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
    
    # Calculate total working days in month (excluding Sundays)
    working_days = _get_working_days_in_month(year, month)
    
    # Get attendance records from attendance table
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_recorded_days,
            SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) as present_days,
            SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as late_days,
            SUM(CASE WHEN status = 'leave' THEN 1 ELSE 0 END) as leave_days,
            SUM(CASE WHEN status = 'on_duty' THEN 1 ELSE 0 END) as on_duty_days,
            SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as explicitly_absent_days
        FROM attendance
        WHERE staff_id = ? AND date BETWEEN ? AND ?
    ''', (staff_id, first_day, last_day)).fetchone()

    # Get approved leave applications for current month and count leave days
    approved_leaves = db.execute('''
        SELECT start_date, end_date
        FROM leave_applications
        WHERE staff_id = ? 
          AND status = 'approved'
          AND ((start_date BETWEEN ? AND ?) 
               OR (end_date BETWEEN ? AND ?)
               OR (start_date <= ? AND end_date >= ?))
    ''', (staff_id, first_day, last_day, first_day, last_day, first_day, last_day)).fetchall()

    # Count leave days from approved leave applications (avoid double counting)
    leave_days_from_applications = 0
    dates_with_leave = set()
    
    for leave in approved_leaves:
        start_date = _as_date(leave['start_date'])
        end_date = _as_date(leave['end_date'])
        if not start_date or not end_date:
            continue
        
        # Iterate through each day in the leave period
        current_date = max(start_date, first_day)
        end_date_capped = min(end_date, last_day)
        
        while current_date <= end_date_capped:
            # Only count weekdays (Monday-Saturday), skip Sundays
            if current_date.weekday() < 6:  # 0=Monday, 6=Sunday
                dates_with_leave.add(current_date)
            current_date += datetime.timedelta(days=1)
    
    leave_days_from_applications = len(dates_with_leave)

    # Calculate accurate counts
    present_days = (attendance_summary['present_days'] or 0)
    late_days = (attendance_summary['late_days'] or 0)
    on_duty_days = (attendance_summary['on_duty_days'] or 0)
    leave_days = (attendance_summary['leave_days'] or 0) + leave_days_from_applications
    
    # Total effective present days (present + late + on_duty)
    total_effective_present = present_days + late_days + on_duty_days
    
    # Absent days = Working days - (effective present days + leave days)
    absent_days = working_days - (total_effective_present + leave_days)
    
    # Ensure absent days is never negative
    absent_days = max(0, absent_days)
    
    # Calculate holiday days in the current month
    from database import is_holiday
    
    # Get staff department and school info for holiday checking
    staff_info = db.execute('SELECT department, school_id FROM staff WHERE id = ?', (staff_id,)).fetchone()
    staff_department = staff_info['department'] if staff_info else None
    staff_school_id = staff_info['school_id'] if staff_info else None
    
    # Count holidays in the current month
    holiday_days = 0
    current_date = first_day
    
    while current_date <= last_day:
        # Only count holidays on working days (Monday-Saturday)
        if current_date.weekday() < 6:  # 0=Monday, 6=Sunday
            if is_holiday(current_date, department=staff_department, school_id=staff_school_id):
                holiday_days += 1
        current_date += datetime.timedelta(days=1)

    return {
        'present_days': present_days,
        'absent_days': absent_days,
        'late_days': late_days,
        'leave_days': leave_days,
        'on_duty_days': on_duty_days,
        'holiday_days': holiday_days,
        'working_days': working_days,
        'total_recorded_days': attendance_summary['total_recorded_days'] or 0
    }

@app.route('/staff/profile')
def staff_profile_page():
    """Optimized staff profile page with combined queries and improved performance"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    db = get_db()
    staff_id = session['user_id']
    today = datetime.date.today()
    one_week_ago = today - datetime.timedelta(days=7)
    one_week_ago_str = one_week_ago.strftime('%Y-%m-%d')

    # 🚀 OPTIMIZATION 1: Single combined query for staff info and basic data
    staff_basic_data = db.execute('''
        SELECT s.*, 
               COUNT(DISTINCT la.id) as total_leave_applications,
               COUNT(DISTINCT oda.id) as total_od_applications,
               COUNT(DISTINCT pa.id) as total_permission_applications
        FROM staff s
        LEFT JOIN leave_applications la ON s.id = la.staff_id
        LEFT JOIN on_duty_applications oda ON s.id = oda.staff_id  
        LEFT JOIN permission_applications pa ON s.id = pa.staff_id
        WHERE s.id = ?
        GROUP BY s.id
    ''', (staff_id,)).fetchone()

    if not staff_basic_data:
        return redirect(url_for('index'))

    # Convert Row to dict for easier handling
    staff = dict(staff_basic_data)

    # 🚀 OPTIMIZATION 2: Combined applications query with UNION for efficiency
    all_applications = db.execute('''
         SELECT *
         FROM (
             SELECT 'leave' as app_type, id, leave_type as type_detail, start_date, end_date,
                 NULL as location, reason, status, applied_at, NULL as duration_hours,
                 COALESCE(withdrawn, 0) as withdrawn
             FROM leave_applications
             WHERE staff_id = ?

             UNION ALL

             SELECT 'on_duty' as app_type, id, duty_type as type_detail, start_date, end_date,
                 location, COALESCE(reason, purpose) as reason, status, applied_at, NULL as duration_hours,
                 COALESCE(withdrawn, 0) as withdrawn
             FROM on_duty_applications
             WHERE staff_id = ?

             UNION ALL

             SELECT 'permission' as app_type, id, permission_type as type_detail,
                 permission_date as start_date, permission_date as end_date, NULL as location,
                 reason, status, applied_at, duration_hours, COALESCE(withdrawn, 0) as withdrawn
             FROM permission_applications
             WHERE staff_id = ?
         ) recent_apps
         WHERE DATE(COALESCE(applied_at, start_date)) >= DATE(?)
         ORDER BY COALESCE(applied_at, start_date) DESC
        LIMIT 30
        ''', (staff_id, staff_id, staff_id, one_week_ago_str)).fetchall()

    # Separate applications by type for template
    leave_applications = []
    on_duty_applications = []
    permission_applications = []

    for app in all_applications:
        app_dict = dict(app)
        if app['app_type'] == 'leave':
            leave_applications.append({
                'id': app['id'],
                'leave_type': app['type_detail'], 
                'start_date': app['start_date'],
                'end_date': app['end_date'],
                'reason': app['reason'],
                'status': 'withdrawn' if app['withdrawn'] else app['status'],
                'applied_at': app['applied_at']
            })
        elif app['app_type'] == 'on_duty':
            on_duty_applications.append({
                'id': app['id'],
                'duty_type': app['type_detail'],
                'start_date': app['start_date'], 
                'end_date': app['end_date'],
                'location': app['location'],
                'purpose': app['reason'],  # Map reason to purpose for template compatibility
                'reason': app['reason'],
                'status': 'withdrawn' if app['withdrawn'] else app['status'],
                'applied_at': app['applied_at']
            })
        elif app['app_type'] == 'permission':
            permission_applications.append({
                'id': app['id'],
                'permission_type': app['type_detail'],
                'permission_date': app['start_date'],
                'start_time': None,  # These would need separate query if needed
                'end_time': None,
                'duration_hours': app['duration_hours'],
                'reason': app['reason'],
                'status': 'withdrawn' if app['withdrawn'] else app['status'],
                'applied_at': app['applied_at']
            })

    # Limit each type to 10 as per original logic
    leave_applications = leave_applications[:10]
    on_duty_applications = on_duty_applications[:10] 
    permission_applications = permission_applications[:10]

    # 🚀 OPTIMIZATION 3: Simplified biometric verifications query (remove complex subquery)
    recent_verifications = db.execute('''
        SELECT verification_type, verification_time, biometric_method, verification_status
        FROM biometric_verifications
        WHERE staff_id = ?
          AND DATE(verification_time) >= DATE('now', '-14 days')
        ORDER BY verification_time DESC
        LIMIT 20
    ''', (staff_id,)).fetchall()

    # 🚀 OPTIMIZATION 3B: Get ALL attendance records (including web punches: id_scan, otp, dynamic_qr)
    # This combines with biometric verifications to show complete daily attendance
    all_attendance_records = db.execute('''
        SELECT date, time_in, time_out, status, late_duration_minutes, 
               early_departure_minutes, shift_type, shift_start_time, shift_end_time, notes
        FROM attendance
        WHERE staff_id = ? AND date >= DATE('now', '-30 days')
        ORDER BY date DESC
        LIMIT 60
    ''', (staff_id,)).fetchall()

    # 🚀 OPTIMIZATION 4: Get attendance summary with optimized function call
    attendance_summary_dict = _get_cached_attendance_summary(staff_id, today.year, today.month)

    # 🚀 OPTIMIZATION 5: Get quota summary for current year
    quota_year = today.year
    quota_summary = get_staff_quota_summary(staff_id, quota_year)

    # Get module settings for this school
    school_id = session.get('school_id')
    module_enabled = get_module_enabled(school_id) if school_id else {}

    return render_template('staff_my_profile.html',
                         staff=staff,
                         attendance_summary=attendance_summary_dict,
                         leave_applications=leave_applications,
                         on_duty_applications=on_duty_applications,
                         permission_applications=permission_applications,
                         recent_verifications=recent_verifications,
                         all_attendance_records=all_attendance_records,
                         today=today,
                         current_month=today.strftime('%B %Y'),
                         quota_summary=quota_summary,
                         quota_year=quota_year,
                         module_enabled=module_enabled)

@app.route('/staff/profile/async_data')
def get_staff_profile_async_data():
    """Get heavy profile data asynchronously for improved performance"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    today = datetime.date.today()
    
    try:
        # Get the data that takes longer to load
        db = get_db()
        
        # Get detailed biometric verifications (the expensive query)
        detailed_verifications = db.execute('''
            SELECT verification_type, verification_time, biometric_method, verification_status
            FROM biometric_verifications bv1
            WHERE staff_id = ?
              AND verification_time = (
                SELECT MAX(verification_time)
                FROM biometric_verifications bv2
                WHERE bv2.staff_id = bv1.staff_id
                  AND bv2.verification_type = bv1.verification_type
                  AND DATE(bv2.verification_time) = DATE(bv1.verification_time)
              )
            ORDER BY verification_time DESC
            LIMIT 20
        ''', (staff_id,)).fetchall()

        # Get recent detailed attendance records
        attendance_records = db.execute('''
            SELECT date, time_in, time_out, status, late_duration_minutes, 
                   shift_type, shift_start_time, shift_end_time,
                   work_hours, overtime_hours
            FROM attendance
            WHERE staff_id = ? AND date >= DATE('now', '-30 days')
            ORDER BY date DESC
            LIMIT 30
        ''', (staff_id,)).fetchall()

        return jsonify({
            'success': True,
            'detailed_verifications': [serialize_row(row) for row in detailed_verifications],
            'attendance_records': [serialize_row(row) for row in attendance_records]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/get_attendance_summary')
def get_staff_attendance_summary():
    """Get staff attendance summary for current month (dynamic API) - optimized"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    
    try:
        # Get cached attendance summary for current month
        today = datetime.date.today()
        attendance_summary = _get_cached_attendance_summary(staff_id, today.year, today.month)

        return jsonify({
            'success': True,
            'present_days': attendance_summary['present_days'],
            'absent_days': attendance_summary['absent_days'],
            'late_days': attendance_summary['late_days'],
            'leave_days': attendance_summary['leave_days'],
            'on_duty_days': attendance_summary['on_duty_days'],
            'holiday_days': attendance_summary['holiday_days'],
            'working_days': attendance_summary['working_days'],
            'total_recorded_days': attendance_summary['total_recorded_days'],
            'current_month': today.strftime('%B %Y')
        })
    except Exception as e:
        print(f"Error fetching attendance summary: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/update_profile', methods=['POST'])
def update_staff_profile():
    """Update staff profile information"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    date_of_birth = request.form.get('date_of_birth')
    gender = request.form.get('gender')
    email = request.form.get('email')
    phone = request.form.get('phone')

    db = get_db()

    try:
        # Get staff info for filename
        staff_info = db.execute('SELECT full_name, staff_id FROM staff WHERE id = ?', (staff_id,)).fetchone()
        
        # Handle photo upload
        photo_url = None
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo.filename != '' and allowed_file(photo.filename):
                try:
                    # Ensure uploads directory exists
                    upload_dir = os.path.join(app.static_folder, 'uploads')
                    os.makedirs(upload_dir, exist_ok=True)

                    # Generate filename with staff name
                    ext = os.path.splitext(photo.filename)[1]
                    # Replace spaces and special characters with underscores
                    safe_name = staff_info['full_name'].replace(' ', '_').replace('/', '_').replace('\\', '_')
                    filename = f"{safe_name}_{staff_info['staff_id']}{ext}"
                    
                    # Delete old photo if exists
                    if staff_info:
                        old_photo = db.execute('SELECT photo_url FROM staff WHERE id = ?', (staff_id,)).fetchone()
                        if old_photo and old_photo['photo_url']:
                            old_photo_path = os.path.join(app.static_folder, old_photo['photo_url'])
                            if os.path.exists(old_photo_path):
                                os.remove(old_photo_path)
                    
                    photo.save(os.path.join(upload_dir, filename))
                    photo_url = f"uploads/{filename}"
                except Exception as e:
                    return jsonify({'success': False, 'error': 'Error saving photo'})

        # Build update query dynamically based on available columns
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        column_names = [col['name'] for col in columns]

        update_parts = []
        update_values = []

        if 'first_name' in column_names and first_name:
            update_parts.append('first_name = ?')
            update_values.append(first_name)
        if 'last_name' in column_names and last_name:
            update_parts.append('last_name = ?')
            update_values.append(last_name)
        if 'date_of_birth' in column_names and date_of_birth:
            update_parts.append('date_of_birth = ?')
            update_values.append(date_of_birth)
        if 'gender' in column_names and gender:
            update_parts.append('gender = ?')
            update_values.append(gender)
        if 'email' in column_names:
            update_parts.append('email = ?')
            update_values.append(email)
        if 'phone' in column_names:
            update_parts.append('phone = ?')
            update_values.append(phone)
        if 'photo_url' in column_names and photo_url:
            update_parts.append('photo_url = ?')
            update_values.append(photo_url)

        # Update full_name if first_name or last_name changed
        if first_name or last_name:
            # Get current names if only one is being updated
            current_staff = db.execute('SELECT first_name, last_name FROM staff WHERE id = ?', (staff_id,)).fetchone()
            current_first = current_staff['first_name'] if current_staff else ''
            current_last = current_staff['last_name'] if current_staff else ''

            new_first = first_name if first_name else current_first
            new_last = last_name if last_name else current_last
            full_name = f"{new_first} {new_last}".strip()

            update_parts.append('full_name = ?')
            update_values.append(full_name)

        if update_parts:
            # Add WHERE clause value
            update_values.append(staff_id)

            # Build and execute the query
            query = f"UPDATE staff SET {', '.join(update_parts)} WHERE id = ?"

            db.execute(query, update_values)

        db.commit()
        return jsonify({'success': True, 'message': 'Profile updated successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/change_password', methods=['POST'])
def change_staff_password():
    """Change staff password"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if new_password != confirm_password:
        return jsonify({'success': False, 'error': 'New passwords do not match'})

    db = get_db()

    # Get current password hash
    staff = db.execute('SELECT password_hash FROM staff WHERE id = ?', (staff_id,)).fetchone()

    if not staff or not check_password_hash(staff['password_hash'], current_password):
        return jsonify({'success': False, 'error': 'Current password is incorrect'})

    try:
        # Update password
        new_password_hash = generate_password_hash(new_password)
        db.execute('''
            UPDATE staff SET password_hash = ?
            WHERE id = ?
        ''', (new_password_hash, staff_id))

        db.commit()
        return jsonify({'success': True, 'message': 'Password changed successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/attendance_calendar')
def staff_attendance_calendar():
    """Get attendance data for calendar view"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    start_date = request.args.get('start')
    end_date = request.args.get('end')

    db = get_db()

    # Get attendance records for the date range with enhanced information
    attendance = db.execute('''
        SELECT a.date, a.time_in, a.time_out, a.status, a.notes,
               s.shift_type, a.late_duration_minutes, a.early_departure_minutes,
               a.shift_start_time, a.shift_end_time
        FROM attendance a
        JOIN staff s ON a.staff_id = s.id
        WHERE a.staff_id = ? AND a.date BETWEEN ? AND ?
        ORDER BY a.date
    ''', (staff_id, start_date, end_date)).fetchall()

    # Get leave applications for the date range
    leaves = db.execute('''
        SELECT start_date, end_date, leave_type, status
        FROM leave_applications
        WHERE staff_id = ? AND status = 'approved'
        AND ((start_date BETWEEN ? AND ?) OR (end_date BETWEEN ? AND ?)
        OR (start_date <= ? AND end_date >= ?))
    ''', (staff_id, start_date, end_date, start_date, end_date, start_date, end_date)).fetchall()

    # Format attendance times to 12-hour format
    formatted_attendance = [format_attendance_times_to_12hr(dict(a)) for a in attendance]

    return jsonify({
        'success': True,
        'attendance': formatted_attendance,
        'leaves': [dict(l) for l in leaves]
    })

@app.route('/staff/download_pay_slip', methods=['POST'])
def staff_download_pay_slip():
    """Generate and download pay slip for staff member"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_db_id = session['user_id']
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not all([year, month]):
        return jsonify({'success': False, 'error': 'Year and month are required'})

    try:
        db = get_db()
        school_id = session.get('school_id')
        
        # Get staff information including staff_id
        staff_info = db.execute('''
            SELECT id, staff_id, full_name, department, position, basic_salary
            FROM staff WHERE id = ? AND school_id = ?
        ''', (staff_db_id, school_id)).fetchone()
        
        if not staff_info:
            return jsonify({'success': False, 'error': 'Staff information not found'})

        payroll_review_status = 'pending'
        try:
            db.execute('''
                CREATE TABLE IF NOT EXISTS payroll_review_status (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    school_id INTEGER NOT NULL,
                    staff_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    status TEXT NOT NULL DEFAULT 'pending',
                    reviewed_by TEXT,
                    reviewed_at TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(school_id, staff_id, year, month)
                )
            ''')

            status_row = db.execute('''
                SELECT status
                FROM payroll_review_status
                WHERE school_id = ? AND staff_id = ? AND year = ? AND month = ?
            ''', (school_id, staff_db_id, year, month)).fetchone()

            if status_row and status_row['status']:
                normalized = str(status_row['status']).strip().lower()
                if normalized in ('review', 'reviewed'):
                    payroll_review_status = 'reviewed'
                elif normalized in ('complete', 'completed'):
                    payroll_review_status = 'completed'
                elif normalized in ('pay released', 'pay_released', 'released', 'paid'):
                    payroll_review_status = 'pay_released'
                else:
                    payroll_review_status = 'pending'
        except Exception:
            payroll_review_status = 'pending'
        
        # Initialize salary calculator
        salary_calculator = SalaryCalculator(school_id=school_id)
        
        # Generate salary data for the staff member
        result = salary_calculator.generate_salary_report(staff_db_id, year, month)
        
        if not result['success']:
            return jsonify({'success': False, 'error': result.get('error', 'Failed to generate salary data')})
        
        # Generate PDF pay slip
        pdf_response = generate_individual_salary_slip_pdf(result, year, month, payroll_review_status=payroll_review_status)
        return pdf_response
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to generate pay slip: {str(e)}'})

@app.route('/staff/preview_pay_slip', methods=['POST'])
def staff_preview_pay_slip():
    """Generate pay slip data for preview by staff member"""
    if 'user_id' not in session or (session.get('user_type') != 'staff' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_db_id = session['user_id']
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not all([year, month]):
        return jsonify({'success': False, 'error': 'Year and month are required'})

    try:
        db = get_db()
        school_id = session.get('school_id')
        
        # Get staff information including staff_id
        staff_info = db.execute('''
            SELECT id, staff_id, full_name, department, position, basic_salary
            FROM staff WHERE id = ? AND school_id = ?
        ''', (staff_db_id, school_id)).fetchone()
        
        if not staff_info:
            return jsonify({'success': False, 'error': 'Staff information not found'})

        payroll_review_status = 'pending'
        try:
            db.execute('''
                CREATE TABLE IF NOT EXISTS payroll_review_status (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    school_id INTEGER NOT NULL,
                    staff_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    status TEXT NOT NULL DEFAULT 'pending',
                    reviewed_by TEXT,
                    reviewed_at TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(school_id, staff_id, year, month)
                )
            ''')

            status_row = db.execute('''
                SELECT status
                FROM payroll_review_status
                WHERE school_id = ? AND staff_id = ? AND year = ? AND month = ?
            ''', (school_id, staff_db_id, year, month)).fetchone()

            if status_row and status_row['status']:
                normalized = str(status_row['status']).strip().lower()
                if normalized in ('review', 'reviewed'):
                    payroll_review_status = 'reviewed'
                elif normalized in ('complete', 'completed'):
                    payroll_review_status = 'completed'
                elif normalized in ('pay released', 'pay_released', 'released', 'paid'):
                    payroll_review_status = 'pay_released'
                else:
                    payroll_review_status = 'pending'
        except Exception:
            payroll_review_status = 'pending'
        
        # Initialize salary calculator
        salary_calculator = SalaryCalculator(school_id=school_id)
        
        # Generate salary data for the staff member
        result = salary_calculator.generate_salary_report(staff_db_id, year, month)
        
        if not result['success']:
            return jsonify({'success': False, 'error': result.get('error', 'Failed to generate salary data')})
        
        # Return JSON data for preview
        return jsonify({
            'success': True,
            'data': {
                'staff_info': result['staff_info'],
                'salary_details': result['salary_breakdown'],
                'salary_summary': result['salary_breakdown'],  # Same data, for compatibility
                'payroll_workflow': {
                    'current_status': payroll_review_status
                }
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to generate pay slip preview: {str(e)}'})

# Test route removed - attendance functionality working correctly

@app.route('/fix_staff_names', methods=['POST'])
def fix_staff_names():
    """
    Fix existing staff records that have full_name but missing first_name/last_name.
    This resolves the 'None None' display issue in Staff Directory.
    """
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    
    try:
        db = get_db()
        
        # Find staff records with full_name but NULL first_name or last_name
        problematic_staff = db.execute('''
            SELECT id, staff_id, full_name, first_name, last_name 
            FROM staff 
            WHERE school_id = ? 
            AND full_name IS NOT NULL 
            AND full_name != ''
            AND (first_name IS NULL OR last_name IS NULL OR first_name = '' OR last_name = '')
        ''', (school_id,)).fetchall()
        
        if not problematic_staff:
            return jsonify({
                'success': True,
                'message': 'No staff records need fixing. All names are properly set.',
                'fixed_count': 0
            })
        
        fixed_count = 0
        errors = []
        
        for staff in problematic_staff:
            try:
                # Parse the full name
                first_name, last_name = parse_full_name(staff['full_name'])
                
                # Update the record
                db.execute('''
                    UPDATE staff 
                    SET first_name = ?, last_name = ? 
                    WHERE id = ? AND school_id = ?
                ''', (first_name, last_name, staff['id'], school_id))
                
                fixed_count += 1
                print(f"Fixed staff {staff['staff_id']}: '{staff['full_name']}' -> '{first_name}' '{last_name}'")
                
            except Exception as e:
                error_msg = f"Failed to fix staff {staff['staff_id']}: {str(e)}"
                errors.append(error_msg)
                print(error_msg)
        
        db.commit()
        
        message = f"Successfully fixed {fixed_count} staff records"
        if errors:
            message += f". {len(errors)} errors occurred."
        
        return jsonify({
            'success': True,
            'message': message,
            'fixed_count': fixed_count,
            'total_found': len(problematic_staff),
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Database error: {str(e)}'})

def parse_full_name(full_name):
    """
    Parse a full name into first_name and last_name components.
    Handles various name formats and provides fallbacks.
    
    Args:
        full_name (str): The full name to parse
    
    Returns:
        tuple: (first_name, last_name)
    """
    if not full_name or not full_name.strip():
        return 'Unknown', ''
    
    # Clean the input
    full_name = full_name.strip()
    
    # Split by spaces
    name_parts = [part.strip() for part in full_name.split() if part.strip()]
    
    if not name_parts:
        return 'Unknown', ''
    elif len(name_parts) == 1:
        # Single name - use as first name, leave last name empty
        return name_parts[0], ''
    elif len(name_parts) == 2:
        # Two parts - first and last
        return name_parts[0], name_parts[1]
    else:
        # Multiple parts - first is first_name, rest combine to last_name
        return name_parts[0], ' '.join(name_parts[1:])

def fix_user_suffix_in_staff_names():
    """
    Fix existing staff records that have 'User' as their last_name.
    This addresses the bug where biometric imported staff show ' user' suffix.
    """
    try:
        db = get_db()
        
        # Find staff records with 'User' as last_name (case-insensitive)
        staff_with_user_suffix = db.execute('''
            SELECT id, staff_id, full_name, first_name, last_name 
            FROM staff 
            WHERE LOWER(last_name) = 'user'
        ''').fetchall()
        
        fixed_count = 0
        for staff in staff_with_user_suffix:
            # Update the record to remove 'User' from last_name
            db.execute('''
                UPDATE staff 
                SET last_name = '', 
                    full_name = ?
                WHERE id = ?
            ''', (staff['first_name'], staff['id']))
            
            fixed_count += 1
            print(f"Fixed staff record: {staff['staff_id']} - '{staff['first_name']} User' → '{staff['first_name']}'")
        
        db.commit()
        print(f"✅ Fixed {fixed_count} staff records with 'User' suffix")
        return fixed_count
        
    except Exception as e:
        print(f"❌ Error fixing user suffix in staff names: {e}")
        return 0

@app.route('/admin/fix_staff_user_suffix', methods=['POST'])
def fix_staff_user_suffix_route():
    """Admin route to manually fix staff names with 'User' suffix"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        fixed_count = fix_user_suffix_in_staff_names()
        return jsonify({
            'success': True, 
            'message': f'Successfully fixed {fixed_count} staff records with "User" suffix',
            'fixed_count': fixed_count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/create_staff_from_device_user', methods=['POST'])
def create_staff_from_device_user():
    """Create staff account for user who already exists on biometric device"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    
    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    device_user_id = request.form.get('device_user_id')
    full_name = request.form.get('full_name')
    # Defer hashing until we know if device user has a password
    provided_password = request.form.get('password')
    email = request.form.get('email', '')
    phone = request.form.get('phone', '')
    department = request.form.get('department', '')
    position = request.form.get('position', '')

    if not device_user_id or not full_name:
        return jsonify({'success': False, 'error': 'Device User ID and full name are required'})

    try:
        # Verify user exists on device
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        users = zk_device.get_users()
        device_user = None
        for user in users:
            if str(user['user_id']) == str(device_user_id):
                device_user = user
                break

        zk_device.disconnect()

        if not device_user:
            return jsonify({'success': False, 'error': f'User {device_user_id} not found on biometric device'})

        # Choose password: prefer the biometric device user's password if present
        device_password = ''
        if device_user:
            # Handle both dict and object formats gracefully
            try:
                device_password = device_user.get('password') if isinstance(device_user, dict) else getattr(device_user, 'password', '')
            except Exception:
                device_password = ''

        selected_password = device_password or provided_password or 'password123'
        password_hash = generate_password_hash(selected_password)

        # Check if staff already exists in database
        db = get_db()
        existing_staff = db.execute('SELECT id FROM staff WHERE staff_id = ? AND school_id = ?',
                                  (device_user_id, school_id)).fetchone()

        if existing_staff:
            return jsonify({'success': False, 'error': f'Staff with ID {device_user_id} already exists in database'})

        # Parse full name into first and last names
        first_name, last_name = parse_full_name(full_name)
        
        # Create staff account with proper name fields
        db.execute('''
            INSERT INTO staff
            (school_id, staff_id, password_hash, full_name, first_name, last_name, email, phone, department, position)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (school_id, device_user_id, password_hash, full_name, first_name, last_name, email, phone, department, position))

        db.commit()

        return jsonify({
            'success': True,
            'message': f'Staff account created successfully for biometric user {device_user_id} ({full_name})',
            'details': {
                'full_name': full_name,
                'first_name': first_name,
                'last_name': last_name,
                'staff_id': device_user_id
            }
        })

    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'Staff ID already exists'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/resolve_user_conflict', methods=['POST'])
def resolve_user_conflict():
    """Resolve conflicts when user ID already exists on biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get device for this institution
    device_ip, device_port = get_institution_device()
    if not device_ip:
        return jsonify({'success': False, 'error': 'No biometric device configured'})
    
    action = request.form.get('action')  # 'overwrite', 'use_different_id', 'create_from_existing'

    if action == 'overwrite':
        # Overwrite existing user on device
        user_id = request.form.get('user_id')
        name = request.form.get('name')
        privilege = int(request.form.get('privilege', 0))

        try:
            zk_device = ZKBiometricDevice(device_ip)
            if zk_device.connect():
                result = zk_device.enroll_user(user_id, name, privilege, overwrite=True)
                zk_device.disconnect()

                if result['success']:
                    return jsonify({
                        'success': True,
                        'message': f'User {user_id} has been overwritten on the device. You can now proceed with biometric enrollment.',
                        'action': 'overwritten'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'message': result['message']
                    })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to connect to biometric device'
                })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error overwriting user: {str(e)}'
            })

    elif action == 'create_from_existing':
        # Create staff account using existing biometric user
        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'error': 'School ID not found'})

        existing_user_id = request.form.get('existing_user_id')
        full_name = request.form.get('full_name')
        password = generate_password_hash(request.form.get('password', 'default123'))
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        department = request.form.get('department', '')
        position = request.form.get('position', '')

        try:
            db = get_db()

            # Create staff account
            db.execute('''
                INSERT INTO staff
                (school_id, staff_id, password_hash, full_name, email, phone, department, position)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (school_id, existing_user_id, password, full_name, email, phone, department, position))

            db.commit()

            return jsonify({
                'success': True,
                'message': f'Staff account created for existing biometric user {existing_user_id}',
                'action': 'created_from_existing'
            })

        except sqlite3.IntegrityError:
            return jsonify({'success': False, 'error': 'Staff ID already exists in database'})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    else:
        return jsonify({
            'success': False,
            'error': 'Invalid action specified'
        })

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# Forgot Password Routes
@app.route('/forgot_password_request', methods=['POST'])
def forgot_password_request():
    """Handle staff forgot password requests"""
    try:
        school_id = request.form.get('school_id')
        user_id = request.form.get('user_id')
        reason = request.form.get('reason', '')
        
        if not school_id or not user_id:
            return jsonify({'success': False, 'error': 'Institution and User ID are required'})
        
        db = get_db()
        
        # Check if staff exists
        staff = db.execute('''
            SELECT id, full_name, staff_id FROM staff 
            WHERE school_id = ? AND staff_id = ?
        ''', (school_id, user_id)).fetchone()
        
        if not staff:
            return jsonify({'success': False, 'error': 'Staff member not found with the provided details'})
        
        # Check if there's already a pending request
        existing_request = db.execute('''
            SELECT id FROM password_reset_requests 
            WHERE staff_id = ? AND status = 'pending'
        ''', (staff['id'],)).fetchone()
        
        if existing_request:
            return jsonify({'success': False, 'error': 'You already have a pending password reset request'})
        
        # Create password reset request
        db.execute('''
            INSERT INTO password_reset_requests (staff_id, school_id, staff_name, staff_user_id, reason, status, created_at)
            VALUES (?, ?, ?, ?, ?, 'pending', datetime('now'))
        ''', (staff['id'], school_id, staff['full_name'], staff['staff_id'], reason))
        db.commit()
        
        return jsonify({'success': True, 'message': 'Password reset request submitted successfully'})
        
    except Exception as e:
        print(f"Error in forgot password request: {e}")
        return jsonify({'success': False, 'error': 'An error occurred while processing your request'})

@app.route('/admin/password_reset_requests')
def admin_password_reset_requests():
    """Admin page to view and manage password reset requests with month filtering"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))
    
    try:
        from datetime import datetime, timedelta
        import calendar
        
        db = get_db()
        school_id = session.get('school_id')
        
        # Get filter parameters
        month_filter = request.args.get('month_filter', 'current')  # Default to current month
        custom_month = request.args.get('custom_month', '')
        
        # Build the base query
        base_query = '''
            SELECT pr.*, s.department, s.position
            FROM password_reset_requests pr
            LEFT JOIN staff s ON pr.staff_id = s.id
            WHERE pr.school_id = ?
        '''
        query_params = [school_id]
        
        # Add month filtering
        if month_filter == 'current':
            # Current month
            now = datetime.now()
            start_date = datetime(now.year, now.month, 1)
            # Get last day of current month
            last_day = calendar.monthrange(now.year, now.month)[1]
            end_date = datetime(now.year, now.month, last_day, 23, 59, 59)
            
            base_query += ' AND pr.created_at >= ? AND pr.created_at <= ?'
            query_params.extend([start_date.strftime('%Y-%m-%d %H:%M:%S'), 
                               end_date.strftime('%Y-%m-%d %H:%M:%S')])
            
        elif month_filter == 'previous':
            # Previous month
            now = datetime.now()
            if now.month == 1:
                prev_month = 12
                prev_year = now.year - 1
            else:
                prev_month = now.month - 1
                prev_year = now.year
                
            start_date = datetime(prev_year, prev_month, 1)
            # Get last day of previous month
            last_day = calendar.monthrange(prev_year, prev_month)[1]
            end_date = datetime(prev_year, prev_month, last_day, 23, 59, 59)
            
            base_query += ' AND pr.created_at >= ? AND pr.created_at <= ?'
            query_params.extend([start_date.strftime('%Y-%m-%d %H:%M:%S'), 
                               end_date.strftime('%Y-%m-%d %H:%M:%S')])
            
        elif month_filter == 'custom' and custom_month:
            # Custom month from input (format: YYYY-MM)
            try:
                year, month = map(int, custom_month.split('-'))
                start_date = datetime(year, month, 1)
                # Get last day of selected month
                last_day = calendar.monthrange(year, month)[1]
                end_date = datetime(year, month, last_day, 23, 59, 59)
                
                base_query += ' AND pr.created_at >= ? AND pr.created_at <= ?'
                query_params.extend([start_date.strftime('%Y-%m-%d %H:%M:%S'), 
                                   end_date.strftime('%Y-%m-%d %H:%M:%S')])
            except (ValueError, IndexError):
                flash('Invalid custom month format', 'error')
                return redirect(url_for('admin_dashboard'))
        
        # Add ordering
        base_query += ' ORDER BY pr.created_at DESC'
        
        # Execute query
        requests = db.execute(base_query, query_params).fetchall()
        
        return render_template('admin_password_reset_requests.html', 
                             requests=requests, 
                             month_filter=month_filter, 
                             custom_month=custom_month)
        
    except Exception as e:
        print(f"Error loading password reset requests: {e}")
        flash('Error loading password reset requests', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/approve_password_reset', methods=['POST'])
def approve_password_reset():
    """Approve or reject password reset request"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        request_id = request.form.get('request_id')
        action = request.form.get('action')  # 'approve' or 'reject'
        new_password = request.form.get('new_password', '')
        
        if not request_id or not action:
            return jsonify({'success': False, 'error': 'Missing parameters'})
        
        db = get_db()
        school_id = session.get('school_id')
        
        # Get the request details
        reset_request = db.execute('''
            SELECT * FROM password_reset_requests 
            WHERE id = ? AND school_id = ?
        ''', (request_id, school_id)).fetchone()
        
        if not reset_request:
            return jsonify({'success': False, 'error': 'Request not found'})
        
        if action == 'approve':
            if not new_password:
                return jsonify({'success': False, 'error': 'New password is required for approval'})
            
            # Update staff password
            password_hash = generate_password_hash(new_password)
            db.execute('''
                UPDATE staff SET password_hash = ? WHERE id = ?
            ''', (password_hash, reset_request['staff_id']))
            
            # Update request status
            db.execute('''
                UPDATE password_reset_requests 
                SET status = 'approved', approved_at = datetime('now'), approved_by = ?
                WHERE id = ?
            ''', (session['user_id'], request_id))
            
            message = f"Password reset approved and new password set for {reset_request['staff_name']}"
            
        elif action == 'reject':
            # Update request status
            db.execute('''
                UPDATE password_reset_requests 
                SET status = 'rejected', approved_at = datetime('now'), approved_by = ?
                WHERE id = ?
            ''', (session['user_id'], request_id))
            
            message = f"Password reset request rejected for {reset_request['staff_name']}"
        
        else:
            return jsonify({'success': False, 'error': 'Invalid action'})
        
        db.commit()
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        print(f"Error in approve password reset: {e}")
        return jsonify({'success': False, 'error': 'An error occurred while processing the request'})

@app.route('/api/password_reset_count')
def get_password_reset_count():
    """API endpoint to get count of pending password reset requests"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        db = get_db()
        school_id = session.get('school_id')
        
        count = db.execute('''
            SELECT COUNT(*) as count FROM password_reset_requests 
            WHERE school_id = ? AND status = 'pending'
        ''', (school_id,)).fetchone()
        
        return jsonify({'success': True, 'count': count['count'] if count else 0})
        
    except Exception as e:
        print(f"Error getting password reset count: {e}")
        return jsonify({'success': False, 'error': 'An error occurred'})

@app.route('/api/password_reset_requests')
def get_password_reset_requests():
    """API endpoint to get password reset requests for AJAX loading with month filtering"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        from datetime import datetime, timedelta
        import calendar
        
        db = get_db()
        school_id = session.get('school_id')
        
        # Get filter parameters
        month_filter = request.args.get('month_filter', '')
        custom_month = request.args.get('custom_month', '')
        
        # Build the base query
        base_query = '''
            SELECT pr.*, s.department, s.position
            FROM password_reset_requests pr
            LEFT JOIN staff s ON pr.staff_id = s.id
            WHERE pr.school_id = ?
        '''
        query_params = [school_id]
        
        # Add month filtering
        if month_filter == 'current':
            # Current month
            now = datetime.now()
            start_date = datetime(now.year, now.month, 1)
            # Get last day of current month
            last_day = calendar.monthrange(now.year, now.month)[1]
            end_date = datetime(now.year, now.month, last_day, 23, 59, 59)
            
            base_query += ' AND pr.created_at >= ? AND pr.created_at <= ?'
            query_params.extend([start_date.strftime('%Y-%m-%d %H:%M:%S'), 
                               end_date.strftime('%Y-%m-%d %H:%M:%S')])
            
        elif month_filter == 'previous':
            # Previous month
            now = datetime.now()
            if now.month == 1:
                prev_month = 12
                prev_year = now.year - 1
            else:
                prev_month = now.month - 1
                prev_year = now.year
                
            start_date = datetime(prev_year, prev_month, 1)
            # Get last day of previous month
            last_day = calendar.monthrange(prev_year, prev_month)[1]
            end_date = datetime(prev_year, prev_month, last_day, 23, 59, 59)
            
            base_query += ' AND pr.created_at >= ? AND pr.created_at <= ?'
            query_params.extend([start_date.strftime('%Y-%m-%d %H:%M:%S'), 
                               end_date.strftime('%Y-%m-%d %H:%M:%S')])
            
        elif month_filter == 'custom' and custom_month:
            # Custom month from input (format: YYYY-MM)
            try:
                year, month = map(int, custom_month.split('-'))
                start_date = datetime(year, month, 1)
                # Get last day of selected month
                last_day = calendar.monthrange(year, month)[1]
                end_date = datetime(year, month, last_day, 23, 59, 59)
                
                base_query += ' AND pr.created_at >= ? AND pr.created_at <= ?'
                query_params.extend([start_date.strftime('%Y-%m-%d %H:%M:%S'), 
                                   end_date.strftime('%Y-%m-%d %H:%M:%S')])
            except (ValueError, IndexError):
                return jsonify({'success': False, 'error': 'Invalid custom month format'})
        
        # Add ordering
        base_query += ' ORDER BY pr.created_at DESC'
        
        # Execute query
        requests = db.execute(base_query, query_params).fetchall()
        
        # Convert to list of dictionaries for JSON response
        requests_list = []
        for req in requests:
            requests_list.append({
                'id': req['id'],
                'staff_id': req['staff_id'],
                'staff_name': req['staff_name'],
                'staff_user_id': req['staff_user_id'],
                'department': req['department'],
                'position': req['position'],
                'reason': req['reason'],
                'status': req['status'],
                'created_at': req['created_at'],
                'approved_at': req['approved_at']
            })
        
        return jsonify({'success': True, 'requests': requests_list})
        
    except Exception as e:
        print(f"Error loading password reset requests: {e}")
        return jsonify({'success': False, 'error': 'An error occurred while loading requests'})

# Database migration - Add created_at column to staff table (commented out to prevent errors)
# This should be run only once during initial setup
# import sqlite3
# conn = sqlite3.connect('vishnorex.db')
# try:
#     conn.execute("ALTER TABLE staff ADD COLUMN created_at TEXT")
#     conn.commit()
# except sqlite3.OperationalError:
#     pass  # Column already exists
# conn.close()

# Backup and Data Management Routes
@app.route('/create_backup', methods=['POST'])
def create_backup():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    backup_name = request.form.get('backup_name')
    include_logs = request.form.get('include_logs', True, type=bool)

    backup_manager = BackupManager()
    result = backup_manager.create_database_backup(backup_name, include_logs)

    return jsonify(result)

@app.route('/get_backup_list')
def get_backup_list():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        backup_manager = BackupManager()
        backup_dir = backup_manager.backup_dir

        backups = []
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.db'):
                    backup_name = filename[:-3]  # Remove .db extension
                    backup_path = os.path.join(backup_dir, filename)
                    metadata_path = os.path.join(backup_dir, f"{backup_name}_metadata.json")

                    backup_info = {
                        'name': backup_name,
                        'size': os.path.getsize(backup_path),
                        'created_at': datetime.datetime.fromtimestamp(os.path.getctime(backup_path)).isoformat()
                    }

                    # Load metadata if available
                    if os.path.exists(metadata_path):
                        try:
                            with open(metadata_path, 'r') as f:
                                metadata = json.load(f)
                                backup_info.update(metadata)
                        except:
                            pass

                    backups.append(backup_info)

        # Sort by creation date (newest first)
        backups.sort(key=lambda x: x['created_at'], reverse=True)

        return jsonify({'success': True, 'backups': backups})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export_data_backup', methods=['POST'])
def export_data_backup():
    if 'user_id' not in session or (session.get('user_type') != 'admin' and not session.get('is_sub_admin')):
        return jsonify({'success': False, 'error': 'Unauthorized'})

    export_type = request.form.get('export_type', 'excel')
    school_id = session.get('school_id')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    tables = request.form.getlist('tables')

    backup_manager = BackupManager()
    result = backup_manager.export_data(export_type, school_id, start_date, end_date, tables)

    if result['success']:
        # Return file for download
        from flask import send_file
        return send_file(result['export_path'], as_attachment=True)
    else:
        return jsonify(result)

# Salary Calculation Routes
@app.route('/calculate_salary', methods=['POST'])
def calculate_salary():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id', type=int)
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not all([staff_id, year, month]):
        return jsonify({'success': False, 'error': 'Staff ID, year, and month are required'})

    school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=school_id)
    result = salary_calculator.calculate_monthly_salary(staff_id, year, month)

    return jsonify(result)

@app.route('/generate_salary_report', methods=['POST'])
def generate_salary_report():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id', type=int)
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not all([staff_id, year, month]):
        return jsonify({'success': False, 'error': 'Staff ID, year, and month are required'})

    school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=school_id)
    result = salary_calculator.generate_salary_report(staff_id, year, month)

    return jsonify(result)

@app.route('/download_salary_slip', methods=['POST'])
def download_salary_slip():
    """Generate and download individual salary slip PDF"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id', type=int)
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not all([staff_id, year, month]):
        return jsonify({'success': False, 'error': 'Staff ID, year, and month are required'})

    try:
        school_id = session.get('school_id')
        salary_calculator = SalaryCalculator(school_id=school_id)
        
        # Get salary data
        result = salary_calculator.generate_salary_report(staff_id, year, month)
        
        if not result['success']:
            return jsonify({'success': False, 'error': result.get('error', 'Failed to generate salary data')})
        
        # Generate PDF salary slip
        pdf_response = generate_individual_salary_slip_pdf(result, year, month)
        return pdf_response
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to generate salary slip: {str(e)}'})

@app.route('/bulk_salary_calculation', methods=['POST'])
def bulk_salary_calculation():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)
    department = request.form.get('department')
    shift_type = request.form.get('shift')
    gender = request.form.get('gender')
    search = request.form.get('search')
    school_id = session.get('school_id')

    if not all([year, month]):
        return jsonify({'success': False, 'error': 'Year and month are required'})

    try:
        # Persist latest salary calculation filters so Payroll Review can load matching records.
        session['last_salary_calc_filters'] = {
            'year': year,
            'month': month,
            'department': (department or '').strip(),
            'shift': (shift_type or '').strip(),
            'gender': (gender or '').strip(),
            'search': (search or '').strip()
        }
        session['last_salary_calc_meta'] = {
            'calculated_at': datetime.datetime.now().isoformat(timespec='microseconds'),
            'calculated_by': session.get('full_name') or session.get('username') or session.get('user_id')
        }

        db = get_db()
        staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]

        # Get staff list based on filters
        query = 'SELECT id, staff_id, full_name, department FROM staff WHERE school_id = ?'
        params = [school_id]

        if 'is_active' in staff_columns:
            query += ' AND COALESCE(is_active, 1) = 1'
        elif 'status' in staff_columns:
            query += " AND LOWER(COALESCE(status, 'active')) = 'active'"

        if department:
            query += ' AND department = ?'
            params.append(department)
            
        query += ' ORDER BY CAST(staff_id AS INTEGER) ASC'

        staff_list = db.execute(query, params).fetchall()

        school_id = session.get('school_id')
        salary_calculator = SalaryCalculator(school_id=school_id)
        results = []

        for staff in staff_list:
            salary_result = salary_calculator.calculate_monthly_salary(staff['id'], year, month)
            if salary_result['success']:
                breakdown = salary_result['salary_breakdown']
                earnings = breakdown.get('earnings', {})
                deductions = breakdown.get('deductions', {})
                manual_bonus = float(breakdown.get('manual_bonus', 0) or 0)
                manual_deduction = float(breakdown.get('manual_deduction', 0) or 0)
                adjusted_total_earnings = float(breakdown.get('adjusted_total_earnings', earnings.get('total_earnings', 0)) or 0)
                adjusted_total_deductions = float(breakdown.get('adjusted_total_deductions', deductions.get('total_deductions', 0)) or 0)
                adjusted_net_salary = float(adjusted_total_earnings - adjusted_total_deductions)

                results.append({
                    'id': staff['id'],  # Add database ID
                    'staff_id': staff['staff_id'],
                    'staff_name': staff['full_name'],
                    'department': staff['department'],
                    'net_salary': adjusted_net_salary,
                    'total_earnings': adjusted_total_earnings,
                    'total_deductions': adjusted_total_deductions,
                    'manual_bonus': manual_bonus,
                    'manual_deduction': manual_deduction,
                    'present_days': breakdown['attendance_summary']['present_days'],
                    'absent_days': breakdown['attendance_summary']['absent_days']
                })

        total_earnings = sum(float(row.get('total_earnings', 0) or 0) for row in results)
        total_deductions = sum(float(row.get('total_deductions', 0) or 0) for row in results)
        total_net_salary = sum(float(row.get('net_salary', 0) or 0) for row in results)

        db.execute('''
            CREATE TABLE IF NOT EXISTS payroll_calculation_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                school_id INTEGER NOT NULL,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                calculated_at TIMESTAMP NOT NULL,
                calculated_by TEXT,
                total_staff INTEGER NOT NULL DEFAULT 0,
                total_earnings REAL NOT NULL DEFAULT 0,
                total_deductions REAL NOT NULL DEFAULT 0,
                total_net_salary REAL NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(school_id, year, month)
            )
        ''')

        calculated_at = datetime.datetime.now().isoformat(timespec='microseconds')
        calculated_by = session.get('full_name') or session.get('username') or str(session.get('user_id'))
        existing_month_run = db.execute('''
            SELECT id
            FROM payroll_calculation_runs
            WHERE school_id = ? AND year = ? AND month = ?
        ''', (school_id, year, month)).fetchone()

        if existing_month_run:
            db.execute('''
                UPDATE payroll_calculation_runs
                SET calculated_at = ?,
                    calculated_by = ?,
                    total_staff = ?,
                    total_earnings = ?,
                    total_deductions = ?,
                    total_net_salary = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (
                calculated_at,
                str(calculated_by),
                len(results),
                total_earnings,
                total_deductions,
                total_net_salary,
                existing_month_run['id']
            ))
        else:
            db.execute('''
                INSERT INTO payroll_calculation_runs
                (school_id, year, month, calculated_at, calculated_by, total_staff, total_earnings, total_deductions, total_net_salary)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                school_id,
                year,
                month,
                calculated_at,
                str(calculated_by),
                len(results),
                total_earnings,
                total_deductions,
                total_net_salary
            ))

        db.commit()

        return jsonify({
            'success': True,
            'calculation_period': f"{calendar.month_name[month]} {year}",
            'total_staff': len(results),
            'salary_calculations': results
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update_salary_rules', methods=['POST'])
def update_salary_rules():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        new_rules = {}

        # Get salary rule updates from form
        rule_fields = [
            'early_arrival_bonus_per_hour',
            'early_departure_penalty_per_hour',
            'late_arrival_penalty_per_hour',
            'single_punch_penalty_rate',
            'absent_day_deduction_rate',
            'overtime_rate_multiplier',
            'on_duty_rate',
            'bonus_rate_percentage',
            'minimum_hours_for_bonus'
        ]

        for field in rule_fields:
            value = request.form.get(field, type=float)
            if value is not None:
                new_rules[field] = value

        school_id = session.get('school_id')
        salary_calculator = SalaryCalculator(school_id=school_id)
        result = salary_calculator.update_salary_rules(new_rules)

        return jsonify(result)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/calculate_hourly_rate', methods=['POST'])
def calculate_hourly_rate_api():
    """Calculate hourly rate from base monthly salary"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        from database import calculate_hourly_rate

        base_salary = request.json.get('base_salary', 0)
        if not base_salary or base_salary <= 0:
            return jsonify({'success': False, 'error': 'Valid base salary is required'})

        result = calculate_hourly_rate(base_salary)
        return jsonify({'success': True, 'data': result})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/calculate_enhanced_salary', methods=['POST'])
def calculate_enhanced_salary_api():
    """Calculate enhanced salary based on actual hours worked"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        staff_id = request.json.get('staff_id')
        year = request.json.get('year')
        month = request.json.get('month')

        if not all([staff_id, year, month]):
            return jsonify({'success': False, 'error': 'Staff ID, year, and month are required'})

        school_id = session.get('school_id')
        salary_calculator = SalaryCalculator(school_id=school_id)
        result = salary_calculator.calculate_enhanced_monthly_salary(int(staff_id), int(year), int(month))

        return jsonify(result)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/get_staff_hourly_rate/<int:staff_id>')
def get_staff_hourly_rate(staff_id):
    """Get hourly rate for a specific staff member"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        from database import calculate_hourly_rate

        db = get_db()
        school_id = session['school_id']

        # Get staff basic salary
        staff = db.execute('''
            SELECT basic_salary, full_name
            FROM staff
            WHERE id = ? AND school_id = ?
        ''', (staff_id, school_id)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff member not found'})

        basic_salary = float(staff['basic_salary'] or 0)
        if basic_salary <= 0:
            return jsonify({'success': False, 'error': 'Base salary not set for this staff member'})

        hourly_rate_info = calculate_hourly_rate(basic_salary)

        return jsonify({
            'success': True,
            'staff_name': staff['full_name'],
            'basic_salary': basic_salary,
            'hourly_rate_info': hourly_rate_info
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_salary_rules')
def get_salary_rules():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=school_id)

    return jsonify({
        'success': True,
        'salary_rules': salary_calculator.salary_rules
    })

@app.route('/export_salary_calculation_results', methods=['POST'])
def export_salary_calculation_results():
    """Export salary calculation results to Excel"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        # Get the calculation parameters from the request
        year = request.form.get('year', type=int)
        month = request.form.get('month', type=int)
        department = request.form.get('department')
        export_format = request.form.get('format', 'excel').lower()

        if not year or not month:
            return jsonify({'success': False, 'error': 'Year and month are required'})

        school_id = session.get('school_id')

        # Re-run the bulk salary calculation to get fresh data
        salary_calculator = SalaryCalculator(school_id=school_id)

        db = get_db()
        staff_columns = [col['name'] for col in db.execute("PRAGMA table_info(staff)").fetchall()]

        # Get staff list based on filters
        query = 'SELECT id, staff_id, full_name, department FROM staff WHERE school_id = ?'
        params = [school_id]

        if 'is_active' in staff_columns:
            query += ' AND COALESCE(is_active, 1) = 1'
        elif 'status' in staff_columns:
            query += " AND LOWER(COALESCE(status, 'active')) = 'active'"

        if department:
            query += ' AND department = ?'
            params.append(department)

        staff_list = db.execute(query, params).fetchall()

        # Calculate salaries for all staff
        salary_results = []
        for staff in staff_list:
            salary_result = salary_calculator.calculate_monthly_salary(staff['id'], year, month)
            if salary_result['success']:
                salary_results.append({
                    'id': staff['id'],
                    'staff_id': staff['staff_id'],
                    'staff_name': staff['full_name'],
                    'department': staff['department'],
                    'salary_data': salary_result
                })

        # Generate report in requested format
        if export_format == 'excel':
            return generate_salary_calculation_excel(salary_results, year, month, department)
        elif export_format == 'csv':
            return generate_salary_calculation_csv(salary_results, year, month, department)
        elif export_format == 'pdf':
            return generate_salary_calculation_pdf(salary_results, year, month, department)
        else:
            return jsonify({'success': False, 'error': 'Supported formats: excel, csv, pdf'})

    except Exception as e:
        return jsonify({'success': False, 'error': f'Export failed: {str(e)}'})

def generate_salary_calculation_excel(salary_results, year, month, department=None):
    """Generate Excel report for salary calculation results"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import io
    import calendar

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Calculation Results"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    title = f"Salary Calculation Results - {calendar.month_name[month]} {year}"
    if department:
        title += f" - {department} Department"

    ws.merge_cells('A1:P1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        'Staff ID', 'Staff Name', 'Department', 'Base Salary', 'Present Days', 'Absent Days',
        'Total Earnings', 'Bonuses', 'Allowances', 'Total Deductions', 'Penalties',
        'Other Deductions', 'Net Salary', 'Hourly Rate', 'Hours Worked', 'Attendance Rate'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = 4
    total_earnings = 0
    total_deductions = 0
    total_net_salary = 0

    for result in salary_results:
        salary_data = result['salary_data']
        breakdown = salary_data['salary_breakdown']
        attendance = breakdown['attendance_summary']
        earnings = breakdown['earnings']
        deductions = breakdown['deductions']

        # Calculate attendance rate
        total_days = attendance['present_days'] + attendance['absent_days']
        attendance_rate = (attendance['present_days'] / total_days * 100) if total_days > 0 else 0

        # Row data
        row_data = [
            result['staff_id'],
            result['staff_name'],
            result['department'],
            breakdown.get('base_salary', 0),
            attendance['present_days'],
            attendance['absent_days'],
            earnings['total_earnings'],
            earnings.get('bonuses', 0),
            earnings.get('allowances', 0),
            deductions['total_deductions'],
            deductions.get('penalties', 0),
            deductions.get('other_deductions', 0),
            breakdown['net_salary'],
            salary_data.get('hourly_rate', 0),
            salary_data.get('actual_hours_worked', 0),
            f"{attendance_rate:.1f}%"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col >= 4 and col <= 15:  # Numeric columns
                cell.alignment = Alignment(horizontal='right')
                if isinstance(value, (int, float)) and col != 16:  # Not percentage
                    cell.number_format = '#,##0.00'

        # Update totals
        total_earnings += earnings['total_earnings']
        total_deductions += deductions['total_deductions']
        total_net_salary += breakdown['net_salary']
        row += 1

    # Summary row
    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=summary_row, column=7, value=total_earnings).font = Font(bold=True)
    ws.cell(row=summary_row, column=10, value=total_deductions).font = Font(bold=True)
    ws.cell(row=summary_row, column=13, value=total_net_salary).font = Font(bold=True)

    # Format summary row
    for col in [7, 10, 13]:
        cell = ws.cell(row=summary_row, column=col)
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'salary_calculation_results_{year}_{month:02d}'
    if department:
        filename += f'_{department.replace(" ", "_")}'
    filename += f'_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def generate_salary_calculation_csv(salary_results, year, month, department=None):
    """Generate CSV report for salary calculation results"""
    import csv
    import io
    import calendar

    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)

    # Title
    title = f"Salary Calculation Results - {calendar.month_name[month]} {year}"
    if department:
        title += f" - {department} Department"
    writer.writerow([title])
    writer.writerow([])  # Empty row

    # Headers
    headers = [
        'Staff ID', 'Staff Name', 'Department', 'Base Salary', 'Present Days', 'Absent Days',
        'Total Earnings', 'Bonuses', 'Allowances', 'Total Deductions', 'Penalties',
        'Other Deductions', 'Net Salary', 'Hourly Rate', 'Hours Worked', 'Attendance Rate'
    ]
    writer.writerow(headers)

    # Data rows
    total_earnings = 0
    total_deductions = 0
    total_net_salary = 0

    for result in salary_results:
        salary_data = result['salary_data']
        breakdown = salary_data['salary_breakdown']
        attendance = breakdown['attendance_summary']
        earnings = breakdown['earnings']
        deductions = breakdown['deductions']

        # Calculate attendance rate
        total_days = attendance['present_days'] + attendance['absent_days']
        attendance_rate = (attendance['present_days'] / total_days * 100) if total_days > 0 else 0

        # Row data
        row_data = [
            result['staff_id'],
            result['staff_name'],
            result['department'],
            breakdown.get('base_salary', 0),
            attendance['present_days'],
            attendance['absent_days'],
            earnings['total_earnings'],
            earnings.get('bonuses', 0),
            earnings.get('allowances', 0),
            deductions['total_deductions'],
            deductions.get('penalties', 0),
            deductions.get('other_deductions', 0),
            breakdown['net_salary'],
            salary_data.get('hourly_rate', 0),
            salary_data.get('actual_hours_worked', 0),
            f"{attendance_rate:.1f}%"
        ]

        writer.writerow(row_data)

        # Update totals
        total_earnings += earnings['total_earnings']
        total_deductions += deductions['total_deductions']
        total_net_salary += breakdown['net_salary']

    # Summary row
    writer.writerow([])  # Empty row
    summary_row = ['TOTALS', '', '', '', '', '', total_earnings, '', '', total_deductions, '', '', total_net_salary, '', '', '']
    writer.writerow(summary_row)

    # Create response
    csv_content = output.getvalue()
    output.close()

    response = make_response(csv_content)
    filename = f'salary_calculation_results_{year}_{month:02d}'
    if department:
        filename += f'_{department.replace(" ", "_")}'
    filename += f'_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def generate_salary_calculation_pdf(salary_results, year, month, department=None):
    """Generate PDF report for salary calculation results"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io
        import calendar

        # Create PDF buffer
        buffer = io.BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )

        # Title
        title = f"Salary Calculation Results<br/>{calendar.month_name[month]} {year}"
        if department:
            title += f"<br/>{department} Department"

        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 12))

        # Prepare table data
        table_data = [
            ['Staff ID', 'Name', 'Dept', 'Base Salary', 'Present', 'Absent', 'Earnings', 'Deductions', 'Net Salary']
        ]

        total_earnings = 0
        total_deductions = 0
        total_net_salary = 0

        for result in salary_results:
            salary_data = result['salary_data']
            breakdown = salary_data['salary_breakdown']
            attendance = breakdown['attendance_summary']
            earnings = breakdown['earnings']
            deductions = breakdown['deductions']

            row_data = [
                result['staff_id'],
                result['staff_name'][:15] + '...' if len(result['staff_name']) > 15 else result['staff_name'],
                result['department'][:8] + '...' if len(result['department']) > 8 else result['department'],
                f"₹{breakdown.get('base_salary', 0):,.0f}",
                str(attendance['present_days']),
                str(attendance['absent_days']),
                f"₹{earnings['total_earnings']:,.0f}",
                f"₹{deductions['total_deductions']:,.0f}",
                f"₹{breakdown['net_salary']:,.0f}"
            ]

            table_data.append(row_data)

            # Update totals
            total_earnings += earnings['total_earnings']
            total_deductions += deductions['total_deductions']
            total_net_salary += breakdown['net_salary']

        # Add totals row
        table_data.append([
            'TOTALS', '', '', '', '', '',
            f"₹{total_earnings:,.0f}",
            f"₹{total_deductions:,.0f}",
            f"₹{total_net_salary:,.0f}"
        ])

        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        # Get PDF content
        pdf_content = buffer.getvalue()
        buffer.close()

        # Create response
        response = make_response(pdf_content)
        filename = f'salary_calculation_results_{year}_{month:02d}'
        if department:
            filename += f'_{department.replace(" ", "_")}'
        filename += f'_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

        return response

    except ImportError:
        # If reportlab is not installed, return error
        return jsonify({
            'success': False,
            'error': 'PDF generation requires reportlab library. Please install it with: pip install reportlab'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

def generate_individual_salary_slip_pdf(salary_data, year, month, payroll_review_status='pending'):
    """Generate PDF salary slip for an individual employee"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io
        import calendar

        # Extract data
        staff_info = salary_data['staff_info']
        breakdown = salary_data['salary_breakdown']
        attendance = breakdown['attendance_summary']
        earnings = breakdown['earnings']
        deductions = breakdown['deductions']

        # Ensure core values exist even if an older salary breakdown is returned
        if 'per_day_salary' not in breakdown:
            gross_salary = (
                float(staff_info.get('basic_salary', 0) or 0) +
                float(staff_info.get('hra', 0) or 0) +
                float(staff_info.get('transport_allowance', 0) or 0) +
                float(staff_info.get('other_allowances', 0) or 0)
            )
            working_days = breakdown.get('working_days') or 0
            breakdown['per_day_salary'] = round(gross_salary / working_days, 2) if working_days else 0.0
        if 'per_hour_salary' not in breakdown:
            breakdown['per_hour_salary'] = round((breakdown.get('per_day_salary', 0) or 0) / 8, 2)
        
        # Keep salary slip math aligned with Payroll Processing & Review.
        total_earnings = float(earnings.get('total_earnings', 0) or 0)
        total_deductions = float(deductions.get('total_deductions', 0) or 0)
        manual_bonus = float(breakdown.get('manual_bonus', 0) or 0)
        manual_deduction = float(breakdown.get('manual_deduction', 0) or 0)
        adjusted_total_earnings = total_earnings + manual_bonus
        adjusted_total_deductions = total_deductions + manual_deduction
        net_salary = adjusted_total_earnings - adjusted_total_deductions

        # Create PDF buffer
        buffer = io.BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center alignment
            textColor=colors.darkblue
        )
        
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            textColor=colors.darkblue
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )

        # Title
        month_name = calendar.month_name[month]
        title = f"SALARY SLIP<br/>{month_name} {year}"
        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 8))

        workflow_steps = [
            ('auto_generated', 'Auto Generated'),
            ('reviewed', 'Reviewed'),
            ('completed', 'Completed'),
            ('pay_released', 'Pay Released')
        ]
        status_value = str(payroll_review_status or 'pending').strip().lower()
        if status_value in ('review', 'reviewed'):
            status_value = 'reviewed'
        elif status_value in ('complete', 'completed'):
            status_value = 'completed'
        elif status_value in ('pay released', 'pay_released', 'released', 'paid'):
            status_value = 'pay_released'
        else:
            status_value = 'pending'

        active_counts = {
            'pending': 1,
            'reviewed': 2,
            'completed': 3,
            'pay_released': 4
        }
        active_count = active_counts.get(status_value, 1)
        workflow_labels = [label for _, label in workflow_steps]
        workflow_table = Table([workflow_labels], colWidths=[1.35 * inch, 1.15 * inch, 1.1 * inch, 1.1 * inch])

        workflow_style_cmds = [
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#64748b')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f1f5f9')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]

        for i in range(active_count):
            workflow_style_cmds.append(('BACKGROUND', (i, 0), (i, 0), colors.HexColor('#dcfce7')))
            workflow_style_cmds.append(('TEXTCOLOR', (i, 0), (i, 0), colors.HexColor('#166534')))

        workflow_table.setStyle(TableStyle(workflow_style_cmds))
        elements.append(workflow_table)
        elements.append(Spacer(1, 20))

        # Employee Information Section
        emp_info_data = [
            ['Employee Information', ''],
            ['Name:', staff_info['full_name']],
            ['Staff ID:', staff_info['staff_id']],
            ['Department:', staff_info['department']],
            ['Position:', staff_info.get('position', 'N/A')],
            ['Bank A/C Name:', staff_info.get('bank_account_name') or 'N/A'],
            ['Bank Name:', staff_info.get('bank_name') or 'N/A'],
            ['Account Number:', staff_info.get('bank_account_number') or 'N/A'],
            ['IFSC Code:', staff_info.get('ifsc_code') or 'N/A'],
            ['Calculation Period:', breakdown['calculation_period']]
        ]

        emp_table = Table(emp_info_data, colWidths=[2*inch, 3*inch])
        emp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(emp_table)
        elements.append(Spacer(1, 20))

        # Attendance Summary
        attendance_data = [
            ['Attendance Summary', ''],
            ['Working Days:', str(breakdown['working_days'])],
            ['Present Days:', str(attendance['present_days'])],
            ['Absent Days:', str(attendance['absent_days'])],
            ['On Duty Days:', str(attendance['on_duty_days'])],
            ['Leave Days:', str(attendance['leave_days'])],
            ['Admin Assigned Holidays:', str(attendance.get('holiday_days', 0))]
        ]

        attendance_table = Table(attendance_data, colWidths=[2*inch, 3*inch])
        attendance_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(attendance_table)
        elements.append(Spacer(1, 20))

        # Holiday Details Section (if holidays exist)
        if attendance.get('holiday_days', 0) > 0 and attendance.get('holiday_list'):
            holiday_detail_data = [['Holiday Details', '']]
            
            for holiday in attendance['holiday_list']:
                holiday_name = holiday['name'][:30] + ('...' if len(holiday['name']) > 30 else '')  # Truncate long names
                holiday_type = holiday['type']
                holiday_detail_data.append([f"• {holiday_name}", holiday_type])
            
            holiday_table = Table(holiday_detail_data, colWidths=[3*inch, 2*inch])
            holiday_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(holiday_table)
            elements.append(Spacer(1, 20))

        # Earnings and Deductions Table with safe property access
        salary_data = [
            ['EARNINGS', 'Amount (₹)', 'DEDUCTIONS', 'Amount (₹)'],
            ['Basic Salary', f'{earnings.get("basic_salary", 0):,.2f}', 'Absent Days Deduction', f'{deductions.get("absent_deduction", 0):,.2f}'],
            ['HRA', f'{earnings.get("hra", 0):,.2f}', 'PF Deduction', f'{deductions.get("pf_deduction", 0):,.2f}'],
            ['Transport Allowance', f'{earnings.get("transport_allowance", 0):,.2f}', 'ESI Deduction', f'{deductions.get("esi_deduction", 0):,.2f}'],
            ['Other Allowances', f'{earnings.get("other_allowances", 0):,.2f}', 'Professional Tax', f'{deductions.get("professional_tax", 0):,.2f}'],
            ['Present Days Pay', f'{earnings.get("present_pay", 0):,.2f}', 'Single Punch Penalty', f'{deductions.get("single_punch_penalty", 0):,.2f}'],
            ['', '', 'Early Departure Penalty', f'{deductions.get("early_departure_penalty", 0):,.2f}'],
            ['On Duty Pay', f'{earnings.get("on_duty_pay", 0):,.2f}', 'Late Arrival Penalty', f'{deductions.get("late_arrival_penalty", 0):,.2f}'],
            ['Leave Pay', f'{earnings.get("leave_pay", 0):,.2f}', 'Other Deductions', f'{deductions.get("other_deductions", 0):,.2f}']
        ]

        # Add bonuses if they exist
        if earnings.get("early_arrival_bonus", 0) > 0:
            salary_data.append(['Early Arrival Bonus', f'{earnings.get("early_arrival_bonus", 0):,.2f}', '', ''])
        if earnings.get("overtime_pay", 0) > 0:
            salary_data.append(['Overtime Pay', f'{earnings.get("overtime_pay", 0):,.2f}', '', ''])
        if manual_bonus > 0:
            salary_data.append(['Manual Bonus', f'{manual_bonus:,.2f}', '', ''])
        if manual_deduction > 0:
            salary_data.append(['', '', 'Manual Deduction', f'{manual_deduction:,.2f}'])

        # Add totals with calculated net salary
        salary_data.extend([
            ['', '', '', ''],
            ['TOTAL EARNINGS', f'{adjusted_total_earnings:,.2f}', 'TOTAL DEDUCTIONS', f'{adjusted_total_deductions:,.2f}'],
            ['', '', '', ''],
            [f'NET SALARY: {net_salary:,.2f}', '', '', '']
        ])

        salary_table = Table(salary_data, colWidths=[2.5*inch, 1*inch, 2.5*inch, 1*inch])
        salary_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Total rows
            ('BACKGROUND', (0, -4), (-1, -4), colors.lightgrey),
            ('FONTNAME', (0, -4), (-1, -4), 'Helvetica-Bold'),
            
            # Net salary row - span across all columns
            ('BACKGROUND', (0, -1), (-1, -1), colors.darkgreen),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('SPAN', (0, -1), (-1, -1)),
            ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ]))

        elements.append(salary_table)
        elements.append(Spacer(1, 30))

        # Footer
        footer_text = f"Generated on: {datetime.datetime.now().strftime('%d/%m/%Y at %I:%M %p')}<br/>This is a computer-generated salary slip and does not require a signature."
        footer_para = Paragraph(footer_text, normal_style)
        elements.append(footer_para)

        # Build PDF
        doc.build(elements)

        # Get PDF content
        pdf_content = buffer.getvalue()
        buffer.close()

        # Create response
        response = make_response(pdf_content)
        
        # Generate filename
        staff_name = staff_info['full_name'].replace(' ', '_')
        filename = f'SalarySlip_{staff_name}_{month_name}{year}.pdf'

        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

        return response

    except ImportError:
        return jsonify({
            'success': False,
            'error': 'PDF generation requires reportlab library. Please install it with: pip install reportlab'
        })
    except Exception as e:
        print(f"ERROR in PDF generation: {str(e)}")
        print(f"Salary data structure: {salary_data}")
        return jsonify({'success': False, 'error': f'Salary slip generation failed: {str(e)}'})

@app.route('/get_staff_count')
def get_staff_count():
    """Get total staff count for sidebar stats"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        db = get_db()
        school_id = session.get('school_id', 1)

        count = db.execute(
            'SELECT COUNT(*) as total FROM staff WHERE school_id = ?',
            (school_id,)
        ).fetchone()

        return jsonify({
            'success': True,
            'count': count['total'] if count else 0
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/reports')
@requires_permission('reports', 'view')
def admin_reports():
    """Admin reports page"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and session.get('user_type') != 'company_admin' and not session.get('is_sub_admin')):
        return redirect(url_for('index'))

    school_id = session.get('school_id')
    today = datetime.datetime.now()
    
    # Get module settings for navigation
    module_enabled = get_module_enabled(school_id) if school_id else {}
    
    return render_template('admin_reports.html', current_year=today.year, module_enabled=module_enabled)

@app.route('/admin/settings')
def admin_settings():
    """Admin settings page"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('index'))

    return render_template('admin_settings.html')

@app.route('/api/get_institution_timings', methods=['GET'])
def get_institution_timings():
    """Get current institution check-in and check-out timings from the general shift definition."""
    try:
        db = get_db()
        school_id = session.get('school_id')
        if not school_id:
            return jsonify({
                'success': False,
                'message': 'School context is required'
            }), 400

        # Read directly from shift_definitions (general shift is the single source of truth)
        row = db.execute("""
            SELECT start_time, end_time FROM shift_definitions
            WHERE school_id = ? AND shift_type = 'general' AND is_active = 1
            LIMIT 1
        """, (school_id,)).fetchone()

        if row:
            # Strip seconds part if present (HH:MM:SS -> HH:MM)
            checkin = row['start_time'][:5] if row['start_time'] else '09:00'
            checkout = row['end_time'][:5] if row['end_time'] else '17:00'
        else:
            checkin, checkout = '09:00', '17:00'

        return jsonify({
            'success': True,
            'checkin_time': checkin,
            'checkout_time': checkout
        })

    except Exception as e:
        print(f"Error getting institution timings: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to retrieve institution timings'
        }), 500

@app.route('/api/update_institution_timings', methods=['POST'])
def update_institution_timings():
    """Update institution check-in and check-out timings"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        checkin_time = request.form.get('checkin_time')
        checkout_time = request.form.get('checkout_time')

        # Validate inputs
        if not checkin_time or not checkout_time:
            return jsonify({
                'success': False,
                'message': 'Both check-in and check-out times are required'
            }), 400

        # Validate time format (HH:MM)
        import re
        time_pattern = re.compile(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$')

        if not time_pattern.match(checkin_time) or not time_pattern.match(checkout_time):
            return jsonify({
                'success': False,
                'message': 'Invalid time format. Please use HH:MM format'
            }), 400

        # Validate that checkout is after checkin
        from datetime import datetime
        checkin_dt = datetime.strptime(checkin_time, '%H:%M')
        checkout_dt = datetime.strptime(checkout_time, '%H:%M')

        if checkout_dt <= checkin_dt:
            return jsonify({
                'success': False,
                'message': 'Check-out time must be later than check-in time'
            }), 400

        db = get_db()
        school_id = session.get('school_id')
        if not school_id:
            return jsonify({
                'success': False,
                'message': 'School context is required'
            }), 400

        # Write directly to shift_definitions (general shift) — single source of truth
        db.execute("""
            UPDATE shift_definitions
            SET start_time = ?, end_time = ?
            WHERE school_id = ? AND shift_type = 'general' AND is_active = 1
        """, (checkin_time + ':00', checkout_time + ':00', school_id))

        if db.execute("SELECT changes()").fetchone()[0] == 0:
            # Row didn't exist yet — insert it
            db.execute("""
                INSERT INTO shift_definitions
                (school_id, shift_type, start_time, end_time, grace_period_minutes, description, is_active)
                VALUES (?, 'general', ?, ?, 15, 'General Institution Shift', 1)
            """, (school_id, checkin_time + ':00', checkout_time + ':00'))

        db.commit()
        print(f"General shift timing updated: {checkin_time} - {checkout_time}")

        # Notify all systems to refresh their configurations
        try:
            # Reload shift manager if it exists
            from shift_management import ShiftManager
            if hasattr(app, 'shift_manager'):
                app.shift_manager.reload_shift_definitions()
            else:
                # Create new shift manager to ensure latest timings are loaded
                app.shift_manager = ShiftManager()

            print(f"✅ Institution timings updated and synced across all systems")

        except Exception as reload_error:
            print(f"⚠️ Warning: Could not reload shift manager: {reload_error}")

        return jsonify({
            'success': True,
            'message': 'Institution timings updated successfully',
            'checkin_time': checkin_time,
            'checkout_time': checkout_time
        })

    except Exception as e:
        print(f"Error updating institution timings: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to update institution timings'
        }), 500

@app.route('/api/debug_session', methods=['GET'])
def debug_session():
    """Debug route to check session status and sub-admin permissions"""
    try:
        db = get_db()
        # Get all permissions from DB
        all_perms = db.execute('SELECT * FROM sub_admin_permissions').fetchall()
        perms_list = [dict(row) for row in all_perms]
        
        # Get all staff for comparison
        all_staff = db.execute('SELECT id, school_id, staff_id, full_name FROM staff LIMIT 10').fetchall()
        staff_list = [dict(row) for row in all_staff]
        
        return jsonify({
            'success': True,
            'session_data': {
                'user_id': session.get('user_id'),
                'user_type': session.get('user_type'),
                'full_name': session.get('full_name'),
                'school_id': session.get('school_id'),
                'is_sub_admin': session.get('is_sub_admin'),
                'permissions': session.get('permissions', {}),
                'has_session': 'user_id' in session
            },
            'db_sub_admin_permissions': perms_list,
            'db_staff_sample': staff_list
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/test_timing_sync', methods=['GET'])
def test_timing_sync():
    """Test route to verify that all systems are using the same institution timings"""
    try:
        from database import get_institution_timings, calculate_attendance_status
        from shift_management import ShiftManager
        import datetime

        # Get institution timings
        institution_timings = get_institution_timings()

        # Get shift manager timings
        shift_manager = ShiftManager()
        general_shift = shift_manager.get_shift_info('general')

        # Test attendance calculation
        test_time = datetime.time(9, 30)  # 9:30 AM
        status = calculate_attendance_status(test_time, 'check-in')

        return jsonify({
            'success': True,
            'sync_check': {
                'institution_checkin': institution_timings['checkin_time'].strftime('%H:%M'),
                'institution_checkout': institution_timings['checkout_time'].strftime('%H:%M'),
                'institution_is_custom': institution_timings['is_custom'],
                'shift_manager_checkin': general_shift['start_time'].strftime('%H:%M') if general_shift else 'Not found',
                'shift_manager_checkout': general_shift['end_time'].strftime('%H:%M') if general_shift else 'Not found',
                'attendance_status_test': f"9:30 AM check-in status: {status}",
                'systems_synced': (
                    institution_timings['checkin_time'] == general_shift['start_time'] and
                    institution_timings['checkout_time'] == general_shift['end_time']
                ) if general_shift else False
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Holiday Management API Routes

@app.route('/api/holidays', methods=['GET'])
def get_holidays_api():
    """Get holidays for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import get_holidays
        import json

        # Get query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        department = request.args.get('department')

        # Get holidays
        holidays = get_holidays(
            start_date=start_date,
            end_date=end_date,
            department=department
        )

        # Convert to list of dictionaries for JSON serialization
        holidays_list = []
        for holiday in holidays:
            holiday_dict = dict(holiday)
            # Parse departments JSON if present
            if holiday_dict.get('departments'):
                try:
                    holiday_dict['departments'] = json.loads(holiday_dict['departments'])
                except (json.JSONDecodeError, TypeError):
                    holiday_dict['departments'] = []
            else:
                holiday_dict['departments'] = []
            holidays_list.append(holiday_dict)

        return jsonify({
            'success': True,
            'holidays': holidays_list,
            'count': len(holidays_list)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to retrieve holidays'
        }), 500


@app.route('/api/holidays', methods=['POST'])
def create_holiday_api():
    """Create a new holiday"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import create_holiday

        # Get form data
        holiday_data = {
            'holiday_name': request.form.get('holiday_name'),
            'start_date': request.form.get('start_date'),
            'end_date': request.form.get('end_date'),
            'holiday_type': request.form.get('holiday_type', 'institution_wide'),
            'description': request.form.get('description', ''),
            'is_recurring': bool(request.form.get('is_recurring')),
            'recurring_type': request.form.get('recurring_type')
        }

        # Handle departments for department-specific holidays
        if holiday_data['holiday_type'] == 'department_specific':
            departments = request.form.getlist('departments')
            if not departments:
                departments_str = request.form.get('departments')
                if departments_str:
                    departments = [dept.strip() for dept in departments_str.split(',')]
            holiday_data['departments'] = departments

        # Validate required fields
        if not holiday_data['holiday_name'] or not holiday_data['start_date'] or not holiday_data['end_date']:
            return jsonify({
                'success': False,
                'message': 'Holiday name, start date, and end date are required'
            }), 400

        # Create holiday
        result = create_holiday(holiday_data)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to create holiday'
        }), 500


@app.route('/api/holidays/<int:holiday_id>', methods=['PUT'])
def update_holiday_api(holiday_id):
    """Update an existing holiday"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import update_holiday

        # Get form data
        holiday_data = {
            'holiday_name': request.form.get('holiday_name'),
            'start_date': request.form.get('start_date'),
            'end_date': request.form.get('end_date'),
            'holiday_type': request.form.get('holiday_type'),
            'description': request.form.get('description'),
            'is_recurring': bool(request.form.get('is_recurring')),
            'recurring_type': request.form.get('recurring_type')
        }

        # Handle departments for department-specific holidays
        if holiday_data['holiday_type'] == 'department_specific':
            departments = request.form.getlist('departments')
            if not departments:
                departments_str = request.form.get('departments')
                if departments_str:
                    departments = [dept.strip() for dept in departments_str.split(',')]
            holiday_data['departments'] = departments

        # Update holiday
        result = update_holiday(holiday_id, holiday_data)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to update holiday'
        }), 500


@app.route('/api/holidays/<int:holiday_id>', methods=['DELETE'])
def delete_holiday_api(holiday_id):
    """Delete a holiday"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import delete_holiday

        # Delete holiday
        result = delete_holiday(holiday_id)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to delete holiday'
        }), 500


@app.route('/api/departments', methods=['GET'])
def get_departments_api():
    """Get list of departments for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import get_departments_list

        departments = get_departments_list()

        return jsonify({
            'success': True,
            'departments': departments,
            'count': len(departments)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to retrieve departments'
        }), 500


# Weekly Off Configuration API Routes

@app.route('/api/weekly_off_config', methods=['GET'])
def get_weekly_off_config_api():
    """Get weekly off configuration for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import get_weekly_off_config

        result = get_weekly_off_config()

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to get weekly off configuration'
        }), 500


@app.route('/api/weekly_off_config', methods=['POST'])
def save_weekly_off_config_api():
    """Save weekly off configuration for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        # CSRF Protection
        try:
            csrf.protect()
        except ValidationError as e:
            return jsonify({
                'success': False,
                'message': 'CSRF token validation failed',
                'error': str(e)
            }), 400

        from database import save_weekly_off_config

        # Get weekly off days from form data
        weekly_off_days = []

        # Check for Sunday off
        if request.form.get('sunday_off_enabled') == 'true':
            weekly_off_days.append('sunday')

        result = save_weekly_off_config(weekly_off_days)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to save weekly off configuration'
        }), 500


@app.route('/api/staff/holidays', methods=['GET'])
def get_staff_holidays_api():
    """Get holidays applicable to a specific staff member based on their department"""
    try:
        # Check authorization - both admin and staff can access this
        if 'user_id' not in session:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        # Get staff_id - either from parameter (admin) or session (staff)
        staff_id = request.args.get('staff_id')
        if session['user_type'] == 'staff':
            staff_id = session['user_id']
        elif session['user_type'] in ['admin', 'company_admin'] and staff_id:
            staff_id = staff_id
        else:
            return jsonify({
                'success': False,
                'message': 'Staff ID required'
            }), 400

        # Get date range parameters
        start_date = request.args.get('start_date')  # YYYY-MM-DD format
        end_date = request.args.get('end_date')      # YYYY-MM-DD format

        db = get_db()

        # Get staff information including department
        staff_info = db.execute('''
            SELECT id, staff_id, full_name, department
            FROM staff
            WHERE id = ?
        ''', (staff_id,)).fetchone()

        if not staff_info:
            return jsonify({
                'success': False,
                'message': 'Staff member not found'
            }), 404

        staff_department = staff_info['department'] or ''

        # Build query to get applicable holidays
        query_conditions = ['h.is_active = 1']
        query_params = []

        # Base query for holidays
        base_query = '''
            SELECT h.id, h.holiday_name, h.start_date, h.end_date, h.holiday_type,
                   h.description, h.is_recurring, h.recurring_type, h.created_at, h.departments
            FROM holidays h
        '''

        # Add date range filter if provided
        if start_date and end_date:
            query_conditions.append('''
                ((h.start_date BETWEEN ? AND ?) OR
                 (h.end_date BETWEEN ? AND ?) OR
                 (h.start_date <= ? AND h.end_date >= ?))
            ''')
            query_params.extend([start_date, end_date, start_date, end_date, start_date, end_date])

        # Combine query
        full_query = base_query + ' WHERE ' + ' AND '.join(query_conditions)
        full_query += ' ORDER BY h.start_date ASC'

        # Execute query to get all holidays in date range
        holidays_data = db.execute(full_query, query_params).fetchall()

        # Filter holidays applicable to this staff member
        holidays = []
        import json
        for holiday in holidays_data:
            # Always include institution_wide and common_leave holidays
            if holiday['holiday_type'] in ['institution_wide', 'common_leave']:
                holiday_dict = {
                    'id': holiday['id'],
                    'name': holiday['holiday_name'],  # Use 'name' to match JS expectations
                    'holiday_name': holiday['holiday_name'],
                    'start_date': holiday['start_date'],
                    'end_date': holiday['end_date'],
                    'type': holiday['holiday_type'],  # Use 'type' to match JS expectations
                    'holiday_type': holiday['holiday_type'],
                    'description': holiday['description'],
                    'is_recurring': bool(holiday['is_recurring']),
                    'recurring_type': holiday['recurring_type'],
                    'created_at': holiday['created_at']
                }
                holidays.append(holiday_dict)

            # For department-specific holidays, check if staff's department is included
            elif holiday['holiday_type'] == 'department_specific' and staff_department:
                try:
                    departments = json.loads(holiday['departments'] or '[]')
                    if staff_department in departments:
                        holiday_dict = {
                            'id': holiday['id'],
                            'name': holiday['holiday_name'],  # Use 'name' to match JS expectations
                            'holiday_name': holiday['holiday_name'],
                            'start_date': holiday['start_date'],
                            'end_date': holiday['end_date'],
                            'type': holiday['holiday_type'],  # Use 'type' to match JS expectations
                            'holiday_type': holiday['holiday_type'],
                            'description': holiday['description'],
                            'is_recurring': bool(holiday['is_recurring']),
                            'recurring_type': holiday['recurring_type'],
                            'created_at': holiday['created_at']
                        }
                        holidays.append(holiday_dict)
                except (json.JSONDecodeError, TypeError):
                    # Skip holidays with invalid department JSON
                    continue

        return jsonify({
            'success': True,
            'holidays': holidays,
            'count': len(holidays),
            'staff_info': {
                'id': staff_info['id'],
                'staff_id': staff_info['staff_id'],
                'full_name': staff_info['full_name'],
                'department': staff_department
            },
            'date_range': {
                'start_date': start_date,
                'end_date': end_date
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to retrieve holidays for staff member'
        }), 500


@app.route('/admin/holiday_management')
def holiday_management():
    """Holiday Management page for administrators"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('index'))

    return render_template('holiday_management.html')


# ===========================
# QUOTA MANAGEMENT FUNCTIONS
# ===========================

def get_staff_quota_summary(staff_id, quota_year=None, school_id=None):
    """
    Get comprehensive quota summary for a staff member
    
    Args:
        staff_id (int): Staff ID
        quota_year (int, optional): Year for quota. Defaults to current year
        school_id (int, optional): School ID. Defaults to session school_id
    
    Returns:
        dict: Quota summary with leave, OD, and permission balances
    """
    import datetime
    
    if quota_year is None:
        quota_year = datetime.datetime.now().year
    
    if school_id is None:
        school_id = session.get('school_id')
    
    if not school_id:
        return {'error': 'School ID required'}
    
    try:
        db = get_db()

        # Keep displayed balances in sync with current application states
        # (approved/pending/withdrawn effects).
        update_quota_usage(staff_id, school_id, quota_year)
        
        # Get leave quotas
        leave_quotas = db.execute('''
            SELECT leave_type, allocated_days, used_days,
                   (allocated_days - used_days) as remaining_days
            FROM staff_leave_quotas
            WHERE staff_id = ? AND school_id = ? AND quota_year = ?
        ''', (staff_id, school_id, quota_year)).fetchall()
        
        # Get OD quota
        od_quota = db.execute('''
            SELECT allocated_days, used_days,
                   (allocated_days - used_days) as remaining_days
            FROM staff_od_quotas
            WHERE staff_id = ? AND school_id = ? AND quota_year = ?
        ''', (staff_id, school_id, quota_year)).fetchone()
        
        # Get Permission quota
        permission_quota = db.execute('''
            SELECT allocated_hours, used_hours,
                   (allocated_hours - used_hours) as remaining_hours
            FROM staff_permission_quotas
            WHERE staff_id = ? AND school_id = ? AND quota_year = ?
        ''', (staff_id, school_id, quota_year)).fetchone()
        
        # Format response
        quota_summary = {
            'staff_id': staff_id,
            'quota_year': quota_year,
            'leave_quotas': {},
            'od_quota': {
                'allocated_days': 0,
                'used_days': 0,
                'remaining_days': 0
            },
            'permission_quota': {
                'allocated_hours': 0.0,
                'used_hours': 0.0,
                'remaining_hours': 0.0
            }
        }
        
        # Process leave quotas
        for leave in leave_quotas:
            quota_summary['leave_quotas'][leave['leave_type']] = {
                'allocated_days': leave['allocated_days'],
                'used_days': leave['used_days'],
                'remaining_days': leave['remaining_days']
            }
        
        # Process OD quota
        if od_quota:
            quota_summary['od_quota'] = {
                'allocated_days': od_quota['allocated_days'],
                'used_days': od_quota['used_days'],
                'remaining_days': od_quota['remaining_days']
            }
        
        # Process Permission quota
        if permission_quota:
            quota_summary['permission_quota'] = {
                'allocated_hours': float(permission_quota['allocated_hours']),
                'used_hours': float(permission_quota['used_hours']),
                'remaining_hours': float(permission_quota['remaining_hours'])
            }
        
        return quota_summary
        
    except Exception as e:
        print(f"Error getting quota summary: {e}")
        return {'error': str(e)}


def calculate_used_quotas(staff_id, quota_year, school_id):
    """
    Calculate used quotas based on approved and pending applications.

    Pending requests are treated as reserved balance so staff cannot over-apply
    before admin acts on earlier requests.
    
    Args:
        staff_id (int): Staff ID
        quota_year (int): Year to calculate for
        school_id (int): School ID
    
    Returns:
        dict: Calculated usage for leave, OD, and permission
    """
    import datetime
    
    try:
        db = get_db()
        year_start = f"{quota_year}-01-01"
        year_end = f"{quota_year}-12-31"
        
        # Calculate leave usage by type
        leave_usage = {}
        leave_types = ['CL', 'SL', 'EL', 'ML']
        
        for leave_type in leave_types:
            usage = db.execute('''
                SELECT COALESCE(SUM(
                    julianday(end_date) - julianday(start_date) + 1
                ), 0) as total_days
                FROM leave_applications
                WHERE staff_id = ? AND school_id = ? AND leave_type = ?
                AND status IN ('approved', 'pending')
                AND start_date BETWEEN ? AND ?
            ''', (staff_id, school_id, leave_type, year_start, year_end)).fetchone()
            
            leave_usage[leave_type] = int(usage['total_days']) if usage else 0
        
        # Calculate OD usage
        od_usage = db.execute('''
            SELECT COALESCE(SUM(
                julianday(end_date) - julianday(start_date) + 1
            ), 0) as total_days
            FROM on_duty_applications
            WHERE staff_id = ? AND school_id = ? AND status IN ('approved', 'pending')
            AND start_date BETWEEN ? AND ?
        ''', (staff_id, school_id, year_start, year_end)).fetchone()
        
        # Calculate Permission usage (in hours)
        permission_usage = db.execute('''
            SELECT COALESCE(SUM(duration_hours), 0) as total_hours
            FROM permission_applications
            WHERE staff_id = ? AND school_id = ? AND status IN ('approved', 'pending')
            AND permission_date BETWEEN ? AND ?
        ''', (staff_id, school_id, year_start, year_end)).fetchone()
        
        return {
            'leave_usage': leave_usage,
            'od_usage': int(od_usage['total_days']) if od_usage else 0,
            'permission_usage': float(permission_usage['total_hours']) if permission_usage else 0.0
        }
        
    except Exception as e:
        print(f"Error calculating used quotas: {e}")
        return {
            'leave_usage': {lt: 0 for lt in ['CL', 'SL', 'EL', 'ML']},
            'od_usage': 0,
            'permission_usage': 0.0
        }


def update_quota_usage(staff_id, school_id, quota_year=None):
    """
    Update quota usage based on approved applications
    
    Args:
        staff_id (int): Staff ID
        school_id (int): School ID
        quota_year (int, optional): Year to update. Defaults to current year
    
    Returns:
        dict: Success status and updated usage
    """
    import datetime
    
    if quota_year is None:
        quota_year = datetime.datetime.now().year
    
    try:
        db = get_db()
        
        # Calculate current usage
        usage = calculate_used_quotas(staff_id, quota_year, school_id)
        
        # Update leave quotas
        for leave_type, used_days in usage['leave_usage'].items():
            db.execute('''
                UPDATE staff_leave_quotas
                SET used_days = ?, updated_at = CURRENT_TIMESTAMP
                WHERE staff_id = ? AND school_id = ? AND quota_year = ? AND leave_type = ?
            ''', (used_days, staff_id, school_id, quota_year, leave_type))
        
        # Update OD quota
        db.execute('''
            UPDATE staff_od_quotas
            SET used_days = ?, updated_at = CURRENT_TIMESTAMP
            WHERE staff_id = ? AND school_id = ? AND quota_year = ?
        ''', (usage['od_usage'], staff_id, school_id, quota_year))
        
        # Update Permission quota
        db.execute('''
            UPDATE staff_permission_quotas
            SET used_hours = ?, updated_at = CURRENT_TIMESTAMP
            WHERE staff_id = ? AND school_id = ? AND quota_year = ?
        ''', (usage['permission_usage'], staff_id, school_id, quota_year))
        
        db.commit()
        
        return {
            'success': True,
            'message': 'Quota usage updated successfully',
            'usage': usage
        }
        
    except Exception as e:
        db.rollback()
        print(f"Error updating quota usage: {e}")
        return {
            'success': False,
            'error': str(e)
        }


def set_staff_quotas(staff_id, school_id, quotas, quota_year=None):
    """
    Set quotas for a staff member
    
    Args:
        staff_id (int): Staff ID
        school_id (int): School ID
        quotas (dict): Quota allocations
        quota_year (int, optional): Year for quota. Defaults to current year
    
    Returns:
        dict: Success status and message
    """
    import datetime
    
    if quota_year is None:
        quota_year = datetime.datetime.now().year
    
    try:
        db = get_db()
        
        # Set leave quotas
        if 'leave' in quotas:
            for leave_type, allocated_days in quotas['leave'].items():
                if leave_type in ['CL', 'SL', 'EL', 'ML']:
                    # Insert or update leave quota (MySQL compatible)
                    db.execute('''
                        INSERT INTO staff_leave_quotas
                        (staff_id, school_id, quota_year, leave_type, allocated_days, used_days)
                        VALUES (?, ?, ?, ?, ?, 0)
                        ON DUPLICATE KEY UPDATE allocated_days = VALUES(allocated_days)
                    ''', (staff_id, school_id, quota_year, leave_type, allocated_days))
        
        # Set OD quota
        if 'od' in quotas:
            allocated_days = quotas['od']
            db.execute('''
                INSERT INTO staff_od_quotas
                (staff_id, school_id, quota_year, allocated_days, used_days)
                VALUES (?, ?, ?, ?, 0)
                ON DUPLICATE KEY UPDATE allocated_days = VALUES(allocated_days)
            ''', (staff_id, school_id, quota_year, allocated_days))
        
        # Set Permission quota
        if 'permission' in quotas:
            allocated_hours = quotas['permission']
            db.execute('''
                INSERT INTO staff_permission_quotas
                (staff_id, school_id, quota_year, allocated_hours, used_hours)
                VALUES (?, ?, ?, ?, 0.0)
                ON DUPLICATE KEY UPDATE allocated_hours = VALUES(allocated_hours)
            ''', (staff_id, school_id, quota_year, allocated_hours))
        
        db.commit()
        
        # Update usage based on existing applications
        update_quota_usage(staff_id, school_id, quota_year)
        
        return {
            'success': True,
            'message': 'Staff quotas set successfully'
        }
        
    except Exception as e:
        db.rollback()
        print(f"Error setting staff quotas: {e}")
        return {
            'success': False,
            'error': str(e)
        }


@app.route('/api/department_staff_count', methods=['GET'])
def get_department_staff_count():
    """Get the count of staff in a specific department"""
    try:
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({'success': False, 'message': 'Unauthorized access'}), 403

        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'message': 'School ID not found in session'}), 400

        department = request.args.get('department')
        if not department:
            return jsonify({'success': False, 'message': 'Department parameter is required'}), 400

        db = get_db()
        count_result = db.execute('''
            SELECT COUNT(*) as count FROM staff WHERE school_id = ? AND department = ?
        ''', (school_id, department)).fetchone()

        count = count_result['count'] if count_result else 0

        return jsonify({
            'success': True,
            'count': count,
            'department': department
        })

    except Exception as e:
        print(f"Error in get_department_staff_count: {e}")
        return jsonify({'success': False, 'error': str(e), 'message': 'Failed to get staff count'}), 500


@app.route('/admin/bulk_quota_assignment', methods=['POST'])
def bulk_quota_assignment():
    """Bulk assign leave/OD/permission quotas to all staff in a department for a given year."""
    try:
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({'success': False, 'message': 'Unauthorized access'}), 403

        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'message': 'School ID not found in session'}), 400

        # Accept form-encoded or JSON
        data = request.form if request.form else request.get_json(force=True, silent=True) or {}

        department = data.get('department')
        quota_year = data.get('quota_year') or data.get('year')

        # Leave quotas
        try:
            cl = int(data.get('cl') or data.get('clQuota') or 0)
        except:
            cl = 0
        try:
            sl = int(data.get('sl') or data.get('slQuota') or 0)
        except:
            sl = 0
        try:
            el = int(data.get('el') or data.get('elQuota') or 0)
        except:
            el = 0
        try:
            ml = int(data.get('ml') or data.get('mlQuota') or 0)
        except:
            ml = 0

        # OD and Permission
        try:
            od_days = int(data.get('od') or data.get('odQuota') or 0)
        except:
            od_days = 0
        try:
            permission_hours = float(data.get('permission') or data.get('permissionQuota') or 0.0)
        except:
            permission_hours = 0.0

        if not department:
            return jsonify({'success': False, 'message': 'Department is required'}), 400

        if not quota_year:
            import datetime
            quota_year = datetime.datetime.now().year
        else:
            try:
                quota_year = int(quota_year)
            except:
                quota_year = int(str(quota_year))

        db = get_db()

        # Find staff in department
        staff_rows = db.execute('''
            SELECT id FROM staff WHERE school_id = ? AND department = ?
        ''', (school_id, department)).fetchall()

        staff_ids = [r['id'] for r in staff_rows] if staff_rows else []

        if len(staff_ids) == 0:
            return jsonify({'success': True, 'message': 'No staff found in department', 'updated': 0}), 200

        quotas = {
            'leave': {
                'CL': cl,
                'SL': sl,
                'EL': el,
                'ML': ml
            },
            'od': od_days,
            'permission': permission_hours
        }

        updated = 0
        failed = []

        for sid in staff_ids:
            try:
                # Inline SQL to avoid per-staff commits and lock contention
                for leave_type in ['CL', 'SL', 'EL', 'ML']:
                    allocated = quotas['leave'].get(leave_type, 0)
                    db.execute('''
                        INSERT INTO staff_leave_quotas
                        (staff_id, school_id, quota_year, leave_type, allocated_days, used_days)
                        VALUES (?, ?, ?, ?, ?, 0)
                        ON DUPLICATE KEY UPDATE allocated_days = VALUES(allocated_days)
                    ''', (sid, school_id, quota_year, leave_type, allocated))

                # OD quota
                db.execute('''
                    INSERT INTO staff_od_quotas
                    (staff_id, school_id, quota_year, allocated_days, used_days)
                    VALUES (?, ?, ?, ?, 0)
                    ON DUPLICATE KEY UPDATE allocated_days = VALUES(allocated_days)
                ''', (sid, school_id, quota_year, od_days))

                # Permission quota
                db.execute('''
                    INSERT INTO staff_permission_quotas
                    (staff_id, school_id, quota_year, allocated_hours, used_hours)
                    VALUES (?, ?, ?, ?, 0.0)
                    ON DUPLICATE KEY UPDATE allocated_hours = VALUES(allocated_hours)
                ''', (sid, school_id, quota_year, permission_hours))

                updated += 1
            except Exception as e:
                failed.append({'staff_id': sid, 'error': str(e)})

        db.commit()

        return jsonify({
            'success': True,
            'message': f'Bulk quota assignment completed. Updated {updated} staff.',
            'updated': updated,
            'failed': failed
        })

    except Exception as e:
        print(f"Error in bulk_quota_assignment: {e}")
        return jsonify({'success': False, 'error': str(e), 'message': 'Bulk quota assignment failed'}), 500


def validate_application_against_quota(staff_id, school_id, application_type, application_data):
    """
    Validate an application against available quota
    
    Args:
        staff_id (int): Staff ID
        school_id (int): School ID  
        application_type (str): 'leave', 'od', or 'permission'
        application_data (dict): Application details
    
    Returns:
        dict: Validation result with success status and message
    """
    import datetime
    from datetime import datetime as dt
    
    try:
        quota_year = dt.now().year
        quota_summary = get_staff_quota_summary(staff_id, quota_year, school_id)
        
        if 'error' in quota_summary:
            return {
                'valid': False,
                'message': 'Unable to retrieve quota information'
            }
        
        if application_type == 'leave':
            leave_type = application_data.get('leave_type')
            start_date = dt.strptime(application_data.get('start_date'), '%Y-%m-%d').date()
            end_date = dt.strptime(application_data.get('end_date'), '%Y-%m-%d').date()
            
            # Calculate days requested
            days_requested = (end_date - start_date).days + 1
            
            # Check quota
            if leave_type not in quota_summary['leave_quotas']:
                return {
                    'valid': False,
                    'message': f'No quota assigned for {leave_type} leave'
                }
            
            remaining_days = quota_summary['leave_quotas'][leave_type]['remaining_days']
            
            if days_requested > remaining_days:
                return {
                    'valid': False,
                    'message': f'Insufficient {leave_type} leave balance. Requested: {days_requested} days, Available: {remaining_days} days'
                }
            
            return {
                'valid': True,
                'message': f'Application valid. {remaining_days - days_requested} days will remain after approval'
            }
        
        elif application_type == 'od':
            start_date = dt.strptime(application_data.get('start_date'), '%Y-%m-%d').date()
            end_date = dt.strptime(application_data.get('end_date'), '%Y-%m-%d').date()
            
            # Calculate days requested
            days_requested = (end_date - start_date).days + 1
            
            remaining_days = quota_summary['od_quota']['remaining_days']
            
            if days_requested > remaining_days:
                return {
                    'valid': False,
                    'message': f'Insufficient OD balance. Requested: {days_requested} days, Available: {remaining_days} days'
                }
            
            return {
                'valid': True,
                'message': f'Application valid. {remaining_days - days_requested} days will remain after approval'
            }
        
        elif application_type == 'permission':
            start_time = dt.strptime(application_data.get('start_time'), '%H:%M').time()
            end_time = dt.strptime(application_data.get('end_time'), '%H:%M').time()
            
            # Calculate hours requested
            start_dt = dt.combine(dt.today(), start_time)
            end_dt = dt.combine(dt.today(), end_time)
            
            if end_dt < start_dt:
                end_dt += datetime.timedelta(days=1)  # Next day
            
            hours_requested = (end_dt - start_dt).total_seconds() / 3600
            
            remaining_hours = quota_summary['permission_quota']['remaining_hours']
            
            if hours_requested > remaining_hours:
                return {
                    'valid': False,
                    'message': f'Insufficient permission balance. Requested: {hours_requested:.2f} hours, Available: {remaining_hours:.2f} hours'
                }
            
            return {
                'valid': True,
                'message': f'Application valid. {remaining_hours - hours_requested:.2f} hours will remain after approval'
            }
        
        else:
            return {
                'valid': False,
                'message': 'Invalid application type'
            }
    
    except Exception as e:
        print(f"Error validating application: {e}")
        return {
            'valid': False,
            'message': f'Validation error: {str(e)}'
        }


def get_default_quotas(school_id):
    """
    Get default quota configurations for a school
    
    Args:
        school_id (int): School ID
    
    Returns:
        dict: Default quotas configuration
    """
    try:
        db = get_db()
        
        defaults = db.execute('''
            SELECT quota_type, leave_type, default_allocation
            FROM default_quota_config
            WHERE school_id = ? AND is_active = 1
        ''', (school_id,)).fetchall()
        
        config = {
            'leave': {},
            'od': 0,
            'permission': 0
        }
        
        for default in defaults:
            if default['quota_type'] == 'leave':
                config['leave'][default['leave_type']] = default['default_allocation']
            elif default['quota_type'] == 'od':
                config['od'] = default['default_allocation']
            elif default['quota_type'] == 'permission':
                config['permission'] = default['default_allocation']
        
        return config
        
    except Exception as e:
        print(f"Error getting default quotas: {e}")
        return {
            'leave': {'CL': 12, 'SL': 7, 'EL': 21, 'ML': 180},
            'od': 10,
            'permission': 15
        }


# ===========================
# QUOTA MANAGEMENT API ENDPOINTS
# ===========================

@app.route('/api/staff/list')
def api_staff_list():
    """Get list of all staff members for quota management"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    # Accept school_id from query parameter or session (consistent with timetable APIs)
    school_id = request.args.get('school_id') or session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'School ID required'}), 400
    
    try:
        db = get_db()
        staff = db.execute('''
            SELECT id, staff_id, full_name, department, position
            FROM staff
            WHERE school_id = ?
            ORDER BY CAST(staff_id AS INTEGER) ASC
        ''', (school_id,)).fetchall()
        
        staff_list = []
        for s in staff:
            staff_list.append({
                'id': s['id'],
                'staff_id': s['staff_id'],
                'full_name': s['full_name'],
                'department': s['department'] or 'N/A',
                'position': s['position'] or 'N/A'
            })
        
        return jsonify({
            'success': True,
            'staff': staff_list,
            'count': len(staff_list)
        })
        
    except Exception as e:
        print(f"Error getting staff list: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to retrieve staff list'
        }), 500


@app.route('/api/staff/<int:staff_id>/quotas/<int:quota_year>')
def api_get_staff_quotas(staff_id, quota_year):
    """Get quota summary for a specific staff member and year"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'School ID required'}), 400
    
    try:
        # Ensure default quotas exist before fetching
        initialize_staff_quotas(staff_id, school_id, quota_year)
        
        quota_summary = get_staff_quota_summary(staff_id, quota_year, school_id)
        
        if 'error' in quota_summary:
            return jsonify({
                'success': False,
                'message': quota_summary['error']
            }), 500
        
        return jsonify({
            'success': True,
            'quotas': quota_summary
        })
        
    except Exception as e:
        print(f"Error getting staff quotas: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to retrieve staff quotas'
        }), 500


@app.route('/api/staff/<int:staff_id>/quotas', methods=['POST'])
def api_set_staff_quotas(staff_id):
    """Set quotas for a specific staff member"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'School ID required'}), 400
    
    try:
        quotas_json = request.form.get('quotas')
        quota_year = int(request.form.get('quota_year', datetime.datetime.now().year))
        
        if not quotas_json:
            return jsonify({
                'success': False,
                'message': 'Quotas data required'
            }), 400
        
        quotas = json.loads(quotas_json)
        
        # Validate quota data
        if not isinstance(quotas, dict):
            return jsonify({
                'success': False,
                'message': 'Invalid quota data format'
            }), 400
        
        # Set the quotas
        result = set_staff_quotas(staff_id, school_id, quotas, quota_year)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message']
            })
        else:
            return jsonify({
                'success': False,
                'message': result['error']
            }), 500
        
    except Exception as e:
        print(f"Error setting staff quotas: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to set staff quotas'
        }), 500


@app.route('/api/default_quotas')
def api_get_default_quotas():
    """Get default quota configurations"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'School ID required'}), 400
    
    try:
        defaults = get_default_quotas(school_id)
        
        return jsonify({
            'success': True,
            'defaults': defaults
        })
        
    except Exception as e:
        print(f"Error getting default quotas: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to retrieve default quotas'
        }), 500


@app.route('/api/staff/<int:staff_id>/quota_validation', methods=['POST'])
def api_validate_quota(staff_id):
    """Validate an application against staff quota"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'School ID required'}), 400
    
    try:
        application_type = request.form.get('application_type')
        application_data = {}
        
        if application_type == 'leave':
            application_data = {
                'leave_type': request.form.get('leave_type'),
                'start_date': request.form.get('start_date'),
                'end_date': request.form.get('end_date')
            }
        elif application_type == 'od':
            application_data = {
                'start_date': request.form.get('start_date'),
                'end_date': request.form.get('end_date')
            }
        elif application_type == 'permission':
            application_data = {
                'start_time': request.form.get('start_time'),
                'end_time': request.form.get('end_time')
            }
        
        validation = validate_application_against_quota(staff_id, school_id, application_type, application_data)
        
        return jsonify({
            'success': True,
            'validation': validation
        })
        
    except Exception as e:
        print(f"Error validating quota: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to validate application against quota'
        }), 500


@app.route('/get_leave_usage_summary', methods=['POST'])
def get_leave_usage_summary():
    """Get comprehensive leave usage summary for a staff member"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'School ID required'}), 400
    
    try:
        data = request.get_json()
        staff_id = data.get('staff_id')
        year = data.get('year', datetime.datetime.now().year)
        
        if not staff_id:
            return jsonify({'success': False, 'message': 'Staff ID required'}), 400
        
        db = get_db()
        cursor = db.cursor()
        
        # Get staff name
        cursor.execute("SELECT full_name FROM staff WHERE id = ? AND school_id = ?", (staff_id, school_id))
        staff_result = cursor.fetchone()
        if not staff_result:
            return jsonify({'success': False, 'message': 'Staff member not found'}), 404
        
        staff_name = staff_result[0]
        
        # First, let's recalculate and update the actual usage from applications
        # Get all quota types for this school
        cursor.execute("SELECT id, name, unit, default_value FROM quota_types WHERE school_id = ?", (school_id,))
        quota_type_details = cursor.fetchall()
        
        for qt_id, qt_name, qt_unit, qt_default in quota_type_details:
            actual_used = 0
            
            if qt_unit.lower() == 'days':
                if qt_name == 'On Duty':
                    # Calculate OD usage
                    cursor.execute("""
                        SELECT SUM(
                            CASE 
                                WHEN start_date = end_date THEN 1
                                ELSE julianday(end_date) - julianday(start_date) + 1
                            END
                        ) FROM on_duty_applications 
                        WHERE staff_id = ? AND school_id = ? 
                        AND strftime('%Y', start_date) = ?
                        AND status = 'approved'
                    """, (staff_id, school_id, str(year)))
                else:
                    # Calculate leave usage for this specific type
                    cursor.execute("""
                        SELECT SUM(
                            CASE 
                                WHEN start_date = end_date THEN 1
                                ELSE julianday(end_date) - julianday(start_date) + 1
                            END
                        ) FROM leave_applications 
                        WHERE staff_id = ? AND school_id = ? 
                        AND leave_type = ?
                        AND strftime('%Y', start_date) = ?
                        AND status = 'approved'
                    """, (staff_id, school_id, qt_name, str(year)))
                    
                result = cursor.fetchone()[0]
                actual_used = result if result else 0
                
            elif qt_unit.lower() == 'hours' and qt_name == 'Permission':
                # Calculate permission usage
                cursor.execute("""
                    SELECT SUM(duration_hours) 
                    FROM permission_applications 
                    WHERE staff_id = ? AND school_id = ? 
                    AND strftime('%Y', permission_date) = ?
                    AND status = 'approved'
                """, (staff_id, school_id, str(year)))
                
                result = cursor.fetchone()[0]
                actual_used = result if result else 0
            
            # Update or insert the staff quota with actual usage
            cursor.execute("""
                INSERT OR REPLACE INTO staff_quotas 
                (staff_id, quota_type_id, year, allocated_quota, used_quota)
                VALUES (?, ?, ?, ?, ?)
            """, (staff_id, qt_id, year, qt_default, actual_used))
        
        db.commit()
        
        # Now get the updated quotas
        cursor.execute("""
            SELECT qt.name, qt.unit, 
                   COALESCE(sq.allocated_quota, qt.default_value) as allocated,
                   COALESCE(sq.used_quota, 0) as used,
                   (COALESCE(sq.allocated_quota, qt.default_value) - COALESCE(sq.used_quota, 0)) as remaining
            FROM quota_types qt
            LEFT JOIN staff_quotas sq ON qt.id = sq.quota_type_id AND sq.staff_id = ? AND sq.year = ?
            WHERE qt.school_id = ?
            ORDER BY qt.name
        """, (staff_id, year, school_id))
        
        quota_types = []
        total_used_days = 0
        total_remaining_days = 0
        total_used_hours = 0
        total_remaining_hours = 0
        
        for row in cursor.fetchall():
            quota_data = {
                'type_name': row[0],
                'unit': row[1],
                'total_quota': row[2],
                'used': row[3],
                'remaining': row[4]
            }
            quota_types.append(quota_data)
            
            # Separate days and hours for proper totaling
            if row[1].lower() == 'days':
                total_used_days += row[3]
                total_remaining_days += row[4]
            elif row[1].lower() == 'hours':
                total_used_hours += row[3]
                total_remaining_hours += row[4]
        
        # Get usage records for the year
        cursor.execute("""
            SELECT 
                la.start_date, la.end_date, 
                CASE 
                    WHEN la.start_date = la.end_date THEN 1
                    ELSE julianday(la.end_date) - julianday(la.start_date) + 1
                END as days_used,
                la.reason, la.status, la.leave_type as quota_type
            FROM leave_applications la
            WHERE la.staff_id = ? AND la.school_id = ? 
            AND strftime('%Y', la.start_date) = ?
            ORDER BY la.start_date DESC
        """, (staff_id, school_id, str(year)))
        
        usage_records = []
        for row in cursor.fetchall():
            record = {
                'start_date': row[0],
                'end_date': row[1],
                'days_used': row[2],
                'reason': row[3],
                'status': row[4],
                'quota_type': row[5]
            }
            usage_records.append(record)
        
        # Get OD records as well
        cursor.execute("""
            SELECT 
                oda.start_date, oda.end_date,
                CASE 
                    WHEN oda.start_date = oda.end_date THEN 1
                    ELSE julianday(oda.end_date) - julianday(oda.start_date) + 1
                END as days_used,
                oda.reason, oda.status, 'On Duty' as quota_type
            FROM on_duty_applications oda
            WHERE oda.staff_id = ? AND oda.school_id = ? 
            AND strftime('%Y', oda.start_date) = ?
            ORDER BY oda.start_date DESC
        """, (staff_id, school_id, str(year)))
        
        for row in cursor.fetchall():
            record = {
                'start_date': row[0],
                'end_date': row[1],
                'days_used': row[2],
                'reason': row[3],
                'status': row[4],
                'quota_type': row[5]
            }
            usage_records.append(record)
        
        # Sort all records by start date (most recent first)
        usage_records.sort(key=lambda x: x['start_date'], reverse=True)
        
        # Calculate actual usage from approved applications for better accuracy
        actual_used_days = 0
        actual_used_hours = 0
        
        # Calculate from leave applications
        cursor.execute("""
            SELECT SUM(
                CASE 
                    WHEN la.start_date = la.end_date THEN 1
                    ELSE julianday(la.end_date) - julianday(la.start_date) + 1
                END
            ) FROM leave_applications la
            WHERE la.staff_id = ? AND la.school_id = ? 
            AND strftime('%Y', la.start_date) = ?
            AND la.status = 'approved'
        """, (staff_id, school_id, str(year)))
        
        leave_days_used = cursor.fetchone()[0] or 0
        actual_used_days += leave_days_used
        
        # Calculate from OD applications
        cursor.execute("""
            SELECT SUM(
                CASE 
                    WHEN oda.start_date = oda.end_date THEN 1
                    ELSE julianday(oda.end_date) - julianday(oda.start_date) + 1
                END
            ) FROM on_duty_applications oda
            WHERE oda.staff_id = ? AND oda.school_id = ? 
            AND strftime('%Y', oda.start_date) = ?
            AND oda.status = 'approved'
        """, (staff_id, school_id, str(year)))
        
        od_days_used = cursor.fetchone()[0] or 0
        actual_used_days += od_days_used
        
        # Calculate from permission applications (hours)
        cursor.execute("""
            SELECT SUM(pa.duration_hours) 
            FROM permission_applications pa
            WHERE pa.staff_id = ? AND pa.school_id = ? 
            AND strftime('%Y', pa.permission_date) = ?
            AND pa.status = 'approved'
        """, (staff_id, school_id, str(year)))
        
        permission_hours_used = cursor.fetchone()[0] or 0
        actual_used_hours += permission_hours_used
        
        # Create better summary with separate units
        summary_stats = {
            'total_used': f"{int(actual_used_days)} days" + (f", {int(actual_used_hours)} hours" if actual_used_hours > 0 else ""),
            'total_remaining': f"{int(total_remaining_days)} days" + (f", {int(total_remaining_hours)} hours" if total_remaining_hours > 0 else ""),
            'used_days': int(actual_used_days),
            'remaining_days': int(total_remaining_days),
            'used_hours': int(actual_used_hours),
            'remaining_hours': int(total_remaining_hours)
        }
        
        usage_data = {
            'staff_name': staff_name,
            'year': year,
            'quota_types': quota_types,
            'usage_records': usage_records,
            'summary_stats': summary_stats
        }
        
        return jsonify({
            'success': True,
            'usage_data': usage_data
        })
        
    except Exception as e:
        print(f"Error getting leave usage summary: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to retrieve usage summary'
        }), 500


@app.route('/generate_usage_report', methods=['POST'])
def generate_usage_report():
    """Generate Excel report for leave usage summary"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    school_id = session.get('school_id')
    if not school_id:
        return jsonify({'success': False, 'message': 'School ID required'}), 400
    
    try:
        staff_id = request.form.get('staff_id')
        year = request.form.get('year', datetime.datetime.now().year)
        
        if not staff_id:
            return jsonify({'success': False, 'message': 'Staff ID required'}), 400
        
        # Get staff data
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("SELECT full_name, staff_id, department FROM staff WHERE id = ? AND school_id = ?", (staff_id, school_id))
        staff_result = cursor.fetchone()
        if not staff_result:
            return jsonify({'success': False, 'message': 'Staff member not found'}), 404
        
        staff_name, employee_id, department = staff_result
        
        # Use Excel report generator
        excel_generator = ExcelReportGenerator()
        
        # Recalculate and update quotas first (same logic as summary function)
        cursor.execute("SELECT id, name, unit, default_value FROM quota_types WHERE school_id = ?", (school_id,))
        quota_type_details = cursor.fetchall()
        
        for qt_id, qt_name, qt_unit, qt_default in quota_type_details:
            actual_used = 0
            
            if qt_unit.lower() == 'days':
                if qt_name == 'On Duty':
                    cursor.execute("""
                        SELECT SUM(
                            CASE 
                                WHEN start_date = end_date THEN 1
                                ELSE julianday(end_date) - julianday(start_date) + 1
                            END
                        ) FROM on_duty_applications 
                        WHERE staff_id = ? AND school_id = ? 
                        AND strftime('%Y', start_date) = ?
                        AND status = 'approved'
                    """, (staff_id, school_id, str(year)))
                else:
                    cursor.execute("""
                        SELECT SUM(
                            CASE 
                                WHEN start_date = end_date THEN 1
                                ELSE julianday(end_date) - julianday(start_date) + 1
                            END
                        ) FROM leave_applications 
                        WHERE staff_id = ? AND school_id = ? 
                        AND leave_type = ?
                        AND strftime('%Y', start_date) = ?
                        AND status = 'approved'
                    """, (staff_id, school_id, qt_name, str(year)))
                    
                result = cursor.fetchone()[0]
                actual_used = result if result else 0
                
            elif qt_unit.lower() == 'hours' and qt_name == 'Permission':
                cursor.execute("""
                    SELECT SUM(duration_hours) 
                    FROM permission_applications 
                    WHERE staff_id = ? AND school_id = ? 
                    AND strftime('%Y', permission_date) = ?
                    AND status = 'approved'
                """, (staff_id, school_id, str(year)))
                
                result = cursor.fetchone()[0]
                actual_used = result if result else 0
            
            cursor.execute("""
                INSERT OR REPLACE INTO staff_quotas 
                (staff_id, quota_type_id, year, allocated_quota, used_quota)
                VALUES (?, ?, ?, ?, ?)
            """, (staff_id, qt_id, year, qt_default, actual_used))
        
        db.commit()
        
        # Get the updated usage data
        cursor.execute("""
            SELECT qt.name, qt.unit, 
                   COALESCE(sq.allocated_quota, qt.default_value) as allocated,
                   COALESCE(sq.used_quota, 0) as used,
                   (COALESCE(sq.allocated_quota, qt.default_value) - COALESCE(sq.used_quota, 0)) as remaining
            FROM quota_types qt
            LEFT JOIN staff_quotas sq ON qt.id = sq.quota_type_id AND sq.staff_id = ? AND sq.year = ?
            WHERE qt.school_id = ?
            ORDER BY qt.name
        """, (staff_id, year, school_id))
        
        quota_data = cursor.fetchall()
        
        # Get usage records
        cursor.execute("""
            SELECT 
                la.start_date, la.end_date,
                CASE 
                    WHEN la.start_date = la.end_date THEN 1
                    ELSE julianday(la.end_date) - julianday(la.start_date) + 1
                END as days_used,
                la.reason, la.status, la.leave_type as quota_type, 'Leave' as application_type
            FROM leave_applications la
            WHERE la.staff_id = ? AND la.school_id = ? 
            AND strftime('%Y', la.start_date) = ?
            
            UNION ALL
            
            SELECT 
                oda.start_date, oda.end_date,
                CASE 
                    WHEN oda.start_date = oda.end_date THEN 1
                    ELSE julianday(oda.end_date) - julianday(oda.start_date) + 1
                END as days_used,
                oda.reason, oda.status, 'On Duty' as quota_type, 'OD' as application_type
            FROM on_duty_applications oda
            WHERE oda.staff_id = ? AND oda.school_id = ? 
            AND strftime('%Y', oda.start_date) = ?
            
            ORDER BY start_date DESC
        """, (staff_id, school_id, str(year), staff_id, school_id, str(year)))
        
        usage_records = cursor.fetchall()
        
        # Generate the report
        report_data = {
            'staff_name': staff_name,
            'employee_id': employee_id,
            'department': department,
            'year': year,
            'quota_data': quota_data,
            'usage_records': usage_records
        }
        
        filename = f"Leave_Usage_Report_{staff_name.replace(' ', '_')}_{year}.xlsx"
        
        response = excel_generator.generate_usage_report(report_data, filename)
        
        return response
        
    except Exception as e:
        print(f"Error generating usage report: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to generate usage report'
        }), 500


# =======================================
# QUOTA MANAGEMENT API ENDPOINTS
# =======================================

@app.route('/api/staff/<int:staff_id>/quotas')
def get_staff_quotas(staff_id):
    """Get staff quota information for current year"""
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        
        # Ensure staff can only access their own quotas
        if session['user_type'] == 'staff' and session['user_id'] != staff_id:
            return jsonify({'success': False, 'error': 'Access denied'}), 403
        
        db = get_db()
        current_year = datetime.datetime.now().year
        
        # Get staff info
        staff = db.execute("""
            SELECT staff_id, full_name, school_id FROM staff WHERE id = ?
        """, (staff_id,)).fetchone()
        
        if not staff:
            return jsonify({'success': False, 'error': 'Staff not found'}), 404
        
        school_id = staff['school_id']
        
        # Initialize quotas if not exists
        initialize_staff_quotas(staff_id, school_id, current_year)
        
        # Get quota information
        quotas = {}
        
        # Leave quotas (by leave type)
        leave_quotas = db.execute("""
            SELECT leave_type, allocated_days, used_days,
                   (allocated_days - used_days) as remaining_days
            FROM staff_leave_quotas 
            WHERE staff_id = ? AND quota_year = ?
        """, (staff_id, current_year)).fetchall()
        
        quotas['leave'] = {}
        for quota in leave_quotas:
            quotas['leave'][quota['leave_type']] = {
                'allocated': quota['allocated_days'],
                'used': quota['used_days'],
                'remaining': quota['remaining_days']
            }
        
        # On Duty quota
        od_quota = db.execute("""
            SELECT allocated_days, used_days,
                   (allocated_days - used_days) as remaining_days
            FROM staff_od_quotas 
            WHERE staff_id = ? AND quota_year = ?
        """, (staff_id, current_year)).fetchone()
        
        if od_quota:
            quotas['od'] = {
                'allocated': od_quota['allocated_days'],
                'used': od_quota['used_days'],
                'remaining': od_quota['remaining_days']
            }
        
        # Permission quota
        permission_quota = db.execute("""
            SELECT allocated_hours, used_hours,
                   (allocated_hours - used_hours) as remaining_hours
            FROM staff_permission_quotas 
            WHERE staff_id = ? AND quota_year = ?
        """, (staff_id, current_year)).fetchone()
        
        if permission_quota:
            quotas['permission'] = {
                'allocated': float(permission_quota['allocated_hours']),
                'used': float(permission_quota['used_hours']),
                'remaining': float(permission_quota['remaining_hours'])
            }
        
        return jsonify({
            'success': True,
            'quotas': quotas,
            'staff_info': {
                'staff_id': staff['staff_id'],
                'name': staff['full_name']
            }
        })
        
    except Exception as e:
        print(f"Error fetching staff quotas: {e}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch quota information'
        }), 500


def initialize_staff_quotas(staff_id, school_id, year):
    """Initialize default quotas for staff if not exists"""
    db = get_db()
    
    try:
        # Initialize leave quotas
        leave_types = ['CL', 'SL', 'EL', 'ML']
        default_allocations = {'CL': 12, 'SL': 10, 'EL': 21, 'ML': 90}
        
        for leave_type in leave_types:
            existing = db.execute("""
                SELECT id FROM staff_leave_quotas 
                WHERE staff_id = ? AND quota_year = ? AND leave_type = ?
            """, (staff_id, year, leave_type)).fetchone()
            
            if not existing:
                db.execute("""
                    INSERT INTO staff_leave_quotas 
                    (staff_id, school_id, quota_year, leave_type, allocated_days, used_days)
                    VALUES (?, ?, ?, ?, ?, 0)
                """, (staff_id, school_id, year, leave_type, default_allocations.get(leave_type, 12)))
        
        # Initialize OD quota
        existing_od = db.execute("""
            SELECT id FROM staff_od_quotas 
            WHERE staff_id = ? AND quota_year = ?
        """, (staff_id, year)).fetchone()
        
        if not existing_od:
            db.execute("""
                INSERT INTO staff_od_quotas 
                (staff_id, school_id, quota_year, allocated_days, used_days)
                VALUES (?, ?, ?, 20, 0)
            """, (staff_id, school_id, year))
        
        # Initialize permission quota
        existing_permission = db.execute("""
            SELECT id FROM staff_permission_quotas 
            WHERE staff_id = ? AND quota_year = ?
        """, (staff_id, year)).fetchone()
        
        if not existing_permission:
            db.execute("""
                INSERT INTO staff_permission_quotas 
                (staff_id, school_id, quota_year, allocated_hours, used_hours)
                VALUES (?, ?, ?, 40.0, 0.0)
            """, (staff_id, school_id, year))
        
        db.commit()
        
    except Exception as e:
        print(f"Error initializing quotas: {e}")


# =============================================================================
# LOCAL AGENT API ENDPOINTS (Desktop Bridge Software)
# =============================================================================

@app.route('/api/agent/register', methods=['POST'])
@csrf.exempt
def register_agent():
    """
    Register a new Local Agent (Desktop Bridge Software)
    
    Expected JSON:
    {
        "school_id": 1,
        "agent_name": "Front Desk PC",
        "admin_token": "secure_token"  # For authentication
    }
    
    Returns agent_id and api_key
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        school_id = data.get('school_id')
        agent_name = data.get('agent_name')
        admin_token = data.get('admin_token')
        
        if not all([school_id, agent_name]):
            return jsonify({
                'success': False,
                'error': 'school_id and agent_name are required'
            }), 400
        
        # Optional: Validate admin_token here if you want extra security
        # For now, we'll allow registration with valid school_id
        
        # Verify school exists
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute('SELECT id FROM schools WHERE id = ?', (school_id,))
        school = cursor.fetchone()
        
        if not school:
            return jsonify({
                'success': False,
                'error': f'School ID {school_id} not found'
            }), 404
        
        # Create agent using database function
        from database import create_biometric_agent
        
        result = create_biometric_agent(school_id, agent_name)
        
        if result['success']:
            return jsonify({
                'success': True,
                'agent_id': result['agent_id'],
                'api_key': result['api_key'],
                'message': result['message']
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': result['message']
            }), 400
        
    except Exception as e:
        logger.error(f"Error registering agent: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/agent/heartbeat', methods=['POST'])
@csrf.exempt
def agent_heartbeat():
    """
    Agent heartbeat endpoint - agents should call this every 30-60 seconds
    
    Requires: Authorization header with API key
    
    Returns agent status and any pending commands
    """
    try:
        # Get API key from header
        auth_header = request.headers.get('Authorization')
        
        if not auth_header:
            return jsonify({
                'success': False,
                'error': 'Authorization header required'
            }), 401
        
        # Extract API key (support both "Bearer TOKEN" and plain "TOKEN")
        api_key = auth_header.replace('Bearer ', '').strip()
        
        # Update heartbeat using database function
        from database import update_agent_heartbeat
        
        result = update_agent_heartbeat(api_key)
        
        if not result['success']:
            return jsonify({
                'success': False,
                'error': result.get('message', 'Invalid API key')
            }), 401
        
        # Return status and any commands (future: commands for device management)
        return jsonify({
            'success': True,
            'agent_id': result['agent_id'],
            'school_id': result['school_id'],
            'timestamp': datetime.datetime.utcnow().isoformat() + 'Z',
            'last_heartbeat': result.get('last_heartbeat'),
            'commands': []  # Future: Return pending commands
        })
        
    except Exception as e:
        print(f"Error processing heartbeat: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/agent/info', methods=['GET'])
@csrf.exempt
def agent_info():
    """Get agent information from API key"""
    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'success': False, 'error': 'Authorization header required'}), 401
        
        api_key = auth_header.replace('Bearer ', '').strip()
        
        from database import verify_agent_api_key
        agent_info = verify_agent_api_key(api_key)
        
        if not agent_info:
            return jsonify({'success': False, 'error': 'Invalid API key'}), 401
        
        return jsonify({
            'success': True,
            'agent_id': agent_info['id'],
            'agent_name': agent_info['agent_name'],
            'school_id': agent_info['school_id'],
            'is_active': agent_info['is_active']
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agent/push_logs', methods=['POST'])
@csrf.exempt
def agent_push_logs():
    """
    Receive attendance logs from Local Agent
    
    Expected JSON:
    {
        "device_id": 1,  # From biometric_devices table
        "records": [
            {
                "user_id": "101",
                "timestamp": "2025-12-02T09:15:23",
                "punch_code": 0,
                "verify_method": 1
            }
        ]
    }
    
    Requires: Authorization header with API key
    """
    try:
        # Authenticate agent
        auth_header = request.headers.get('Authorization')
        
        if not auth_header:
            return jsonify({
                'success': False,
                'error': 'Authorization header required'
            }), 401
        
        api_key = auth_header.replace('Bearer ', '').strip()
        
        # Verify API key
        from database import verify_agent_api_key
        
        agent_info = verify_agent_api_key(api_key)
        
        if not agent_info:
            return jsonify({
                'success': False,
                'error': 'Invalid API key'
            }), 401
        
        agent_id = agent_info['id']
        agent_school_id = agent_info['school_id']
        
        # Parse request data
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 400
        
        device_id = data.get('device_id')
        records = data.get('records', [])
        
        if not device_id:
            return jsonify({
                'success': False,
                'error': 'device_id required'
            }), 400
        
        if not records:
            return jsonify({
                'success': False,
                'error': 'No records provided'
            }), 400
        
        # Verify device belongs to agent
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute('''
            SELECT d.id, d.school_id, d.device_name, d.is_active
            FROM biometric_devices d
            WHERE d.id = ? AND d.agent_id = ? AND d.connection_type = 'Agent_LAN'
        ''', (device_id, agent_id))
        
        device = cursor.fetchone()
        
        if not device:
            # Check if device exists but is not assigned to this agent
            cursor.execute('SELECT device_name, agent_id FROM biometric_devices WHERE id = ?', (device_id,))
            device_check = cursor.fetchone()
            
            if device_check:
                device_name_check, current_agent_id = device_check
                if current_agent_id is None:
                    error_msg = f'Device "{device_name_check}" (ID: {device_id}) exists but is not assigned to any agent. Your agent ID is {agent_id}. Please edit the device in the web interface and assign it to agent ID {agent_id}.'
                else:
                    error_msg = f'Device "{device_name_check}" (ID: {device_id}) is assigned to agent ID {current_agent_id}, but your API key belongs to agent ID {agent_id}. Please edit the device and change the agent assignment to ID {agent_id}.'
            else:
                error_msg = f'Device ID {device_id} not found in database.'
            
            return jsonify({
                'success': False,
                'error': error_msg
            }), 403
        
        device_school_id = device[1]
        device_name = device[2]
        is_active = device[3]
        
        if not is_active:
            return jsonify({
                'success': False,
                'error': f'Device {device_name} is inactive'
            }), 403
        
        # Security check: Ensure device school matches agent school
        if device_school_id != agent_school_id:
            return jsonify({
                'success': False,
                'error': 'Device institution mismatch'
            }), 403
        
        # Process records using UnifiedAttendanceProcessor
        from zk_biometric import UnifiedAttendanceProcessor
        
        processor = UnifiedAttendanceProcessor()
        punches = []
        
        for record in records:
            try:
                # Parse timestamp (support ISO format)
                timestamp_str = record.get('timestamp')
                
                # Try ISO format first
                try:
                    from datetime import datetime as dt
                    timestamp = dt.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                except:
                    # Fallback to standard format
                    from datetime import datetime as dt
                    timestamp = dt.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                
                # Map verify method
                verify_method_map = {
                    1: 'fingerprint',
                    2: 'face',
                    3: 'password',
                    4: 'card'
                }
                verify_method = verify_method_map.get(record.get('verify_method', 1), 'fingerprint')
                
                punches.append({
                    'user_id': record.get('user_id'),
                    'timestamp': timestamp,
                    'punch_code': record.get('punch_code', 0),
                    'verification_method': verify_method
                })
                
            except Exception as e:
                print(f"Error parsing agent record: {e}")
                continue
        
        # Process all punches
        result = processor.process_batch_punches(device_id, punches)
        
        # Update device sync status
        from database import update_device_sync_status
        update_device_sync_status(device_id, datetime.datetime.now(), 'success')
        
        # Update agent heartbeat
        from database import update_agent_heartbeat
        update_agent_heartbeat(api_key)
        
        print(f"✓ Agent '{agent_info['agent_name']}' pushed {len(records)} records from device '{device_name}' (ID:{device_id})")
        print(f"  → Processed: {result['processed']}, Rejected: {result['rejected']}, Ignored: {result['ignored']}")
        
        return jsonify({
            'success': True,
            'device_id': device_id,
            'device_name': device_name,
            'agent_id': agent_id,
            'records_received': len(records),
            'processed': result['processed'],
            'rejected': result['rejected'],
            'ignored': result['ignored'],
            'message': f'Successfully processed {result["processed"]} attendance record(s)'
        })
        
    except Exception as e:
        print(f"Error processing agent logs: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/agent/devices', methods=['GET'])
def get_agent_devices():
    """
    Get devices assigned to the authenticated agent
    
    Requires: Authorization header with API key
    """
    try:
        # Authenticate agent
        auth_header = request.headers.get('Authorization')
        
        if not auth_header:
            return jsonify({
                'success': False,
                'error': 'Authorization header required'
            }), 401
        
        api_key = auth_header.replace('Bearer ', '').strip()
        
        # Verify API key
        from database import verify_agent_api_key
        
        agent_info = verify_agent_api_key(api_key)
        
        if not agent_info:
            return jsonify({
                'success': False,
                'error': 'Invalid API key'
            }), 401
        
        agent_id = agent_info['id']
        
        # Get devices for this agent
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute('''
            SELECT id, device_name, ip_address, port, is_active, last_sync, sync_status
            FROM biometric_devices
            WHERE agent_id = ? AND connection_type = 'Agent_LAN'
            ORDER BY device_name
        ''', (agent_id,))
        
        devices = []
        for row in cursor.fetchall():
            devices.append({
                'id': row[0],
                'device_name': row[1],
                'ip_address': row[2],
                'port': row[3],
                'is_active': bool(row[4]),
                'last_sync': row[5],
                'sync_status': row[6]
            })
        
        return jsonify({
            'success': True,
            'agent_id': agent_id,
            'agent_name': agent_info['agent_name'],
            'school_id': agent_info['school_id'],
            'device_count': len(devices),
            'devices': devices
        })
        
    except Exception as e:
        logger.error(f"Error getting agent devices: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


########################################
# AUTOMATIC SYNC SCHEDULER
########################################

def _parse_sync_timestamp(value):
    """Parse DB timestamp values used for incremental sync windows."""
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S.%f'):
            try:
                return datetime.datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.datetime.fromisoformat(text)
        except ValueError:
            return None
    return None


def _get_school_staff_id_set(school_id):
    """Return normalized staff IDs for a school for fast attendance filtering."""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        SELECT staff_id
        FROM staff
        WHERE school_id = ?
          AND staff_id IS NOT NULL
          AND TRIM(staff_id) <> ''
    ''', (school_id,))
    return {str(row[0]).strip() for row in cursor.fetchall() if row[0] and str(row[0]).strip()}


def _format_system_time_display(value):
    """Format a stored timestamp using the server/system local time representation."""
    if not value:
        return None

    if isinstance(value, datetime.datetime):
        dt_value = value
    else:
        text = str(value).strip()
        if not text:
            return None

        dt_value = None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S.%f'):
            try:
                dt_value = datetime.datetime.strptime(text, fmt)
                break
            except ValueError:
                continue

        if dt_value is None:
            try:
                dt_value = datetime.datetime.fromisoformat(text)
            except ValueError:
                return text

    return dt_value.strftime('%-m/%-d/%Y, %-I:%M:%S %p') if os.name != 'nt' else dt_value.strftime('%#m/%#d/%Y, %#I:%M:%S %p')

def auto_sync_direct_lan_devices():
    """Background job to automatically sync all Direct LAN devices"""
    logger.info("Starting automatic sync for Direct LAN devices")

    try:
        with app.app_context():
            db = get_db()
            cursor = db.cursor()

            # Get all active Direct_LAN devices
            cursor.execute("""
                SELECT d.id, d.device_name, d.ip_address, d.port, d.school_id, s.name as school_name, d.last_sync
                FROM biometric_devices d
                LEFT JOIN schools s ON d.school_id = s.id
                WHERE d.is_active = 1 AND d.connection_type = 'Direct_LAN'
            """)

            devices = cursor.fetchall()
            logger.info(f"Found {len(devices)} Direct LAN devices to sync")

            from database import update_device_sync_status
            from zk_biometric import UnifiedAttendanceProcessor

            staff_id_cache = {}

            for device in devices:
                device_id = device[0]
                device_name = device[1]
                ip_address = device[2]
                port = device[3] or 4370
                school_id = device[4]
                school_name = device[5]
                last_sync = _parse_sync_timestamp(device[6])

                try:
                    logger.info(f"Syncing device: {device_name} ({ip_address}:{port}) - {school_name}")

                    if school_id not in staff_id_cache:
                        staff_id_cache[school_id] = _get_school_staff_id_set(school_id)
                    allowed_staff_ids = staff_id_cache[school_id]

                    if not allowed_staff_ids:
                        logger.info(f"Skipping device {device_name}: no staff IDs configured for school {school_id}")
                        update_device_sync_status(device_id)
                        continue

                    # Connect to device
                    zk_device = ZKBiometricDevice(ip_address, port)

                    if zk_device.connect():
                        # Get only records for this school's staff since last sync
                        attendance_records = zk_device.get_attendance_records(
                            allowed_user_ids=allowed_staff_ids,
                            since_timestamp=last_sync
                        )

                        if attendance_records:
                            # Process records using UnifiedAttendanceProcessor
                            processor = UnifiedAttendanceProcessor()

                            # Convert records to format expected by processor
                            punches = []
                            for record in attendance_records:
                                punches.append({
                                    'user_id': str(record['user_id']),
                                    'timestamp': record['timestamp'],
                                    'punch_code': record.get('punch', 0),
                                    'verification_method': record.get('verification_type', 'fingerprint')
                                })

                            results = processor.process_batch_punches(device_id, punches)

                            logger.info(
                                f"Device {device_name}: Processed {results['processed']} records, "
                                f"Rejected {results['rejected']}, Ignored {results.get('ignored', 0)}"
                            )

                            # Update last sync time
                            update_device_sync_status(device_id)
                        else:
                            logger.info(f"No new attendance records for {device_name}")

                        zk_device.disconnect()
                    else:
                        logger.error(f"Failed to connect to device {device_name} ({ip_address}:{port})")
                        update_device_sync_status(device_id, sync_status='failed')

                except Exception as e:
                    logger.error(f"Error syncing device {device_name}: {str(e)}")
                    update_device_sync_status(device_id, sync_status='failed')

            db.commit()
            logger.info("Automatic sync completed")

    except Exception as e:
        logger.error(f"Error in auto_sync_direct_lan_devices: {str(e)}")

# Schedule the sync job with configurable interval
def schedule_auto_sync():
    """Schedule or reschedule the auto-sync job"""
    global AUTO_SYNC_INTERVAL_MINUTES
    AUTO_SYNC_INTERVAL_MINUTES = _load_auto_sync_interval(AUTO_SYNC_INTERVAL_MINUTES)

    scheduler.add_job(
        func=auto_sync_direct_lan_devices,
        trigger="interval",
        minutes=AUTO_SYNC_INTERVAL_MINUTES,
        id='auto_sync_biometric',
        name='Auto Sync Direct LAN Devices',
        replace_existing=True
    )

schedule_auto_sync()

########################################
# AUTO-SYNC CONFIGURATION
########################################

@app.route('/api/biometric/sync-config', methods=['GET', 'POST'])
@csrf.exempt
def api_sync_config():
    """Get or update auto-sync interval configuration"""
    global AUTO_SYNC_INTERVAL_MINUTES
    
    if request.method == 'GET':
        return jsonify({
            'success': True,
            'interval_minutes': AUTO_SYNC_INTERVAL_MINUTES,
            'interval_display': f"{AUTO_SYNC_INTERVAL_MINUTES} minutes"
        })
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            new_interval = data.get('interval_minutes')
            
            # Validate interval
            if not new_interval or not isinstance(new_interval, (int, float)):
                return jsonify({'success': False, 'message': 'Invalid interval value'}), 400
            
            if new_interval < 1:
                return jsonify({'success': False, 'message': 'Interval must be at least 1 minute'}), 400
            
            if new_interval > 1440:
                return jsonify({'success': False, 'message': 'Interval cannot exceed 1440 minutes (24 hours)'}), 400
            
            # Update interval and reschedule
            AUTO_SYNC_INTERVAL_MINUTES = int(new_interval)
            
            # Save to database for persistence
            from database import set_system_setting
            set_system_setting('auto_sync_interval_minutes', AUTO_SYNC_INTERVAL_MINUTES, 'Auto-sync interval in minutes')
            
            schedule_auto_sync()
            
            return jsonify({
                'success': True,
                'message': f'Auto-sync interval updated to {AUTO_SYNC_INTERVAL_MINUTES} minutes',
                'interval_minutes': AUTO_SYNC_INTERVAL_MINUTES
            })
            
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

########################################
# DEVICE MANAGEMENT UI ROUTES
########################################

@app.route('/biometric_devices')
@requires_permission('biometric_devices', 'view')
def biometric_devices_page():
    """Render the device management page"""
    if 'user_id' not in session or (session.get('user_type') != 'admin' and session.get('user_type') != 'company_admin' and not session.get('is_sub_admin')):
        flash('Please login as administrator', 'error')
        return redirect(url_for('index'))
    
    school_id = session.get('school_id')
    # Get module settings for navigation
    module_enabled = get_module_enabled(school_id) if school_id else {}
    
    return render_template('biometric_device_management.html', module_enabled=module_enabled)


@app.route('/api/schools', methods=['GET'])
@csrf.exempt
def api_get_schools():
    """Get list of all schools for dropdowns"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        from database import get_all_schools
        schools = get_all_schools()
        return jsonify({
            'success': True,
            'schools': schools
        })
    except Exception as e:
        print(f"Error fetching schools: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/devices/list', methods=['GET'])
@csrf.exempt
def api_list_devices():
    """Get list of all devices for the logged-in institution"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        school_id = session.get('school_id')
        
        db = get_db()
        cursor = db.cursor()
        
        # Get devices for this institution with agent info
        cursor.execute('''
            SELECT 
                d.id,
                d.device_name,
                d.connection_type,
                d.ip_address,
                d.port,
                d.serial_number,
                d.sync_status,
                d.last_sync,
                d.school_id,
                s.name as school_name,
                a.agent_name
            FROM biometric_devices d
            LEFT JOIN schools s ON d.school_id = s.id
            LEFT JOIN biometric_agents a ON d.agent_id = a.id
            WHERE d.school_id = ? AND d.is_active = 1
            ORDER BY d.device_name
        ''', (school_id,))
        
        devices = []
        for row in cursor.fetchall():
            last_sync_display = _format_system_time_display(row[7]) if row[7] else None
            devices.append({
                'id': row[0],
                'device_name': row[1],
                'connection_type': row[2],
                'ip_address': row[3],
                'port': row[4],
                'serial_number': row[5],
                'sync_status': row[6],
                'last_sync': row[7],
                'last_sync_display': last_sync_display,
                'school_id': row[8],
                'school_name': row[9],
                'agent_name': row[10]
            })
        
        return jsonify({
            'success': True,
            'devices': devices
        })
        
    except Exception as e:
        print(f"Error listing devices: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/devices/add', methods=['POST'])
@csrf.exempt
def api_add_device():
    """Add a new biometric device"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        from database import create_biometric_device
        data = request.json
        
        # Validate required fields
        required = ['device_name', 'school_id', 'connection_type']
        if not all(field in data for field in required):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        # Connection-specific validation
        connection_type = data['connection_type']
        if connection_type == 'Direct_LAN':
            if not data.get('ip_address'):
                return jsonify({'success': False, 'error': 'IP address required for Direct LAN'}), 400
        elif connection_type == 'ADMS':
            if not data.get('serial_number'):
                return jsonify({'success': False, 'error': 'Serial number required for ADMS'}), 400
        elif connection_type == 'Agent_LAN':
            if not data.get('agent_id') or not data.get('ip_address'):
                return jsonify({'success': False, 'error': 'Agent and IP address required for Agent LAN'}), 400
        
        # Create device
        result = create_biometric_device(
            school_id=data['school_id'],
            device_name=data['device_name'],
            connection_type=connection_type,
            ip_address=data.get('ip_address'),
            port=data.get('port', 4370),
            serial_number=data.get('serial_number'),
            agent_id=data.get('agent_id')
        )
        
        if not result.get('success'):
            return jsonify({'success': False, 'error': result.get('message', 'Unknown error')}), 400
        
        print(f"Device created: ID={result['device_id']}, Name={data['device_name']}, Type={connection_type}")
        
        return jsonify({
            'success': True,
            'device_id': result['device_id'],
            'message': result.get('message', 'Device added successfully')
        })
        
    except Exception as e:
        print(f"Error adding device: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/devices/<int:device_id>', methods=['PUT'])
@csrf.exempt
def api_update_device(device_id):
    """Update an existing device"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        from database import get_device_for_institution, update_biometric_device
        data = request.json
        school_id = session.get('school_id')
        
        # Verify ownership
        device = get_device_for_institution(school_id, device_id)
        if not device:
            return jsonify({'success': False, 'error': 'Device not found or access denied'}), 404
        
        # Update device
        result = update_biometric_device(
            device_id=device_id,
            school_id=school_id,
            device_name=data.get('device_name'),
            ip_address=data.get('ip_address'),
            port=data.get('port'),
            serial_number=data.get('serial_number'),
            agent_id=data.get('agent_id')
        )
        
        if not result.get('success'):
            return jsonify({'success': False, 'error': result.get('message', 'Unknown error')}), 500
        
        print(f"Device updated: ID={device_id}")
        
        return jsonify({
            'success': True,
            'message': 'Device updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating device: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/devices/<int:device_id>', methods=['DELETE'])
@csrf.exempt
def api_delete_device(device_id):
    """Soft delete a device"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        from database import get_device_for_institution, delete_biometric_device
        school_id = session.get('school_id')
        
        # Verify ownership
        device = get_device_for_institution(school_id, device_id)
        if not device:
            return jsonify({'success': False, 'error': 'Device not found or access denied'}), 404
        
        # Soft delete
        result = delete_biometric_device(device_id, school_id)
        
        if not result.get('success'):
            return jsonify({'success': False, 'error': result.get('message', 'Unknown error')}), 500
        
        print(f"Device deleted: ID={device_id}")
        
        return jsonify({
            'success': True,
            'message': 'Device deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting device: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agents/list', methods=['GET'])
@csrf.exempt
def api_list_agents():
    """Get list of all agents for the logged-in institution"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        school_id = session.get('school_id')
        user_type = session.get('user_type')
        
        db = get_db()
        cursor = db.cursor()
        
        # Company admins can view all active agents; admins are scoped to own institution.
        if user_type == 'company_admin':
            cursor.execute('''
                SELECT 
                    a.id,
                    a.agent_name,
                    a.last_heartbeat,
                    a.is_active,
                    a.school_id,
                    s.name as school_name,
                    COUNT(d.id) as device_count
                FROM biometric_agents a
                LEFT JOIN schools s ON a.school_id = s.id
                LEFT JOIN biometric_devices d ON d.agent_id = a.id AND d.is_active = 1
                WHERE a.is_active IN (1, '1')
                GROUP BY a.id
                ORDER BY a.school_id, a.agent_name
            ''')
        else:
            cursor.execute('''
                SELECT 
                    a.id,
                    a.agent_name,
                    a.last_heartbeat,
                    a.is_active,
                    a.school_id,
                    s.name as school_name,
                    COUNT(d.id) as device_count
                FROM biometric_agents a
                LEFT JOIN schools s ON a.school_id = s.id
                LEFT JOIN biometric_devices d ON d.agent_id = a.id AND d.is_active = 1
                WHERE a.school_id = ? AND a.is_active IN (1, '1')
                GROUP BY a.id
                ORDER BY a.agent_name
            ''', (school_id,))
        
        agents = []
        for row in cursor.fetchall():
            # Convert SQLite timestamp to ISO format with UTC timezone for JavaScript
            last_heartbeat = row[2]
            if last_heartbeat:
                try:
                    # Parse SQLite timestamp (stored as UTC) and add Z suffix for UTC
                    dt = datetime.datetime.strptime(last_heartbeat, '%Y-%m-%d %H:%M:%S')
                    last_heartbeat = dt.isoformat() + 'Z'
                except:
                    # If already in ISO format or other format, keep as is
                    pass
            
            agents.append({
                'id': row[0],
                'agent_name': row[1],
                'last_heartbeat': last_heartbeat,
                'is_active': row[3],
                'school_id': row[4],
                'school_name': row[5],
                'device_count': row[6]
            })
        
        return jsonify({
            'success': True,
            'agents': agents
        })
        
    except Exception as e:
        print(f"Error listing agents: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/devices/<int:device_id>/sync', methods=['POST'])
@csrf.exempt
def api_sync_device(device_id):
    """Manually trigger sync for a specific device"""
    try:
        if session.get('user_type') not in ['admin', 'company_admin']:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        
        school_id = session.get('school_id')
        
        from database import update_device_sync_status
        from zk_biometric import UnifiedAttendanceProcessor
        
        db = get_db()
        cursor = db.cursor()
        
        # Get device details
        cursor.execute("""
            SELECT device_name, connection_type, ip_address, port, school_id, serial_number, agent_id, last_sync
            FROM biometric_devices
            WHERE id = ? AND school_id = ? AND is_active = 1
        """, (device_id, school_id))
        
        device = cursor.fetchone()
        
        if not device:
            return jsonify({'success': False, 'error': 'Device not found or access denied'}), 404
        
        device_name, connection_type, ip_address, port, dev_school_id, serial_number, agent_id, last_sync = device
        
        # Only Direct_LAN can be manually synced from web interface
        if connection_type != 'Direct_LAN':
            return jsonify({
                'success': False, 
                'error': f'{connection_type} devices sync automatically. No manual sync needed.'
            }), 400
        
        # Connect and sync
        zk_device = ZKBiometricDevice(ip_address, port or 4370)
        
        if not zk_device.connect():
            update_device_sync_status(device_id, sync_status='failed')
            return jsonify({
                'success': False, 
                'error': f'Failed to connect to device at {ip_address}:{port}. Check IP address and network connectivity.'
            }), 500
        
        try:
            allowed_staff_ids = _get_school_staff_id_set(dev_school_id)
            if not allowed_staff_ids:
                zk_device.disconnect()
                update_device_sync_status(device_id)
                return jsonify({
                    'success': True,
                    'message': 'No staff IDs configured for this institution',
                    'processed': 0,
                    'inserted': 0,
                    'rejected': 0
                })

            attendance_records = zk_device.get_attendance_records(
                allowed_user_ids=allowed_staff_ids,
                since_timestamp=_parse_sync_timestamp(last_sync)
            )
            
            if not attendance_records:
                zk_device.disconnect()
                update_device_sync_status(device_id)
                return jsonify({
                    'success': True, 
                    'message': 'No new attendance records found',
                    'processed': 0,
                    'inserted': 0,
                    'rejected': 0
                })
            
            # Process records
            processor = UnifiedAttendanceProcessor()
            
            # Convert records to format expected by processor
            punches = []
            for record in attendance_records:
                punches.append({
                    'user_id': str(record['user_id']),
                    'timestamp': record['timestamp'],
                    'punch_code': record.get('punch', 0),
                    'verification_method': record.get('verification_type', 'fingerprint')
                })
            
            results = processor.process_batch_punches(device_id, punches)
            
            # Update sync status
            update_device_sync_status(device_id)
            db.commit()
            
            zk_device.disconnect()
            
            # Analyze rejection reasons
            rejection_summary = {}
            for detail in results.get('details', []):
                if detail['action'] == 'rejected':
                    reason = detail.get('reason', 'unknown')
                    rejection_summary[reason] = rejection_summary.get(reason, 0) + 1
            
            # Build detailed message
            message_parts = [f"Processed: {results['processed']}"]
            if results['rejected'] > 0:
                message_parts.append(f"Rejected: {results['rejected']}")
                for reason, count in rejection_summary.items():
                    if reason == 'institution_mismatch':
                        message_parts.append(f"  • {count} staff IDs not found in your institution")
                    elif reason == 'duplicate':
                        message_parts.append(f"  • {count} duplicate records")
                    else:
                        message_parts.append(f"  • {count} {reason}")
            
            return jsonify({
                'success': True,
                'message': '\n'.join(message_parts),
                'processed': results['processed'],
                'inserted': results['processed'],
                'rejected': results['rejected'],
                'rejection_summary': rejection_summary,
                'details': results.get('details', [])[:10]  # First 10 for debugging
            })
            
        except Exception as e:
            zk_device.disconnect()
            update_device_sync_status(device_id, sync_status='failed')
            return jsonify({'success': False, 'error': f'Error processing attendance: {str(e)}'}), 500
            
    except Exception as e:
        print(f"Error in api_sync_device: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/devices/<int:device_id>/diagnose', methods=['POST'])
@csrf.exempt
def api_diagnose_device(device_id):
    """Diagnose sync issues - compare device users with database staff"""
    try:
        if session.get('user_type') not in ['admin', 'company_admin']:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        
        school_id = session.get('school_id')
        
        db = get_db()
        cursor = db.cursor()
        
        # Get device details
        cursor.execute("""
            SELECT device_name, connection_type, ip_address, port, school_id
            FROM biometric_devices
            WHERE id = ? AND school_id = ? AND is_active = 1
        """, (device_id, school_id))
        
        device = cursor.fetchone()
        
        if not device:
            return jsonify({'success': False, 'error': 'Device not found'}), 404
        
        device_name, connection_type, ip_address, port, dev_school_id = device
        
        if connection_type != 'Direct_LAN':
            return jsonify({
                'success': False,
                'error': f'{connection_type} devices cannot be diagnosed from web interface'
            }), 400
        
        # Connect to device and get users
        zk_device = ZKBiometricDevice(ip_address, port or 4370)
        
        if not zk_device.connect():
            return jsonify({
                'success': False,
                'error': f'Failed to connect to device at {ip_address}:{port}'
            }), 500
        
        try:
            # Get users from device
            device_users = zk_device.get_users()
            device_user_ids = set()
            
            for u in device_users:
                # Handle both dict and object formats
                if isinstance(u, dict):
                    device_user_ids.add(str(u.get('user_id', '')))
                else:
                    device_user_ids.add(str(getattr(u, 'user_id', '')))
            
            # Remove empty strings
            device_user_ids.discard('')
            
            zk_device.disconnect()
            
            # Get staff from database for this institution
            cursor.execute("""
                SELECT staff_id, full_name, department 
                FROM staff 
                WHERE school_id = ?
                ORDER BY staff_id
            """, (dev_school_id,))
            
            db_staff = cursor.fetchall()
            db_staff_ids = set([row[0] for row in db_staff])
            
            # Compare
            matched = device_user_ids.intersection(db_staff_ids)
            in_device_not_db = device_user_ids - db_staff_ids
            in_db_not_device = db_staff_ids - device_user_ids
            
            # Get details for mismatches
            device_only_details = []
            for uid in sorted(in_device_not_db):
                # Find user in device_users list
                user = None
                for u in device_users:
                    if isinstance(u, dict):
                        if str(u.get('user_id', '')) == uid:
                            user = u
                            break
                    else:
                        if str(getattr(u, 'user_id', '')) == uid:
                            user = u
                            break
                
                if user:
                    if isinstance(user, dict):
                        device_only_details.append({
                            'user_id': uid,
                            'name': user.get('name', 'Unknown')
                        })
                    else:
                        device_only_details.append({
                            'user_id': uid,
                            'name': getattr(user, 'name', 'Unknown')
                        })
            
            db_only_details = []
            for staff in db_staff:
                if staff[0] in in_db_not_device:
                    db_only_details.append({
                        'staff_id': staff[0],
                        'name': staff[1],
                        'department': staff[2]
                    })
            
            return jsonify({
                'success': True,
                'device_name': device_name,
                'institution_id': dev_school_id,
                'summary': {
                    'device_users': len(device_user_ids),
                    'database_staff': len(db_staff_ids),
                    'matched': len(matched),
                    'device_only': len(in_device_not_db),
                    'database_only': len(in_db_not_device)
                },
                'matched_ids': sorted(list(matched)),
                'device_only': device_only_details[:20],  # Limit to 20
                'database_only': db_only_details[:20]
            })
            
        except Exception as e:
            zk_device.disconnect()
            return jsonify({'success': False, 'error': f'Error getting device users: {str(e)}'}), 500
            
    except Exception as e:
        print(f"Error in api_diagnose_device: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/devices/<int:device_id>/test', methods=['POST'])
@csrf.exempt
def api_test_device_connection(device_id):
    """Test connection to a specific device"""
    try:
        if session.get('user_type') not in ['admin', 'company_admin']:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        
        school_id = session.get('school_id')
        
        db = get_db()
        cursor = db.cursor()
        
        # Get device details
        cursor.execute("""
            SELECT device_name, connection_type, ip_address, port
            FROM biometric_devices
            WHERE id = ? AND school_id = ? AND is_active = 1
        """, (device_id, school_id))
        
        device = cursor.fetchone()
        
        if not device:
            return jsonify({'success': False, 'error': 'Device not found'}), 404
        
        device_name, connection_type, ip_address, port = device
        
        if connection_type != 'Direct_LAN':
            return jsonify({
                'success': False,
                'error': f'{connection_type} devices cannot be tested from web interface'
            }), 400
        
        # Test connection
        zk_device = ZKBiometricDevice(ip_address, port or 4370)
        
        if zk_device.connect():
            try:
                # Try to get device info
                users = zk_device.conn.get_users() if hasattr(zk_device.conn, 'get_users') else []
                device_info = {
                    'firmware_version': getattr(zk_device.conn, 'firmware_version', 'Unknown'),
                    'platform': getattr(zk_device.conn, 'platform', 'Unknown'),
                    'device_name': getattr(zk_device.conn, 'device_name', device_name),
                    'user_count': len(users)
                }
            except:
                device_info = {
                    'firmware_version': 'Unknown',
                    'platform': 'Unknown',
                    'device_name': device_name,
                    'user_count': 0
                }
            
            zk_device.disconnect()
            
            return jsonify({
                'success': True,
                'message': f'Successfully connected to {device_name}',
                'device_info': device_info
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Failed to connect to {ip_address}:{port}. Check device power, network, and firewall settings.'
            }), 500
            
    except Exception as e:
        print(f"Error testing device: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/manage_sub_admins')
def manage_sub_admins():
    """View list of all sub-admins"""
    if 'user_id' not in session or session.get('user_type') != 'admin' or session.get('is_sub_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized, only Super Admin can manage Sub-Admins'}), 401
    
    school_id = session.get('school_id')
    db = get_db()
    
    # Get all staff who have entries in sub_admin_permissions
    sub_admins = db.execute('''
        SELECT s.id, s.staff_id, s.full_name, s.department, p.module_name, p.can_view, p.can_edit, p.can_delete
        FROM staff s
        JOIN sub_admin_permissions p ON s.staff_id = p.staff_id AND s.school_id = p.school_id
        WHERE s.school_id = ?
        ORDER BY s.full_name, p.module_name
    ''', (school_id,)).fetchall()
    
    # Group by staff_id
    grouped_sub_admins = {}
    for sa in sub_admins:
        sid = sa['staff_id']
        if sid not in grouped_sub_admins:
            grouped_sub_admins[sid] = {
                'id': sa['id'],
                'staff_id': sid,
                'full_name': sa['full_name'],
                'department': sa['department'],
                'modules': []
            }
        grouped_sub_admins[sid]['modules'].append({
            'module_name': sa['module_name'],
            'can_view': bool(sa['can_view']),
            'can_edit': bool(sa['can_edit']),
            'can_delete': bool(sa['can_delete'])
        })
    
    # Get module settings for navigation
    module_enabled = get_module_enabled(school_id)
        
    return render_template('manage_sub_admins.html', sub_admins=list(grouped_sub_admins.values()),
                          module_enabled=module_enabled)


@app.route('/add_sub_admin', methods=['GET', 'POST'])
def add_sub_admin():
    """Add a new sub-admin or update an existing one"""
    if 'user_id' not in session or session.get('user_type') != 'admin' or session.get('is_sub_admin'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    school_id = session.get('school_id')
    db = get_db()
    
    if request.method == 'GET':
        module_enabled = get_module_enabled(school_id)
        staff_list = db.execute('SELECT id, staff_id, full_name, department FROM staff WHERE school_id = ? ORDER BY full_name', (school_id,)).fetchall()
        return render_template('add_sub_admin.html', staff_list=staff_list, existing_permissions=None, edit_staff_id=None, module_enabled=module_enabled)
        
    if request.method == 'POST':
        staff_id = request.form.get('staff_id') # The actual staff username/id string
        # Modules to assign
        modules = ['staff_management', 'shift_management', 'salary_management', 'timetable_management', 'reports', 'biometric_devices']
        
        try:
            # First, delete existing permissions for this staff member to overwrite
            db.execute('DELETE FROM sub_admin_permissions WHERE staff_id = ? AND school_id = ?', (staff_id, school_id))
            
            # Insert new permissions based on checkboxes
            for mod in modules:
                # Checkbox name format: module_name_view, module_name_edit, module_name_delete
                can_view = 1 if request.form.get(f'{mod}_view') else 0
                can_edit = 1 if request.form.get(f'{mod}_edit') else 0
                can_delete = 1 if request.form.get(f'{mod}_delete') else 0
                
                # Only insert if at least view is checked
                if can_view or can_edit or can_delete:
                    # Enforce view if edit/delete is checked
                    if can_edit or can_delete:
                        can_view = 1
                        
                    db.execute('''
                        INSERT INTO sub_admin_permissions (staff_id, module_name, school_id, can_view, can_edit, can_delete)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (staff_id, mod, school_id, can_view, can_edit, can_delete))
            
            db.commit()
            flash('Sub-admin permissions updated successfully.', 'success')
            return redirect(url_for('manage_sub_admins'))
            
        except sqlite3.Error as e:
            db.rollback()
            return jsonify({'success': False, 'error': f'Database error: {str(e)}'}), 500


@app.route('/edit_sub_admin/<staff_id>', methods=['GET'])
def edit_sub_admin(staff_id):
    """Edit an existing sub-admin's permissions"""
    if 'user_id' not in session or session.get('user_type') != 'admin' or session.get('is_sub_admin'):
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('manage_sub_admins'))
    
    school_id = session.get('school_id')
    db = get_db()
    
    # Get existing permissions for this staff member
    permissions = db.execute('''
        SELECT module_name, can_view, can_edit, can_delete
        FROM sub_admin_permissions
        WHERE staff_id = ? AND school_id = ?
    ''', (staff_id, school_id)).fetchall()
    
    # Convert to dictionary for easy lookup in template
    existing_permissions = {}
    for perm in permissions:
        existing_permissions[perm['module_name']] = {
            'view': bool(perm['can_view']),
            'edit': bool(perm['can_edit']),
            'delete': bool(perm['can_delete'])
        }
    
    # Get staff list for the dropdown
    staff_list = db.execute('SELECT id, staff_id, full_name, department FROM staff WHERE school_id = ? ORDER BY full_name', (school_id,)).fetchall()
    
    module_enabled = get_module_enabled(school_id)
    
    return render_template('add_sub_admin.html', 
                         staff_list=staff_list, 
                         existing_permissions=existing_permissions,
                         edit_staff_id=staff_id,
                         module_enabled=module_enabled)


@app.route('/delete_sub_admin', methods=['POST'])
def delete_sub_admin():
    """Revoke all sub-admin permissions for a staff member"""
    if 'user_id' not in session or session.get('user_type') != 'admin' or session.get('is_sub_admin'):
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('manage_sub_admins'))
        
    staff_id = request.form.get('staff_id')
    school_id = session.get('school_id')
    
    db = get_db()
    try:
        db.execute('DELETE FROM sub_admin_permissions WHERE staff_id = ? AND school_id = ?', (staff_id, school_id))
        db.commit()
        flash('Sub-Admin access revoked successfully.', 'success')
    except sqlite3.Error as e:
        flash(f'Database error: {str(e)}', 'danger')
    
    return redirect(url_for('manage_sub_admins'))


@app.route('/api/agents/create', methods=['POST'])
@csrf.exempt
def api_create_agent():
    """Create a new local agent and generate API key"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        from database import create_biometric_agent
        data = request.json
        
        if not data.get('agent_name') or not data.get('school_id'):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        # Create agent
        result = create_biometric_agent(
            school_id=data['school_id'],
            agent_name=data['agent_name']
        )
        
        if not result.get('success'):
            return jsonify({'success': False, 'error': result.get('message', 'Unknown error')}), 400
        
        print(f"Agent created: ID={result['agent_id']}, Name={data['agent_name']}")
        
        return jsonify({
            'success': True,
            'agent_id': result['agent_id'],
            'api_key': result['api_key'],
            'message': result.get('message', 'Agent created successfully')
        })
        
    except Exception as e:
        print(f"Error creating agent: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/agents/<int:agent_id>', methods=['DELETE'])
@csrf.exempt
def api_deactivate_agent(agent_id):
    """Deactivate a local agent"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        from database import deactivate_biometric_agent
        school_id = session.get('school_id')
        user_type = session.get('user_type')
        
        # Verify ownership
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT school_id FROM biometric_agents WHERE id = ?', (agent_id,))
        row = cursor.fetchone()
        
        if not row:
            return jsonify({'success': False, 'error': 'Agent not found or access denied'}), 404

        target_school_id = row[0]
        if user_type != 'company_admin' and target_school_id != school_id:
            return jsonify({'success': False, 'error': 'Agent not found or access denied'}), 404
        
        # Deactivate
        result = deactivate_biometric_agent(agent_id, target_school_id)
        if not result.get('success'):
            return jsonify({'success': False, 'error': result.get('message', 'Unknown error')}), 500
        
        print(f"Agent deactivated: ID={agent_id}")
        
        return jsonify({
            'success': True,
            'message': 'Agent deactivated successfully'
        })
        
    except Exception as e:
        print(f"Error deactivating agent: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ===============================================================================
# STUDENT PORTAL ROUTES
# ===============================================================================

@app.route('/student/login', methods=['GET', 'POST'])
def student_login():
    """Redirect to main login page - students can login through the same interface"""
    return redirect(url_for('index'))


@app.route('/student/dashboard')
def student_dashboard():
    """Student dashboard with attendance display"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return redirect(url_for('student_login'))
    
    db = get_db()
    
    # Get student details
    student = db.execute('''
        SELECT * FROM students WHERE id = ?
    ''', (session['student_id'],)).fetchone()
    
    if not student:
        session.clear()
        return redirect(url_for('student_login'))
    
    # Get school name
    school = db.execute('SELECT name FROM schools WHERE id = ?', 
                       (student['school_id'],)).fetchone()
    
    # Get today's attendance
    today = datetime.date.today().strftime('%Y-%m-%d')
    attendance = db.execute('''
        SELECT sa.*, 
               s1.full_name as morning_marked_by_name,
               s2.full_name as afternoon_marked_by_name
        FROM student_attendance sa
        LEFT JOIN staff s1 ON sa.morning_marked_by = s1.id
        LEFT JOIN staff s2 ON sa.afternoon_marked_by = s2.id
        WHERE sa.student_id = ? AND sa.attendance_date = ?
    ''', (session['student_id'], today)).fetchone()

    # Apply holiday planner for student-facing status (bulk or class/section).
    holiday_record = None
    try:
        holiday_record = db.execute('''
            SELECT holiday_title, holiday_start_date, holiday_end_date, description
            FROM student_holidays
            WHERE school_id = ?
              AND holiday_start_date <= ?
              AND holiday_end_date >= ?
              AND (
                    target_mode = 'bulk'
                    OR (
                        target_mode = 'class_section'
                        AND target_class = ?
                        AND (
                            target_section IS NULL
                            OR TRIM(target_section) = ''
                            OR target_section = ?
                        )
                    )
              )
            ORDER BY id DESC
            LIMIT 1
        ''', (student['school_id'], today, today, (student['class'] or '').strip(), (student['section'] or '').strip())).fetchone()
    except Exception:
        holiday_record = None

    if attendance:
        attendance = dict(attendance)

    if holiday_record:
        if not attendance:
            attendance = {
                'morning_status': 'holiday',
                'afternoon_status': 'holiday',
                'morning_time': None,
                'afternoon_time': None,
                'morning_notes': holiday_record['description'] or holiday_record['holiday_title'],
                'afternoon_notes': holiday_record['description'] or holiday_record['holiday_title'],
                'morning_marked_by_name': None,
                'afternoon_marked_by_name': None
            }
        else:
            if not attendance.get('morning_status'):
                attendance['morning_status'] = 'holiday'
                if not attendance.get('morning_notes'):
                    attendance['morning_notes'] = holiday_record['description'] or holiday_record['holiday_title']
            if not attendance.get('afternoon_status'):
                attendance['afternoon_status'] = 'holiday'
                if not attendance.get('afternoon_notes'):
                    attendance['afternoon_notes'] = holiday_record['description'] or holiday_record['holiday_title']
    
    return render_template('student/student_dashboard.html',
                         student=student,
                         school_name=school['name'] if school else 'N/A',
                         attendance=attendance,
                         current_holiday=dict(holiday_record) if holiday_record else None,
                         current_date=datetime.date.today().strftime('%d %B, %Y'))


@app.route('/student/attendance-history')
def student_attendance_history():
    """Student attendance history page"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return redirect(url_for('student_login'))
    
    db = get_db()
    student = db.execute('SELECT * FROM students WHERE id = ?', 
                        (session['student_id'],)).fetchone()
    
    # Get attendance history (last 30 days)
    today_obj = datetime.date.today()
    start_date_obj = today_obj - datetime.timedelta(days=29)
    start_date = start_date_obj.strftime('%Y-%m-%d')
    end_date = today_obj.strftime('%Y-%m-%d')

    attendance_records = db.execute('''
        SELECT sa.*, 
               s1.full_name as morning_marked_by_name,
               s2.full_name as afternoon_marked_by_name
        FROM student_attendance sa
        LEFT JOIN staff s1 ON sa.morning_marked_by = s1.id
        LEFT JOIN staff s2 ON sa.afternoon_marked_by = s2.id
        WHERE sa.student_id = ? AND sa.attendance_date BETWEEN ? AND ?
        ORDER BY sa.attendance_date DESC
    ''', (session['student_id'], start_date, end_date)).fetchall()

    attendance_map = {}
    for row in attendance_records:
        row_dict = dict(row)
        date_key = str(row_dict.get('attendance_date') or '')
        if date_key and not row_dict.get('date'):
            row_dict['date'] = date_key
        if date_key:
            attendance_map[date_key] = row_dict

    holiday_rows = []
    try:
        holiday_rows = db.execute('''
            SELECT holiday_title, holiday_start_date, holiday_end_date, description,
                   target_mode, target_class, target_section
            FROM student_holidays
            WHERE school_id = ?
              AND holiday_end_date >= ?
              AND holiday_start_date <= ?
              AND (
                    target_mode = 'bulk'
                    OR (
                        target_mode = 'class_section'
                        AND target_class = ?
                        AND (
                            target_section IS NULL
                            OR TRIM(target_section) = ''
                            OR target_section = ?
                        )
                    )
              )
            ORDER BY holiday_start_date DESC, id DESC
        ''', (
            student['school_id'],
            start_date,
            end_date,
            (student['class'] or '').strip(),
            (student['section'] or '').strip()
        )).fetchall()
    except Exception:
        holiday_rows = []

    for holiday in holiday_rows:
        try:
            h_start = datetime.date.fromisoformat(str(holiday['holiday_start_date']))
            h_end = datetime.date.fromisoformat(str(holiday['holiday_end_date']))
        except Exception:
            continue

        current = max(h_start, start_date_obj)
        final = min(h_end, today_obj)
        while current <= final:
            key = current.strftime('%Y-%m-%d')
            if key in attendance_map:
                if not attendance_map[key].get('morning_status'):
                    attendance_map[key]['morning_status'] = 'holiday'
                    attendance_map[key]['morning_notes'] = attendance_map[key].get('morning_notes') or holiday['description'] or holiday['holiday_title']
                if not attendance_map[key].get('afternoon_status'):
                    attendance_map[key]['afternoon_status'] = 'holiday'
                    attendance_map[key]['afternoon_notes'] = attendance_map[key].get('afternoon_notes') or holiday['description'] or holiday['holiday_title']
            else:
                attendance_map[key] = {
                    'attendance_date': key,
                    'date': key,
                    'morning_status': 'holiday',
                    'morning_time': None,
                    'morning_notes': holiday['description'] or holiday['holiday_title'],
                    'morning_marked_by_name': None,
                    'afternoon_status': 'holiday',
                    'afternoon_time': None,
                    'afternoon_notes': holiday['description'] or holiday['holiday_title'],
                    'afternoon_marked_by_name': None
                }
            current += datetime.timedelta(days=1)

    attendance_records = sorted(
        attendance_map.values(),
        key=lambda item: str(item.get('attendance_date') or item.get('date') or ''),
        reverse=True
    )[:30]
    
    school = db.execute('SELECT name FROM schools WHERE id = ?', 
                       (student['school_id'],)).fetchone()
    
    return render_template('student/student_attendance_history.html',
                         student=student,
                         school_name=school['name'] if school else 'N/A',
                         attendance_records=attendance_records)


@app.route('/student/profile')
def student_profile():
    """Student profile page"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return redirect(url_for('student_login'))
    
    db = get_db()
    student = db.execute('SELECT * FROM students WHERE id = ?', 
                        (session['student_id'],)).fetchone()
    
    school = db.execute('SELECT name FROM schools WHERE id = ?', 
                       (student['school_id'],)).fetchone()
    
    return render_template('student/student_profile.html',
                         student=student,
                         school_name=school['name'] if school else 'N/A')


@app.route('/student/homework-exam')
def student_homework_exam():
    """Student homework and exam corner"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return redirect(url_for('student_login'))

    db = get_db()
    student = db.execute('SELECT * FROM students WHERE id = ?',
                        (session['student_id'],)).fetchone()

    if not student:
        return redirect(url_for('student_login'))

    school = db.execute('SELECT name FROM schools WHERE id = ?',
                       (student['school_id'],)).fetchone()

    db.execute('''
        CREATE TABLE IF NOT EXISTS student_homework_completion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            homework_id INTEGER NOT NULL,
            student_id INTEGER NOT NULL,
            school_id INTEGER NOT NULL,
            staff_id INTEGER NOT NULL,
            completed_at TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(homework_id, student_id)
        )
    ''')

    # Load assigned exams for this student's class
    student_class = (student['class'] or '').strip()
    student_section = (student['section'] or '').strip()
    today_iso = datetime.date.today().strftime('%Y-%m-%d')
    upcoming_exams = []
    pending_homework = []
    completed_homework = []

    if student_class:
        try:
            upcoming_exams = db.execute('''
                SELECT id, class_name, exam_title, exam_date, description, created_at
                FROM exam_assignments
                WHERE school_id = ?
                  AND class_name = ?
                  AND exam_date >= ?
                ORDER BY exam_date ASC, id ASC
            ''', (student['school_id'], student_class, today_iso)).fetchall()
        except Exception:
            # Table may not exist yet in a fresh environment.
            upcoming_exams = []

        # Load homework assigned by staff for this student's class/section.
    try:
        if student_section:
            pending_homework = db.execute('''
                                SELECT sha.id, sha.class_name, sha.section_name, sha.subject_name, sha.period_number,
                                             sha.homework_text, sha.due_date, sha.assigned_date, sha.attachment_path
                                FROM staff_homework_assignments sha
                                WHERE sha.school_id = ?
                                    AND LOWER(TRIM(sha.class_name)) = LOWER(TRIM(?))
                                    AND LOWER(TRIM(sha.section_name)) = LOWER(TRIM(?))
                                    AND NOT EXISTS (
                                            SELECT 1
                                            FROM student_homework_completion shc
                                            WHERE shc.homework_id = sha.id
                                                AND shc.student_id = ?
                                    )
                ORDER BY
                                        CASE WHEN sha.due_date IS NULL THEN 1 ELSE 0 END,
                                        sha.due_date ASC,
                                        sha.assigned_date DESC,
                                        sha.id DESC
                        ''', (student['school_id'], student_class, student_section, student['id'])).fetchall()

            completed_homework = db.execute('''
                                SELECT sha.id, sha.class_name, sha.section_name, sha.subject_name, sha.period_number,
                                             sha.homework_text, sha.due_date, sha.assigned_date, sha.attachment_path,
                                             shc.completed_at
                                FROM staff_homework_assignments sha
                                JOIN student_homework_completion shc
                                    ON shc.homework_id = sha.id
                                 AND shc.student_id = ?
                                WHERE sha.school_id = ?
                                    AND LOWER(TRIM(sha.class_name)) = LOWER(TRIM(?))
                                    AND LOWER(TRIM(sha.section_name)) = LOWER(TRIM(?))
                                ORDER BY shc.completed_at DESC, sha.assigned_date DESC, sha.id DESC
                LIMIT 30
                        ''', (student['id'], student['school_id'], student_class, student_section)).fetchall()
        else:
            pending_homework = db.execute('''
                                SELECT sha.id, sha.class_name, sha.section_name, sha.subject_name, sha.period_number,
                                             sha.homework_text, sha.due_date, sha.assigned_date, sha.attachment_path
                                FROM staff_homework_assignments sha
                                WHERE sha.school_id = ?
                                    AND LOWER(TRIM(sha.class_name)) = LOWER(TRIM(?))
                                    AND NOT EXISTS (
                                            SELECT 1
                                            FROM student_homework_completion shc
                                            WHERE shc.homework_id = sha.id
                                                AND shc.student_id = ?
                                    )
                ORDER BY
                                        CASE WHEN sha.due_date IS NULL THEN 1 ELSE 0 END,
                                        sha.due_date ASC,
                                        sha.assigned_date DESC,
                                        sha.id DESC
                        ''', (student['school_id'], student_class, student['id'])).fetchall()

            completed_homework = db.execute('''
                                SELECT sha.id, sha.class_name, sha.section_name, sha.subject_name, sha.period_number,
                                             sha.homework_text, sha.due_date, sha.assigned_date, sha.attachment_path,
                                             shc.completed_at
                                FROM staff_homework_assignments sha
                                JOIN student_homework_completion shc
                                    ON shc.homework_id = sha.id
                                 AND shc.student_id = ?
                                WHERE sha.school_id = ?
                                    AND LOWER(TRIM(sha.class_name)) = LOWER(TRIM(?))
                                ORDER BY shc.completed_at DESC, sha.assigned_date DESC, sha.id DESC
                LIMIT 30
                        ''', (student['id'], student['school_id'], student_class)).fetchall()
    except Exception:
        # Homework table may not exist yet in a fresh environment.
        pending_homework = []
        completed_homework = []

    return render_template('student/student_homework_exam.html',
                         student=student,
                         school_name=school['name'] if school else 'N/A',
                         upcoming_exams=upcoming_exams,
                         pending_homework=pending_homework,
                         completed_homework=completed_homework)


@app.route('/student/fees')
def student_fees():
    """Student fees and payment history page"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return redirect(url_for('student_login'))
    
    db = get_db()
    student = db.execute('SELECT * FROM students WHERE id = ?', 
                        (session['student_id'],)).fetchone()
    
    if not student:
        return redirect(url_for('student_login'))
    
    school = db.execute('SELECT name FROM schools WHERE id = ?', 
                       (student['school_id'],)).fetchone()

    # Fetch fees for this student
    fee_rows = db.execute('''
        SELECT sf.id, ft.name as fee_type_name, sf.amount,
               COALESCE(sf.paid_amount, 0) as paid_amount,
               sf.due_date, sf.status, sf.paid_date, sf.payment_mode, sf.notes, sf.created_at
        FROM student_fees sf
        JOIN fee_types ft ON sf.fee_type_id = ft.id
        WHERE sf.student_db_id = ?
        ORDER BY sf.created_at DESC
    ''', (session['student_id'],)).fetchall()

    fees = [dict(r) for r in fee_rows]
    # Compute balance for each fee
    for f in fees:
        f['balance'] = round(float(f['amount']) - float(f['paid_amount']), 2)

    total_fees      = sum(f['amount'] for f in fees)
    paid_amount     = sum(f['paid_amount'] for f in fees)
    balance_amount  = round(total_fees - paid_amount, 2)
    partial_fees    = [f for f in fees if f['status'] == 'partial']
    pending_amount  = sum(f['balance'] for f in fees if f['status'] in ('pending', 'partial', 'overdue'))
    overdue_amount  = sum(f['balance'] for f in fees if f['status'] == 'overdue')

    return render_template('student/student_fees.html',
                         student=student,
                         school_name=school['name'] if school else 'N/A',
                         fees=fees,
                         total_fees=total_fees,
                         paid_amount=paid_amount,
                         balance_amount=balance_amount,
                         pending_amount=pending_amount,
                         overdue_amount=overdue_amount,
                         partial_count=len(partial_fees))


@app.route('/student/leave')
def student_leave():
    """Student leave application and history page"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return redirect(url_for('student_login'))
    
    db = get_db()
    student = db.execute('SELECT * FROM students WHERE id = ?', 
                        (session['student_id'],)).fetchone()
    
    if not student:
        return redirect(url_for('student_login'))
    
    school = db.execute('SELECT name FROM schools WHERE id = ?', 
                       (student['school_id'],)).fetchone()
    
    # Get leave applications (placeholder for now - can be extended later)
    # You can add actual leave tables later
    
    return render_template('student/student_leave.html',
                         student=student,
                         school_name=school['name'] if school else 'N/A')


@app.route('/student/timetable')
def student_timetable():
    """Student timetable/class schedule page"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return redirect(url_for('student_login'))
    
    db = get_db()
    student = db.execute('SELECT * FROM students WHERE id = ?', 
                        (session['student_id'],)).fetchone()
    
    if not student:
        return redirect(url_for('student_login'))
    
    school = db.execute('SELECT name FROM schools WHERE id = ?', 
                       (student['school_id'],)).fetchone()
    
    # Get timetable periods for student's class and section
    school_id = student['school_id']
    student_class = student['class']
    student_section = student['section']
    
    # Query timetable_periods table for periods configured for this class/section
    cursor = db.cursor()
    
    # Get academic level and section IDs
    level_id = None
    section_id = None
    
    # Try to get level_id from timetable_academic_levels
    level_row = cursor.execute('''
        SELECT id FROM timetable_academic_levels 
        WHERE school_id = ? AND level_name = ? AND is_active = 1
    ''', (school_id, student_class)).fetchone()
    
    if level_row:
        level_id = level_row['id']
        
        # Get section_id from timetable_sections
        section_row = cursor.execute('''
            SELECT id FROM timetable_sections 
            WHERE school_id = ? AND level_id = ? AND section_name = ? AND is_active = 1
        ''', (school_id, level_id, student_section)).fetchone()
        
        if section_row:
            section_id = section_row['id']
    
    # Fetch all periods configured for this class/section or school-wide
    periods = cursor.execute('''
        SELECT DISTINCT period_number, period_name, start_time, end_time, duration_minutes
        FROM timetable_periods
        WHERE school_id = ? 
            AND (level_id IS NULL OR level_id = ?)
            AND (section_id IS NULL OR section_id = ?)
            AND is_active = 1
        ORDER BY period_number
    ''', (school_id, level_id, section_id)).fetchall()
    
    # Fetch timetable assignments for this class/section (only show if staff is actually assigned)
    timetable_assignments = []
    if level_id and section_id:
        timetable_assignments = cursor.execute('''
            SELECT ha.day_of_week, ha.period_number, ha.subject_name, 
                   ha.room_number, s.full_name as teacher_name,
                   tp.start_time, tp.end_time, tp.period_name, ha.staff_id
            FROM timetable_hierarchical_assignments ha
            LEFT JOIN staff s ON ha.staff_id = s.id
            LEFT JOIN timetable_periods tp 
                ON ha.school_id = tp.school_id 
                AND ha.period_number = tp.period_number
                AND (tp.level_id IS NULL OR tp.level_id = ?)
                AND (tp.section_id IS NULL OR tp.section_id = ?)
            WHERE ha.school_id = ? 
                AND ha.level_id = ?
                AND ha.section_id = ?
                AND ha.staff_id IS NOT NULL
            ORDER BY ha.day_of_week, ha.period_number
        ''', (level_id, section_id, school_id, level_id, section_id)).fetchall()
    
    # Convert to list of dicts for easier template access
    periods_list = [dict(row) for row in periods]
    assignments_list = [dict(row) for row in timetable_assignments]
    
    # Build timetable grid structure (day_of_week -> period_number -> assignment)
    timetable_grid = {}
    for assignment in assignments_list:
        day = assignment['day_of_week']
        period = assignment['period_number']
        if day not in timetable_grid:
            timetable_grid[day] = {}
        timetable_grid[day][period] = assignment
    
    return render_template('student/student_timetable.html',
                         student=student,
                         school_name=school['name'] if school else 'N/A',
                         periods=periods_list,
                         timetable_grid=timetable_grid)


@app.route('/student/update-theme', methods=['POST'])
@csrf.exempt
def student_update_theme():
    """Update student theme preference"""
    if 'student_id' not in session or session.get('user_type') != 'student':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    theme = data.get('theme', 'light')
    
    if theme not in ['light', 'dark', 'blue', 'green']:
        return jsonify({'success': False, 'error': 'Invalid theme'}), 400
    
    db = get_db()
    db.execute('''
        UPDATE students SET theme_preference = ? WHERE id = ?
    ''', (theme, session['student_id']))
    db.commit()
    
    return jsonify({'success': True})


@app.route('/student/logout')
def student_logout():
    """Student logout"""
    session.clear()
    return redirect(url_for('student_login'))


# Staff route for marking student attendance
@app.route('/staff/mark-student-attendance', methods=['GET', 'POST'])
def staff_mark_student_attendance():
    """Staff interface for marking student attendance"""
    if 'user_id' not in session or session.get('user_type') not in ['admin', 'staff']:
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']
    staff_id = session['user_id']
    today = datetime.date.today().strftime('%Y-%m-%d')
    is_admin_user = session.get('user_type') == 'admin'

    def _to_day_of_week(date_str):
        """Map date to timetable day index (Mon=1 ... Sat=6, Sun=0)."""
        date_obj = datetime.date.fromisoformat(date_str)
        day_number = date_obj.weekday() + 1
        return 0 if day_number == 7 else day_number

    def _get_attendance_rule_for_school():
        """Return attendance scope rule: all time slots or selected time slots."""
        default_rule = {'mode': 'all', 'time_slot_ids': []}
        if is_admin_user:
            return default_rule

        setting_key = f"student_attendance_rule_school_{school_id}"
        raw_setting = get_system_setting(setting_key, None)
        if not raw_setting:
            return default_rule

        try:
            parsed = json.loads(raw_setting)
            mode = str(parsed.get('mode', 'all')).lower()
            slot_ids = parsed.get('time_slot_ids', None)
            if slot_ids is None:
                slot_ids = parsed.get('period_numbers', [])

            if mode not in ['all', 'selected'] or not isinstance(slot_ids, list):
                return default_rule

            cleaned_slot_ids = []
            for item in slot_ids:
                try:
                    value = int(item)
                except (TypeError, ValueError):
                    continue
                if value not in cleaned_slot_ids:
                    cleaned_slot_ids.append(value)

            return {'mode': mode, 'time_slot_ids': cleaned_slot_ids}
        except (TypeError, ValueError, json.JSONDecodeError):
            return default_rule

    def _get_allowed_period_numbers(attendance_rule):
        """Resolve selected time slots into timetable period numbers."""
        if attendance_rule['mode'] != 'selected':
            return None

        selected_slot_ids = attendance_rule.get('time_slot_ids', [])
        if not selected_slot_ids:
            return []

        slot_placeholders = ','.join('?' for _ in selected_slot_ids)
        sequence_rows = db.execute(f'''
            SELECT id, period_sequence
            FROM timetable_period_timings
            WHERE school_id = ?
              AND id IN ({slot_placeholders})
              AND is_active = 1
        ''', (school_id, *selected_slot_ids)).fetchall()

        selected_sequences = []
        for row in sequence_rows:
            try:
                sequence_value = int(row['period_sequence'])
            except (TypeError, ValueError):
                continue
            if sequence_value not in selected_sequences:
                selected_sequences.append(sequence_value)

        fallback_sql = ''
        fallback_params = []
        if selected_sequences:
            sequence_placeholders = ','.join('?' for _ in selected_sequences)
            fallback_sql = f' OR (time_slot_id IS NULL AND period_number IN ({sequence_placeholders})) '
            fallback_params = selected_sequences

        rows = db.execute(f'''
            SELECT DISTINCT period_number
            FROM timetable_periods
            WHERE school_id = ?
              AND is_active = 1
              AND (time_slot_id IN ({slot_placeholders}){fallback_sql})
            ORDER BY period_number
        ''', (school_id, *selected_slot_ids, *fallback_params)).fetchall()

        return [int(row['period_number']) for row in rows]

    def _get_accessible_students(date_str):
        """Return students this user can mark attendance for on the given date."""
        if is_admin_user:
            all_students = db.execute('''
                SELECT id, student_id, full_name, `class`, section, roll_number
                FROM students
                WHERE school_id = ?
                ORDER BY `class`, section, roll_number
            ''', (school_id,)).fetchall()
            return all_students, 'Admin scope: all classes'

        attendance_rule = _get_attendance_rule_for_school()
        allowed_period_numbers = _get_allowed_period_numbers(attendance_rule)

        if attendance_rule['mode'] == 'selected' and not allowed_period_numbers:
            return [], 'Selected time slot rule has no mapped periods for this school'

        period_filter_sql = ''
        period_params = []
        if allowed_period_numbers is not None:
            placeholders = ','.join('?' for _ in allowed_period_numbers)
            period_filter_sql = f' AND ha.period_number IN ({placeholders}) '
            period_params = allowed_period_numbers

        students = db.execute(f'''
            SELECT DISTINCT s.id, s.student_id, s.full_name, s.`class`, s.section, s.roll_number
            FROM students s
            JOIN timetable_academic_levels tal
              ON tal.school_id = s.school_id
             AND tal.level_name = s.`class`
             AND tal.is_active = 1
            JOIN timetable_sections ts
              ON ts.school_id = s.school_id
             AND ts.level_id = tal.id
             AND ts.section_name = s.section
             AND ts.is_active = 1
            JOIN timetable_hierarchical_assignments ha
              ON ha.school_id = s.school_id
             AND ha.level_id = tal.id
             AND ha.section_id = ts.id
            WHERE s.school_id = ?
              AND ha.staff_id = ?
              {period_filter_sql}
            ORDER BY s.`class`, s.section, s.roll_number
        ''', (school_id, staff_id, *period_params)).fetchall()

        assigned_classes = db.execute(f'''
            SELECT DISTINCT tal.level_name AS class_name, ts.section_name AS section_name
            FROM timetable_hierarchical_assignments ha
            JOIN timetable_academic_levels tal ON ha.level_id = tal.id
            JOIN timetable_sections ts ON ha.section_id = ts.id
            WHERE ha.school_id = ?
              AND ha.staff_id = ?
              {period_filter_sql}
            ORDER BY tal.level_number, ts.section_name
        ''', (school_id, staff_id, *period_params)).fetchall()

        rule_label = 'all time slots' if attendance_rule['mode'] == 'all' else 'selected time slots'
        if assigned_classes:
            class_text = ', '.join([f"{row['class_name']}-{row['section_name']}" for row in assigned_classes])
            scope_text = f"Scope: {rule_label}; assigned classes: {class_text}"
        else:
            scope_text = f"Scope: {rule_label}; no classes assigned to you"

        return students, scope_text

    selected_date = request.args.get('date', today)
    try:
        selected_date = datetime.date.fromisoformat(selected_date).strftime('%Y-%m-%d')
    except (TypeError, ValueError):
        selected_date = today

    selected_session = request.args.get('session_type', 'morning')
    if selected_session not in ['morning', 'afternoon']:
        selected_session = 'morning'

    if request.method == 'POST':
        date = request.form.get('date', today)
        try:
            date = datetime.date.fromisoformat(date).strftime('%Y-%m-%d')
        except (TypeError, ValueError):
            date = today

        session_type = request.form.get('session_type', 'morning')
        if session_type not in ['morning', 'afternoon']:
            flash('Please select a valid session', 'error')
            return redirect(url_for('staff_mark_student_attendance', date=date, session_type='morning'))

        notes = request.form.get('notes', '').strip()
        attendance_payload = request.form.get('attendance_payload', '')

        try:
            attendance_map = json.loads(attendance_payload) if attendance_payload else {}
        except json.JSONDecodeError:
            attendance_map = None

        if not isinstance(attendance_map, dict):
            flash('Attendance list could not be processed. Please try again.', 'error')
            return redirect(url_for('staff_mark_student_attendance', date=date, session_type=session_type))

        current_time = datetime.datetime.now().strftime('%H:%M:%S')

        students, _ = _get_accessible_students(date)
        student_ids = [student['id'] for student in students]

        if not student_ids:
            flash('No eligible students are assigned for your attendance scope', 'error')
            return redirect(url_for('staff_mark_student_attendance', date=date, session_type=session_type))

        placeholders = ','.join('?' for _ in student_ids)
        existing_records = db.execute(f'''
            SELECT id, student_id
            FROM student_attendance
            WHERE school_id = ?
              AND attendance_date = ?
              AND student_id IN ({placeholders})
        ''', (school_id, date, *student_ids)).fetchall()
        existing_map = {record['student_id']: record['id'] for record in existing_records}

        marked_count = 0
        for student_id in student_ids:
            raw_status = str(attendance_map.get(str(student_id), 'present')).strip().lower()
            status = 'absent' if raw_status == 'absent' else 'present'

            if student_id in existing_map:
                if session_type == 'morning':
                    db.execute('''
                        UPDATE student_attendance
                        SET morning_status = ?, morning_time = ?,
                            morning_marked_by = ?, morning_notes = ?,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE student_id = ? AND attendance_date = ?
                    ''', (status, current_time, staff_id, notes, student_id, date))
                else:
                    db.execute('''
                        UPDATE student_attendance
                        SET afternoon_status = ?, afternoon_time = ?,
                            afternoon_marked_by = ?, afternoon_notes = ?,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE student_id = ? AND attendance_date = ?
                    ''', (status, current_time, staff_id, notes, student_id, date))
            else:
                if session_type == 'morning':
                    db.execute('''
                        INSERT INTO student_attendance (
                            student_id, school_id, attendance_date,
                            morning_status, morning_time, morning_marked_by, morning_notes
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, school_id, date, status, current_time, staff_id, notes))
                else:
                    db.execute('''
                        INSERT INTO student_attendance (
                            student_id, school_id, attendance_date,
                            afternoon_status, afternoon_time, afternoon_marked_by, afternoon_notes
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, school_id, date, status, current_time, staff_id, notes))

            marked_count += 1

        db.commit()
        flash(f'Attendance saved for {marked_count} students', 'success')
        return redirect(url_for('staff_mark_student_attendance', date=date, session_type=session_type))

    # GET request - show form
    students, attendance_scope_text = _get_accessible_students(selected_date)
    student_ids = [student['id'] for student in students]

    if student_ids:
        placeholders = ','.join('?' for _ in student_ids)
        attendance_records = db.execute(f'''
            SELECT sa.*, s.student_id, s.full_name, s.`class`, s.section, s.roll_number
            FROM student_attendance sa
            JOIN students s ON sa.student_id = s.id
            WHERE sa.school_id = ?
              AND sa.attendance_date = ?
              AND s.id IN ({placeholders})
            ORDER BY s.`class`, s.section, s.roll_number
        ''', (school_id, selected_date, *student_ids)).fetchall()
    else:
        attendance_records = []

    attendance_map = {record['student_id']: record for record in attendance_records}
    prepared_students = []
    for student in students:
        student_data = dict(student)
        attendance_record = attendance_map.get(student['id'])
        student_data['morning_status'] = attendance_record['morning_status'] if attendance_record else None
        student_data['afternoon_status'] = attendance_record['afternoon_status'] if attendance_record else None
        current_status = student_data[f'{selected_session}_status']
        student_data['initial_status'] = 'absent' if current_status == 'absent' else 'present'
        prepared_students.append(student_data)

    return render_template('student/mark_student_attendance.html',
                         students=prepared_students,
                         today_attendance=attendance_records,
                         current_date=selected_date,
                         selected_session=selected_session,
                         max_date=today,
                         attendance_scope_text=attendance_scope_text)


# ===============================================================================



if __name__ == '__main__':
    init_db(app)
    app.run(debug=True, host='0.0.0.0', port=5500)
